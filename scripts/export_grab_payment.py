#!/usr/bin/env python3
"""导出华莱士门店4月份Grab支付方式Excel报表"""
import sys
sys.path.insert(0, '/home/weifashi/hwt/analysis')

from bq_reports.utils.bq_client import get_bq_client
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

client = get_bq_client()
BQ_LOCATION = "asia-southeast1"

START_TS = 1774976400
END_TS = 1777568400

# ========== 1. 从 saas 获取所有CFG总部下的TH门店 ==========
job = client.query('''
    SELECT c.uuid, c.name, cs.erpnext_company_abbr
    FROM `diyl-407103`.`saas`.`ttpos_company` c
    LEFT JOIN `diyl-407103`.`saas`.`ttpos_company_setting` cs
      ON cs.company_uuid = c.uuid AND cs.delete_time = 0
    WHERE c.delete_time = 0
      AND cs.headquarter_uuid = 5080409448448000
      AND cs.erpnext_company_abbr LIKE 'TH%'
    ORDER BY cs.erpnext_company_abbr
''', location=BQ_LOCATION)

all_th_stores = []
for row in job.result():
    all_th_stores.append({
        'uuid': str(row.uuid),
        'name': row.name,
        'abbr': row.erpnext_company_abbr or '',
        'dataset': f"shop{row.uuid}",
    })

datasets = list(client.list_datasets())
shop_dataset_ids = set(d.dataset_id for d in datasets if d.dataset_id.startswith('shop'))
stores = [s for s in all_th_stores if s['dataset'] in shop_dataset_ids]

# ========== 2. 查询Grab数据 ==========
results = []
for s in stores:
    dataset = s['dataset']

    pos_rows = []
    try:
        job = client.query(f"""
            SELECT DATE(TIMESTAMP_SECONDS(sp.complete_time), 'Asia/Bangkok') AS pay_date,
                pm.payment_name AS method_name, COUNT(*) AS bill_cnt,
                ROUND(SUM(sp.payment_amount), 2) AS total_amount
            FROM `diyl-407103`.`{dataset}`.`ttpos_statistics_payment` sp
            LEFT JOIN `diyl-407103`.`{dataset}`.`ttpos_payment_method` pm
                ON pm.uuid = sp.payment_method_uuid
            WHERE sp.delete_time = 0 AND sp.complete_time >= {START_TS}
                AND sp.complete_time < {END_TS}
                AND LOWER(IFNULL(pm.payment_name, '')) LIKE '%grab%'
            GROUP BY pay_date, method_name ORDER BY pay_date
        """, location=BQ_LOCATION)
        for row in job.result():
            pos_rows.append({'date': row.pay_date.isoformat() if row.pay_date else None,
                             'method': row.method_name, 'count': row.bill_cnt,
                             'amount': float(row.total_amount) if row.total_amount else 0})
    except Exception:
        pass

    to_rows = []
    try:
        job = client.query(f"""
            SELECT DATE(TIMESTAMP_SECONDS(
                CASE WHEN order_state = 40 AND completed_time > 0 THEN completed_time
                ELSE accepted_time END), 'Asia/Bangkok') AS pay_date,
                COUNT(*) AS bill_cnt, ROUND(SUM(platform_total), 2) AS total_amount
            FROM `diyl-407103`.`{dataset}`.`ttpos_takeout_order`
            WHERE delete_time = 0 AND order_state IN (10, 20, 30, 40)
                AND platform = 'grab' AND accepted_time > 0
                AND (CASE WHEN order_state = 40 AND completed_time > 0 THEN completed_time
                    ELSE accepted_time END) >= {START_TS}
                AND (CASE WHEN order_state = 40 AND completed_time > 0 THEN completed_time
                    ELSE accepted_time END) < {END_TS}
            GROUP BY pay_date ORDER BY pay_date
        """, location=BQ_LOCATION)
        for row in job.result():
            to_rows.append({'date': row.pay_date.isoformat() if row.pay_date else None,
                            'count': row.bill_cnt,
                            'amount': float(row.total_amount) if row.total_amount else 0})
    except Exception:
        pass

    results.append({'store': s, 'pos': pos_rows, 'takeout': to_rows})

# ========== 3. 生成Excel ==========
wb = Workbook()

# 样式定义
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

# --- Sheet1: 门店汇总 ---
ws1 = wb.active
ws1.title = '门店汇总'
ws1.append(['门店编号', '门店名称', 'POS Grab笔数', 'POS Grab金额', '外卖Grab笔数', '外卖Grab金额', '总笔数', '总金额', '是否有数据'])

for r in results:
    s = r['store']
    pos_cnt = sum(d['count'] for d in r['pos'])
    pos_amt = sum(d['amount'] for d in r['pos'])
    to_cnt = sum(d['count'] for d in r['takeout'])
    to_amt = sum(d['amount'] for d in r['takeout'])
    total_cnt = pos_cnt + to_cnt
    total_amt = pos_amt + to_amt
    has_data = '是' if (pos_cnt > 0 or to_cnt > 0) else '否'
    ws1.append([s['abbr'], s['name'], pos_cnt, pos_amt, to_cnt, to_amt, total_cnt, total_amt, has_data])

# 合计行
total_pos_cnt = sum(d['count'] for r in results for d in r['pos'])
total_pos_amt = sum(d['amount'] for r in results for d in r['pos'])
total_to_cnt = sum(d['count'] for r in results for d in r['takeout'])
total_to_amt = sum(d['amount'] for r in results for d in r['takeout'])
ws1.append(['合计', '', total_pos_cnt, total_pos_amt, total_to_cnt, total_to_amt,
            total_pos_cnt + total_to_cnt, total_pos_amt + total_to_amt, ''])

# 样式
for col in range(1, 10):
    cell = ws1.cell(row=1, column=col)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row in range(2, ws1.max_row + 1):
    for col in range(1, 10):
        cell = ws1.cell(row=row, column=col)
        cell.border = thin_border
        if col >= 3:
            cell.alignment = center_align

for col_idx in range(1, 10):
    ws1.column_dimensions[get_column_letter(col_idx)].width = 16
ws1.column_dimensions['B'].width = 40

# 合计行加粗
for col in range(1, 10):
    ws1.cell(row=ws1.max_row, column=col).font = Font(bold=True)

# --- Sheet2: 按日期汇总 ---
ws2 = wb.create_sheet('按日期汇总')
ws2.append(['日期', 'POS Grab门店数', 'POS Grab笔数', 'POS Grab金额',
            '外卖Grab门店数', '外卖Grab笔数', '外卖Grab金额', '总门店数', '总笔数', '总金额'])

all_dates = sorted(set(
    d['date'] for r in results for d in r['pos']
) | set(
    d['date'] for r in results for d in r['takeout']
))

for d in all_dates:
    pos_stores = len([r for r in results if any(x['date'] == d for x in r['pos'])])
    pos_cnt = sum(x['count'] for r in results for x in r['pos'] if x['date'] == d)
    pos_amt = sum(x['amount'] for r in results for x in r['pos'] if x['date'] == d)
    to_stores = len([r for r in results if any(x['date'] == d for x in r['takeout'])])
    to_cnt = sum(x['count'] for r in results for x in r['takeout'] if x['date'] == d)
    to_amt = sum(x['amount'] for r in results for x in r['takeout'] if x['date'] == d)
    total_stores = len([r for r in results if any(x['date'] == d for x in r['pos'] + r['takeout'])])
    ws2.append([d, pos_stores, pos_cnt, pos_amt, to_stores, to_cnt, to_amt,
                total_stores, pos_cnt + to_cnt, pos_amt + to_amt])

# 合计行
ws2.append(['合计', '', total_pos_cnt, total_pos_amt, '', total_to_cnt, total_to_amt,
            '', total_pos_cnt + total_to_cnt, total_pos_amt + total_to_amt])

for col in range(1, 11):
    cell = ws2.cell(row=1, column=col)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row in range(2, ws2.max_row + 1):
    for col in range(1, 11):
        cell = ws2.cell(row=row, column=col)
        cell.border = thin_border
        if col >= 2:
            cell.alignment = center_align

for col_idx in range(1, 11):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 16
ws2.column_dimensions['A'].width = 14

for col in range(1, 11):
    ws2.cell(row=ws2.max_row, column=col).font = Font(bold=True)

# --- Sheet3: POS Grab 明细 ---
ws3 = wb.create_sheet('POS Grab明细')
ws3.append(['门店编号', '门店名称', '日期', '支付方式', '笔数', '金额'])

for r in results:
    s = r['store']
    for d in r['pos']:
        ws3.append([s['abbr'], s['name'], d['date'], d['method'], d['count'], d['amount']])

for col in range(1, 7):
    cell = ws3.cell(row=1, column=col)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row in range(2, ws3.max_row + 1):
    for col in range(1, 7):
        cell = ws3.cell(row=row, column=col)
        cell.border = thin_border
        if col >= 5:
            cell.alignment = center_align

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 14

# --- Sheet4: 外卖 Grab 明细 ---
ws4 = wb.create_sheet('外卖Grab明细')
ws4.append(['门店编号', '门店名称', '日期', '笔数', '金额'])

for r in results:
    s = r['store']
    for d in r['takeout']:
        ws4.append([s['abbr'], s['name'], d['date'], d['count'], d['amount']])

for col in range(1, 6):
    cell = ws4.cell(row=1, column=col)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row in range(2, ws4.max_row + 1):
    for col in range(1, 6):
        cell = ws4.cell(row=row, column=col)
        cell.border = thin_border
        if col >= 4:
            cell.alignment = center_align

ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 40
ws4.column_dimensions['C'].width = 14
ws4.column_dimensions['D'].width = 10
ws4.column_dimensions['E'].width = 14

# 保存
output_path = '/home/weifashi/hwt/analysis/exports/grab_payment_2026-04.xlsx'
wb.save(output_path)
print(f'报表已导出: {output_path}')
print(f'  Sheet1: 门店汇总 ({len(results)}行)')
print(f'  Sheet2: 按日期汇总 ({len(all_dates)}行)')
print(f'  Sheet3: POS Grab明细 ({sum(len(r["pos"]) for r in results)}行)')
print(f'  Sheet4: 外卖Grab明细 ({sum(len(r["takeout"]) for r in results)}行)')
