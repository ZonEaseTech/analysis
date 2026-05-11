#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
列出 2026-04 各商家有消耗记录的所有原料 code（合集去重）

用途：作为 report_material_stats_bq.py 的 --materials 输入

Usage:
    venv/bin/python scripts/list_april_materials.py
    venv/bin/python scripts/list_april_materials.py --month 2026-04 --output exports/april_materials.txt
"""

import argparse
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from bq_reports.utils.bq_client import get_bq_client, setup_proxy, PROJECT_ID


SQL = """
SELECT DISTINCT m.code AS code, m.name AS name
FROM `{project}.{dataset}.ttpos_sale_order_material` som
JOIN `{project}.{dataset}.ttpos_material` m
  ON m.uuid = som.material_uuid
JOIN `{project}.{dataset}.ttpos_sale_bill` sb
  ON sb.uuid = som.sale_bill_uuid AND sb.delete_time = 0
WHERE som.delete_time = 0
  AND sb.create_time >= {start_ts}
  AND sb.create_time <  {end_ts}
  AND m.code IS NOT NULL
  AND m.code != ''
"""


def month_range(month: str):
    y, mo = int(month[:4]), int(month[5:7])
    start = int(datetime(y, mo, 1, tzinfo=timezone.utc).timestamp())
    if mo == 12:
        end = int(datetime(y + 1, 1, 1, tzinfo=timezone.utc).timestamp())
    else:
        end = int(datetime(y, mo + 1, 1, tzinfo=timezone.utc).timestamp())
    return start, end


def load_merchants(xlsx: str):
    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 3 and row[1] and row[2]:
            out.append((str(row[1]).strip(), str(row[2]).strip()))
    wb.close()
    return out


def query_one(client, project, uuid_str, start_ts, end_ts):
    sql = SQL.format(
        project=project, dataset=f"shop{uuid_str}",
        start_ts=start_ts, end_ts=end_ts,
    )
    rows = list(client.query(sql).result())
    return [(r["code"], r["name"]) for r in rows]


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--month", default="2026-04")
    p.add_argument("--merchants", default="resources/wallace.20260506/华莱士商家56家ID.xlsx")
    p.add_argument("--workers", type=int, default=8)
    p.add_argument("--output", default="exports/april_materials.txt")
    args = p.parse_args()

    setup_proxy()
    client = get_bq_client()

    start_ts, end_ts = month_range(args.month)
    merchants = load_merchants(args.merchants)
    print(f"商家数: {len(merchants)} | 月份: {args.month} | 窗口: [{start_ts}, {end_ts})")

    code_to_name: dict[str, str] = {}
    failed = []
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        fut_map = {
            ex.submit(query_one, client, PROJECT_ID, uuid_str, start_ts, end_ts): account
            for account, uuid_str in merchants
        }
        done_count = 0
        for fut in as_completed(fut_map):
            account = fut_map[fut]
            done_count += 1
            try:
                pairs = fut.result()
                for code, name in pairs:
                    if code not in code_to_name:
                        code_to_name[code] = name or ""
                print(f"  [{done_count}/{len(merchants)}] {account}: +{len(pairs)} codes (合集 {len(code_to_name)})")
            except Exception as e:
                failed.append((account, str(e)))
                print(f"  [{done_count}/{len(merchants)}] {account}: ERROR {e}")

    print(f"\n=== 4 月独立原料 code: {len(code_to_name)} ===")
    if failed:
        print(f"失败 {len(failed)} 家: {failed[:5]}")

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        f.write(f"# 2026-04 全 {len(merchants)} 家有销售消耗的独立原料 code: {len(code_to_name)} 个\n")
        f.write("# 格式: code\tname\n")
        for code in sorted(code_to_name.keys()):
            f.write(f"{code}\t{code_to_name[code]}\n")

    csv_arg = ",".join(sorted(code_to_name.keys()))
    csv_path = Path(args.output).with_suffix(".csv-arg.txt")
    csv_path.write_text(csv_arg, encoding="utf-8")

    print(f"已写入: {args.output}")
    print(f"CLI 参数串: {csv_path} (长度 {len(csv_arg)} 字符)")


if __name__ == "__main__":
    main()
