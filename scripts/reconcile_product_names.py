#!/usr/bin/env python3
"""
简易商品名对账脚本

比较销售报表中的商品名和 BQ menu 中的商品名，输出差异。

Usage:
    venv/bin/python scripts/reconcile_product_names.py \
        --input /path/to/销售报表.xlsx \
        --sheet-name "未设置BOM商品销量" \
        --name-column 2
"""

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
DEFAULT_DATASET = "shop1958987436032000"

BQ_SQL = """
SELECT
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pp.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS name_zh,
  JSON_EXTRACT_SCALAR(pp.name, '$.en') AS name_en,
  JSON_EXTRACT_SCALAR(pp.name, '$.th') AS name_th,
  CASE WHEN pp.product_type = 0 THEN '单品' WHEN pp.product_type = 1 THEN '套餐' ELSE '其他' END AS ptype
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
WHERE pp.delete_time = 0
UNION ALL
SELECT
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(ps.name, '$.zh'),
    JSON_EXTRACT_SCALAR(ps.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS name_zh,
  JSON_EXTRACT_SCALAR(ps.name, '$.en') AS name_en,
  JSON_EXTRACT_SCALAR(ps.name, '$.th') AS name_th,
  '加料' AS ptype
FROM `{project}`.`{dataset}`.`ttpos_product_sauce` ps
WHERE ps.delete_time = 0
"""


def load_excel_names(input_path: str, sheet_name: str, name_col: int):
    wb = load_workbook(input_path)
    ws = wb[sheet_name]

    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > name_col and row[name_col]:
            names.add(str(row[name_col]).strip())

    return names


def query_bq_names(client, dataset: str):
    sql = BQ_SQL.format(project=client.project, dataset=dataset)
    rows = list(client.query(sql).result())
    return {r.name_zh: r for r in rows if r.name_zh}


def main():
    p = argparse.ArgumentParser(description="商品名对账：销售报表 vs BQ menu")
    p.add_argument("--input", required=True, help="销售报表 Excel 路径")
    p.add_argument("--sheet-name", default="未设置BOM商品销量", help="Sheet 名")
    p.add_argument("--name-column", type=int, default=2, help="商品名列索引（0-based）")
    p.add_argument("--project", default=PROJECT_ID, help="GCP 项目 ID")
    p.add_argument("--dataset", default=DEFAULT_DATASET, help="BQ dataset")
    p.add_argument("--output", help="输出差异到文件（可选）")
    args = p.parse_args()

    setup_proxy()

    print(f"[读取] {args.input} [{args.sheet_name}]")
    excel_names = load_excel_names(args.input, args.sheet_name, args.name_column)
    print(f"[Excel] 共 {len(excel_names)} 种商品")

    client = get_bq_client(project_id=args.project)
    bq_map = query_bq_names(client, args.dataset)
    print(f"[BQ] 共 {len(bq_map)} 条商品")

    matched = excel_names & set(bq_map.keys())
    only_in_excel = excel_names - set(bq_map.keys())
    only_in_bq = set(bq_map.keys()) - excel_names

    print()
    print("=" * 60)
    print(f"匹配成功: {len(matched)} 种")
    print(f"仅在 Excel 中: {len(only_in_excel)} 种")
    print(f"仅在 BQ 中: {len(only_in_bq)} 种")
    print("=" * 60)

    if only_in_excel:
        print(f"\n仅在 Excel 中（BQ 找不到）:")
        for name in sorted(only_in_excel):
            print(f"  - {name}")

    if args.output and only_in_excel:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write("仅在 Excel 中（BQ 找不到）:\n")
            for name in sorted(only_in_excel):
                f.write(f"  - {name}\n")
        print(f"\n差异已写入: {args.output}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
