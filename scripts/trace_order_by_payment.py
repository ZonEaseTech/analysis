#!/usr/bin/env python3
"""
按支付方式反向追溯订单 —— 商家 → 支付方式 → 金额 → 商品 → 差异对比

排查流程：
  1. 找到对应商家（支持模糊匹配商家名称）
  2. 根据商家 + 支付方式关键词筛选订单
  3. 按金额精确匹配或范围过滤
  4. 输出订单汇总 + 商品明细
  5. 自动对比 Grab/平台金额 与 TTPOS 订单总价，标记差异

支持两种订单来源：
  - POS 内支付（ttpos_statistics_payment + ttpos_order）
  - 外卖平台（ttpos_takeout_order）

输出 Excel（三 Sheet）：
  Sheet1「订单汇总」：Grab/平台金额 | TTPOS订单号 | TTPOS订单总价 | 差异金额 | 差异状态
  Sheet2「商品明细」：每笔订单的商品清单
  Sheet3「差异订单」：只展示 Grab/平台金额 ≠ TTPOS 订单总价的订单（红色高亮）

差异状态说明：
  一致（绿色）    ：Grab/平台金额 = TTPOS 订单总价
  微小差异（黄色）：|差异| ≤ 1（四舍五入导致）
  差异（红色）    ：|差异| > 1（需人工核查）

Usage:
    # 按金额精确查找 Grab 支付订单
    venv/bin/python scripts/trace_order_by_payment.py --store "SE05" --payment grab --amount 125.50 --date 2026-04-01

    # 金额范围查找
    venv/bin/python scripts/trace_order_by_payment.py --store "SE05" --payment grab --amount-min 100 --amount-max 200 --date 2026-04-01

    # 只查 POS 侧
    venv/bin/python scripts/trace_order_by_payment.py --store "SE05" --payment grab --amount 125.50 --source pos

    # 只查外卖侧
    venv/bin/python scripts/trace_order_by_payment.py --store "SE05" --payment grab --amount 125.50 --source takeout
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, '/home/weifashi/hwt/analysis')
from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
BQ_LOCATION = "asia-southeast1"

THAILAND_TZ_OFFSET = 7 * 3600  # UTC+7


def _parse_date(date_str: str) -> tuple[int, int]:
    """解析日期字符串为当日开始/结束的 Unix 时间戳（泰国时间）"""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    start_ts = int(dt.replace(tzinfo=timezone.utc).timestamp()) - THAILAND_TZ_OFFSET
    end_ts = start_ts + 86400
    return start_ts, end_ts


def _sanitize_kw(kw: str) -> str:
    """过滤关键词，只允许字母数字和常见符号，防止 SQL 注入"""
    return ''.join(c for c in kw if c.isalnum() or c in ' ._-')


def find_stores(client, store_keyword: str) -> list[dict]:
    """模糊匹配商家名称，返回匹配的门店列表"""
    job = client.query('''
        SELECT c.uuid, c.name, cs.erpnext_company_abbr
        FROM `diyl-407103`.`saas`.`ttpos_company` c
        LEFT JOIN `diyl-407103`.`saas`.`ttpos_company_setting` cs
          ON cs.company_uuid = c.uuid AND cs.delete_time = 0
        WHERE c.delete_time = 0
          AND cs.headquarter_uuid = 5080409448448000
          AND cs.erpnext_company_abbr LIKE 'TH%%'
        ORDER BY cs.erpnext_company_abbr
    ''', location=BQ_LOCATION)

    all_stores = []
    for row in job.result():
        name = row.name or ""
        abbr = row.erpnext_company_abbr or ""
        kw = store_keyword.lower()
        if kw in name.lower() or kw in abbr.lower() or kw in str(row.uuid):
            all_stores.append({
                'uuid': str(row.uuid),
                'name': name,
                'abbr': abbr,
                'dataset': f"shop{row.uuid}",
            })

    # 过滤掉没有 BQ dataset 的
    datasets = list(client.list_datasets())
    shop_dataset_ids = set(d.dataset_id for d in datasets if d.dataset_id.startswith('shop'))
    return [s for s in all_stores if s['dataset'] in shop_dataset_ids]


def query_pos_orders(client, store: dict, payment_kw: str, start_ts: int, end_ts: int,
                     amount: float | None, amount_min: float | None, amount_max: float | None) -> list[dict]:
    """查询 POS 侧的支付订单"""
    dataset = store['dataset']

    amount_where = ""
    if amount is not None:
        amount_where = f"AND ROUND(sp.payment_amount, 2) = {amount}"
    elif amount_min is not None or amount_max is not None:
        if amount_min is not None:
            amount_where += f"AND sp.payment_amount >= {amount_min} "
        if amount_max is not None:
            amount_where += f"AND sp.payment_amount <= {amount_max} "

    safe_kw = _sanitize_kw(payment_kw).lower()
    sql = f"""
    SELECT
      sp.uuid AS payment_uuid,
      sp.order_uuid,
      sp.payment_amount,
      sp.complete_time,
      sp.serial_number,
      pm.payment_name AS method_name,
      o.order_state,
      o.total_price AS order_total,
      o.table_number,
      o.create_time AS order_create_time
    FROM `{PROJECT_ID}`.`{dataset}`.`ttpos_statistics_payment` sp
    LEFT JOIN `{PROJECT_ID}`.`{dataset}`.`ttpos_payment_method` pm
      ON pm.uuid = sp.payment_method_uuid
    LEFT JOIN `{PROJECT_ID}`.`{dataset}`.`ttpos_order` o
      ON o.uuid = sp.order_uuid AND o.delete_time = 0
    WHERE sp.delete_time = 0
      AND sp.complete_time >= {start_ts}
      AND sp.complete_time < {end_ts}
      AND LOWER(IFNULL(pm.payment_name, '')) LIKE '%{safe_kw}%'
      {amount_where}
    ORDER BY sp.complete_time DESC
    LIMIT 100
    """

    rows = []
    try:
        for r in client.query(sql, location=BQ_LOCATION).result():
            rows.append({
                'payment_uuid': r.payment_uuid,
                'order_uuid': r.order_uuid,
                'payment_amount': float(r.payment_amount) if r.payment_amount else 0,
                'complete_time': r.complete_time,
                'serial_number': r.serial_number,
                'method_name': r.method_name or '',
                'order_state': r.order_state,
                'order_total': float(r.order_total) if r.order_total else 0,
                'table_number': r.table_number or '',
                'order_create_time': r.order_create_time,
            })
    except Exception as e:
        if '404 Not found' in str(e) or 'was not found' in str(e):
            pass  # 表不存在，忽略
        else:
            print(f"  [POS查询失败] {store['name']}: {e}")
    return rows


def query_pos_order_items(client, store: dict, order_uuids: list[str]) -> dict[str, list[dict]]:
    """查询 POS 订单的商品明细"""
    if not order_uuids:
        return {}

    dataset = store['dataset']
    uuid_list = ", ".join(f'"{u}"' for u in order_uuids)

    sql = f"""
    SELECT
      oi.order_uuid,
      REGEXP_REPLACE(COALESCE(
        JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
        JSON_EXTRACT_SCALAR(pp.name, '$.en'),
        ''
      ), r'^\\s+|\\s+$', '') AS product_name,
      oi.num AS quantity,
      oi.price,
      oi.total_price,
      oi.is_combo,
      oi.package_uuid
    FROM `{PROJECT_ID}`.`{dataset}`.`ttpos_order_item` oi
    LEFT JOIN `{PROJECT_ID}`.`{dataset}`.`ttpos_product_package` pp
      ON pp.uuid = oi.package_uuid
    WHERE oi.order_uuid IN ({uuid_list})
      AND oi.delete_time = 0
    ORDER BY oi.order_uuid, oi.id
    """

    items_by_order = {}
    try:
        for r in client.query(sql, location=BQ_LOCATION).result():
            ouid = r.order_uuid
            if ouid not in items_by_order:
                items_by_order[ouid] = []
            items_by_order[ouid].append({
                'product_name': r.product_name or '',
                'quantity': float(r.quantity) if r.quantity else 0,
                'price': float(r.price) if r.price else 0,
                'total_price': float(r.total_price) if r.total_price else 0,
                'is_combo': bool(r.is_combo) if r.is_combo is not None else False,
            })
    except Exception as e:
        print(f"  [商品查询失败] {store['name']}: {e}")
    return items_by_order


def query_takeout_orders(client, store: dict, platform_kw: str, start_ts: int, end_ts: int,
                         amount: float | None, amount_min: float | None, amount_max: float | None) -> list[dict]:
    """查询外卖平台订单"""
    dataset = store['dataset']

    tto = "tto"  # 表别名，避免用 to（SQL 关键字）
    amount_where = ""
    if amount is not None:
        amount_where = f"AND ROUND({tto}.platform_total, 2) = {amount}"
    elif amount_min is not None or amount_max is not None:
        if amount_min is not None:
            amount_where += f"AND {tto}.platform_total >= {amount_min} "
        if amount_max is not None:
            amount_where += f"AND {tto}.platform_total <= {amount_max} "

    safe_kw = _sanitize_kw(platform_kw).lower()
    sql = f"""
    SELECT
      {tto}.uuid AS order_uuid,
      {tto}.platform_order_id AS order_number,
      {tto}.platform,
      {tto}.platform_total,
      {tto}.subtotal,
      {tto}.order_state,
      CASE WHEN {tto}.order_state = 40 AND {tto}.completed_time > 0 THEN {tto}.completed_time
           ELSE {tto}.accepted_time END AS pay_time,
      {tto}.accepted_time,
      {tto}.completed_time
    FROM `{PROJECT_ID}`.`{dataset}`.`ttpos_takeout_order` AS {tto}
    WHERE {tto}.delete_time = 0
      AND {tto}.order_state IN (10, 20, 30, 40)
      AND LOWER({tto}.platform) LIKE '%{safe_kw}%'
      AND (CASE WHEN {tto}.order_state = 40 AND {tto}.completed_time > 0 THEN {tto}.completed_time
                ELSE {tto}.accepted_time END) >= {start_ts}
      AND (CASE WHEN {tto}.order_state = 40 AND {tto}.completed_time > 0 THEN {tto}.completed_time
                ELSE {tto}.accepted_time END) < {end_ts}
      {amount_where}
    ORDER BY pay_time DESC
    LIMIT 100
    """

    rows = []
    try:
        for r in client.query(sql, location=BQ_LOCATION).result():
            rows.append({
                'order_uuid': r.order_uuid,
                'order_number': r.order_number or '',
                'platform': r.platform or '',
                'platform_total': float(r.platform_total) if r.platform_total else 0,
                'subtotal': float(r.subtotal) if r.subtotal else 0,
                'order_state': r.order_state,
                'pay_time': r.pay_time,
            })
    except Exception as e:
        if '404 Not found' in str(e) or 'was not found' in str(e):
            pass  # 表不存在，忽略
        else:
            print(f"  [外卖查询失败] {store['name']}: {e}")
    return rows


def query_takeout_order_items(client, store: dict, order_uuids: list[str]) -> dict[str, list[dict]]:
    """查询外卖订单的商品明细"""
    if not order_uuids:
        return {}

    dataset = store['dataset']
    uuid_list = ", ".join(f'"{u}"' for u in order_uuids)

    sql = f"""
    SELECT
      toi.order_uuid,
      REGEXP_REPLACE(COALESCE(
        JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
        JSON_EXTRACT_SCALAR(pp.name, '$.en'),
        ''
      ), r'^\\s+|\\s+$', '') AS product_name,
      toi.num AS quantity,
      toi.price,
      toi.total_price,
      toi.is_combo
    FROM `{PROJECT_ID}`.`{dataset}`.`ttpos_takeout_order_item` toi
    LEFT JOIN `{PROJECT_ID}`.`{dataset}`.`ttpos_product_package` pp
      ON pp.uuid = toi.package_uuid
    WHERE toi.order_uuid IN ({uuid_list})
      AND toi.delete_time = 0
    ORDER BY toi.order_uuid, toi.id
    """

    items_by_order = {}
    try:
        for r in client.query(sql, location=BQ_LOCATION).result():
            ouid = r.order_uuid
            if ouid not in items_by_order:
                items_by_order[ouid] = []
            items_by_order[ouid].append({
                'product_name': r.product_name or '',
                'quantity': float(r.quantity) if r.quantity else 0,
                'price': float(r.price) if r.price else 0,
                'total_price': float(r.total_price) if r.total_price else 0,
                'is_combo': bool(r.is_combo) if r.is_combo is not None else False,
            })
    except Exception as e:
        print(f"  [外卖商品查询失败] {store['name']}: {e}")
    return items_by_order


def _ts_to_bkk_str(ts: int) -> str:
    """Unix 时间戳转泰国时间字符串"""
    if not ts:
        return ""
    dt = datetime.fromtimestamp(ts + THAILAND_TZ_OFFSET, tz=timezone.utc)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _state_name(state: int, source: str) -> str:
    """订单状态码转可读名称"""
    if source == 'pos':
        names = {0: "待下单", 1: "待支付", 2: "已支付", 3: "已上菜", 4: "已完成", 5: "已撤单"}
    else:
        names = {10: "待处理", 20: "处理中", 30: "待配送", 40: "已完成"}
    return names.get(state, f"未知({state})")


def _calc_diff(o: dict) -> tuple[float, float, float, str, str]:
    """计算订单金额差异。

    Returns:
        (grab_price, ttpos_price, diff, diff_status, order_no)
    """
    if o['source'] == 'pos':
        grab_price = round(o.get('payment_amount', 0), 2)
        ttpos_price = round(o.get('order_total', 0), 2)
        order_no = o.get('serial_number', '')
    else:
        grab_price = round(o.get('platform_total', 0), 2)
        ttpos_price = round(o.get('platform_actual_price', 0), 2)
        order_no = o.get('order_number', '')

    diff = round(grab_price - ttpos_price, 2)
    if abs(diff) < 0.01:
        diff_status = "一致"
    elif abs(diff) <= 1:
        diff_status = "微小差异"
    else:
        diff_status = "差异"

    return grab_price, ttpos_price, diff, diff_status, order_no


def export_excel(results: list[dict], output_path: str):
    """导出 Excel：Sheet1 订单汇总，Sheet2 商品明细，Sheet3 差异订单"""
    wb = Workbook()

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 差异高亮样式
    diff_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    diff_red_font = Font(color='9C0006')
    diff_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    diff_green_font = Font(color='006100')
    diff_yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    diff_yellow_font = Font(color='9C5700')

    # ===== Sheet1: 订单汇总 =====
    ws1 = wb.active
    ws1.title = "订单汇总"
    headers1 = ["门店编号", "门店名称", "来源", "支付时间", "支付方式/平台",
                "Grab/平台金额", "TTPOS订单号", "订单状态", "TTPOS订单总价",
                "差异金额", "差异状态", "桌号", "备注"]
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row_idx = 2
    diff_rows = []  # 收集有差异的行号
    for r in results:
        store = r['store']
        for o in r['orders']:
            grab_price, ttpos_price, diff, diff_status, order_no = _calc_diff(o)

            ws1.cell(row=row_idx, column=1, value=store['abbr']).alignment = center_align
            ws1.cell(row=row_idx, column=2, value=store['name']).alignment = center_align
            ws1.cell(row=row_idx, column=3, value="POS" if o['source'] == 'pos' else "外卖").alignment = center_align
            ws1.cell(row=row_idx, column=4, value=_ts_to_bkk_str(o.get('complete_time') or o.get('pay_time'))).alignment = center_align
            ws1.cell(row=row_idx, column=5, value=o.get('method_name') or o.get('platform', '')).alignment = center_align
            ws1.cell(row=row_idx, column=6, value=grab_price).alignment = center_align
            ws1.cell(row=row_idx, column=7, value=order_no).alignment = center_align
            ws1.cell(row=row_idx, column=8, value=_state_name(o.get('order_state', 0), o['source'])).alignment = center_align
            ws1.cell(row=row_idx, column=9, value=ttpos_price).alignment = center_align
            ws1.cell(row=row_idx, column=10, value=diff).alignment = center_align
            ws1.cell(row=row_idx, column=11, value=diff_status).alignment = center_align
            ws1.cell(row=row_idx, column=12, value=o.get('table_number', '')).alignment = center_align
            ws1.cell(row=row_idx, column=13, value=o.get('remark', ''))

            # 差异状态高亮
            if diff_status == "一致":
                ws1.cell(row=row_idx, column=11).fill = diff_green_fill
                ws1.cell(row=row_idx, column=11).font = diff_green_font
            elif diff_status == "微小差异":
                ws1.cell(row=row_idx, column=11).fill = diff_yellow_fill
                ws1.cell(row=row_idx, column=11).font = diff_yellow_font
                diff_rows.append((row_idx, o, store))
            else:
                ws1.cell(row=row_idx, column=11).fill = diff_red_fill
                ws1.cell(row=row_idx, column=11).font = diff_red_font
                diff_rows.append((row_idx, o, store))

            for col in range(1, 14):
                ws1.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    for col_idx, w in enumerate([12, 35, 8, 20, 18, 14, 20, 12, 14, 12, 12, 10, 25], 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = w
    ws1.freeze_panes = "A2"

    # ===== Sheet2: 商品明细 =====
    ws2 = wb.create_sheet("商品明细")
    headers2 = ["门店编号", "门店名称", "订单号", "商品名称", "数量", "单价", "小计", "是否套餐"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row_idx = 2
    for r in results:
        store = r['store']
        for o in r['orders']:
            items = o.get('items', [])
            _, _, _, _, order_no = _calc_diff(o)
            for item in items:
                ws2.cell(row=row_idx, column=1, value=store['abbr']).alignment = center_align
                ws2.cell(row=row_idx, column=2, value=store['name']).alignment = center_align
                ws2.cell(row=row_idx, column=3, value=order_no).alignment = center_align
                ws2.cell(row=row_idx, column=4, value=item['product_name'])
                ws2.cell(row=row_idx, column=5, value=item['quantity']).alignment = center_align
                ws2.cell(row=row_idx, column=6, value=item['price']).alignment = center_align
                ws2.cell(row=row_idx, column=7, value=item['total_price']).alignment = center_align
                ws2.cell(row=row_idx, column=8, value="是" if item['is_combo'] else "否").alignment = center_align
                for col in range(1, 9):
                    ws2.cell(row=row_idx, column=col).border = thin_border
                row_idx += 1
            if not items:
                ws2.cell(row=row_idx, column=1, value=store['abbr']).alignment = center_align
                ws2.cell(row=row_idx, column=2, value=store['name']).alignment = center_align
                ws2.cell(row=row_idx, column=3, value=order_no).alignment = center_align
                ws2.cell(row=row_idx, column=4, value="(无商品明细)").alignment = center_align
                for col in range(1, 9):
                    ws2.cell(row=row_idx, column=col).border = thin_border
                row_idx += 1

    for col_idx, w in enumerate([12, 35, 20, 35, 8, 10, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.freeze_panes = "A2"

    # ===== Sheet3: 差异订单 =====
    ws3 = wb.create_sheet("差异订单")
    headers3 = ["门店编号", "门店名称", "来源", "支付时间", "支付方式/平台",
                "Grab/平台金额", "TTPOS订单号", "订单状态", "TTPOS订单总价",
                "差异金额", "差异状态", "商品摘要", "备注"]
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row_idx = 2
    for dr, o, store in diff_rows:
        grab_price, ttpos_price, diff, diff_status, order_no = _calc_diff(o)
        item_names = ", ".join(i['product_name'] for i in o.get('items', [])[:3])
        if len(o.get('items', [])) > 3:
            item_names += f" 等{len(o['items'])}项"

        ws3.cell(row=row_idx, column=1, value=store['abbr']).alignment = center_align
        ws3.cell(row=row_idx, column=2, value=store['name']).alignment = center_align
        ws3.cell(row=row_idx, column=3, value="POS" if o['source'] == 'pos' else "外卖").alignment = center_align
        ws3.cell(row=row_idx, column=4, value=_ts_to_bkk_str(o.get('complete_time') or o.get('pay_time'))).alignment = center_align
        ws3.cell(row=row_idx, column=5, value=o.get('method_name') or o.get('platform', '')).alignment = center_align
        ws3.cell(row=row_idx, column=6, value=grab_price).alignment = center_align
        ws3.cell(row=row_idx, column=7, value=order_no).alignment = center_align
        ws3.cell(row=row_idx, column=8, value=_state_name(o.get('order_state', 0), o['source'])).alignment = center_align
        ws3.cell(row=row_idx, column=9, value=ttpos_price).alignment = center_align
        ws3.cell(row=row_idx, column=10, value=diff).alignment = center_align
        ws3.cell(row=row_idx, column=11, value=diff_status).alignment = center_align
        ws3.cell(row=row_idx, column=12, value=item_names)
        ws3.cell(row=row_idx, column=13, value=o.get('remark', ''))

        # 高亮（微小差异黄色，差异红色）
        if diff_status == "微小差异":
            ws3.cell(row=row_idx, column=11).fill = diff_yellow_fill
            ws3.cell(row=row_idx, column=11).font = diff_yellow_font
        else:
            ws3.cell(row=row_idx, column=11).fill = diff_red_fill
            ws3.cell(row=row_idx, column=11).font = diff_red_font
        for col in range(1, 14):
            ws3.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    for col_idx, w in enumerate([12, 35, 8, 20, 18, 14, 20, 12, 14, 12, 12, 40, 20], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.freeze_panes = "A2"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n报表已导出: {output_path}")
    print(f"  Sheet1「订单汇总」: {ws1.max_row - 1} 行")
    print(f"  Sheet2「商品明细」: {ws2.max_row - 1} 行")
    print(f"  Sheet3「差异订单」: {ws3.max_row - 1} 行")


def main():
    parser = argparse.ArgumentParser(description="按支付方式反向追溯订单")
    parser.add_argument("--store", required=True, help="商家名称关键词（模糊匹配）")
    parser.add_argument("--payment", default="grab", help="支付方式关键词（默认 grab）")
    parser.add_argument("--amount", type=float, help="精确金额匹配")
    parser.add_argument("--amount-min", type=float, help="金额下限")
    parser.add_argument("--amount-max", type=float, help="金额上限")
    parser.add_argument("--date", required=True, help="日期（格式 YYYY-MM-DD，泰国时间）")
    parser.add_argument("--source", choices=["pos", "takeout", "all"], default="all",
                        help="订单来源：pos=POS内支付, takeout=外卖平台, all=全部（默认）")
    parser.add_argument("--output", default="exports/trace_order_by_payment.xlsx", help="输出路径")
    args = parser.parse_args()

    if args.amount is None and args.amount_min is None and args.amount_max is None:
        print("错误: 必须指定 --amount 或 --amount-min/--amount-max")
        return 1

    safe_payment = _sanitize_kw(args.payment)
    if not safe_payment:
        print("错误: 支付方式关键词无效")
        return 1

    setup_proxy()
    client = get_bq_client(project_id=PROJECT_ID)

    # 1. 找商家
    print(f"[1/4] 搜索商家: 关键词 '{args.store}'")
    stores = find_stores(client, args.store)
    if not stores:
        print(f"  未找到匹配商家")
        return 1
    print(f"  找到 {len(stores)} 家匹配门店:")
    for s in stores:
        print(f"    - {s['abbr']} | {s['name']}")

    # 2-3. 按支付方式 + 金额找订单
    start_ts, end_ts = _parse_date(args.date)
    print(f"\n[2/4] 查询订单: 日期 {args.date} ({_ts_to_bkk_str(start_ts)} ~ {_ts_to_bkk_str(end_ts)})")
    print(f"[3/4] 支付方式: '{safe_payment}'")
    if args.amount is not None:
        print(f"       金额: = {args.amount}")
    else:
        print(f"       金额: {args.amount_min or '不限'} ~ {args.amount_max or '不限'}")

    results = []
    total_orders = 0

    for store in stores:
        store_result = {'store': store, 'orders': []}

        if args.source in ('pos', 'all'):
            pos_orders = query_pos_orders(
                client, store, safe_payment, start_ts, end_ts,
                args.amount, args.amount_min, args.amount_max
            )
            if pos_orders:
                order_uuids = [o['order_uuid'] for o in pos_orders if o['order_uuid']]
                items = query_pos_order_items(client, store, order_uuids)
                for o in pos_orders:
                    o['source'] = 'pos'
                    o['items'] = items.get(o['order_uuid'], [])
                store_result['orders'].extend(pos_orders)
                print(f"  [{store['abbr']}] POS: {len(pos_orders)} 笔")

        if args.source in ('takeout', 'all'):
            takeout_orders = query_takeout_orders(
                client, store, safe_payment, start_ts, end_ts,
                args.amount, args.amount_min, args.amount_max
            )
            if takeout_orders:
                order_uuids = [o['order_uuid'] for o in takeout_orders]
                items = query_takeout_order_items(client, store, order_uuids)
                for o in takeout_orders:
                    o['source'] = 'takeout'
                    o['items'] = items.get(o['order_uuid'], [])
                store_result['orders'].extend(takeout_orders)
                print(f"  [{store['abbr']}] 外卖: {len(takeout_orders)} 笔")

        if store_result['orders']:
            results.append(store_result)
            total_orders += len(store_result['orders'])

    print(f"\n  合计: {total_orders} 笔订单")

    if total_orders == 0:
        print("\n未找到匹配订单")
        return 0

    # 4. 输出商品明细
    print(f"\n[4/4] 导出 Excel...")
    export_excel(results, args.output)

    # 打印摘要
    print(f"\n{'='*60}")
    print("排查摘要")
    print(f"{'='*60}")
    diff_count = 0
    for r in results:
        store = r['store']
        pos_cnt = sum(1 for o in r['orders'] if o['source'] == 'pos')
        to_cnt = sum(1 for o in r['orders'] if o['source'] == 'takeout')
        store_diff = sum(1 for o in r['orders'] if _calc_diff(o)[3] != "一致")
        diff_count += store_diff
        print(f"\n{store['abbr']} | {store['name']} (差异 {store_diff} 笔)")
        for o in r['orders']:
            grab_price, ttpos_price, diff, diff_status, order_no = _calc_diff(o)
            time_str = _ts_to_bkk_str(o.get('complete_time') or o.get('pay_time'))
            item_names = ", ".join(i['product_name'] for i in o.get('items', [])[:3])
            if len(o.get('items', [])) > 3:
                item_names += f" 等{len(o['items'])}项"
            src = "POS" if o['source'] == 'pos' else "外卖"
            flag = " ⚠️" if diff_status != "一致" else ""
            print(f"  [{src}] {time_str} | Grab¥{grab_price:.2f} / TTPOS¥{ttpos_price:.2f} | 差异¥{diff:.2f}{flag} | {order_no}")
            if item_names:
                print(f"         商品: {item_names}")

    print(f"\n{'='*60}")
    print(f"总计 {total_orders} 笔订单，其中差异 {diff_count} 笔")
    print(f"差异订单已在 Sheet3「差异订单」中单独列出")
    print(f"{'='*60}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
