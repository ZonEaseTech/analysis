#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
华莱士门店《营业数据汇总》导出（店 × 日，纯 Python，单连接跨库）。

谁问的: 何伟涛 / 2026-06-26 —— 把同事手工导的《营业数据汇总》做成可重跑自助脚本。
口径: 全部对齐 ttpos 后端 statistics.go + 同事 payment.sh 验证过的支付/外卖通道。

输出列:
  营业日 | 店铺名称 | 总营业额 | 实收金额 | 营业收入 | TC | AC
        | 现金TC | 现金金额 | 现金AC | 到店业绩 | 到店订单数 | 外卖业绩 | 外卖数

口径来源（与 BQ 实测 2026-06-25 全 62 店 62/62 一致）:
  总营业额 = Σ(product_price+product_tax+service_fee+service_tax+payment_fee+extend_price)   [ttpos_statistics_sale]
  实收金额 = Σ(payment_amount-refund_amount-payment_balance)                                  [ttpos_statistics_sale]
  营业收入 = Σ(payment_amount-refund_amount-refund_payment_balance-product_tax-service_tax+refund_tax) [不含税]
  TC       = Σ(is_meger=0)
  AC       = 总营业额 / TC
  现金*    = ttpos_payment_order 里 payment_method.name='Cash'（扣退款 return_order_amount）
  到店业绩 = 非外卖支付方式的实收（扣退款）
  外卖业绩 = Grab/Lineman/Shopee/Robinhood 支付方式 + ttpos_takeout_order 平台单(platform_total)

用法:
  python3 business_summary.py [YYYY-MM-DD] [--env .env] [--output 路径] [--no-feishu]
  不传日期 → 取泰国时区"昨天"。.env 与 payment.sh 同一份(DB_HOST/DB_PORT/DB_USER/DB_PASS)。
"""
import argparse
import hashlib
import os
import subprocess
import sys
from datetime import datetime, timedelta

try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except ImportError:
    from backports.zoneinfo import ZoneInfo  # Python 3.6-3.8

BKK = ZoneInfo("Asia/Bangkok")

# 外卖支付方式归类（按名称，与 payment.sh 一致）
PAYMENT_CASE = """
  CASE
    WHEN pm.name = 'Grab' OR pm.name LIKE '%Grab%' THEN 'DELIVERY'
    WHEN pm.name LIKE '%LINE%MAN%' OR pm.name LIKE '%LINEMAN%' OR pm.name LIKE '%Lineman%' THEN 'DELIVERY'
    WHEN pm.name LIKE '%Shopee%' OR pm.name LIKE '%shopee%' THEN 'DELIVERY'
    WHEN pm.name LIKE '%Robinhood%' OR pm.name LIKE '%robinhood%' THEN 'DELIVERY'
    ELSE 'INSTORE'
  END
"""


# ───────────────────────── 依赖 ─────────────────────────
def _ensure(pkg, pip_name=None):
    try:
        return __import__(pkg)
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "--quiet",
                        pip_name or pkg], check=False)
        try:
            return __import__(pkg)
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "--quiet",
                            "--break-system-packages", pip_name or pkg], check=False)
            return __import__(pkg)


pymysql = _ensure("pymysql")
openpyxl = _ensure("openpyxl")
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


def load_env(path):
    env = {}
    if not os.path.isfile(path):
        sys.exit(f"错误: 未找到 .env 文件 ({path})；需含 DB_HOST/DB_PORT/DB_USER/DB_PASS")
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def day_bounds(biz_date, tz):
    y, m, d = (int(x) for x in biz_date.split("-"))
    try:
        zone = ZoneInfo(tz)
    except Exception:
        zone = BKK
    start = int(datetime(y, m, d, 0, 0, 0, tzinfo=zone).timestamp())
    end = int(datetime(y, m, d, 23, 59, 59, tzinfo=zone).timestamp())
    return start, end


# ───────────────────────── SQL ─────────────────────────
def q_totals(uuid, start, end):
    return f"""
      SELECT
        COALESCE(ROUND(SUM(product_price+product_tax+service_fee+service_tax+payment_fee+extend_price),2),0),
        COALESCE(ROUND(SUM(payment_amount-refund_amount-payment_balance),2),0),
        COALESCE(ROUND(SUM(payment_amount-refund_amount-refund_payment_balance-product_tax-service_tax+refund_tax),2),0),
        COALESCE(SUM(IF(is_meger=0,1,0)),0)
      FROM `shop{uuid}`.ttpos_statistics_sale
      WHERE complete_time >= {start} AND complete_time <= {end}"""


def q_payments(uuid, start, end):
    return f"""
      -- 计数口径 = 支付笔数 COUNT(*)（与原表一致：一单拆两笔现金算两笔），非去重订单数
      SELECT
        COALESCE(ROUND(SUM(IF(t.is_cash=1, t.net, 0)),2),0),
        SUM(IF(t.is_cash=1, 1, 0)),
        COALESCE(ROUND(SUM(IF(t.cat='INSTORE', t.net, 0)),2),0),
        SUM(IF(t.cat='INSTORE', 1, 0)),
        COALESCE(ROUND(SUM(IF(t.cat='DELIVERY', t.net, 0)),2),0),
        SUM(IF(t.cat='DELIVERY', 1, 0))
      FROM (
        SELECT {PAYMENT_CASE} AS cat,
               IF(pm.name='Cash',1,0) AS is_cash,
               (po.amount - IFNULL(roa.refund_amount,0)) AS net,
               po.related_uuid AS rid
        FROM `shop{uuid}`.ttpos_payment_order po
        LEFT JOIN `shop{uuid}`.ttpos_payment_method pm ON po.payment_method_uuid = pm.uuid
        LEFT JOIN `shop{uuid}`.ttpos_sale_order so ON po.related_uuid = so.uuid AND so.delete_time=0
        LEFT JOIN `shop{uuid}`.ttpos_sale_bill sb ON so.sale_bill_uuid = sb.uuid AND sb.delete_time=0
        LEFT JOIN (
          SELECT payment_order_uuid, SUM(amount) AS refund_amount
          FROM `shop{uuid}`.ttpos_return_order_amount
          WHERE delete_time=0 AND refund_status=1 GROUP BY payment_order_uuid
        ) roa ON roa.payment_order_uuid = po.uuid
        WHERE po.delete_time=0 AND po.related_type=0 AND po.status=1
          AND sb.finish_time >= {start} AND sb.finish_time <= {end}
          AND sb.status=1 AND so.status=1
          AND (sb.bill_type != 2 OR EXISTS (
            SELECT 1 FROM `shop{uuid}`.ttpos_member_sale_order mso
            WHERE mso.sale_order_uuid = so.uuid AND mso.delete_time=0 AND mso.status=7))
      ) t"""


def q_takeout(uuid, start, end):
    return f"""
      SELECT
        COALESCE(ROUND(SUM(platform_total),2),0),
        COUNT(*)
      FROM `shop{uuid}`.ttpos_takeout_order
      WHERE delete_time=0 AND platform IN ('grab','lineman','shopee')
        AND order_state IN (10,20,30,40) AND accepted_time>0
        AND ((order_state=40 AND completed_time >= {start} AND completed_time <= {end})
          OR (order_state!=40 AND accepted_time >= {start} AND accepted_time <= {end}))"""


def one_row(cur, sql, default):
    try:
        cur.execute(sql)
        r = cur.fetchone()
        return [float(x) if x is not None else 0 for x in r] if r else list(default)
    except Exception as e:
        print(f"   ⚠️ 查询失败: {str(e)[:120]}")
        return list(default)


# ───────────────────────── 门店筛选 ─────────────────────────
def list_stores(cur, saas_db, hq_uuid):
    cur.execute(f"""
      SELECT c.uuid, c.name, IFNULL(cs.timezone,'Asia/Bangkok')
      FROM `{saas_db}`.ttpos_company c
      INNER JOIN `{saas_db}`.ttpos_company_setting cs
        ON cs.company_uuid=c.uuid AND cs.delete_time=0
      WHERE c.delete_time=0 AND c.status=1 AND cs.headquarter_uuid={hq_uuid}
      ORDER BY c.uuid""")
    candidates = cur.fetchall()
    cur.execute("SELECT SCHEMA_NAME FROM information_schema.SCHEMATA")
    dbs = {r[0] for r in cur.fetchall()}

    stores, skip_db, skip_test, skip_code = [], 0, 0, 0
    for uuid, name, tz in candidates:
        db = f"shop{uuid}"
        if db not in dbs:
            skip_db += 1
            continue
        cur.execute(f"SELECT COUNT(*) FROM `{db}`.ttpos_business_status_period WHERE delete_time=0 AND end_time=0")
        if cur.fetchone()[0] != 0:
            skip_test += 1
            continue
        cur.execute(f"SELECT JSON_UNQUOTE(JSON_EXTRACT(`values`,'$.store_code')) "
                    f"FROM `{db}`.ttpos_setting WHERE `key`='store' AND delete_time=0 LIMIT 1")
        row = cur.fetchone()
        code = row[0] if row else None
        if code and str(code).lower() != "null":
            stores.append((uuid, name, tz, str(code)))
        else:
            skip_code += 1
    return stores, (skip_test, skip_code, skip_db)


# ───────────────────────── 主流程 ─────────────────────────
def main():
    ap = argparse.ArgumentParser(description="营业数据汇总导出（店×日）")
    ap.add_argument("date", nargs="?", help="营业日 YYYY-MM-DD（默认泰国昨天）")
    ap.add_argument("--env", default=os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
    ap.add_argument("--output", default=None)
    ap.add_argument("--no-feishu", action="store_true")
    args = ap.parse_args()

    env = load_env(args.env)
    biz_date = args.date or (datetime.now(BKK) - timedelta(days=1)).strftime("%Y-%m-%d")
    saas_db = env.get("SAAS_DB", "saas")
    hq_uuid = env.get("HQ_UUID", "5080409448448000")
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = args.output or os.path.join(out_dir, f"营业数据汇总-{biz_date}.xlsx")

    conn = pymysql.connect(
        host=env["DB_HOST"], port=int(env.get("DB_PORT", 3306)),
        user=env["DB_USER"], password=env.get("DB_PASS", ""),
        charset="utf8mb4", connect_timeout=15, read_timeout=120,
    )
    cur = conn.cursor()
    print(f"连接成功 {env['DB_HOST']}:{env.get('DB_PORT')}")

    stores, (sk_test, sk_code, sk_db) = list_stores(cur, saas_db, hq_uuid)
    print(f"营业日: {biz_date}   正式营业门店: {len(stores)} 家"
          f"（跳过 测试中:{sk_test} 无code:{sk_code} 无库:{sk_db}）\n")

    rows = []
    for i, (uuid, name, tz, code) in enumerate(stores, 1):
        start, end = day_bounds(biz_date, tz)
        tot_rev, recv, biz_inc, tc = one_row(cur, q_totals(uuid, start, end), (0, 0, 0, 0))
        cash_amt, cash_cnt, ins_amt, ins_cnt, dpay_amt, dpay_cnt = one_row(
            cur, q_payments(uuid, start, end), (0, 0, 0, 0, 0, 0))
        tko_amt, tko_cnt = one_row(cur, q_takeout(uuid, start, end), (0, 0))

        deliv_amt = round(dpay_amt + tko_amt, 2)
        deliv_cnt = int(dpay_cnt + tko_cnt)
        rows.append(dict(
            code=code, name=f"{code} {name}",
            tot_rev=tot_rev, recv=recv, biz_inc=biz_inc, tc=int(tc),
            ac=round(tot_rev / tc, 2) if tc else 0,
            cash_cnt=int(cash_cnt), cash_amt=cash_amt,
            cash_ac=round(cash_amt / cash_cnt, 2) if cash_cnt else 0,
            ins_amt=ins_amt, ins_cnt=int(ins_cnt),
            deliv_amt=deliv_amt, deliv_cnt=deliv_cnt))
        print(f" [{i}/{len(stores)}] {code} {name[:24]:24s} "
              f"总营业额={tot_rev:>10} 实收={recv:>10} 到店={ins_amt:>9} 外卖={deliv_amt:>9}")

    cur.close()
    conn.close()

    rows.sort(key=lambda x: x["tot_rev"])
    write_excel(rows, biz_date, out_path)

    size = os.path.getsize(out_path)
    md5 = hashlib.md5(open(out_path, "rb").read()).hexdigest()

    # 对账校验：到店+外卖 应≈实收
    print("\n===== 对账校验（到店+外卖 vs 实收）=====")
    worst = []
    for d in rows:
        diff = round(d["ins_amt"] + d["deliv_amt"] - d["recv"], 2)
        if abs(diff) >= 1:
            worst.append((d["name"][:20], d["recv"], round(d["ins_amt"] + d["deliv_amt"], 2), diff))
    worst.sort(key=lambda x: -abs(x[3]))
    if worst:
        print(f"  ⚠️ {len(worst)}/{len(rows)} 店 到店+外卖 与实收差异≥1（看外卖平台单/退款口径）:")
        for w in worst[:10]:
            print(f"     {w[0]:20s} 实收={w[1]} 到店+外卖={w[2]} 差={w[3]}")
    else:
        print(f"  ✅ 全部 {len(rows)} 店 到店+外卖 = 实收")

    print("\n===== 交付 =====")
    print(f"输出: {out_path}")
    print(f"  门店: {len(rows)}  营业日: {biz_date}")
    print(f"  大小: {size} bytes")
    print(f"  MD5:  {md5}")

    if not args.no_feishu:
        maybe_send_feishu(env, out_path, biz_date)


def write_excel(rows, biz_date, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    hfill = PatternFill("solid", fgColor="4472C4")
    hfont = Font(bold=True, color="FFFFFF")
    tfill = PatternFill("solid", fgColor="E2EFDA")
    tfont = Font(bold=True)
    border = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    CN = ["营业日", "店铺名称", "总营业额", "实收金额", "营业收入", "TC", "AC",
          "现金TC", "现金金额", "现金AC", "到店业绩", "到店订单数", "外卖业绩", "外卖数"]
    EN = ["Business Day", "Store Name", "Total Revenue", "Net Receipt", "Business Income", "TC", "AC",
          "Cash TC", "Cash Amount", "Cash AC", "In-Store Performance", "In-Store Orders",
          "Takeout Performance", "Takeout Orders"]
    for ci, (c, e) in enumerate(zip(CN, EN), 1):
        for ri, v in ((1, c), (2, e)):
            cell = ws.cell(ri, ci, v)
            cell.border = border
            cell.alignment = Alignment("center", "center", wrap_text=True)
            if ri == 1:
                cell.font = hfont
                cell.fill = hfill

    def put(r, c, v, fmt=None, bold=False):
        cell = ws.cell(r, c, v)
        cell.border = border
        if fmt:
            cell.number_format = fmt
        if bold:
            cell.font = tfont
        return cell

    money, intf = "#,##0.00", "#,##0"
    DATA_FIRST = 3
    r = DATA_FIRST
    for d in rows:
        put(r, 1, biz_date)
        put(r, 2, d["name"])
        put(r, 3, d["tot_rev"], money)
        put(r, 4, d["recv"], money)
        put(r, 5, d["biz_inc"], money)
        put(r, 6, d["tc"], intf)
        put(r, 7, d["ac"], money)
        put(r, 8, d["cash_cnt"], intf)
        put(r, 9, d["cash_amt"], money)
        put(r, 10, d["cash_ac"], money)
        put(r, 11, d["ins_amt"], money)
        put(r, 12, d["ins_cnt"], intf)
        put(r, 13, d["deliv_amt"], money)
        put(r, 14, d["deliv_cnt"], intf)
        r += 1
    DATA_LAST = r - 1

    for label, fn, rr in (("总值", "SUM", r), ("平均值", "AVERAGE", r + 1)):
        put(rr, 1, label, bold=True)
        put(rr, 2, "-" if label == "平均值" else "", bold=True)
        for col in range(3, 15):
            L = get_column_letter(col)
            put(rr, col, f"={fn}({L}{DATA_FIRST}:{L}{DATA_LAST})",
                intf if col in (6, 8, 12, 14) else money, bold=True)
        ws.cell(rr, 1).fill = tfill
        ws.cell(rr, 2).fill = tfill
    last_row = r + 1

    for ci in range(1, 15):
        mx = max(sum(2 if ord(ch) > 127 else 1 for ch in str(ws.cell(rr, ci).value or ""))
                 for rr in range(1, last_row + 1))
        ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 46 if ci == 2 else 16)
    ws.freeze_panes = "C3"

    note = wb.create_sheet("说明")
    note["A1"] = f"营业数据汇总  营业日 {biz_date}  门店 {len(rows)} 家  生成自 business_summary.py"
    note["A1"].font = Font(bold=True, color="FF0000")
    note["A2"] = "总营业额/实收/营业收入/TC 来自 ttpos_statistics_sale（对齐 statistics.go）;"
    note["A3"] = "现金/到店/外卖 来自 ttpos_payment_order(扣退款)+ttpos_takeout_order; 外卖=Grab/Lineman/Shopee/Robinhood+平台单。"
    wb.save(out_path)


def maybe_send_feishu(env, out_path, biz_date):
    app_id = env.get("FEISHU_APP_ID")
    app_secret = env.get("FEISHU_APP_SECRET")
    chat_id = env.get("FEISHU_CHAT_ID")
    if not (app_id and app_secret and chat_id):
        print("飞书未配置，跳过发送")
        return
    try:
        requests = _ensure("requests")
        base = "https://open.larksuite.com/open-apis"
        tok = requests.post(f"{base}/auth/v3/tenant_access_token/internal",
                            json={"app_id": app_id, "app_secret": app_secret},
                            timeout=15).json().get("tenant_access_token")
        if not tok:
            print("飞书: 获取 token 失败")
            return
        with open(out_path, "rb") as f:
            up = requests.post(f"{base}/im/v1/files",
                               headers={"Authorization": f"Bearer {tok}"},
                               data={"file_type": "xlsx", "file_name": os.path.basename(out_path)},
                               files={"file": f}, timeout=30).json()
        file_key = up.get("data", {}).get("file_key")
        if not file_key:
            print(f"飞书: 上传失败 {up}")
            return
        import json
        resp = requests.post(f"{base}/im/v1/messages?receive_id_type=chat_id",
                             headers={"Authorization": f"Bearer {tok}", "Content-Type": "application/json"},
                             json={"receive_id": chat_id, "msg_type": "file",
                                   "content": json.dumps({"file_key": file_key})}, timeout=15).json()
        print("飞书发送成功 ✓" if resp.get("code") == 0 else f"飞书发送失败 {resp}")
    except Exception as e:
        print(f"飞书发送异常: {str(e)[:150]}")


if __name__ == "__main__":
    main()
