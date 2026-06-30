#!/usr/bin/env python3
"""
三方对账单与 TTPOS 订单匹配 —— 在原始表上加两列：TTPOS订单号、TTPOS金额

匹配逻辑：商家 → 支付方式(grab) → 金额

Usage:
    # Wallace SHP（无分店信息，需手动指定门店）
    venv/bin/python scripts/reconcile_statement.py \
        --input wallace_tr.xlsx --store "TH0001" \
        --date-col "นที่เกิดรายการ" --amount-col "ยอดเงิน" --order-col "หมายเลข" \
        --output exports/wallace_tr_reconciled.xlsx

    # RBH（有分店信息，自动检测）
    venv/bin/python scripts/reconcile_statement.py \
        --input rbh.xlsx --skip-rows 3 \
        --output exports/rbh_reconciled.xlsx
"""

import argparse
import re
import sys
from datetime import datetime, timezone
from datetime import date as Date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, '/home/weifashi/hwt/analysis')
sys.path.insert(0, '/home/weifashi/hwt/analysis/scripts')
from bq_reports.utils.bq_client import get_bq_client, setup_proxy
from trace_order_by_payment import query_pos_orders, query_takeout_orders

PROJECT_ID = "diyl-407103"
BQ_LOCATION = "asia-southeast1"
THAILAND_TZ_OFFSET = 7 * 3600

STORE_CANDIDATES = [
    'store', 'merchant', 'outlet', 'branch', 'restaurant',
    '门店', '商家', '店铺', '分店',
    'ชื่อร้านค้า', 'ชื่อร้านค้า/สาขา', 'branch_name', 'shop name',
    'shop name', 'shopname', 'store name',
]
DATE_CANDIDATES = [
    'date', 'transaction date', 'order date', 'settlement date',
    '日期', '交易日期', '订单日期',
    'นที่เกิดรายการ', 'วันที่เกิดรายการ', 'order completed time',
]
AMOUNT_CANDIDATES = [
    'amount', 'total', 'price', 'gross amount', 'net amount',
    '金额', '总价', '订单金额',
    'ยอดเงิน', 'total_revenue', '应收',
]
ORDER_CANDIDATES = [
    'order number', 'order id', 'order no', 'transaction id', 'reference',
    '订单号', '交易号',
    'หมายเลข', 'order no.', 'order_no',
]


def _detect_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    col_map = {str(c).lower().strip(): c for c in df.columns}
    for c in candidates:
        if c in col_map:
            return col_map[c]
    for col in df.columns:
        col_lower = str(col).lower().strip()
        for c in candidates:
            if c in col_lower or col_lower in c:
                return col
    return None


def _parse_amount(val) -> float:
    if pd.isna(val):
        return 0.0
    s = str(val).strip()
    s = re.sub(r'[฿$€¥£]', '', s)
    s = s.replace(' ', '')
    if ',' in s and '.' in s:
        last_comma = s.rfind(',')
        last_dot = s.rfind('.')
        if last_comma > last_dot:
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        parts = s.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0


def _parse_date(val, fmt: str | None = None) -> Date | None:
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if fmt:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y%m%d']:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return pd.to_datetime(s).date()
    except (ValueError, pd.errors.OutOfBoundsDatetime):
        return None


def _date_to_ts_range(d: Date) -> tuple[int, int]:
    dt_start = datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
    start_ts = int(dt_start.timestamp()) - THAILAND_TZ_OFFSET
    end_ts = start_ts + 86400
    return start_ts, end_ts


def find_all_stores(client) -> list[dict]:
    job = client.query('''
        SELECT c.uuid, c.name, cs.erpnext_company_abbr
        FROM `diyl-407103`.`saas`.`ttpos_company` c
        LEFT JOIN `diyl-407103`.`saas`.`ttpos_company_setting` cs
          ON cs.company_uuid = c.uuid AND cs.delete_time = 0
        WHERE c.delete_time = 0
          AND cs.headquarter_uuid = 5080409448448000
          AND cs.erpnext_company_abbr LIKE 'TH%%'
        ORDER BY cs.erpnext_company_abbr
    ''', location=BQ_LOCATION)

    stores = []
    datasets = list(client.list_datasets())
    shop_dataset_ids = set(d.dataset_id for d in datasets if d.dataset_id.startswith('shop'))

    for row in job.result():
        dataset = f"shop{row.uuid}"
        if dataset in shop_dataset_ids:
            stores.append({
                'uuid': str(row.uuid),
                'name': row.name or '',
                'abbr': row.erpnext_company_abbr or '',
                'dataset': dataset,
            })
    return stores


def match_store(store_name: str, all_stores: list[dict]) -> dict | None:
    if not store_name or pd.isna(store_name):
        return None
    s_lower = str(store_name).lower().strip()

    for s in all_stores:
        if s_lower == s['name'].lower() or s_lower == s['abbr'].lower():
            return s

    for s in all_stores:
        if s_lower in s['name'].lower() or s['name'].lower() in s_lower:
            return s
        if s['abbr'] and (s_lower in s['abbr'].lower() or s['abbr'].lower() in s_lower):
            return s

    # 4. 泰文部分匹配（RBH 格式如 "Wallace (วอลเลส ) - ลาซาล" → 提取 "ลาซาล" 匹配 TTPOS "Lasalle ลาซาล"）
    thai_parts = re.findall(r'[฀-๿]+', store_name)
    for part in thai_parts:
        part_lower = part.lower().strip()
        if len(part_lower) >= 2:
            for s in all_stores:
                if part_lower in s['name'].lower():
                    return s

    # 5. 关键词匹配（去掉常见后缀）
    s_clean = re.sub(r'\s*(wallace|burger|restaurant|store|branch|\d+)$', '', s_lower).strip()
    if s_clean and s_clean != s_lower:
        for s in all_stores:
            name_clean = re.sub(r'\s*(wallace|burger|restaurant|store|branch|\d+)$', '', s['name'].lower()).strip()
            if s_clean in name_clean or name_clean in s_clean:
                return s

    return None


def fetch_orders(client, store: dict, date_val: Date) -> list[dict]:
    start_ts, end_ts = _date_to_ts_range(date_val)

    pos = query_pos_orders(client, store, 'grab', start_ts, end_ts, None, None, None)
    for o in pos:
        o['source'] = 'pos'
        o['ttpos_amount'] = round(o.get('payment_amount', 0), 2)
        o['ttpos_order_no'] = o.get('serial_number', '')

    to = query_takeout_orders(client, store, 'grab', start_ts, end_ts, None, None, None)
    for o in to:
        o['source'] = 'takeout'
        o['ttpos_amount'] = round(float(o.get('platform_total', 0)), 2)
        o['ttpos_order_no'] = o.get('order_number', '') or o.get('platform_order_id', '')

    return pos + to


def find_best_match(amount: float, orders: list[dict], tolerance: float = 1.0) -> dict | None:
    if not orders:
        return None
    best = None
    best_diff = float('inf')
    for o in orders:
        diff = abs(o['ttpos_amount'] - amount)
        if diff < best_diff:
            best_diff = diff
            best = o
    if best and best_diff <= tolerance:
        return best
    return None


def export(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "对账结果"

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(vertical='center')

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="三方对账单与 TTPOS 订单匹配")
    parser.add_argument("--input", required=True, help="对账单文件路径（Excel/CSV）")
    parser.add_argument("--store", help="手动指定门店关键词（如 TH0001）")
    parser.add_argument("--store-col", help="门店名称列名（默认自动检测）")
    parser.add_argument("--date-col", help="日期列名（默认自动检测）")
    parser.add_argument("--amount-col", help="金额列名（默认自动检测）")
    parser.add_argument("--order-col", help="订单号列名（可选）")
    parser.add_argument("--date-format", help="日期格式（如 %%d/%%m/%%Y）")
    parser.add_argument("--tolerance", type=float, default=1.0, help="金额容差（默认 ±1.0）")
    parser.add_argument("--skip-rows", type=int, default=0, help="跳过前 N 行")
    parser.add_argument("--output", default="exports/reconciled.xlsx", help="输出路径")
    args = parser.parse_args()

    # 1. 读取
    print(f"[1/4] 读取: {args.input}")
    kwargs = {'header': args.skip_rows}
    df = pd.read_excel(args.input, **kwargs) if args.input.endswith('.xlsx') else pd.read_csv(args.input, **kwargs)
    print(f"  {len(df)} 行 x {len(df.columns)} 列")

    # 2. 检测列名
    print("[2/4] 检测列名...")
    store_col = args.store_col or _detect_column(df, STORE_CANDIDATES)
    date_col = args.date_col or _detect_column(df, DATE_CANDIDATES)
    amount_col = args.amount_col or _detect_column(df, AMOUNT_CANDIDATES)
    order_col = args.order_col or _detect_column(df, ORDER_CANDIDATES)

    if not store_col and not args.store:
        print(f"  可用列: {list(df.columns)}")
        print("  错误: 未检测到门店列，请用 --store 或 --store-col 指定")
        return 1
    if not date_col:
        print("  错误: 未检测到日期列，请用 --date-col 指定")
        return 1
    if not amount_col:
        print("  错误: 未检测到金额列，请用 --amount-col 指定")
        return 1

    print(f"  门店: {store_col or ('--store ' + args.store)}")
    print(f"  日期: {date_col}")
    print(f"  金额: {amount_col}")
    if order_col:
        print(f"  订单号: {order_col}")

    # 3. 解析
    df['_amount'] = df[amount_col].apply(_parse_amount)
    df['_date'] = df[date_col].apply(lambda x: _parse_date(x, args.date_format))

    valid_mask = (df['_amount'] > 0) & (df['_date'].notna())
    df = df[valid_mask].copy()
    print(f"  有效数据: {len(df)} 行")

    if len(df) == 0:
        print("  错误: 无有效数据")
        return 1

    # 4. 连接 BQ
    print("\n[3/4] 查询 TTPOS...")
    setup_proxy()
    client = get_bq_client(project_id=PROJECT_ID)
    all_stores = find_all_stores(client)
    print(f"  TTPOS 共 {len(all_stores)} 家门店")

    if args.store:
        matched = match_store(args.store, all_stores)
        if matched:
            print(f"  门店: {matched['abbr']} | {matched['name']}")
            df['_store'] = matched
        else:
            print(f"  错误: --store '{args.store}' 未匹配")
            return 1
    elif store_col:
        df['_store'] = df[store_col].apply(lambda x: match_store(x, all_stores))
        unmatched = df['_store'].isna().sum()
        if unmatched > 0:
            print(f"  警告: {unmatched} 行未匹配门店")

    # 按 (门店, 日期) 分组查询
    groups = {}
    for _, row in df.iterrows():
        store = row['_store']
        if store is None:
            continue
        d = row['_date']
        key = (store['uuid'], d)
        if key not in groups:
            groups[key] = {'store': store, 'date': d, 'orders': []}

    print(f"  查询 {len(groups)} 个组合...")
    cache = {}
    for i, (key, g) in enumerate(groups.items(), 1):
        orders = fetch_orders(client, g['store'], g['date'])
        cache[key] = orders
        print(f"  [{i}/{len(groups)}] {g['store']['abbr']} {g['date']} → {len(orders)} 笔")

    # 5. 匹配
    print("\n[4/4] 匹配订单...")
    results = []
    matched_count = 0
    for _, row in df.iterrows():
        store = row['_store']
        d = row['_date']
        amt = row['_amount']

        if store is None or d is None:
            results.append({'TTPOS订单号': '', 'TTPOS金额': ''})
            continue

        key = (store['uuid'], d)
        orders = cache.get(key, [])
        best = find_best_match(amt, orders, tolerance=args.tolerance)

        if best:
            results.append({
                'TTPOS订单号': best['ttpos_order_no'],
                'TTPOS金额': best['ttpos_amount'],
            })
            matched_count += 1
            # 移除已匹配，防重复
            orders.remove(best)
        else:
            results.append({'TTPOS订单号': '', 'TTPOS金额': ''})

    result_df = pd.DataFrame(results)
    df = df.reset_index(drop=True)
    df['TTPOS订单号'] = result_df['TTPOS订单号']
    df['TTPOS金额'] = result_df['TTPOS金额']

    # 清理临时列
    df = df.drop(columns=['_amount', '_date', '_store'], errors='ignore')

    export(df, args.output)
    print(f"\n输出: {args.output}")
    print(f"  总记录: {len(df)} 行")
    print(f"  匹配成功: {matched_count} 行")
    print(f"  未匹配: {len(df) - matched_count} 行")

    return 0


if __name__ == "__main__":
    sys.exit(main())
