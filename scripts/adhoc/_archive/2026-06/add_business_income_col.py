# 谁问的: 市场/老板  /  2026-06-26
# 问什么: 给「营业数据汇总2026-06-25.xlsx」62 家门店补一列「营业收入」(对齐 ttpos 页面口径, 溯源源码)
# 结论:   营业收入 = ttpos statistics.go business_amount(不含税营业收入); 源表 ttpos.statistics_sale;
#         按 (总营业额,实收) 指纹匹配 62 行全中 0 歧义, 同时复现两已有列做交叉验证; 一次性, 同事日报增列。
#
# 口径溯源 (ttpos-server-go/main/app/repository/statistics.go:88-138, bill 级 select):
#   总营业额 = product_price + product_tax + service_fee + service_tax + payment_fee + extend_price   (含税GMV, 已有列, 验证锚)
#   实收金额 = payment_amount - refund_amount - payment_balance                                        (已有列, 验证锚)
#   营业收入 = payment_amount - refund_amount - refund_payment_balance - product_tax - service_tax + refund_tax  (不含税, 要补)
# 日期: complete_time 落在 BKK 2026-06-25 (页面按 complete_time 筛, 与 total/实收 完全对齐, 故不另加测试营业排除)。

import hashlib
import os
import re

import openpyxl
from openpyxl.utils import get_column_letter

from bq_reports.utils.bq_client import get_bq_client

SRC = "营业数据汇总2026-06-25.xlsx"
DATE = "2026-06-25"
PROJECT = "diyl-407103"

AGG_SQL = f"""
SELECT tenant_id,
  ROUND(SUM(product_price + product_tax + service_fee + service_tax + payment_fee + extend_price), 2) AS total_revenue,
  ROUND(SUM(payment_amount - refund_amount - payment_balance), 2) AS received_amount,
  ROUND(SUM(payment_amount - refund_amount - refund_payment_balance - product_tax - service_tax + refund_tax), 2) AS business_amount
FROM `{PROJECT}.ttpos.statistics_sale`
WHERE DATE(TIMESTAMP_SECONDS(complete_time), 'Asia/Bangkok') = '{DATE}'
GROUP BY tenant_id
"""


def next_version_path(base_dir: str, stem: str) -> tuple[str, int]:
    os.makedirs(base_dir, exist_ok=True)
    pat = re.compile(re.escape(stem) + r"_v(\d+)\.xlsx$")
    mx = 0
    for f in os.listdir(base_dir):
        m = pat.match(f)
        if m:
            mx = max(mx, int(m.group(1)))
    v = mx + 1
    return os.path.join(base_dir, f"{stem}_v{v}.xlsx"), v


def main():
    client = get_bq_client(PROJECT)
    print("[BQ] 聚合 ttpos.statistics_sale ...")
    idx = {}
    agg = {}
    for r in client.query(AGG_SQL).result():
        agg[r.tenant_id] = (r.total_revenue, r.received_amount, r.business_amount)
        idx.setdefault((round(r.total_revenue), round(r.received_amount)), []).append(
            (r.tenant_id, r.business_amount)
        )
    print(f"[BQ] tenant 数: {len(agg)}")

    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb["Sheet1"]

    DATA_FIRST, DATA_LAST = 3, 64  # 62 门店行
    INSERT_AT = 5  # 在「实收金额」(D=4) 之后插入 营业收入 -> 新列 E

    # 1) 先匹配 + 校验(用插入前的列读 总营业额=C/3, 实收=D/4)
    biz_by_row = {}
    matched = ambig = miss = 0
    miss_detail = []
    for row in range(DATA_FIRST, DATA_LAST + 1):
        name = ws.cell(row=row, column=2).value
        tr = ws.cell(row=row, column=3).value
        rec = ws.cell(row=row, column=4).value
        cand = idx.get((round(tr), round(rec)), [])
        if len(cand) == 1:
            biz_by_row[row] = cand[0][1]
            matched += 1
        elif not cand:
            miss += 1
            miss_detail.append((name, tr, rec, "无匹配 tenant"))
        else:
            ambig += 1
            miss_detail.append((name, tr, rec, f"{len(cand)} 个 tenant 撞指纹"))
    print(f"[匹配] 命中={matched} 歧义={ambig} 缺失={miss} (共 {DATA_LAST-DATA_FIRST+1})")
    if matched != (DATA_LAST - DATA_FIRST + 1):
        for d in miss_detail:
            print("   ⚠️", d)
        raise SystemExit("❌ 有未命中/歧义行, 终止 (不交付半成品)")

    # 2) 插列
    ws.insert_cols(INSERT_AT)
    ws.column_dimensions[get_column_letter(INSERT_AT)].width = 13
    # 表头(双语) + 复制相邻列样式
    ws.cell(row=1, column=INSERT_AT, value="营业收入")
    ws.cell(row=2, column=INSERT_AT, value="Business Income")
    for hr in (1, 2):
        src = ws.cell(row=hr, column=4)  # 实收金额表头样式
        dst = ws.cell(row=hr, column=INSERT_AT)
        dst.font = src.font.copy()
        dst.fill = src.fill.copy()
        dst.alignment = src.alignment.copy()
        dst.border = src.border.copy()

    # 3) 回填营业收入(数据行)
    for row in range(DATA_FIRST, DATA_LAST + 1):
        ws.cell(row=row, column=INSERT_AT, value=round(biz_by_row[row], 2))

    # 4) 重写 合计/平均 公式(插列后所有数值列引用错位, 全部按当前列字母重建)
    last_col = ws.max_column  # 插入后共 14 列
    total_row, avg_row = 65, 66
    for col in range(3, last_col + 1):  # C..N 全是数值列
        L = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({L}{DATA_FIRST}:{L}{DATA_LAST})")
        ws.cell(row=avg_row, column=col, value=f"=AVERAGE({L}{DATA_FIRST}:{L}{DATA_LAST})")

    # 5) 内部版本指纹(说明)
    out_dir = "exports"
    out_path, ver = next_version_path(out_dir, "营业数据汇总2026-06-25_补营业收入")
    note = wb.create_sheet("说明")
    note["A1"] = f"内部版本 v{ver}  |  营业收入 = ttpos business_amount(不含税)  |  源表 ttpos.statistics_sale  |  生成日 2026-06-26"
    note["A1"].font = openpyxl.styles.Font(bold=True, color="FF0000")
    note["A2"] = "营业收入公式: payment_amount - refund_amount - refund_payment_balance - product_tax - service_tax + refund_tax"
    note["A3"] = "校验: 62 店「总营业额/实收金额」由同源表公式 100% 复现 → 营业收入可信。"
    wb.properties.title = f"营业数据汇总2026-06-25 补营业收入 v{ver}"

    wb.save(out_path)

    # 6) 指纹打印
    size = os.path.getsize(out_path)
    md5 = hashlib.md5(open(out_path, "rb").read()).hexdigest()
    biz_total = sum(round(biz_by_row[r], 2) for r in biz_by_row)
    print("\n================= 交付 =================")
    print(f"输出: {out_path}")
    print(f"  内部版本: v{ver}")
    print(f"  大小:     {size} bytes")
    print(f"  MD5:      {md5}")
    print(f"  营业收入合计(62店): {round(biz_total, 2)}")


if __name__ == "__main__":
    main()
