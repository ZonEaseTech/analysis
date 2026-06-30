# 谁问的: 市场  /  2026-06-08
# 问什么: 一炉领鲜(2831805321216000) 7 个酒水/饮品分类(Q/O/M/N/P/L/K)的商品目录, 中英泰三语, 带售价和上架状态, 全部商品
# 结论:   按天逐日导出(6/1~6/8 共8天)目录+当日堂食出库量, 8个xlsx入文件夹打zip; 改 STORE/CATS/FIRST_DAY/NUM_DAYS 复用
"""一炉领鲜 酒水饮品分类商品目录导出 (中英泰 + 售价 + 状态)

数据源: shop{store}.ttpos_product_package (商品) JOIN ttpos_product_bom (规格/售价)
        JOIN ttpos_product_category (分类)
多语言: name 字段是 JSON {"zh","th","en"}, JSON_EXTRACT_SCALAR 拆三列.
粒度  : 一行 = 一个规格 (bom). 多规格商品(扎/杯, 奶油顶/无)拆多行, 售价各异.

无销量/金额指标 → 销售类 validators(qty/revenue/refund 恒等式) 不适用.
改跑数据质量 audit: 缺中文名 / 售价<=0 / 分类覆盖.
"""
import hashlib
import re
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bq_reports.utils.bq_client import get_bq_client

BKK_TZ = timezone(timedelta(hours=7))
# 出库口径: 按天逐日导出。默认 2026-06-01 起 8 天 (含 8 号), 可 --start / --days 覆盖。
DEFAULT_FIRST_DAY = "2026-06-01"
DEFAULT_NUM_DAYS = 8
# 运行期由 main() 按参数填充
FIRST_DAY = datetime(2026, 6, 1, tzinfo=BKK_TZ)
NUM_DAYS = 8


def day_windows():
    """逐日 [start_ts, end_ts) + 日期标签, BKK 时区。"""
    for i in range(NUM_DAYS):
        d0 = FIRST_DAY + timedelta(days=i)
        d1 = d0 + timedelta(days=1)
        yield d0.strftime("%Y-%m-%d"), int(d0.timestamp()), int(d1.timestamp())

PROJECT = "diyl-407103"
STORE = "2831805321216000"
DATASET = f"shop{STORE}"
STORE_LABEL = "一炉领鲜"

# 用户给定顺序 + 干净中文标签 (category 表里的 zh 混了泰文, 这里用业务标签)
# category_uuid: 排序序, 标签
CATS = [
    (40, "Q 黄酒"),
    (34, "O 红酒"),
    (32, "M 洋酒"),
    (30, "N 白酒"),
    (28, "P 啤酒"),
    (26, "L 饮料"),
    (24, "K 特色饮品"),
]
CAT_LABEL = {uuid: label for uuid, label in CATS}
CAT_ORDER = {uuid: i for i, (uuid, _) in enumerate(CATS)}
CAT_IDS = ", ".join(str(u) for u, _ in CATS)

COLUMNS = ["一级分类", "商品中文名", "商品英文名", "商品泰文名", "规格", "售价",
           "出库数量", "剩余库存", "状态"]


def _next_version_path(base_dir: Path, prefix: str, suffix: str = ".xlsx") -> Path:
    pattern = re.compile(rf"^{re.escape(prefix)}_v(\d+){re.escape(suffix)}$")
    versions = [int(m.group(1)) for f in base_dir.iterdir()
                if (m := pattern.match(f.name))]
    next_v = (max(versions) + 1) if versions else 1
    return base_dir / f"{prefix}_v{next_v}{suffix}", next_v


def fetch_rows():
    client = get_bq_client(PROJECT)
    sql = f"""
    SELECT
      pp.uuid AS package_uuid,
      pp.category_uuid AS category_uuid,
      JSON_EXTRACT_SCALAR(pp.name, '$.zh') AS name_zh,
      JSON_EXTRACT_SCALAR(pp.name, '$.en') AS name_en,
      JSON_EXTRACT_SCALAR(pp.name, '$.th') AS name_th,
      JSON_EXTRACT_SCALAR(b.name, '$.zh') AS spec_zh,
      b.price AS price,
      b.stock_num AS stock_num,
      b.is_open_stock AS is_open_stock,
      pp.status AS status,
      pp.sort AS sort
    FROM `{PROJECT}`.`{DATASET}`.`ttpos_product_package` pp
    JOIN `{PROJECT}`.`{DATASET}`.`ttpos_product_bom` b
      ON b.product_package_uuid = pp.uuid AND b.delete_time = 0
    WHERE pp.delete_time = 0
      AND pp.category_uuid IN ({CAT_IDS})
    """
    rows = list(client.query(sql).result())
    # Python 侧排序: 分类(用户顺序) → 商品 sort → 中文名 → 售价
    rows.sort(key=lambda r: (
        CAT_ORDER.get(r.category_uuid, 99),
        r.sort if r.sort is not None else 0,
        r.name_zh or "",
        float(r.price) if r.price is not None else 0,
    ))
    return rows


def fetch_period_qty(start_ts, end_ts):
    """区间出库量, 粒度 (product_package_uuid, price)。

    本店 (一炉领鲜) 为火锅堂食店, 无外卖业务 (无 ttpos_takeout_order 表),
    出库 = 堂食实际销量, 直接取自 ttpos_statistics_product (sale_event 堂食支同口径)。
    qty = SUM(product_num) 毛出库量 (与对账锚 sale_event.qty 一致)。返回:
      pkg_sales: {pkg_uuid: [(sale_price, qty), ...]}  # 该商品当日各成交价档的销量
    """
    client = get_bq_client(PROJECT)
    sql = f"""
    SELECT
      sp.product_package_uuid AS item_uuid,
      sp.product_sale_price AS price,
      SUM(sp.product_num) AS qty
    FROM `{PROJECT}`.`{DATASET}`.`ttpos_statistics_product` sp
    WHERE sp.complete_time >= {start_ts}
      AND sp.complete_time < {end_ts}
    GROUP BY item_uuid, price
    """
    pkg_sales = defaultdict(list)
    for r in client.query(sql).result():
        q = float(r.qty) if r.qty is not None else 0.0
        price = float(r.price) if r.price is not None else 0.0
        pkg_sales[int(r.item_uuid)].append((price, q))
    return pkg_sales


def to_records(rows, pkg_sales):
    """把每个商品当日销量分配到目录行 (规格)。

    单规格商品 → 全部销量归该行。
    多规格商品 → 各成交价就近归到「售价最接近」的规格 (改价/促销价不丢数, 总量守恒)。
    """
    # 同一商品的规格行 (保持出现顺序)
    pkg_specs = defaultdict(list)
    for idx, r in enumerate(rows):
        pkg_specs[int(r.package_uuid)].append((idx, float(r.price) if r.price is not None else 0.0))

    out_by_idx = {}
    for pkg, specs in pkg_specs.items():
        sales = pkg_sales.get(pkg, [])
        if len(specs) == 1:
            out_by_idx[specs[0][0]] = sum(q for _, q in sales)
            continue
        # 多规格: 每个成交价就近分配到价格最接近的规格
        acc = {idx: 0.0 for idx, _ in specs}
        for sale_price, q in sales:
            best_idx = min(specs, key=lambda s: abs(s[1] - sale_price))[0]
            acc[best_idx] += q
        out_by_idx.update(acc)

    recs = []
    matched_total = 0.0
    for idx, r in enumerate(rows):
        out_qty = out_by_idx.get(idx, 0.0)
        matched_total += out_qty
        recs.append({
            "一级分类": CAT_LABEL.get(r.category_uuid, str(r.category_uuid)),
            "商品中文名": r.name_zh or "",
            "商品英文名": r.name_en or "",
            "商品泰文名": r.name_th or "",
            "规格": r.spec_zh or "",
            "售价": float(r.price) if r.price is not None else None,
            "出库数量": out_qty,
            # 开放库存(is_open_stock=1) → 不限量; 否则取 POS 实时库存 stock_num 原值
            "剩余库存": ("不限量" if r.is_open_stock == 1
                     else (float(r.stock_num) if r.stock_num is not None else None)),
            "状态": "上架" if r.status == 0 else "下架",
        })
    return recs, matched_total


def audit(recs):
    print("\n===== 交付前 audit =====")
    # 1. 分类覆盖
    by_cat = Counter(r["一级分类"] for r in recs)
    print("分类行数分布:")
    for _, label in CATS:
        print(f"  {label}: {by_cat.get(label, 0)} 行")
    print(f"  合计: {len(recs)} 行")
    # 2. 缺中文名
    miss_zh = [r for r in recs if not r["商品中文名"]]
    print(f"缺中文名: {len(miss_zh)}")
    for r in miss_zh:
        print("   ", r)
    # 3. 缺英文/泰文
    miss_en = sum(1 for r in recs if not r["商品英文名"])
    miss_th = sum(1 for r in recs if not r["商品泰文名"])
    print(f"缺英文名: {miss_en}  缺泰文名: {miss_th}")
    # 4. 售价 <= 0 / 缺失
    bad_price = [r for r in recs if r["售价"] is None or r["售价"] <= 0]
    print(f"售价缺失/<=0: {len(bad_price)}")
    for r in bad_price:
        print("   ", r["一级分类"], r["商品中文名"], r["规格"], r["售价"])
    # 6. 状态分布
    on = sum(1 for r in recs if r["状态"] == "上架")
    print(f"状态: 上架 {on} / 下架 {len(recs) - on}")
    print("========================\n")


def write_xlsx(recs, out_path, day_label):
    wb = Workbook()
    wb.properties.title = f"{STORE_LABEL} 酒水饮品商品目录 数据日期 {day_label}"

    # ---- 数据 sheet ----
    ws = wb.active
    ws.title = "商品目录"
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for c, name in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for ridx, rec in enumerate(recs, 2):
        for c, name in enumerate(COLUMNS, 1):
            v = rec[name]
            cell = ws.cell(ridx, c, v)
            if name == "售价":
                cell.number_format = "#,##0.00"
                cell.alignment = center
            elif name == "出库数量":
                cell.number_format = "#,##0"
                cell.alignment = center
            elif name == "剩余库存":
                if isinstance(v, (int, float)):
                    cell.number_format = "#,##0"
                cell.alignment = center
            elif name in ("状态", "一级分类"):
                cell.alignment = center
            else:
                cell.alignment = left
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(recs) + 1}"
    widths = [14, 34, 34, 38, 14, 10, 12, 12, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- 说明 sheet ----
    info = wb.create_sheet("说明")
    info["A1"] = f"数据日期 {day_label}"
    info["A1"].font = Font(bold=True, color="C00000", size=14)
    notes = [
        "",
        f"门店: {STORE_LABEL}  (ID {STORE})",
        "范围: 全部商品 (含上架 + 下架), 客户自行按「状态」列过滤",
        "分类: Q黄酒 / O红酒 / M洋酒 / N白酒 / P啤酒 / L饮料 / K特色饮品",
        "语言: 名称中英泰三列, 取自系统多语言字段原值 (未做清洗)",
        "粒度: 一行 = 一个规格. 部分商品(扎/杯、奶油顶/无奶油顶)有多规格 → 多行, 售价各异",
        "售价: 取自商品规格(product_bom)的标价, 单位泰铢(THB)",
        f"出库数量: 当日【{day_label}】(BKK时区, 当天0点~次日0点) 堂食实际销量 (本店火锅堂食, 无外卖业务)",
        "  多规格商品按售价匹配到对应规格; 单规格商品取该商品当日总销量",
        "剩余库存: 取自 POS 实时库存(product_bom.stock_num)。",
        "  ⚠️ 这是【导出当下的实时快照】, 非该日结余 —— POS 不留每日库存历史, 故 8 天文件的此列数值相同, 仅「出库数量」按天变。",
        "  「不限量」= 该商品设为开放库存(不计数); 个别商品库存为店家填的虚高初始值或负数(超卖), 均为 POS 原值。",
        "状态: 上架 / 下架 取自商品(product_package).status",
        "",
        "提醒: 「商品中文名」字段系统里部分混入泰文(如「维C鲜橙 วิตามินซีส้มสด」), 为系统原值, 非导出错误",
    ]
    for i, line in enumerate(notes, 2):
        info[f"A{i}"] = line
    info.column_dimensions["A"].width = 90

    wb.save(out_path)
    return out_path


def main():
    import argparse
    import os
    import time
    import zipfile

    global FIRST_DAY, NUM_DAYS
    parser = argparse.ArgumentParser(description="一炉领鲜 酒水饮品 每日出库+库存导出")
    parser.add_argument("--start", default=DEFAULT_FIRST_DAY,
                        help="起始日 YYYY-MM-DD (BKK), 默认 2026-06-01")
    parser.add_argument("--days", type=int, default=DEFAULT_NUM_DAYS,
                        help="导出天数, 默认 8")
    a = parser.parse_args()
    FIRST_DAY = datetime.strptime(a.start, "%Y-%m-%d").replace(tzinfo=BKK_TZ)
    NUM_DAYS = a.days

    base = Path("exports")
    base.mkdir(parents=True, exist_ok=True)
    # 目录 + zip 走版本号, 重跑不覆盖旧版
    folder_prefix = f"酒水饮品商品目录_{STORE_LABEL}_每日_{FIRST_DAY:%Y%m%d}-{NUM_DAYS}天"
    pattern = re.compile(rf"^{re.escape(folder_prefix)}_v(\d+)$")
    vs = [int(m.group(1)) for f in base.iterdir() if (m := pattern.match(f.name))]
    version = (max(vs) + 1) if vs else 1
    folder = base / f"{folder_prefix}_v{version}"
    folder.mkdir(parents=True, exist_ok=True)

    rows = fetch_rows()           # 目录(名称/规格/价/状态) 每天不变, 只查一次
    audit(to_records(rows, {})[0])  # 结构 audit 一次足够

    print(f"\n===== 按天导出 ({NUM_DAYS} 天) =====")
    day_files = []
    for day_label, start_ts, end_ts in day_windows():
        pkg_sales = fetch_period_qty(start_ts, end_ts)
        recs, matched_total = to_records(rows, pkg_sales)
        period_total = sum(q for sales in pkg_sales.values() for _, q in sales)
        out_path = folder / f"酒水饮品商品目录_{STORE_LABEL}_{day_label}.xlsx"
        write_xlsx(recs, out_path, day_label)
        day_files.append(out_path)
        zero = sum(1 for r in recs if not r["出库数量"])
        print(f"  {day_label}: 出库合计 {matched_total:,.0f} "
              f"(0出库 {zero}/{len(recs)} 行) | 全店当日 {period_total:,.0f}")

    # 打 zip
    zip_path = base / f"{folder_prefix}_v{version}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in day_files:
            zf.write(f, arcname=f"{folder.name}/{f.name}")

    data = zip_path.read_bytes()
    md5 = hashlib.md5(data).hexdigest()
    st = os.stat(zip_path)
    print(f"\n输出文件夹: {folder}/  ({len(day_files)} 个 xlsx)")
    print(f"压缩包:     {zip_path}")
    print(f"  内部版本: v{version}")
    print(f"  修改时间: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(st.st_mtime))}")
    print(f"  大小:     {st.st_size} bytes")
    print(f"  MD5:      {md5}")


if __name__ == "__main__":
    main()
