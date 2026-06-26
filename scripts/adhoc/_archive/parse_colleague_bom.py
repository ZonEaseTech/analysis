#!/usr/bin/env python3
"""把同事调整版 BOM xlsx 解析成统一结构化 BOM 表。

源文件: /workspace/data/uploads/01KRFDZSSZC24D72E30Y8293SY.xlsx
源 sheets:
  - ！套餐去重（匹配完整）  6 列: 套餐名 / 物品名 / 物品编码 / 变更消耗 / 单价 / 单位
  - ！单品去重 (匹配完整)   6 列: 单品名 / 物品名 / 物品编码 / 消耗 / 单价 / 单位
  - ！套餐销量合并          8 列: A=旧名 ... H=新名 → A→H 别名映射 (N:1)

输出:
  resources/wallace.20260513/同事调整版_结构化.xlsx
  sheet "BOM配方(解析结果)"，列: 商品名称 / 物品编号 / 物品名称 / 单耗 / 单位 / 类型 / 来源
  (前 5 列对齐 wallace.20260506 老格式，后 2 列仅做审计)

别名处理:
  对每对 A→H, 把 H 的全部 BOM 行复制一份, 商品名称替换为 A。
  这样 BQ 里旧名 SKU (A) 也能直接命中 BOM。
"""
from __future__ import annotations

import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook


SRC = Path("/workspace/data/uploads/01KRFK6S9FNFJPA8DY39A2HWPW.xlsx")  # 2026-05-13 第 2 版
DST = Path("resources/wallace.20260513/profit_by_price_202603_核对版 BOM.xlsx")

# 列单位映射（跟 _load_fallback_boms 保持一致）
UOM_MAP = {"克": "g", "g": "g", "个": "pc", "pc": "pc", "份": "pc"}


def _norm_uom(u):
    if u is None:
        return ""
    s = str(u).strip()
    return UOM_MAP.get(s, s)


def _read_bom_sheet(wb, sheet_name, kind):
    """读 6 列 BOM sheet (含表头)，返回 [(product, code, mat_name, qty, uom, price_or_None, kind)]。

    新版多带 price (第 5 列, col index 4), 用于后续抽取物料→单价 map (单品 sheet 是权威)。
    """
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        product, mat_name, code, qty, price, unit = row[:6]
        if not product or not code:
            continue
        try:
            qty_val = float(qty)
        except (ValueError, TypeError):
            continue
        try:
            price_val = float(price) if price is not None else None
        except (ValueError, TypeError):
            price_val = None
        rows.append((
            str(product).strip(),
            str(code).strip(),
            str(mat_name or "").strip(),
            qty_val,
            _norm_uom(unit),
            price_val,
            kind,
        ))
    return rows


def _read_alias_map(wb, sheet_name):
    """读 A→H 别名映射 (N:1, 无表头, 列 0=A 旧名, 列 7=H 新名)。"""
    ws = wb[sheet_name]
    pairs = set()
    for row in ws.iter_rows(values_only=True):
        if len(row) < 8:
            continue
        a, h = row[0], row[7]
        if a and h:
            pairs.add((str(a).strip(), str(h).strip()))
    # 去 A==H 的无意义自映射
    return {a: h for a, h in pairs if a != h}


def main():
    if not SRC.exists():
        print(f"源文件不存在: {SRC}")
        sys.exit(1)

    wb_src = load_workbook(SRC, read_only=True, data_only=True)

    combo_rows = _read_bom_sheet(wb_src, "！套餐去重（匹配完整）", "套餐")
    single_rows = _read_bom_sheet(wb_src, "！单品去重 (匹配完整)", "单品")
    alias_map = _read_alias_map(wb_src, "！套餐销量合并")

    print(f"套餐 BOM 行: {len(combo_rows)}  (商品 {len({r[0] for r in combo_rows})} 个)")
    print(f"单品 BOM 行: {len(single_rows)}  (商品 {len({r[0] for r in single_rows})} 个)")
    print(f"A→H 别名: {len(alias_map)} 条")

    # 按商品名分组（便于查 H 的 BOM 复制成 A）
    by_name = defaultdict(list)
    for r in combo_rows + single_rows:
        by_name[r[0]].append(r)

    # 生成别名行: A 的每一行 = H 的对应行, 但 product 改成 A, kind 改成 "别名→H"
    #
    # ⚠️ Semantics (β): 当 A 自己在 BOM 表里已经有完整配方时, 不做 alias 注入。
    # 历史 bug: 之前的 "叠加" 语义会把 H 独有的物料拷到 A 名下, 导致 A 的成本被虚增。
    # 典型案例: 超值套餐 4 (带空格, 29 行汉堡套餐 BOM) 被 alias 错误注入了
    # H=超值套餐4 (无空格, 14 行鸡棒套餐) 独有的 7 行物料 → A 的 BOM 变 36 行。
    #
    # 这种 A-also-has-own-BOM 的情况下, 销量合并表里那条 A→H 极可能是误填
    # (因为 A 既然下架了为啥还要维护自己的配方?), 在 BOM 层面保留 A 自己的配方
    # 更保守、可审计。如果客户后续确认 A→H 是有意为之, 可以手工删 A 自己的 29 行。
    alias_rows = []
    missing_h = []
    skipped_self_bom = []  # A 已有自己的 BOM, 跳过 alias 注入
    for a, h in sorted(alias_map.items()):
        if a in by_name:
            skipped_self_bom.append((a, h, len(by_name[a])))
            continue
        h_rows = by_name.get(h)
        if not h_rows:
            missing_h.append((a, h))
            continue
        for _p, code, mat_name, qty, uom, _price, _kind in h_rows:
            # 别名行不带 price (price 用统一的物料→价 map 解析, 不在 BOM 行带)
            alias_rows.append((a, code, mat_name, qty, uom, None, f"别名→{h}"))

    if skipped_self_bom:
        print(f"\n⚠️  {len(skipped_self_bom)} 个 A 自己已有完整 BOM, 跳过 alias (保留 A 自己的配方):")
        for a, h, n in skipped_self_bom:
            print(f"   A={a!r} ({n} 行自有 BOM)  原本要 alias→ H={h!r}")

    if missing_h:
        print(f"\n⚠️  {len(missing_h)} 个 H 在 BOM 中找不到, A 无法生成别名:")
        for a, h in missing_h:
            print(f"   A={a!r}  H={h!r}")

    print(f"\n生成别名 BOM 行: {len(alias_rows)}  "
          f"(覆盖 {len(alias_map) - len(missing_h) - len(skipped_self_bom)} 个旧名, "
          f"跳过 {len(skipped_self_bom)} 个 A 已自有 BOM, 缺 H {len(missing_h)} 个)")

    # 合并输出
    out_rows = combo_rows + single_rows + alias_rows

    # 同名+同编码去重（保留先到的: 套餐 > 单品 > 别名）
    seen = set()
    dedup = []
    for r in out_rows:
        key = (r[0], r[1])  # product + code
        if key in seen:
            continue
        seen.add(key)
        dedup.append(r)
    print(f"去重后 BOM 行: {len(dedup)}  (去掉 {len(out_rows) - len(dedup)} 重复)")

    # === 物料单价 map ===
    # 单品 sheet 是权威 (用户分析: 94 物料里只有 FR05001 一个内部冲突)。
    # 套餐 sheet 单价乱填, 完全忽略。
    # 同一物料同事填 2 个价 → 取出现次数多的 (mode); 平手时取较高那个 (保守, 避免低估成本)。
    price_votes = defaultdict(list)
    for _p, code, _mn, _q, _u, price, _k in single_rows:
        if price is not None:
            price_votes[code].append(price)
    material_prices = {}
    price_conflicts = []
    for code, prices in price_votes.items():
        cnt = Counter(prices)
        most = cnt.most_common()
        if len(most) == 1:
            material_prices[code] = most[0][0]
        else:
            # 多价: 取出现次数最多的; 平手取较高的 (保守)
            top_count = most[0][1]
            top_prices = [p for p, c in most if c == top_count]
            chosen = max(top_prices)
            material_prices[code] = chosen
            price_conflicts.append((code, dict(cnt), chosen))

    print(f"\n从【单品 sheet】抽取物料单价 map: {len(material_prices)} 个物料")
    if price_conflicts:
        print(f"⚠️  {len(price_conflicts)} 个物料单品 sheet 内部有 ≥2 个不同价:")
        for code, votes, chosen in price_conflicts:
            print(f"   {code}  各价票数: {votes}  → 选 ¥{chosen} (出现次数最多, 平手取高价)")

    # === sanity 警告 ===
    # 单位 = g 但单价 > 0.5 → 极可能是 ¥/kg 误填到 ¥/g (米的 ¥1/g bug)
    suspicious = []
    for _p, code, mat, _q, uom, price, _k in single_rows:
        if uom == 'g' and price is not None and price > 0.5:
            suspicious.append((code, mat, price))
    suspicious = list({s[:2]: s for s in suspicious}.values())  # 去重
    if suspicious:
        print(f"\n⚠️  {len(suspicious)} 个物料单价可疑 (单位=g 但单价 > ¥0.5/g):")
        for code, mat, price in suspicious:
            print(f"   {code}  {mat}  单价=¥{price}/g  ← 检查是不是 ¥/kg 填错单位")

    # === 写文件 ===
    DST.parent.mkdir(parents=True, exist_ok=True)
    wb_out = Workbook()

    # Sheet 1: BOM 数量/单位 (报表的 fallback_bom 维度)
    ws1 = wb_out.active
    ws1.title = "BOM配方(解析结果)"
    ws1.append(["商品名称", "物品编号", "物品名称", "单耗", "单位", "类型", "来源"])
    for product, code, mat_name, qty, uom, _price, kind in dedup:
        ws1.append([product, code, mat_name, qty, uom, kind, "同事调整版 20260513"])

    # Sheet 2: 物料单价 map (独立维度, 报表的 material_price_sources 用)
    ws2 = wb_out.create_sheet("物料单价")
    ws2.append(["物品编号", "物品名称", "单价", "单位", "来源"])
    code_to_name = {code: mat_name for _, code, mat_name, _, _, _, _ in single_rows + combo_rows}
    code_to_uom = {code: uom for _, code, _, _, uom, _, _ in single_rows + combo_rows}
    for code, price in sorted(material_prices.items()):
        ws2.append([code, code_to_name.get(code, ""), price,
                    code_to_uom.get(code, ""), "同事调整版_单品 sheet"])

    wb_out.save(DST)
    print(f"\n✅ 写入 {DST}")
    print(f"   sheet1 'BOM配方(解析结果)': {len(dedup)+1} 行")
    print(f"   sheet2 '物料单价':           {len(material_prices)+1} 行")


if __name__ == "__main__":
    main()
