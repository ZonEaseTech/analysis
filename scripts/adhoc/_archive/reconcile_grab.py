#!/usr/bin/env python3
"""
Grab 对账：用 Long Order ID 精确 1:1 匹配 TTPOS takeout_order.platform_order_id

数据源：
  Grab 文件 Sheet 'Transaction_Store_2026-03-01_to'  - 逐笔交易明细
  TTPOS  ttpos_takeout_order WHERE platform='grab'  - 78 家 TH 门店

对账状态：
  已对上          - Long Order ID 在 TTPOS 找到，金额一致 (差 ≤1 THB)
  金额不一致      - 找到但金额对不上
  已取消(无需对账)- Grab Status=Cancelled
  TTPOS漏单       - Grab Transferred 但 TTPOS 没此订单（核心待查项）

Usage:
    venv/bin/python scripts/reconcile_grab.py \\
        --input /workspace/data/uploads/01KQY2SYTDC96NVZSGZY2MG6MF.xlsx \\
        --output exports/grab_reconciled.xlsx
"""

import argparse
import concurrent.futures
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

sys.path.insert(0, '/home/weifashi/hwt/analysis')
from bq_reports.utils.bq_client import get_bq_client, setup_proxy
from scripts.batch_reconcile import export_df, next_versioned_path

PROJECT_ID = "diyl-407103"
BQ_LOCATION = "asia-southeast1"
TH_OFFSET = 7 * 3600


def get_all_stores(client):
    job = client.query('''
        SELECT c.uuid, c.name, cs.erpnext_company_abbr
        FROM `diyl-407103`.`saas`.`ttpos_company` c
        LEFT JOIN `diyl-407103`.`saas`.`ttpos_company_setting` cs
          ON cs.company_uuid = c.uuid AND cs.delete_time = 0
        WHERE c.delete_time = 0
          AND cs.headquarter_uuid = 5080409448448000
          AND cs.erpnext_company_abbr LIKE 'TH%%'
        ORDER BY cs.erpnext_company_abbr
    ''', location=BQ_LOCATION)
    return [{'uuid': str(r.uuid), 'name': r.name or '', 'abbr': r.erpnext_company_abbr or ''}
            for r in job.result()]


def fetch_grab_orders(client, stores, start_ts, end_ts):
    """并发拉取所有门店 Grab 订单。返回 dict: platform_order_id → row"""
    def q(s):
        try:
            sql = f"""
                SELECT '{s['abbr']}' abbr, platform_order_id,
                  CAST(subtotal AS FLOAT64) subtotal,
                  CAST(platform_total AS FLOAT64) platform_total,
                  order_state, payment_type,
                  TIMESTAMP_SECONDS(CASE WHEN order_state=40 AND completed_time>0 THEN completed_time ELSE accepted_time END + 7*3600) ts
                FROM `{PROJECT_ID}`.`shop{s['uuid']}`.`ttpos_takeout_order`
                WHERE delete_time=0 AND platform='grab' AND order_state IN (10,20,30,40)
                  AND (CASE WHEN order_state=40 AND completed_time>0 THEN completed_time ELSE accepted_time END) >= {start_ts}
                  AND (CASE WHEN order_state=40 AND completed_time>0 THEN completed_time ELSE accepted_time END) < {end_ts}
            """
            return [dict(r) for r in client.query(sql, location=BQ_LOCATION).result()]
        except Exception:
            return []

    idx = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as p:
        for rows in p.map(q, stores):
            for r in rows:
                idx[r['platform_order_id']] = r
    return idx


def reconcile(df, ttpos_idx):
    """对每行 Grab 文件：按 Long Order ID 找 TTPOS 对应订单"""
    statuses, ttpos_oids, ttpos_stores, ttpos_amts, ttpos_pays, diffs = [], [], [], [], [], []
    seen = set()

    for _, row in df.iterrows():
        oid = row.get('Long Order ID')
        status_src = row.get('Status')
        amt_grab = row.get('Amount')

        rec = {'TTPOS门店': '', 'TTPOS订单号': '', 'TTPOS金额(毛)': '',
               'TTPOS支付方式': '', '金额差异': '', '对账状态': ''}

        if pd.isna(oid):
            rec['对账状态'] = '源数据无订单号'
        elif status_src == 'Cancelled':
            rec['对账状态'] = '已取消(无需对账)'
        else:
            t = ttpos_idx.get(oid)
            if not t:
                rec['对账状态'] = 'TTPOS漏单'
            else:
                seen.add(oid)
                rec['TTPOS门店'] = t['abbr']
                rec['TTPOS订单号'] = oid
                rec['TTPOS金额(毛)'] = round(t['subtotal'], 2)
                rec['TTPOS支付方式'] = t['payment_type'] or ''
                try:
                    diff = round(float(amt_grab) - t['subtotal'], 2)
                    rec['金额差异'] = diff
                    rec['对账状态'] = '已对上' if abs(diff) <= 1 else '金额不一致'
                except Exception:
                    rec['对账状态'] = '金额无法比较'

        for k, v in rec.items():
            pass
        statuses.append(rec['对账状态'])
        ttpos_oids.append(rec['TTPOS订单号'])
        ttpos_stores.append(rec['TTPOS门店'])
        ttpos_amts.append(rec['TTPOS金额(毛)'])
        ttpos_pays.append(rec['TTPOS支付方式'])
        diffs.append(rec['金额差异'])

    df = df.reset_index(drop=True).copy()
    df['TTPOS门店'] = ttpos_stores
    df['TTPOS订单号'] = ttpos_oids
    df['TTPOS金额(毛)'] = ttpos_amts
    df['TTPOS支付方式'] = ttpos_pays
    df['金额差异'] = diffs
    df['对账状态'] = statuses
    return df, seen


def build_extras(ttpos_idx, seen):
    """TTPOS 多出的订单（Grab 文件没有）→ 单独 sheet"""
    extras = []
    for oid, t in ttpos_idx.items():
        if oid not in seen:
            extras.append({
                '说明': 'TTPOS 中存在但 Grab 对账文件未涵盖',
                'TTPOS门店': t['abbr'],
                'TTPOS订单号': oid,
                'TTPOS金额(毛)': round(t['subtotal'], 2),
                'TTPOS订单时间': t['ts'].strftime('%Y-%m-%d %H:%M:%S') if t.get('ts') else '',
                'TTPOS支付方式': t.get('payment_type') or '',
                'TTPOS订单状态': t.get('order_state'),
            })
    return pd.DataFrame(extras)


def export_with_extras(main_df, extras_df, output_path):
    """同上 export_df 风格，外加 TTPOS多余订单 sheet"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    text_keywords = ('订单号', '小票号', 'Order ID', 'Order no', 'INV#', 'Transaction ID',
                     'Settlement ID', 'Booking ID', 'Merchant ID', 'Store ID', 'Long Order')

    def write_sheet(ws, df):
        text_cols = set()
        for ci, col_name in enumerate(df.columns, 1):
            if any(k in str(col_name) for k in text_keywords):
                text_cols.add(ci)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                if r_idx > 1 and c_idx in text_cols and value is not None and value != '':
                    value = str(value)
                    if value.endswith('.0'):
                        value = value[:-2]
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                else:
                    cell.alignment = Alignment(vertical='center')
                    if c_idx in text_cols:
                        cell.number_format = '@'
        for col_idx in range(1, ws.max_column + 1):
            ml = 0
            for row_idx in range(1, min(ws.max_row + 1, 200)):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v:
                    ml = max(ml, len(str(v)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(ml + 2, 10), 50)
        ws.freeze_panes = "A2"

    # Sheet 1: 完整对账
    ws1 = wb.active
    ws1.title = "Grab对账"
    write_sheet(ws1, main_df)

    # Sheet 2: TTPOS 多出的订单
    if len(extras_df) > 0:
        ws2 = wb.create_sheet("TTPOS多余订单")
        write_sheet(ws2, extras_df)

    # Sheet 3: 状态汇总
    ws3 = wb.create_sheet("对账汇总")
    summary = main_df['对账状态'].value_counts().reset_index()
    summary.columns = ['对账状态', '行数']
    summary['占比'] = (summary['行数'] / len(main_df) * 100).round(2).astype(str) + '%'
    write_sheet(ws3, summary)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--input', required=True)
    p.add_argument('--output', default='exports/grab_reconciled.xlsx')
    p.add_argument('--month', default='2026-03', help='YYYY-MM')
    args = p.parse_args()

    print("=" * 60)
    print(f"Grab 对账: {args.input}")
    print("=" * 60)

    setup_proxy()
    client = get_bq_client(project_id=PROJECT_ID)

    print("\n[1/4] 拉取 TTPOS 全门店列表...")
    stores = get_all_stores(client)
    print(f"  共 {len(stores)} 家 TH 门店")

    yyyy, mm = args.month.split('-')
    yyyy, mm = int(yyyy), int(mm)
    nm = mm + 1 if mm < 12 else 1
    ny = yyyy if mm < 12 else yyyy + 1
    start_ts = int(datetime(yyyy, mm, 1, tzinfo=timezone.utc).timestamp()) - TH_OFFSET
    end_ts = int(datetime(ny, nm, 1, tzinfo=timezone.utc).timestamp()) - TH_OFFSET

    print(f"\n[2/4] 拉取 TTPOS Grab 订单 ({args.month})...")
    ttpos_idx = fetch_grab_orders(client, stores, start_ts, end_ts)
    print(f"  TTPOS Grab 订单总数: {len(ttpos_idx)}")

    print(f"\n[3/4] 读取 Grab 文件...")
    df = pd.read_excel(args.input, sheet_name='Transaction_Store_2026-03-01_to')
    print(f"  Grab 文件行数: {len(df)}")
    print(f"    Transferred: {(df['Status']=='Transferred').sum()}")
    print(f"    Cancelled: {(df['Status']=='Cancelled').sum()}")

    print(f"\n[4/4] 1:1 对账...")
    matched_df, seen = reconcile(df, ttpos_idx)
    extras_df = build_extras(ttpos_idx, seen)

    # 状态分布
    print()
    print("对账状态分布:")
    for k, v in matched_df['对账状态'].value_counts().items():
        print(f"  {k}: {v} 行 ({v*100/len(matched_df):.1f}%)")
    print(f"\nTTPOS 多余订单 (Grab 文件未涵盖): {len(extras_df)} 笔")

    out = next_versioned_path(args.output)
    export_with_extras(matched_df, extras_df, out)
    print(f"\n输出: {out}")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
