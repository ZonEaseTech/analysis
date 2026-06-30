#!/usr/bin/env python3
# 谁问的: 何伟涛 / 2026-05-28
# 问什么: 2026-04 全 TH 门店微信支付逐笔订单明细导出, 供客户拿去 POS/收单后台核实漏记的微信渠道 27,172
# 结论:   183 笔 / 净额 27,172.00(毛 27,533 / 退 361)与客户预期分文不差; 0 笔取不到订单号(全部命中 sale_bill.order_no)。
#         17 家 TH 店无 ttpos_statistics_payment 表(404)整店跳过, 与归档聚合脚本同样跳过且总额仍命中 27,172, 故不含微信。已沉淀(一次性 adhoc)。
"""遍历所有 TH 门店, 从 ttpos_statistics_payment 拉微信渠道(三个碎片名)逐笔支付记录(一笔一行)导 Excel。

口径(严格照用, 与归档聚合脚本 breakdown_wechat_202604.py 同源):
  微信渠道 = ttpos_payment_method.payment_name IN ('WeChatPay','WeChat Pay','Kbank-WeChatPay')
  净额     = payment_amount - refund_amount
  WHERE    sp.delete_time=0 AND complete_time>=START AND <END
  时间     = TIMESTAMP_SECONDS(sp.complete_time) 转 'Asia/Bangkok'
  2026-04 BKK 整月: START_TS=1774976400, END_TS=1777568400
  全门店净额应 = 27,172.00 / 笔数 183

订单号:
  优先 LEFT JOIN ttpos_sale_bill (CAST(sp.sale_bill_uuid AS STRING)=CAST(sb.uuid AS STRING)) 取 sb.order_no
  取不到则用 sp.uuid 兜底(注: 该表无 serial_number 列, 实际字段为 sp.uuid, 已核实 schema)
缺 sale_bill 表的店 try/except 跳过该表关联(降级为纯支付查询), 不报错不漏店。
"""
import sys
import os
import hashlib
import datetime
from collections import defaultdict

sys.path.insert(0, '/home/weifashi/hwt/analysis')

from bq_reports.utils.bq_client import get_bq_client
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BQ_LOCATION = "asia-southeast1"
START_TS = 1774976400
END_TS = 1777568400
EXPECTED_NET = 27172.0
EXPECTED_COUNT = 183

WECHAT_NAMES = ('WeChatPay', 'WeChat Pay', 'Kbank-WeChatPay')
names_in_clause = ", ".join(f"'{n}'" for n in WECHAT_NAMES)

OUT_DIR = "/home/weifashi/hwt/analysis/exports"
OUT_BASENAME = "wechat_orders_2026-04"
INTERNAL_VERSION = "v1"

client = get_bq_client()


def fmt(x):
    return f"{x:,.2f}"


def next_version_path(base_dir, base_name):
    import re
    n = 0
    pat = re.compile(rf"^{re.escape(base_name)}_v(\d+)\.xlsx$")
    for f in os.listdir(base_dir):
        m = pat.match(f)
        if m:
            n = max(n, int(m.group(1)))
    return os.path.join(base_dir, f"{base_name}_v{n + 1}.xlsx"), f"v{n + 1}"


# ========== 1. 门店枚举(照搬参考脚本) ==========
job = client.query('''
    SELECT c.uuid, c.name, cs.erpnext_company_abbr
    FROM `diyl-407103`.`saas`.`ttpos_company` c
    LEFT JOIN `diyl-407103`.`saas`.`ttpos_company_setting` cs
      ON cs.company_uuid = c.uuid AND cs.delete_time = 0
    WHERE c.delete_time = 0
      AND cs.headquarter_uuid = 5080409448448000
      AND cs.erpnext_company_abbr LIKE 'TH%'
    ORDER BY cs.erpnext_company_abbr
''', location=BQ_LOCATION)

all_th_stores = []
for row in job.result():
    all_th_stores.append({
        'uuid': str(row.uuid),
        'name': row.name,
        'abbr': row.erpnext_company_abbr or '',
        'dataset': f"shop{row.uuid}",
    })

datasets = list(client.list_datasets())
shop_dataset_ids = set(d.dataset_id for d in datasets if d.dataset_id.startswith('shop'))
stores = [s for s in all_th_stores if s['dataset'] in shop_dataset_ids]

print(f"TH 门店总数(配置): {len(all_th_stores)}, 实际存在 shop* dataset: {len(stores)}")


# ========== 2. 跨门店拉逐笔微信支付记录 ==========
# 每条: (abbr, name, dt_str, order_no, payment_name, gross, refund, net)
records = []
no_order_no = 0          # 取不到 order_no, 用 sp.uuid 兜底的笔数
no_salebill_stores = []  # 缺 sale_bill 表降级的店

ok = fail = 0
fail_stores = []


def query_store(dataset, with_salebill):
    """返回 RowIterator。with_salebill=False 时退化为纯支付查询(订单号取 sp.uuid)。"""
    if with_salebill:
        sql = f"""
            SELECT
                FORMAT_TIMESTAMP('%Y-%m-%d %H:%M:%S',
                    TIMESTAMP_SECONDS(sp.complete_time), 'Asia/Bangkok') AS dt,
                COALESCE(sb.order_no, CAST(sp.uuid AS STRING)) AS order_no,
                CASE WHEN sb.order_no IS NULL THEN 1 ELSE 0 END AS no_order,
                pm.payment_name AS payment_name,
                sp.payment_amount AS gross,
                sp.refund_amount AS refund
            FROM `diyl-407103`.`{dataset}`.`ttpos_statistics_payment` sp
            LEFT JOIN `diyl-407103`.`{dataset}`.`ttpos_payment_method` pm
                ON pm.uuid = sp.payment_method_uuid
            LEFT JOIN `diyl-407103`.`{dataset}`.`ttpos_sale_bill` sb
                ON CAST(sp.sale_bill_uuid AS STRING) = CAST(sb.uuid AS STRING)
            WHERE sp.delete_time = 0
                AND sp.complete_time >= {START_TS}
                AND sp.complete_time < {END_TS}
                AND pm.payment_name IN ({names_in_clause})
            ORDER BY sp.complete_time
        """
    else:
        sql = f"""
            SELECT
                FORMAT_TIMESTAMP('%Y-%m-%d %H:%M:%S',
                    TIMESTAMP_SECONDS(sp.complete_time), 'Asia/Bangkok') AS dt,
                CAST(sp.uuid AS STRING) AS order_no,
                1 AS no_order,
                pm.payment_name AS payment_name,
                sp.payment_amount AS gross,
                sp.refund_amount AS refund
            FROM `diyl-407103`.`{dataset}`.`ttpos_statistics_payment` sp
            LEFT JOIN `diyl-407103`.`{dataset}`.`ttpos_payment_method` pm
                ON pm.uuid = sp.payment_method_uuid
            WHERE sp.delete_time = 0
                AND sp.complete_time >= {START_TS}
                AND sp.complete_time < {END_TS}
                AND pm.payment_name IN ({names_in_clause})
            ORDER BY sp.complete_time
        """
    return client.query(sql, location=BQ_LOCATION).result()


for s in stores:
    dataset = s['dataset']
    rows = None
    try:
        rows = query_store(dataset, with_salebill=True)
        rows = list(rows)
    except Exception as e:
        # 可能是缺 sale_bill 表; 降级再试一次纯支付查询
        msg = str(e)[:120]
        try:
            rows = list(query_store(dataset, with_salebill=False))
            no_salebill_stores.append(s['abbr'] or dataset)
        except Exception as e2:
            fail += 1
            fail_stores.append((s['abbr'] or dataset, str(e2)[:80]))
            continue

    for row in rows:
        g = float(row.gross) if row.gross is not None else 0.0
        rf = float(row.refund) if row.refund is not None else 0.0
        net = g - rf
        if row.no_order:
            no_order_no += 1
        records.append({
            'abbr': s['abbr'],
            'name': s['name'],
            'dt': row.dt,
            'order_no': row.order_no,
            'payment_name': row.payment_name,
            'gross': g,
            'refund': rf,
            'net': net,
        })
    ok += 1

print(f"查询成功 {ok} 店, 失败/缺表 {fail} 店")
if no_salebill_stores:
    print(f"缺 sale_bill 表降级(订单号用 sp.uuid 兜底)的店: {', '.join(no_salebill_stores)}")
if fail_stores:
    for ab, err in fail_stores:
        print(f"  [跳过] {ab}: {err}")

# 排序: 门店 abbr -> 支付时间
records.sort(key=lambda x: (x['abbr'], x['dt']))

tot_gross = sum(r['gross'] for r in records)
tot_refund = sum(r['refund'] for r in records)
tot_net = sum(r['net'] for r in records)
tot_count = len(records)

# ========== 3. Console 核对 ==========
W = 70
print("\n" + "=" * W)
print("逐笔核对")
print("=" * W)
print(f"总笔数(系统)     = {tot_count}   (客户预期 {EXPECTED_COUNT})")
print(f"净额合计(系统)   = {fmt(tot_net)}   (客户预期 {fmt(EXPECTED_NET)})")
print(f"毛额合计         = {fmt(tot_gross)}")
print(f"退款合计         = {fmt(tot_refund)}")
print(f"净额差额         = {fmt(tot_net - EXPECTED_NET)}")
print(f"笔数差额         = {tot_count - EXPECTED_COUNT}")
print(f"取不到订单号(用 sp.uuid 兜底)笔数 = {no_order_no}")
net_ok = abs(tot_net - EXPECTED_NET) < 0.01
cnt_ok = tot_count == EXPECTED_COUNT
print(f"核对结果: 净额 {'✅' if net_ok else '🔴'}  笔数 {'✅' if cnt_ok else '🔴'}")

# ========== 4. Excel 导出 ==========
out_path, file_ver = next_version_path(OUT_DIR, OUT_BASENAME)

wb = Workbook()
wb.properties.title = f"微信订单明细 2026-04 ({INTERNAL_VERSION})"

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
HDR_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True)
RED_BOLD = Font(bold=True, color="FF0000")
GREY = Font(size=9, color="666666")
MONEY = '#,##0.00'

ws = wb.active
ws.title = "微信订单明细"

# ---- 说明块 ----
NOTE_LINES = [
    "口径: 微信渠道 = payment_name IN ('WeChatPay','WeChat Pay','Kbank-WeChatPay') 三个原始碎片名合并",
    "净额 = payment_amount - refund_amount; WHERE delete_time=0 AND complete_time∈[2026-04 BKK整月)",
    "支付时间按 Asia/Bangkok; 来源 ttpos_statistics_payment join ttpos_payment_method (各门店 shop{uuid} dataset)",
    "订单号: 优先 ttpos_sale_bill.order_no(按 sale_bill_uuid 关联); 取不到则用支付记录ID sp.uuid 兜底",
    f"全门店净额 = {fmt(tot_net)} / 笔数 {tot_count} (客户预期 {fmt(EXPECTED_NET)} / {EXPECTED_COUNT})",
]
r = 1
ws.cell(row=r, column=1, value=f"内部版本 {INTERNAL_VERSION}").font = RED_BOLD
r += 1
for line in NOTE_LINES:
    ws.cell(row=r, column=1, value=line).font = GREY
    r += 1
r += 1

hdr_row = r
headers = ["门店编号", "门店名称", "支付时间", "订单号", "支付方式", "毛额", "退款", "净额"]
for ci, h in enumerate(headers, start=1):
    ws.cell(row=r, column=ci, value=h)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=hdr_row, column=c)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER
r += 1
data_start = r

for rec in records:
    ws.cell(row=r, column=1, value=rec['abbr'])
    ws.cell(row=r, column=2, value=rec['name'])
    ws.cell(row=r, column=3, value=rec['dt'])
    # 订单号设为文本, 避免 Excel 把长数字串当数字
    oc = ws.cell(row=r, column=4, value=str(rec['order_no']))
    oc.number_format = '@'
    ws.cell(row=r, column=5, value=rec['payment_name'])
    ws.cell(row=r, column=6, value=round(rec['gross'], 2)).number_format = MONEY
    ws.cell(row=r, column=7, value=round(rec['refund'], 2)).number_format = MONEY
    ws.cell(row=r, column=8, value=round(rec['net'], 2)).number_format = MONEY
    for ci in range(1, len(headers) + 1):
        ws.cell(row=r, column=ci).border = BORDER
    r += 1

# 合计行
ws.cell(row=r, column=1, value="合计")
ws.cell(row=r, column=4, value=f"{tot_count} 笔")
ws.cell(row=r, column=6, value=round(tot_gross, 2)).number_format = MONEY
ws.cell(row=r, column=7, value=round(tot_refund, 2)).number_format = MONEY
ws.cell(row=r, column=8, value=round(tot_net, 2)).number_format = MONEY
for ci in range(1, len(headers) + 1):
    c = ws.cell(row=r, column=ci)
    c.font = TOTAL_FONT
    c.fill = TOTAL_FILL
    c.border = BORDER

# 冻结表头(冻结到数据首行)
ws.freeze_panes = ws.cell(row=data_start, column=1)

widths = {'A': 12, 'B': 24, 'C': 21, 'D': 24, 'E': 16, 'F': 12, 'G': 12, 'H': 12}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(out_path)

# ========== 5. 指纹 ==========
size = os.path.getsize(out_path)
with open(out_path, 'rb') as f:
    md5 = hashlib.md5(f.read()).hexdigest()
mtime = datetime.datetime.fromtimestamp(os.path.getmtime(out_path)).strftime("%Y-%m-%d %H:%M:%S")

print("\n" + "=" * W)
print("输出指纹")
print("=" * W)
print(f"输出:     {out_path}")
print(f"  文件版本: {file_ver}")
print(f"  内部版本: {INTERNAL_VERSION}")
print(f"  修改时间: {mtime}")
print(f"  大小:     {size} bytes")
print(f"  MD5:      {md5}")
