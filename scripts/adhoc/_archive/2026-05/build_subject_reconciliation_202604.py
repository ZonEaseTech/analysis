# 谁问的: 何伟涛 / 2026-05-29
# 问什么: 重新整理 2026-04 支付科目对账表(客户账上/客户拉的系统/我方真值 三列对照),供发客户
# 结论: 9个科目,客户漏微信27172+多算Grab误录16188;我方真值35,989,495.7 vs 客户到账35,950,205.4 差39,290.3
"""
2026-04 支付科目对账表(整理版)

三列:
  A 客户账上(实际到账)  — 客户自己的账
  B 客户系统(客户自拉)  — 客户从我们系统导的, 漏了微信、多算Grab POS误录
  C 我方系统真值        — 我方完整核验(payment−refund, 含微信, 剔除Grab误录)

数字全部来自本会话已验证结果, 不再查 BQ。
"""
import hashlib
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

VERSION = "v1"

# 科目: (名, 客户账上A, 客户系统B(None=客户没有), 我方真值C, 差异性质, 说明)
ROWS = [
    ("QR",      10000315.40, 10091913.0, 10092190.00, "扫码换标签", "我方记QR, 银行按卡/支付宝入账 → 与EDC/支付宝对冲"),
    ("EDC",       427836.00,   345480.0,   345480.00, "扫码换标签", "银行刷卡入账比系统多, 与QR对冲"),
    ("ALIPAY",     41583.00,    13872.0,    13872.00, "扫码换标签", "银行支付宝入账比系统多, 与QR对冲"),
    ("GRAB",     4081642.00,  4088499.0,  4072311.00, "外卖结算+误录", "真值取外卖平台口径; 客户系统列多算了POS误录16,188"),
    ("LINEMAN",  8167600.00,  8185896.0,  8186400.00, "外卖结算", "平台在途/抽佣调整, 待Grab/LINEMAN结算单"),
    ("SHOPEE",   4898095.00,  4916184.0,  4916184.00, "外卖结算", "平台在途/抽佣调整, 待Shopee结算单"),
    ("RBH",         6690.00,     6853.7,     6853.70, "零头", "Robinhood尾差, 可忽略"),
    ("CASH",     8326444.00,  8328934.0,  8329033.00, "现金抹零", "抹零/找零/盘差, 待门店现金日结"),
    ("WECHAT",         None,       None,    27172.00, "客户漏统计", "★客户账上+系统两列均无此渠道; 我方183笔/27,172, 明细见附件"),
]

TOTAL_A = 35950205.40   # 客户账上合计
TOTAL_B = 35977631.70   # 客户系统合计
TOTAL_C = 35989495.70   # 我方真值合计
GRAB_POS_ERR = 16188.00 # Grab POS误录(不计入真值)

# 样式
HDR = PatternFill("solid", fgColor="2F5496")
HDRF = Font(bold=True, color="FFFFFF", size=11)
SUBF = Font(bold=True, color="FFFFFF", size=10)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")
BD = Border(*[Side(style="thin", color="BFBFBF")] * 4)
BOLD = Font(bold=True)
RED = Font(color="C00000", bold=True)
GRN = Font(color="006100")
NAT_FILL = {
    "扫码换标签":   PatternFill("solid", fgColor="DDEBF7"),
    "外卖结算":     PatternFill("solid", fgColor="FFF2CC"),
    "外卖结算+误录": PatternFill("solid", fgColor="FFF2CC"),
    "现金抹零":     PatternFill("solid", fgColor="FFF2CC"),
    "客户漏统计":   PatternFill("solid", fgColor="FCE4D6"),
    "零头":         PatternFill("solid", fgColor="E2EFDA"),
}


def m(ws, cell, v):
    if v is None:
        ws[cell] = "—"; ws[cell].alignment = CEN
    else:
        ws[cell] = v; ws[cell].number_format = "#,##0.00"; ws[cell].alignment = RGT


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "科目对账表"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"华莱士 2026-04 支付科目对账表（{VERSION}）"
    ws["A1"].font = Font(bold=True, size=14, color="C00000"); ws["A1"].alignment = CEN
    ws.merge_cells("A2:H2")
    ws["A2"] = ("说明：① 客户账上=实际到账　② 客户系统=客户自行从我方系统导出（漏微信、多算Grab POS误录）　"
                "③ 我方真值=完整核验(payment−refund，含微信，剔除Grab误录)。对账请以③为准。")
    ws["A2"].font = Font(size=9, italic=True); ws["A2"].alignment = LFT

    heads = ["科目", "①客户账上\n(实际到账)", "②客户系统\n(客户自拉)", "③我方真值\n(完整核验)",
             "差异\n(③−①)", "性质", "说明"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = CEN; cell.border = BD

    r = 5
    for name, a, b, cc, nat, note in ROWS:
        ws.cell(row=r, column=1, value=name).alignment = CEN
        ws.cell(row=r, column=1).font = BOLD
        m(ws, f"B{r}", a)
        m(ws, f"C{r}", b)
        m(ws, f"D{r}", cc)
        diff = round((cc or 0) - (a or 0), 2)
        m(ws, f"E{r}", diff)
        ws[f"E{r}"].font = RED if diff > 0 else GRN
        ws.cell(row=r, column=6, value=nat).alignment = CEN
        ws.cell(row=r, column=7, value=note).alignment = LFT
        fill = NAT_FILL.get(nat)
        if fill:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BD
        r += 1

    # 合计
    ws.cell(row=r, column=1, value="合计").font = BOLD
    m(ws, f"B{r}", TOTAL_A); m(ws, f"C{r}", TOTAL_B); m(ws, f"D{r}", TOTAL_C)
    m(ws, f"E{r}", round(TOTAL_C - TOTAL_A, 2)); ws[f"E{r}"].font = RED
    ws.cell(row=r, column=6, value="净差").alignment = CEN
    ws.cell(row=r, column=7,
            value=f"我方真值比客户实际到账多 {TOTAL_C-TOTAL_A:,.2f}（占 {(TOTAL_C-TOTAL_A)/TOTAL_A*100:.3f}%）").font = BOLD
    for c in range(1, 8):
        ws.cell(row=r, column=c).font = BOLD if c != 5 else RED
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=r, column=c).border = BD
    r += 2

    # 旁注: Grab POS误录
    ws.cell(row=r, column=1, value="备注").font = BOLD
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(row=r, column=2,
            value=f"Grab POS误录 {GRAB_POS_ERR:,.2f}：收银员把Grab当支付方式手录，非真实渠道，已从我方真值剔除；"
                  f"客户②系统列未剔除，故②的GRAB比③多约16,188。")
    ws.cell(row=r, column=2).alignment = LFT; ws.cell(row=r, column=2).font = Font(size=9, italic=True)
    r += 2

    # 差异归类小结
    ws.cell(row=r, column=1, value="差异归类（金额=事实；归因=推测，待核实）").font = RED; r += 1
    groups = [
        ("微信缺失 【事实】", 27172.00, "客户账上+系统均无此渠道，我方BQ查得183笔可复现。请确认27,172是否已到账→补进对账"),
        ("外卖三平台差额 【金额事实/归因推测】", (4072311.0-4081642.0)+(8186400.0-8167600.0)+(4916184.0-4898095.0),
         "差额是事实；推测为在途+抽佣/补贴，待三平台结算单证实"),
        ("现金差额 【金额事实/归因推测】", 8329033.0-8326444.0, "差额是事实；推测为抹零/找零/盘差，待门店现金日结证实"),
        ("RBH差额 【事实】", 6853.7-6690.0, "金额极小"),
        ("扫码三渠道差额 【金额事实/归因推测】", (10092190.0-10000315.4)+(345480.0-427836.0)+(13872.0-41583.0),
         "三渠道差额是事实；‘换标签对冲’为推测，未证实，待银行流水核渠道归类"),
    ]
    for gn, gv, note in groups:
        ws.cell(row=r, column=1, value=gn).alignment = LFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1)
        m(ws, f"E{r}", round(gv, 2)); ws[f"E{r}"].font = RED if gv > 0 else GRN
        ws.cell(row=r, column=6, value="").alignment = CEN
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        ws.cell(row=r, column=6, value=note).alignment = LFT
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BD
        r += 1
    # 归类合计
    ws.cell(row=r, column=1, value="净差合计").font = BOLD
    m(ws, f"E{r}", round(TOTAL_C - TOTAL_A, 2)); ws[f"E{r}"].font = RED
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BD

    widths = [26, 16, 16, 16, 14, 12, 50]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[4].height = 34
    ws.freeze_panes = "A5"

    # Sheet2 待办
    ws2 = wb.create_sheet("结论与待办")
    items = [
        ("【事实】数据可复现", "9渠道金额、微信183笔、退款额、Grab POS误录16,188、LINEMAN 4/22切换——均BQ逐笔查证，脚本在 scripts/adhoc/_archive/2026-05/，可重跑。"),
        ("【事实】客户漏微信", "客户对账表账上+系统两列均无微信渠道。我方查得微信27,172/183笔。此为查证事实。"),
        ("【事实】真实净差", f"我方真值 {TOTAL_C:,.2f} − 客户实际到账 {TOTAL_A:,.2f} = {TOTAL_C-TOTAL_A:,.2f}（0.109%）。客户原表27,426是漏微信+多Grab误录凑出的偏小值。"),
        ("【推测·未证实】扫码换标签", "−18,192。QR记多、刷卡/支付宝记少，我推测是同批钱换渠道标签对冲——但无银行流水证据，也可能是三个独立差异。待银行流水证实。"),
        ("【推测·未证实】外卖结算差", "+27,558。我推测为平台在途/抽佣/补贴——未见结算单，未证实。待Grab/LINEMAN/Shopee商家结算单。"),
        ("【推测·未证实】现金+RBH", "+2,753。我推测为抹零/找零/盘差——未证实。待门店现金日结。"),
        ("待客户提供(以证实推测)", "1) 银行/收单流水  2) Grab/LINEMAN/Shopee结算单  3) 门店现金日结。收到后才能把上面的推测逐项坐实。"),
        ("附件", "本表 + 微信183笔订单明细(wechat_orders_2026-04) + 对账桥(reconciliation_bridge_2026-04)。"),
    ]
    ws2.merge_cells("A1:B1"); ws2["A1"] = "结论与待办"; ws2["A1"].font = Font(bold=True, size=13, color="C00000")
    for c, h in enumerate(["项", "内容"], 1):
        cell = ws2.cell(row=2, column=c, value=h); cell.fill = HDR; cell.font = HDRF; cell.alignment = CEN; cell.border = BD
    for i, (k, v) in enumerate(items, 3):
        ws2.cell(row=i, column=1, value=k).font = BOLD
        ws2.cell(row=i, column=1).alignment = LFT
        ws2.cell(row=i, column=2, value=v).alignment = LFT
        for c in range(1, 3):
            ws2.cell(row=i, column=c).border = BD
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 100

    out_dir = Path("/home/weifashi/hwt/analysis/exports")
    base = "科目对账表_2026-04"
    n = 1
    while (out_dir / f"{base}_v{n}.xlsx").exists():
        n += 1
    out = out_dir / f"{base}_v{n}.xlsx"
    wb.save(out)
    data = out.read_bytes()
    print(f"输出: {out}")
    print(f"  内部版本: {VERSION}")
    print(f"  大小:     {len(data)} bytes")
    print(f"  MD5:      {hashlib.md5(data).hexdigest()}")
    print(f"\n  校验: 我方真值 {TOTAL_C:,.2f} − 客户到账 {TOTAL_A:,.2f} = {TOTAL_C-TOTAL_A:,.2f}")


if __name__ == "__main__":
    build()
