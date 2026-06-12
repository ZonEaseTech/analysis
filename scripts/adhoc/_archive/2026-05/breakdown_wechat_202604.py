#!/usr/bin/env python3
# 谁问的: 何伟涛 / 2026-05-28
# 问什么: 2026-04 全 TH 门店微信支付净实收按店按天拆, 定位客户银行/平台到账表里缺失的微信资金集中在哪些店/哪些天
# 结论:   净额 27,172.00 / 毛 27,533.00 / 退 361.00 与客户预期分文不差。微信散落 38 家店, 高度碎片化(Top3 仅占 27.9%, 单笔多为 99-3800), 无单一大额集中; 客户到账表缺"微信"栏更像整列科目漏记而非某店漏账。已沉淀(一次性 adhoc, 脚本归档)
"""遍历所有 TH 门店, 从 ttpos_statistics_payment 聚合微信渠道(三个碎片名)净实收, 按店 + 按天拆分导 Excel。

口径(严格照用):
  微信渠道 = ttpos_payment_method.payment_name IN ('WeChatPay','WeChat Pay','Kbank-WeChatPay')
  净实收   = SUM(payment_amount) - SUM(refund_amount)   (refund_amount 独立退款列必须扣)
  WHERE    sp.delete_time=0 AND complete_time>=START AND <END
  日期     = DATE(TIMESTAMP_SECONDS(sp.complete_time), 'Asia/Bangkok')
  2026-04 BKK 整月: START_TS=1774976400, END_TS=1777568400
  全门店净额应 ≈ 27,172 (毛额≈27,533, 退款≈361)
"""
import sys
import os
import hashlib
from collections import defaultdict

sys.path.insert(0, '/home/weifashi/hwt/analysis')

from bq_reports.utils.bq_client import get_bq_client
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BQ_LOCATION = "asia-southeast1"
START_TS = 1774976400
END_TS = 1777568400
EXPECTED_NET = 27172.0

WECHAT_NAMES = ('WeChatPay', 'WeChat Pay', 'Kbank-WeChatPay')

OUT_DIR = "/home/weifashi/hwt/analysis/exports"
OUT_BASENAME = "wechat_breakdown_2026-04"
INTERNAL_VERSION = "v1"  # 内部版本标记 (auto-version 决定文件名 _vN)

client = get_bq_client()


def fmt(x):
    return f"{x:,.2f}"


def next_version_path(base_dir, base_name):
    """扫 exports/ 已有 _vN, max+1。不覆盖旧版。"""
    import re
    n = 0
    pat = re.compile(rf"^{re.escape(base_name)}_v(\d+)\.xlsx$")
    for f in os.listdir(base_dir):
        m = pat.match(f)
        if m:
            n = max(n, int(m.group(1)))
    return os.path.join(base_dir, f"{base_name}_v{n + 1}.xlsx"), f"v{n + 1}"


# ========== 1. 门店枚举(照搬参考脚本) ==========
job = client.query('''
    SELECT c.uuid, c.name, cs.erpnext_company_abbr
    FROM `diyl-407103`.`saas`.`ttpos_company` c
    LEFT JOIN `diyl-407103`.`saas`.`ttpos_company_setting` cs
      ON cs.company_uuid = c.uuid AND cs.delete_time = 0
    WHERE c.delete_time = 0
      AND cs.headquarter_uuid = 5080409448448000
      AND cs.erpnext_company_abbr LIKE 'TH%'
    ORDER BY cs.erpnext_company_abbr
''', location=BQ_LOCATION)

all_th_stores = []
for row in job.result():
    all_th_stores.append({
        'uuid': str(row.uuid),
        'name': row.name,
        'abbr': row.erpnext_company_abbr or '',
        'dataset': f"shop{row.uuid}",
    })

datasets = list(client.list_datasets())
shop_dataset_ids = set(d.dataset_id for d in datasets if d.dataset_id.startswith('shop'))
stores = [s for s in all_th_stores if s['dataset'] in shop_dataset_ids]

print(f"TH 门店总数(配置): {len(all_th_stores)}, 实际存在 shop* dataset: {len(stores)}")

# ========== 2. 跨门店聚合微信渠道 ==========
# 按店: abbr -> {name, gross, refund, count}
store_agg = {}
# 按天: date_str -> {net, count}  (net 已含退款扣减, 在 SQL 里按天算 gross/refund)
day_gross = defaultdict(float)
day_refund = defaultdict(float)
day_count = defaultdict(int)
# 按店×天: (abbr, date) -> net
store_day_net = defaultdict(float)

names_in_clause = ", ".join(f"'{n}'" for n in WECHAT_NAMES)

ok = fail = 0
fail_stores = []

for s in stores:
    dataset = s['dataset']
    try:
        job = client.query(f"""
            SELECT
                DATE(TIMESTAMP_SECONDS(sp.complete_time), 'Asia/Bangkok') AS d,
                COUNT(*) AS cnt,
                SUM(sp.payment_amount) AS gross,
                SUM(sp.refund_amount) AS refund
            FROM `diyl-407103`.`{dataset}`.`ttpos_statistics_payment` sp
            LEFT JOIN `diyl-407103`.`{dataset}`.`ttpos_payment_method` pm
                ON pm.uuid = sp.payment_method_uuid
            WHERE sp.delete_time = 0
                AND sp.complete_time >= {START_TS}
                AND sp.complete_time < {END_TS}
                AND pm.payment_name IN ({names_in_clause})
            GROUP BY d
            ORDER BY d
        """, location=BQ_LOCATION)

        st_gross = st_refund = 0.0
        st_count = 0
        has_rows = False
        for row in job.result():
            has_rows = True
            d = row.d.isoformat()
            g = float(row.gross) if row.gross else 0.0
            r = float(row.refund) if row.refund else 0.0
            c = int(row.cnt) if row.cnt else 0
            st_gross += g
            st_refund += r
            st_count += c
            day_gross[d] += g
            day_refund[d] += r
            day_count[d] += c
            store_day_net[(s['abbr'], s['name'], d)] += (g - r)

        if has_rows:
            store_agg[s['abbr']] = {
                'name': s['name'],
                'gross': st_gross,
                'refund': st_refund,
                'count': st_count,
            }
        ok += 1
    except Exception as e:
        fail += 1
        fail_stores.append((s['abbr'] or s['dataset'], str(e)[:80]))

print(f"查询成功 {ok} 店, 失败/缺表 {fail} 店")
if fail_stores:
    for ab, err in fail_stores:
        print(f"  [跳过] {ab}: {err}")

# ========== 3. Console 输出 ==========
W = 88
print("\n" + "=" * W)
print("A) 按店汇总 (只列有微信交易的店, 按净额降序)")
print("=" * W)
print(f"{'abbr':<12}{'门店名':<22}{'毛额':>14}{'退款':>12}{'净额':>14}{'笔数':>8}")
print("-" * W)

store_rows = []
for abbr, v in store_agg.items():
    net = v['gross'] - v['refund']
    store_rows.append((abbr, v['name'], v['gross'], v['refund'], net, v['count']))
store_rows.sort(key=lambda r: r[4], reverse=True)

tot_gross = tot_refund = tot_net = 0.0
tot_count = 0
for abbr, name, g, r, net, cnt in store_rows:
    print(f"{abbr:<12}{name[:20]:<22}{fmt(g):>14}{fmt(r):>12}{fmt(net):>14}{cnt:>8,}")
    tot_gross += g
    tot_refund += r
    tot_net += net
    tot_count += cnt
print("-" * W)
print(f"{'合计':<12}{'':<22}{fmt(tot_gross):>14}{fmt(tot_refund):>12}{fmt(tot_net):>14}{tot_count:>8,}")

print("\n" + "=" * W)
print("B) 按天汇总 (每天净额 + 笔数)")
print("=" * W)
print(f"{'日期':<14}{'毛额':>14}{'退款':>12}{'净额':>14}{'笔数':>8}")
print("-" * W)
day_rows = []
for d in sorted(day_count.keys()):
    net = day_gross[d] - day_refund[d]
    day_rows.append((d, day_gross[d], day_refund[d], net, day_count[d]))
for d, g, r, net, cnt in day_rows:
    print(f"{d:<14}{fmt(g):>14}{fmt(r):>12}{fmt(net):>14}{cnt:>8,}")

print("\n" + "=" * W)
print("C) 总计核对")
print("=" * W)
print(f"全门店净额(系统计算)   = {fmt(tot_net)}")
print(f"客户预期净额           = {fmt(EXPECTED_NET)}")
print(f"差额                   = {fmt(tot_net - EXPECTED_NET)}")
print(f"毛额                   = {fmt(tot_gross)}  (预期≈27,533)")
print(f"退款                   = {fmt(tot_refund)}  (预期≈361)")

# 一句话定位
print("\n" + "=" * W)
print("D) 一句话定位")
print("=" * W)
if store_rows:
    top3 = store_rows[:3]
    top3_net = sum(r[4] for r in top3)
    top3_share = top3_net / tot_net * 100 if tot_net else 0
    top3_str = ", ".join(f"{r[0]}({fmt(r[4])})" for r in top3)
    print(f"共 {len(store_rows)} 家店有微信交易; Top3: {top3_str}, 合计 {fmt(top3_net)} 占 {top3_share:.1f}%。")
if day_rows:
    top3_days = sorted(day_rows, key=lambda r: r[3], reverse=True)[:3]
    td_net = sum(r[3] for r in top3_days)
    td_share = td_net / tot_net * 100 if tot_net else 0
    td_str = ", ".join(f"{r[0]}({fmt(r[3])})" for r in top3_days)
    print(f"时间上 Top3 日: {td_str}, 合计 {fmt(td_net)} 占 {td_share:.1f}%。")

# ========== 4. Excel 导出 ==========
out_path, file_ver = next_version_path(OUT_DIR, OUT_BASENAME)

wb = Workbook()
wb.properties.title = f"微信支付实收按店按天 2026-04 ({INTERNAL_VERSION})"

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
HDR_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True)
RED_BOLD = Font(bold=True, color="FF0000")
MONEY = '#,##0.00'


def style_header(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


# ---- 说明块 (写在每个 sheet 顶部之前, 单独放 sheet1 顶部) ----
NOTE_LINES = [
    f"内部版本 {INTERNAL_VERSION}",
    "口径: 微信渠道 = payment_name IN ('WeChatPay','WeChat Pay','Kbank-WeChatPay') 三个原始碎片名合并",
    "净实收 = SUM(payment_amount) - SUM(refund_amount); WHERE delete_time=0 AND complete_time∈[2026-04 BKK整月)",
    "日期按 Asia/Bangkok; 来源表 ttpos_statistics_payment join ttpos_payment_method (各门店 shop{uuid} dataset)",
    f"全门店净额 = {fmt(tot_net)} (客户预期≈{fmt(EXPECTED_NET)})",
]

# Sheet1: 按店汇总
ws1 = wb.active
ws1.title = "按店汇总"
r = 1
ws1.cell(row=r, column=1, value=NOTE_LINES[0]).font = RED_BOLD
r += 1
for line in NOTE_LINES[1:]:
    ws1.cell(row=r, column=1, value=line).font = Font(size=9, color="666666")
    r += 1
r += 1
hdr_row = r
headers1 = ["门店abbr", "门店名", "毛额", "退款", "净额", "笔数"]
for ci, h in enumerate(headers1, start=1):
    ws1.cell(row=r, column=ci, value=h)
style_header(ws1, r, len(headers1))
r += 1
for abbr, name, g, rf, net, cnt in store_rows:
    ws1.cell(row=r, column=1, value=abbr)
    ws1.cell(row=r, column=2, value=name)
    ws1.cell(row=r, column=3, value=round(g, 2)).number_format = MONEY
    ws1.cell(row=r, column=4, value=round(rf, 2)).number_format = MONEY
    ws1.cell(row=r, column=5, value=round(net, 2)).number_format = MONEY
    ws1.cell(row=r, column=6, value=cnt)
    for ci in range(1, 7):
        ws1.cell(row=r, column=ci).border = BORDER
    r += 1
# 合计行
ws1.cell(row=r, column=1, value="合计")
ws1.cell(row=r, column=3, value=round(tot_gross, 2)).number_format = MONEY
ws1.cell(row=r, column=4, value=round(tot_refund, 2)).number_format = MONEY
ws1.cell(row=r, column=5, value=round(tot_net, 2)).number_format = MONEY
ws1.cell(row=r, column=6, value=tot_count)
for ci in range(1, 7):
    c = ws1.cell(row=r, column=ci)
    c.font = TOTAL_FONT
    c.fill = TOTAL_FILL
    c.border = BORDER
ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 24
for col in ['C', 'D', 'E', 'F']:
    ws1.column_dimensions[col].width = 14

# Sheet2: 按店×按天 (行=门店+日期)
ws2 = wb.create_sheet("按店x按天")
headers2 = ["门店abbr", "门店名", "日期", "净额"]
for ci, h in enumerate(headers2, start=1):
    ws2.cell(row=1, column=ci, value=h)
style_header(ws2, 1, len(headers2))
r = 2
sd_rows = sorted(store_day_net.items(), key=lambda kv: (kv[0][0], kv[0][2]))
for (abbr, name, d), net in sd_rows:
    ws2.cell(row=r, column=1, value=abbr)
    ws2.cell(row=r, column=2, value=name)
    ws2.cell(row=r, column=3, value=d)
    ws2.cell(row=r, column=4, value=round(net, 2)).number_format = MONEY
    for ci in range(1, 5):
        ws2.cell(row=r, column=ci).border = BORDER
    r += 1
ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 24
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 14

# Sheet3: 按天汇总
ws3 = wb.create_sheet("按天汇总")
headers3 = ["日期", "毛额", "退款", "净额", "笔数"]
for ci, h in enumerate(headers3, start=1):
    ws3.cell(row=1, column=ci, value=h)
style_header(ws3, 1, len(headers3))
r = 2
for d, g, rf, net, cnt in day_rows:
    ws3.cell(row=r, column=1, value=d)
    ws3.cell(row=r, column=2, value=round(g, 2)).number_format = MONEY
    ws3.cell(row=r, column=3, value=round(rf, 2)).number_format = MONEY
    ws3.cell(row=r, column=4, value=round(net, 2)).number_format = MONEY
    ws3.cell(row=r, column=5, value=cnt)
    for ci in range(1, 6):
        ws3.cell(row=r, column=ci).border = BORDER
    r += 1
ws3.cell(row=r, column=1, value="合计")
ws3.cell(row=r, column=2, value=round(tot_gross, 2)).number_format = MONEY
ws3.cell(row=r, column=3, value=round(tot_refund, 2)).number_format = MONEY
ws3.cell(row=r, column=4, value=round(tot_net, 2)).number_format = MONEY
ws3.cell(row=r, column=5, value=tot_count)
for ci in range(1, 6):
    c = ws3.cell(row=r, column=ci)
    c.font = TOTAL_FONT
    c.fill = TOTAL_FILL
    c.border = BORDER
ws3.column_dimensions['A'].width = 14
for col in ['B', 'C', 'D', 'E']:
    ws3.column_dimensions[col].width = 14

wb.save(out_path)

# ========== 5. 指纹 ==========
size = os.path.getsize(out_path)
with open(out_path, 'rb') as f:
    md5 = hashlib.md5(f.read()).hexdigest()
import datetime
mtime = datetime.datetime.fromtimestamp(os.path.getmtime(out_path)).strftime("%Y-%m-%d %H:%M:%S")

print("\n" + "=" * W)
print("E) 输出指纹")
print("=" * W)
print(f"输出:     {out_path}")
print(f"  文件版本: {file_ver}")
print(f"  内部版本: {INTERNAL_VERSION}")
print(f"  修改时间: {mtime}")
print(f"  大小:     {size} bytes")
print(f"  MD5:      {md5}")
