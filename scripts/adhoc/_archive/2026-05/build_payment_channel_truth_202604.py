#!/usr/bin/env python3
# 谁问的: 何伟涛 / 2026-05-28
# 问什么: 逐店重建2026-04渠道净实收系统真值+导出
# 结论: 按已坐实口径逐店重建,POS净实收=payment_amount-refund_amount,外卖按platform,合计见console;Excel可审计
"""逐门店重建 2026-04 全 TH 门店「系统真值」渠道净实收表并导出 Excel(可审计)。"""
import os
import sys
import hashlib
import datetime
from collections import defaultdict

sys.path.insert(0, '/home/weifashi/hwt/analysis')

from bq_reports.utils.bq_client import get_bq_client

client = get_bq_client()
BQ_LOCATION = "asia-southeast1"

START_TS = 1774976400
END_TS = 1777568400

INTERNAL_VERSION = "v1"

# ---- 客户「系统」行 (2026-04 全门店合计) ----
CLIENT = {
    'QR': 10091913.0,
    'EDC': 345480.0,
    'ALIPAY': 13872.0,
    'GRAB': 4088499.0,
    'LINEMAN': 8185896.0,
    'SHOPEE': 4916184.0,
    'RBH': 6853.7,
    'CASH': 8328934.0,
}
CLIENT_TOTAL = 35977631.7

# ---- POS payment_name -> 规范渠道 合并规则 ----
POS_NAME_TO_CHANNEL = {
    'Thai QR': 'QR',
    'Cash': 'CASH',
    'cash': 'CASH',
    'Credit Card': 'EDC',
    'Credit/Debit': 'EDC',
    'Alipay': 'ALIPAY',
    'Robinhood': 'RBH',
    'WeChatPay': 'WECHAT',
    'WeChat Pay': 'WECHAT',
    'Kbank-WeChatPay': 'WECHAT',
    'Shopee': 'SHOPEE',
    'LINEMAN': 'LINEMAN_POS',
    'LINE MAN': 'LINEMAN_POS',
    'Grab': 'GRAB_POS',
    'Grab（自行添加）': 'GRAB_POS',
}

# 规范渠道展示顺序(最终口径表)
FINAL_CHANNELS = ['QR', 'CASH', 'EDC', 'ALIPAY', 'RBH', 'WECHAT', 'SHOPEE', 'GRAB', 'LINEMAN']
# 按店明细额外列
EXTRA_COLS = ['GRAB_POS', 'OTHER']


def fmt(x):
    return f"{x:,.2f}"


# ========== 1. 门店枚举 ==========
job = client.query('''
    SELECT c.uuid, c.name, cs.erpnext_company_abbr
    FROM `diyl-407103`.`saas`.`ttpos_company` c
    LEFT JOIN `diyl-407103`.`saas`.`ttpos_company_setting` cs
      ON cs.company_uuid = c.uuid AND cs.delete_time = 0
    WHERE c.delete_time = 0
      AND cs.headquarter_uuid = 5080409448448000
      AND cs.erpnext_company_abbr LIKE 'TH%'
    ORDER BY cs.erpnext_company_abbr
''', location=BQ_LOCATION)

all_th_stores = []
for row in job.result():
    all_th_stores.append({
        'uuid': str(row.uuid),
        'name': row.name,
        'abbr': row.erpnext_company_abbr or '',
        'dataset': f"shop{row.uuid}",
    })

datasets = list(client.list_datasets())
shop_dataset_ids = set(d.dataset_id for d in datasets if d.dataset_id.startswith('shop'))
stores = [s for s in all_th_stores if s['dataset'] in shop_dataset_ids]

print(f"TH 门店总数(配置): {len(all_th_stores)}, 实际存在 shop* dataset: {len(stores)}")

# ========== 2. 逐店查询 ==========
# 全门店合计
pos_name_agg = defaultdict(lambda: {'gross': 0.0, 'refund': 0.0, 'count': 0})  # 原始 payment_name 分布
# 按店: store_key -> channel -> net
store_channel = defaultdict(lambda: defaultdict(float))
# 全门店规范渠道净实收(POS部分,合并后)
pos_channel_net = defaultdict(float)
# 外卖
to_grab = 0.0
to_lineman = 0.0
store_to = defaultdict(lambda: {'GRAB_TO': 0.0, 'LINEMAN_TO': 0.0})

missing_pos = []   # 缺 statistics_payment 表的店 (abbr)
missing_to = []

pos_ok = to_ok = 0

for s in stores:
    dataset = s['dataset']
    skey = f"{s['abbr']} | {s['name']}"

    # --- POS 净实收: payment_amount - refund_amount, 按 payment_name ---
    try:
        job = client.query(f"""
            SELECT IFNULL(pm.payment_name, '(NULL)') AS method_name,
                COUNT(*) AS bill_cnt,
                SUM(sp.payment_amount) AS gross_amount,
                SUM(sp.refund_amount) AS refund_amount
            FROM `diyl-407103`.`{dataset}`.`ttpos_statistics_payment` sp
            LEFT JOIN `diyl-407103`.`{dataset}`.`ttpos_payment_method` pm
                ON pm.uuid = sp.payment_method_uuid
            WHERE sp.delete_time = 0
                AND sp.complete_time >= {START_TS}
                AND sp.complete_time < {END_TS}
            GROUP BY method_name
        """, location=BQ_LOCATION)
        for row in job.result():
            name = row.method_name
            gross = float(row.gross_amount) if row.gross_amount else 0.0
            refund = float(row.refund_amount) if row.refund_amount else 0.0
            cnt = int(row.bill_cnt) if row.bill_cnt else 0
            net = gross - refund

            pos_name_agg[name]['gross'] += gross
            pos_name_agg[name]['refund'] += refund
            pos_name_agg[name]['count'] += cnt

            channel = POS_NAME_TO_CHANNEL.get(name, 'OTHER')
            pos_channel_net[channel] += net
            store_channel[skey][channel] += net
        pos_ok += 1
    except Exception:
        missing_pos.append(s['abbr'])

    # --- 外卖: platform_total, platform=grab/lineman ---
    try:
        job = client.query(f"""
            SELECT LOWER(IFNULL(platform, '(NULL)')) AS platform,
                SUM(platform_total) AS total_amount
            FROM `diyl-407103`.`{dataset}`.`ttpos_takeout_order`
            WHERE delete_time = 0
                AND order_state IN (10, 20, 30, 40)
                AND accepted_time > 0
                AND (CASE WHEN order_state = 40 AND completed_time > 0 THEN completed_time
                    ELSE accepted_time END) >= {START_TS}
                AND (CASE WHEN order_state = 40 AND completed_time > 0 THEN completed_time
                    ELSE accepted_time END) < {END_TS}
            GROUP BY platform
        """, location=BQ_LOCATION)
        for row in job.result():
            p = row.platform
            amt = float(row.total_amount) if row.total_amount else 0.0
            if p == 'grab':
                to_grab += amt
                store_to[skey]['GRAB_TO'] += amt
            elif p == 'lineman':
                to_lineman += amt
                store_to[skey]['LINEMAN_TO'] += amt
        to_ok += 1
    except Exception:
        missing_to.append(s['abbr'])

print(f"POS 查询成功 {pos_ok} 店, 缺表 {len(missing_pos)} 店")
print(f"外卖查询成功 {to_ok} 店, 缺表 {len(missing_to)} 店")

# ========== 3. 组装规范渠道最终口径(全门店合计) ==========
# QR/CASH/EDC/ALIPAY/RBH/WECHAT/SHOPEE = 各自 POS 净实收
# GRAB = 外卖 GRAB_TO
# LINEMAN = POS LINEMAN_POS 净 + 外卖 LINEMAN_TO
final = {}
final['QR'] = pos_channel_net.get('QR', 0.0)
final['CASH'] = pos_channel_net.get('CASH', 0.0)
final['EDC'] = pos_channel_net.get('EDC', 0.0)
final['ALIPAY'] = pos_channel_net.get('ALIPAY', 0.0)
final['RBH'] = pos_channel_net.get('RBH', 0.0)
final['WECHAT'] = pos_channel_net.get('WECHAT', 0.0)
final['SHOPEE'] = pos_channel_net.get('SHOPEE', 0.0)
final['GRAB'] = to_grab
final['LINEMAN'] = pos_channel_net.get('LINEMAN_POS', 0.0) + to_lineman

grab_pos_net = pos_channel_net.get('GRAB_POS', 0.0)
other_net = pos_channel_net.get('OTHER', 0.0)
lineman_pos_net = pos_channel_net.get('LINEMAN_POS', 0.0)

CHANNEL_SOURCE = {
    'QR':      'POS净实收 (Thai QR)',
    'CASH':    'POS净实收 (Cash/cash)',
    'EDC':     'POS净实收 (Credit Card/Credit/Debit)',
    'ALIPAY':  'POS净实收 (Alipay)',
    'RBH':     'POS净实收 (Robinhood)',
    'WECHAT':  'POS净实收 (WeChatPay/WeChat Pay/Kbank-WeChatPay)',
    'SHOPEE':  'POS净实收 (Shopee)',
    'GRAB':    '外卖 platform=grab (platform_total)',
    'LINEMAN': f'POS净(LINEMAN={fmt(lineman_pos_net)}) + 外卖lineman({fmt(to_lineman)})',
}

final_total = sum(final.values())

# ========== 4. Console 输出 ==========
print("\n" + "=" * 86)
print(f"规范渠道净实收汇总 (全门店合计 / 内部版本 {INTERNAL_VERSION})")
print("=" * 86)
print(f"{'渠道':<10}{'净实收金额':>18}   来源说明")
print("-" * 86)
for ch in FINAL_CHANNELS:
    print(f"{ch:<10}{fmt(final[ch]):>18}   {CHANNEL_SOURCE[ch]}")
print("-" * 86)
print(f"{'合计':<10}{fmt(final_total):>18}")
print()
print(f"[旁注] GRAB_POS (POS零星误录,不计入GRAB) = {fmt(grab_pos_net)}")
print(f"[旁注] 其他/未分类 OTHER 净实收        = {fmt(other_net)}")
if other_net != 0.0:
    print("       未分类原始 payment_name:")
    for name, v in sorted(pos_name_agg.items(), key=lambda kv: -(kv[1]['gross'] - kv[1]['refund'])):
        if POS_NAME_TO_CHANNEL.get(name, 'OTHER') == 'OTHER':
            net = v['gross'] - v['refund']
            print(f"         {name!r:<30} 净={fmt(net):>16} 笔={v['count']:>8,}")

# 对照客户系统
print("\n" + "=" * 86)
print("跟客户系统行对照 (我方净真值 vs 客户系统 vs 差异)")
print("=" * 86)
print(f"{'渠道':<10}{'我方净真值':>18}{'客户系统':>18}{'差异(我-客)':>18}")
print("-" * 86)
# 客户无 WECHAT, 对照时 WECHAT 客户值视为 0
cmp_total_ours = 0.0
cmp_total_client = 0.0
for ch in FINAL_CHANNELS:
    ours = final[ch]
    cust = CLIENT.get(ch, 0.0)
    diff = ours - cust
    cust_disp = fmt(cust) if ch in CLIENT else "(无)"
    print(f"{ch:<10}{fmt(ours):>18}{cust_disp:>18}{fmt(diff):>18}")
    cmp_total_ours += ours
    cmp_total_client += cust
print("-" * 86)
print(f"{'合计':<10}{fmt(cmp_total_ours):>18}{fmt(CLIENT_TOTAL):>18}{fmt(cmp_total_ours - CLIENT_TOTAL):>18}")
print(f"(客户合计基准 {fmt(CLIENT_TOTAL)}; 客户无 WeChat 行)")

if missing_pos:
    print(f"\n缺 ttpos_statistics_payment 表的门店 ({len(missing_pos)}): {', '.join(missing_pos)}")

# ========== 5. Excel 导出 ==========
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

EXPORT_DIR = "/home/weifashi/hwt/analysis/exports"
os.makedirs(EXPORT_DIR, exist_ok=True)
base = os.path.join(EXPORT_DIR, "payment_channel_truth_2026-04_v1.xlsx")
out_path = base
ver = 1
while os.path.exists(out_path):
    ver += 1
    out_path = os.path.join(EXPORT_DIR, f"payment_channel_truth_2026-04_v{ver}.xlsx")

wb = Workbook()

bold = Font(bold=True)
red_bold = Font(bold=True, color="CC0000")
hdr_fill = PatternFill("solid", fgColor="DDEBF7")
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
right = Alignment(horizontal="right")


def style_header(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = bold
        cell.fill = hdr_fill
        cell.border = border


# ---- Sheet1: 渠道汇总 ----
ws1 = wb.active
ws1.title = "渠道汇总"
r = 1
ws1.cell(row=r, column=1, value=f"内部版本号 {INTERNAL_VERSION}").font = red_bold
r += 1
ws1.cell(row=r, column=1,
         value="口径: POS净实收=SUM(payment_amount)-SUM(refund_amount)[delete_time=0,complete_time∈窗]; "
               "外卖=SUM(platform_total)[order_state∈(10,20,30,40),accepted_time>0,时间用CASE]")
r += 1
ws1.cell(row=r, column=1,
         value="GRAB=外卖grab; LINEMAN=POS LINEMAN净+外卖lineman; GRAB_POS=POS零星误录(不计入GRAB); 时间窗=2026-04 BKK整月")
r += 2

hrow = r
heads = ["规范渠道", "我方净真值", "客户系统", "差异(我-客)", "来源说明"]
for i, h in enumerate(heads, 1):
    ws1.cell(row=hrow, column=i, value=h)
style_header(ws1, hrow, len(heads))
r += 1
for ch in FINAL_CHANNELS:
    ours = final[ch]
    cust = CLIENT.get(ch, None)
    ws1.cell(row=r, column=1, value=ch)
    ws1.cell(row=r, column=2, value=round(ours, 2)).alignment = right
    ws1.cell(row=r, column=3, value=(round(cust, 2) if cust is not None else "(无)")).alignment = right
    ws1.cell(row=r, column=4, value=(round(ours - (cust or 0.0), 2))).alignment = right
    ws1.cell(row=r, column=5, value=CHANNEL_SOURCE[ch])
    for c in range(1, 6):
        ws1.cell(row=r, column=c).border = border
    r += 1
# 合计行
ws1.cell(row=r, column=1, value="合计").font = bold
ws1.cell(row=r, column=2, value=round(final_total, 2)).font = bold
ws1.cell(row=r, column=3, value=round(CLIENT_TOTAL, 2)).font = bold
ws1.cell(row=r, column=4, value=round(final_total - CLIENT_TOTAL, 2)).font = bold
for c in range(1, 6):
    ws1.cell(row=r, column=c).border = border
r += 2
ws1.cell(row=r, column=1, value=f"旁注 GRAB_POS (POS零星误录,不计入GRAB) = {fmt(grab_pos_net)}")
r += 1
ws1.cell(row=r, column=1, value=f"旁注 其他/未分类 OTHER 净实收 = {fmt(other_net)}")
r += 1
ws1.cell(row=r, column=1, value=f"客户系统无 WeChat 行; 客户合计基准 = {fmt(CLIENT_TOTAL)}")
r += 2
ws1.cell(row=r, column=1, value=f"缺 ttpos_statistics_payment 表门店 ({len(missing_pos)}):").font = bold
r += 1
ws1.cell(row=r, column=1, value=(', '.join(missing_pos) if missing_pos else "(无)"))

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 60

# ---- Sheet2: 按店明细 ----
ws2 = wb.create_sheet("按店明细")
# 列: 门店 + FINAL_CHANNELS + GRAB_POS + OTHER
detail_cols = ["门店(abbr | name)"] + FINAL_CHANNELS + EXTRA_COLS
for i, h in enumerate(detail_cols, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(detail_cols))

# 每店逐渠道净值: POS渠道直接取, GRAB店内=外卖grab, LINEMAN店内=POS LINEMAN_POS + 外卖lineman
all_store_keys = sorted(set(list(store_channel.keys()) + list(store_to.keys())))
rr = 2
col_totals = defaultdict(float)
for skey in all_store_keys:
    sc = store_channel.get(skey, {})
    st = store_to.get(skey, {'GRAB_TO': 0.0, 'LINEMAN_TO': 0.0})
    vals = {
        'QR': sc.get('QR', 0.0),
        'CASH': sc.get('CASH', 0.0),
        'EDC': sc.get('EDC', 0.0),
        'ALIPAY': sc.get('ALIPAY', 0.0),
        'RBH': sc.get('RBH', 0.0),
        'WECHAT': sc.get('WECHAT', 0.0),
        'SHOPEE': sc.get('SHOPEE', 0.0),
        'GRAB': st['GRAB_TO'],
        'LINEMAN': sc.get('LINEMAN_POS', 0.0) + st['LINEMAN_TO'],
        'GRAB_POS': sc.get('GRAB_POS', 0.0),
        'OTHER': sc.get('OTHER', 0.0),
    }
    ws2.cell(row=rr, column=1, value=skey).border = border
    for i, ch in enumerate(FINAL_CHANNELS + EXTRA_COLS, 2):
        v = round(vals[ch], 2)
        cell = ws2.cell(row=rr, column=i, value=v)
        cell.alignment = right
        cell.border = border
        col_totals[ch] += vals[ch]
    rr += 1
# 合计行
ws2.cell(row=rr, column=1, value="合计").font = bold
for i, ch in enumerate(FINAL_CHANNELS + EXTRA_COLS, 2):
    cell = ws2.cell(row=rr, column=i, value=round(col_totals[ch], 2))
    cell.font = bold
    cell.alignment = right
    cell.border = border
ws2.cell(row=rr, column=1).border = border

ws2.column_dimensions['A'].width = 42
for i in range(2, len(detail_cols) + 1):
    ws2.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = 14

# ---- Sheet3: payment_name 原始分布 ----
ws3 = wb.create_sheet("payment_name原始分布")
heads3 = ["原始 payment_name", "归到渠道", "毛额", "退款额", "净额", "笔数"]
for i, h in enumerate(heads3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(heads3))
rr = 2
tg = tr = tn = tc = 0.0
for name, v in sorted(pos_name_agg.items(), key=lambda kv: -(kv[1]['gross'] - kv[1]['refund'])):
    net = v['gross'] - v['refund']
    ch = POS_NAME_TO_CHANNEL.get(name, 'OTHER')
    ws3.cell(row=rr, column=1, value=name).border = border
    ws3.cell(row=rr, column=2, value=ch).border = border
    ws3.cell(row=rr, column=3, value=round(v['gross'], 2)).border = border
    ws3.cell(row=rr, column=4, value=round(v['refund'], 2)).border = border
    ws3.cell(row=rr, column=5, value=round(net, 2)).border = border
    ws3.cell(row=rr, column=6, value=v['count']).border = border
    for c in (3, 4, 5, 6):
        ws3.cell(row=rr, column=c).alignment = right
    tg += v['gross']; tr += v['refund']; tn += net; tc += v['count']
    rr += 1
ws3.cell(row=rr, column=1, value="合计").font = bold
ws3.cell(row=rr, column=3, value=round(tg, 2)).font = bold
ws3.cell(row=rr, column=4, value=round(tr, 2)).font = bold
ws3.cell(row=rr, column=5, value=round(tn, 2)).font = bold
ws3.cell(row=rr, column=6, value=int(tc)).font = bold
for c in range(1, 7):
    ws3.cell(row=rr, column=c).border = border
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 14
for col in ('C', 'D', 'E', 'F'):
    ws3.column_dimensions[col].width = 16

wb.save(out_path)

# ========== 6. 指纹 ==========
stat = os.stat(out_path)
with open(out_path, 'rb') as f:
    md5 = hashlib.md5(f.read()).hexdigest()
mtime = datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")

print("\n" + "=" * 86)
print("指纹")
print("=" * 86)
print(f"输出路径   : {out_path}")
print(f"内部版本号 : {INTERNAL_VERSION}")
print(f"修改时间   : {mtime}")
print(f"大小       : {stat.st_size:,} bytes")
print(f"MD5        : {md5}")
