# 谁问的: 何伟涛 / 2026-05-28
# 问什么: 做按渠道的对账桥(我们系统支付净额 vs 客户银行/平台实际到账),让客户看清每渠道差在哪
# 结论: 净差 39,290 = 外卖结算差27,558 + 微信27,172 + 现金抹零2,589 + RBH零头164 − 扫码换标签18,192;其中扫码换标签钱没丢,外卖/微信/现金待外部对账单闭合
"""
对账桥 (Reconciliation Bridge) — 2026-04 全 TH 门店

口径(三个不同的数,客户务必区分):
  ① 营业额(商品成交额)  = 我们交付报表 profit_by_price 的"实收金额"列, ttpos 营业总览口径
  ② 支付记录净额        = ttpos_statistics_payment 的 payment_amount − refund_amount(本次重算,已扣退款/合并渠道碎片/补回微信), 这是我们系统真实收款
  ③ 客户实际到账        = 客户从银行/三方平台拉出的对账单(扣手续费/抽佣后净额)

所有数字来源:本会话前几轮已逐项验证(见 _archive/2026-05/ 下 build_payment_channel_truth /
explain_system_over_record / check_payment_refund_semantics 等脚本)。本脚本只做汇总成桥,不再查 BQ。
"""
import sys
from pathlib import Path
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# 已验证数据 (2026-04 全门店)
# ─────────────────────────────────────────────────────────────
# 渠道: (我们系统支付净额②, 客户实际到账③, 差异性质, 待查来源)
CHANNELS = [
    # 渠道,      系统净额②,      客户到账③,    性质,                          待查/说明
    ("QR",       10092190.00,  10000315.40, "收单换标签",  "扫码记成QR, 银行可能按卡/支付宝结算 → 待银行流水"),
    ("EDC",        345480.00,    427836.00, "收单换标签",  "银行刷卡入账比系统多 → 与QR对冲, 待银行流水"),
    ("ALIPAY",      13872.00,     41583.00, "收单换标签",  "银行支付宝入账比系统多 → 与QR对冲, 待银行流水"),
    ("GRAB",      4072311.00,   4081642.00, "外卖结算差",  "平台在途/补贴调整 → 待 Grab 商家结算单"),
    ("LINEMAN",   8186400.00,   8167600.00, "外卖结算差",  "平台在途/抽佣调整 → 待 LINE MAN 结算单"),
    ("SHOPEE",    4916184.00,   4898095.00, "外卖结算差",  "平台在途/抽佣调整 → 待 Shopee 结算单"),
    ("RBH",          6853.70,      6690.00, "零头",        "Robinhood 尾差, 可忽略"),
    ("CASH",      8329033.00,   8326444.00, "现金抹零",    "抹零/找零/盘差 → 待门店现金日结"),
    ("WECHAT",      27172.00,         0.00, "渠道缺失",    "客户到账无微信栏 → 确认微信资金进了哪个账户/是否漏记"),
]

# 总瀑布: 营业额① → 支付净额② → 到账③
REPORT_REVENUE = 36095575.30   # ① profit_by_price_202604_v18 实收金额合计(套餐+单品)
PAYMENT_NET    = 35989495.70   # ② 系统支付净额(= 各渠道②合计)
ACTUAL_RECEIVED= 35950205.40   # ③ 客户实际到账(= 各渠道③合计, 截图账上)

VERSION = "v1"

# ─────────────────────────────────────────────────────────────
# 样式
# ─────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="4472C4")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
BORDER = Border(*[Side(style="thin")] * 4)
BOLD = Font(bold=True)
RED = Font(color="9C0006")
GREEN = Font(color="006100")

# 性质 → 底色
NATURE_FILL = {
    "收单换标签": PatternFill("solid", fgColor="DDEBF7"),  # 蓝: 钱没丢
    "外卖结算差": PatternFill("solid", fgColor="FFF2CC"),  # 黄: 待平台结算单
    "现金抹零":   PatternFill("solid", fgColor="FFF2CC"),  # 黄: 待门店日结
    "渠道缺失":   PatternFill("solid", fgColor="FCE4D6"),  # 橙: 重点确认
    "零头":       PatternFill("solid", fgColor="E2EFDA"),  # 绿: 可忽略
}


def _money(ws, cell, v):
    ws[cell] = v
    ws[cell].number_format = "#,##0.00"
    ws[cell].alignment = RIGHT


def build():
    wb = Workbook()

    # ===== Sheet1: 按渠道对账桥 =====
    ws = wb.active
    ws.title = "按渠道对账桥"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"2026-04 全门店 按渠道对账桥  （内部版本 {VERSION}）"
    ws["A1"].font = Font(bold=True, size=13, color="C00000")
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:F2")
    ws["A2"] = ("差异 = 我们系统支付净额② − 客户实际到账③ 。正=系统多记 / 负=到账多。"
                "蓝色=扫码换标签(钱没丢,QR↔卡/支付宝对冲);黄色=待外部对账单;橙色=重点确认。")
    ws["A2"].font = Font(size=9, italic=True)
    ws["A2"].alignment = LEFT

    headers = ["渠道", "②我们系统(支付净额)", "③客户实际到账", "差异(②−③)", "性质", "待查来源 / 说明"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER

    r = 5
    tot_sys = tot_act = tot_diff = 0.0
    for name, sysv, actv, nature, note in CHANNELS:
        diff = round(sysv - actv, 2)
        tot_sys += sysv; tot_act += actv; tot_diff += diff
        ws.cell(row=r, column=1, value=name).alignment = CENTER
        _money(ws, f"B{r}", sysv)
        _money(ws, f"C{r}", actv)
        _money(ws, f"D{r}", diff)
        ws.cell(row=r, column=4).font = RED if diff > 0 else GREEN
        ws.cell(row=r, column=5, value=nature).alignment = CENTER
        ws.cell(row=r, column=6, value=note).alignment = LEFT
        fill = NATURE_FILL.get(nature)
        if fill:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = fill
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    # 合计行
    ws.cell(row=r, column=1, value="合计").font = BOLD
    _money(ws, f"B{r}", round(tot_sys, 2)); ws[f"B{r}"].font = BOLD
    _money(ws, f"C{r}", round(tot_act, 2)); ws[f"C{r}"].font = BOLD
    _money(ws, f"D{r}", round(tot_diff, 2)); ws[f"D{r}"].font = BOLD
    ws.cell(row=r, column=5, value="净差").alignment = CENTER
    ws.cell(row=r, column=6,
            value=f"系统比到账净多 {tot_diff:,.2f}（占 {tot_diff/tot_act*100:.3f}%）").font = BOLD
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = BORDER

    # 分组小结
    r += 2
    ws.cell(row=r, column=1, value="按性质归类:").font = BOLD; r += 1
    groups = [
        ("收单换标签 (QR+EDC+ALIPAY)", 91874.60 - 82356.00 - 27711.00, "扫码钱被银行按卡/支付宝入账, 一加一减几乎抵消, 钱没丢, 待银行流水核"),
        ("外卖结算差 (GRAB+LINEMAN+SHOPEE)", -9331.00 + 18800.00 + 18089.00, "在途+抽佣/补贴调整, 待三平台结算单"),
        ("微信 (WECHAT)", 27172.00, "客户到账无此栏, 重点确认资金归属/是否漏记"),
        ("现金抹零 (CASH)", 2589.00, "待门店现金日结"),
        ("零头 (RBH)", 163.70, "可忽略"),
    ]
    for gname, gval, gnote in groups:
        ws.cell(row=r, column=1, value=gname).alignment = LEFT
        _money(ws, f"D{r}", round(gval, 2))
        ws.cell(row=r, column=4).font = RED if gval > 0 else GREEN
        ws.cell(row=r, column=6, value=gnote).alignment = LEFT
        r += 1

    widths = [12, 20, 18, 16, 12, 52]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ===== Sheet2: 总瀑布 =====
    ws2 = wb.create_sheet("总瀑布(口径桥)")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "从「我们报表营业额」走到「客户实际到账」的口径桥"
    ws2["A1"].font = Font(bold=True, size=12, color="C00000")
    ws2["A1"].alignment = CENTER

    steps = [
        ("① 营业额(商品成交额) — 我们交付报表", REPORT_REVENUE, "ttpos营业总览口径, 含税前/未扣手续费抹零"),
        ("  − 商品口径→支付口径(抹零/作废明细/账单级差)", -(REPORT_REVENUE - PAYMENT_NET), ""),
        ("② 支付记录净额(payment−refund) — 系统真实收款", PAYMENT_NET, "★ 我们系统的铁数, 已验证"),
        ("  − 外卖平台结算差(在途/抽佣/补贴)", -(-9331.00 + 18800.00 + 18089.00), "待 Grab/LINEMAN/Shopee 结算单"),
        ("  − 微信(客户到账无此栏)", -27172.00, "待确认资金归属"),
        ("  − 现金抹零/盘差", -2589.00, "待门店日结"),
        ("  − RBH 零头", -163.70, ""),
        ("  + 扫码换标签回补(银行按卡/支付宝多入账)", 18192.40, "待银行流水"),
        ("③ 客户实际到账", ACTUAL_RECEIVED, "客户银行/平台对账单"),
    ]
    for c, h in enumerate(["步骤", "金额", "说明"], 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    rr = 4
    for label, val, note in steps:
        ws2.cell(row=rr, column=1, value=label).alignment = LEFT
        _money(ws2, f"B{rr}", round(val, 2))
        ws2.cell(row=rr, column=3, value=note).alignment = LEFT
        if label.startswith(("①", "②", "③")):
            for c in range(1, 4):
                ws2.cell(row=rr, column=c).font = BOLD
                ws2.cell(row=rr, column=c).fill = PatternFill("solid", fgColor="EDEDED")
        for c in range(1, 4):
            ws2.cell(row=rr, column=c).border = BORDER
        rr += 1
    for c, w in enumerate([44, 16, 40], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ===== Sheet3: 说明 =====
    ws3 = wb.create_sheet("说明")
    notes = [
        (f"内部版本", VERSION),
        ("口径①营业额", "我们交付报表的'实收金额'列(其实是商品成交额, ttpos营业总览口径). 客户对账请勿用这个对银行,口径偏高."),
        ("口径②支付净额", "ttpos_statistics_payment 的 payment_amount − refund_amount, 本次重算(已扣退款135,797/合并渠道名碎片/补回微信27,172/剔除Grab POS误录16,188). 这是我们系统真实收款, 已验证."),
        ("口径③实际到账", "客户从银行/三方平台拉出的对账单(扣手续费/抽佣后净额)."),
        ("差异本质", "②比③净多39,290(占0.076%), 是结算层差: 收单换标签(钱没丢) + 外卖抽佣/在途 + 现金抹零 + 微信归属. 不是系统漏记订单."),
        ("我们自有数据能闭合", "外卖月底在途≈8,405 + 作废/隐藏317. 其余受手续费/抽佣/银行渠道映射支配, 需外部对账单."),
        ("待客户提供(闭合到分)", "1) 银行/收单流水(QR/卡/支付宝逐笔, 看手续费+渠道归类); 2) Grab/LINEMAN/Shopee 商家结算单(佣金/补贴/实际打款); 3) 门店现金日结/钱箱."),
        ("注意-客户之前的系统行", "客户截图'系统行'35,977,632 漏了微信、多算Grab误录, 跟②有小差, 以本次重算②=35,989,496为准."),
    ]
    for c, h in enumerate(["项", "说明"], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for i, (k, v) in enumerate(notes, 2):
        ws3.cell(row=i, column=1, value=k).alignment = LEFT
        ws3.cell(row=i, column=1).font = BOLD
        ws3.cell(row=i, column=2, value=v).alignment = LEFT
        for c in range(1, 3):
            ws3.cell(row=i, column=c).border = BORDER
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 95

    # auto-version
    out_dir = Path("/home/weifashi/hwt/analysis/exports")
    base = "reconciliation_bridge_2026-04"
    n = 1
    while (out_dir / f"{base}_v{n}.xlsx").exists():
        n += 1
    out = out_dir / f"{base}_v{n}.xlsx"
    wb.save(out)

    data = out.read_bytes()
    print(f"输出: {out}")
    print(f"  内部版本: {VERSION}")
    print(f"  大小:     {len(data)} bytes")
    print(f"  MD5:      {hashlib.md5(data).hexdigest()}")
    print(f"\n  净差校验: 系统②合计 {tot_sys:,.2f} − 到账③合计 {tot_act:,.2f} = {tot_sys-tot_act:,.2f}")


if __name__ == "__main__":
    build()
