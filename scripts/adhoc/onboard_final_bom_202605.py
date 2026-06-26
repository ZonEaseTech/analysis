#!/usr/bin/env python3
"""接入 2026-05-14 市场"最终 BOM和价格"事实表 — thin wrapper.

核心 diff 逻辑已提取到 bq_reports/shared/onboard_helper.py。
本脚本保留 xlsx 加载 + POS alias + CSV 写出业务逻辑。

一次性脚本, 跑完归档。
"""
from __future__ import annotations
import csv, os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
import openpyxl
from bq_reports.shared.onboard_helper import build_onboard_diff

RAW = "resources/wallace.20260514/最终BOM和价格_原始.xlsx"
OUT_BOM = "resources/wallace.20260514/最终BOM_202605.csv"
OUT_PRICE = "resources/wallace.20260514/最终价格_202605.csv"
DIFF_XLSX = "exports/diff_最终BOM价格_vs_现有.xlsx"

PRODUCT_ALIASES = {
    "辣番茄2": "辣番茄套餐 2",
    "合艾狂热1": "合艾狂热 1",
    "合艾狂热2": "合艾狂热 2",
    "新店开业套餐1": "新店开业促销，满意十足，10%特别折扣",
    "冰淇淋 0泰铢 - Google评论": "冰淇淋 0฿",
    "套餐满足1经典  旧": "套餐满足1经典 旧",
}

def load_new():
    wb = openpyxl.load_workbook(RAW, data_only=True)
    bom = {}
    for sn in ("单品", "套餐"):
        ws = wb[sn]
        cur = None
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not any(c is not None for c in r):
                continue
            if r[0]:
                cur = str(r[0]).strip()
                bom.setdefault(cur, [])
            if cur and r[2]:
                code = str(r[2]).strip()
                if code in {x[0] for x in bom[cur]}:
                    continue
                bom[cur].append((code, str(r[1] or "").strip(),
                                 float(r[3] or 0), str(r[4] or "").strip()))
    price = {}
    ws = wb["价格"]
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r[1]:
            continue
        code = str(r[1]).strip()
        if code and code not in price:
            price[code] = (str(r[0] or "").strip(), r[2], str(r[3] or "").strip())
    for alias, target in PRODUCT_ALIASES.items():
        if target in bom and alias not in bom:
            bom[alias] = list(bom[target])
    return bom, price

def write_csvs(bom, price):
    with open(OUT_BOM, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["商品名称", "物品编号", "物品名称", "单耗", "单位"])
        for prod, recs in bom.items():
            for code, name, qty, unit in recs:
                w.writerow([prod, code, name, qty, unit])
    with open(OUT_PRICE, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["物品编号", "物品名称", "物料单价", "单位"])
        for code, (name, p, unit) in price.items():
            w.writerow([code, name, p, unit])
    print(f"清洗输出: {OUT_BOM} ({sum(len(v) for v in bom.values())} 行 / {len(bom)} 商品)")
    print(f"          {OUT_PRICE} ({len(price)} code)")

def main():
    bom, price = load_new()
    write_csvs(bom, price)
    build_onboard_diff(bom, price, DIFF_XLSX)

if __name__ == "__main__":
    main()
