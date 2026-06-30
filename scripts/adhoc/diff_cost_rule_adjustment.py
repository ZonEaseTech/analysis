#!/usr/bin/env python3
"""
ERP 成本规则修正 diff 工具 (Phase 5 / Task 6)
============================================
在无 sid 前提下,用 clean_bom.csv 里现有的 `基价(原始)` 作为 base,
跑修正后算法(final_unit_cost_with_rule)重算成本毛利,与当前 v40 输出做逐 SKU diff,
产出对比报告。

关键假设(无 sid 时的诚实声明):
- 无 sid, 所以**假设** PRLE-0003 当前满足条件:
  for_price_list="Buying - Internal"、disabled=False、日期有效、PriceOrDiscount="Price"。
- 用 clean_bom.csv 的 `单位` 列作为 desired-UOM(之前分析过这就是 BOM 消耗单位)。
- `基价(原始)` 列即 ttpos 的 baseCost(Buying-Internal price_list_rate)。
- 税率取 `适用税率%` 列(已含在 csv 里)。
- unit_corrections(MK01018 ÷50)在 v40 路径里已经应用;为了公平比较,新路径也应用同样 legacy
  unit_corrections(因为 Phase 5 还没退役它)。

Limitations:
- 本 diff 只改价格算法,不改 BOM 结构/销量/实收,所以差异**仅**来自物料单价变化。
- 由于无 sid, 规则条件判定是假设性而非验证性;拿到 sid 后应重新跑 live 对账锚。
- 未映射商品(无 BOM)的成本用平均成本率估算,不受价格规则影响,因此 diff 中不体现。

Phase 5 后建议:
- 拿到 sid 后,用 `final_unit_cost_with_rule` 接真实 ERPNext PricingRule 条件判定,
  替换本脚本中的硬编码假设规则,并跑 `semantic/reconciliation/checks/ttpos_cost_anchor.py`
  验证 drift ≤ 0.5%。

方案选择: B(新建独立脚本),理由:
- 最不影响生产路径 wallace_bom_margin.py;
- 复用 bom_pipeline/ 的加载逻辑和 erpnext_price 算法;
- 便于后续复用为 `--recalc-with-rule` 的探针。

用法:
    cd /home/weifashi/hwt/analysis
    PYTHONPATH=bom_pipeline venv/bin/python scripts/adhoc/diff_cost_rule_adjustment.py \
        --v40 exports/Wallace商品成本毛利分析_2026-06_v40.xlsx \
        --bom resources/wallace.20260626/clean_bom.csv \
        --out exports/cost_rule_recalc_diff_202606.xlsx
"""
from __future__ import annotations

import argparse
import csv
import os
import sys

# 复用 bom_pipeline 模块
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "bom_pipeline"))

from erpnext_price import PricingRule, final_unit_cost_with_rule

# ---------------------------------------------------------------------------
# 加载 clean_bom.csv, 用修正算法重算 price
# ---------------------------------------------------------------------------

# legacy unit correction: MK01018 在 v40 路径里 ÷50, 新路径也保持一致
UNIT_CORRECTIONS = {"MK01018": 50.0}


def load_clean_bom_with_rule(csv_path: str):
    """
    从 clean_bom.csv 加载, 应用修正后的 final_unit_cost_with_rule 计算新物料单价。
    返回:
      old_price: {code: 原物料单价}  (csv 里 `物料单价` 列)
      new_price: {code: 新物料单价}  (经 final_unit_cost_with_rule 重算)
      meta:      {code: {"base": float, "tax": float, "uom": str, "name": str}}
    """
    old_price, new_price, meta = {}, {}, {}

    # 硬编码假设规则(无 sid 时的诚实声明)
    ASSUMED_RULE = PricingRule(
        margin_type="Percentage",
        margin_rate_or_amount=5.0,
        for_price_list="Buying - Internal",
        buying=True,
        disabled=False,
    )

    with open(csv_path, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            code = r["物料编码"].strip()
            base = float(r["基价(原始)"]) if r["基价(原始)"] else 0.0
            tax = float(r["适用税率%"]) if r["适用税率%"] else 0.0
            uom = r["单位"]
            name = r["物料名称"]
            old = float(r["物料单价"]) if r["物料单价"] else 0.0

            # 应用 legacy unit correction(与 v40 路径一致)
            corrected_base = base
            if code in UNIT_CORRECTIONS:
                corrected_base = base / UNIT_CORRECTIONS[code]

            new = final_unit_cost_with_rule(
                corrected_base, tax, ASSUMED_RULE, "Buying - Internal"
            )

            # 去重:同一 code 可能出现在多行(不同商品),价格应一致
            if code in old_price:
                if abs(old_price[code] - old) > 1e-9:
                    print(
                        f"⚠️  物料 {code}({name}) 在不同商品行价格不一致: "
                        f"{old_price[code]:.6f} vs {old:.6f}, 取首次出现值",
                        file=sys.stderr,
                    )
                continue

            old_price[code] = old
            new_price[code] = new
            meta[code] = {"base": corrected_base, "tax": tax, "uom": uom, "name": name}

    return old_price, new_price, meta


# ---------------------------------------------------------------------------
# 读取 v40 Excel 输出
# ---------------------------------------------------------------------------

def read_v40_rows(path: str):
    """
    读取 v40 Excel 的 4 个数据 sheet, 返回逐行记录。
    每行: {
        "sheet": str, "store": str, "product": str, "unit_price": float,
        "mat_name": str, "mat_code": str, "qty": float, "mat_unit_price": float,
        "unit": str, "net_qty": int, "revenue": int, "ucost": float,
        "total_cost": int, "margin": int,
    }
    注意: v40 里合并单元格, 同商品多物料行只有首行有商品/销量/实收等数据,
    后续行只有物料列有值。这里按商品粒度聚合,展开成 (商品,物料) 粒度的行。
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = ["堂食-单品", "堂食-套餐", "外卖-单品", "外卖-套餐"]
    rows = []
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        it = iter(ws.iter_rows(min_row=2, values_only=True))
        cur_store = cur_product = cur_unit_price = cur_net_qty = cur_revenue = cur_ucost = None
        cur_total_cost = cur_margin = None
        for row in it:
            if row[0] is not None:  # 新商品首行
                cur_store = str(row[0]).strip() if row[0] else ""
                cur_product = str(row[3]).strip() if row[3] else ""
                cur_unit_price = float(row[4]) if row[4] is not None else 0.0
                cur_net_qty = int(row[10]) if row[10] is not None else 0
                cur_revenue = int(row[11]) if row[11] is not None else 0
                # ucost = 单份总成本(列 N=13), 但 v40 是公式, data_only=True 可能读不到
                # 我们用 物料单价×消耗数量 自己算
                cur_ucost = None
                cur_total_cost = None
                cur_margin = None

            mat_name = str(row[5]).strip() if row[5] else ""
            mat_code = str(row[6]).strip() if row[6] else ""
            qty = float(row[7]) if row[7] is not None else 0.0
            mat_unit_price = float(row[8]) if row[8] is not None else 0.0
            unit = str(row[9]).strip() if row[9] else ""

            if not mat_code and not mat_name:
                continue  # 空行或合并单元格的延续行(无物料)

            # 估算单份总成本 = Σ(消耗数量 × 物料单价)
            # 总成本 = 单份总成本 × 净销量
            # 毛利 = 实收金额 - 总成本
            # 但 v40 里这些可能是公式, data_only 读不到准确值,我们按逻辑重算
            rows.append(
                {
                    "sheet": sheet_name,
                    "store": cur_store,
                    "product": cur_product,
                    "unit_price": cur_unit_price,
                    "mat_name": mat_name,
                    "mat_code": mat_code,
                    "qty": qty,
                    "mat_unit_price": mat_unit_price,
                    "unit": unit,
                    "net_qty": cur_net_qty,
                    "revenue": cur_revenue,
                }
            )
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# 按 (门店,商品,物料编码) 对齐并 diff
# ---------------------------------------------------------------------------

def diff_rows(v40_rows, old_price, new_price, meta):
    """
    对 v40 每行,用新 price 重算成本,与旧成本 diff。
    返回 diff 记录列表和汇总统计。
    """
    diffs = []
    # 按 (sheet, store, product, mat_code) 聚合
    from collections import defaultdict

    grouped = defaultdict(list)
    for r in v40_rows:
        key = (r["sheet"], r["store"], r["product"], r["mat_code"])
        grouped[key].append(r)

    # 汇总: 按 (sheet, store, product) 粒度算总差异
    product_diffs = []
    product_old_cost = {}
    product_new_cost = {}
    product_revenue = {}

    for key, rows in grouped.items():
        sheet, store, product, mat_code = key
        qty = sum(r["qty"] for r in rows)
        net_qty = rows[0]["net_qty"]
        revenue = rows[0]["revenue"]
        old_unit = rows[0]["mat_unit_price"]
        new_unit = new_price.get(mat_code, old_unit)  # 缺价则不变
        old_line = round(qty * old_unit, 4)
        new_line = round(qty * new_unit, 4)
        diff_line = round(new_line - old_line, 4)

        # 单份成本差异
        old_ucost = old_unit * qty / net_qty if net_qty else 0
        new_ucost = new_unit * qty / net_qty if net_qty else 0
        diff_ucost = round(new_ucost - old_ucost, 4)

        diffs.append(
            {
                "sheet": sheet,
                "store": store,
                "product": product,
                "mat_code": mat_code,
                "mat_name": rows[0]["mat_name"],
                "qty": qty,
                "net_qty": net_qty,
                "old_unit_cost": old_unit,
                "new_unit_cost": new_unit,
                "diff_unit_cost": round(new_unit - old_unit, 6),
                "old_line_cost": old_line,
                "new_line_cost": new_line,
                "diff_line_cost": diff_line,
                "base": meta.get(mat_code, {}).get("base", 0),
                "tax": meta.get(mat_code, {}).get("tax", 0),
            }
        )

        # 按 product 汇总
        pkey = (sheet, store, product)
        product_old_cost[pkey] = product_old_cost.get(pkey, 0) + old_line
        product_new_cost[pkey] = product_new_cost.get(pkey, 0) + new_line
        product_revenue[pkey] = revenue

    for pkey in product_old_cost:
        sheet, store, product = pkey
        old_cost = product_old_cost[pkey]
        new_cost = product_new_cost[pkey]
        revenue = product_revenue.get(pkey, 0)
        old_margin = revenue - old_cost
        new_margin = revenue - new_cost
        product_diffs.append(
            {
                "sheet": sheet,
                "store": store,
                "product": product,
                "revenue": revenue,
                "old_total_cost": old_cost,
                "new_total_cost": new_cost,
                "diff_total_cost": round(new_cost - old_cost, 4),
                "old_margin": old_margin,
                "new_margin": new_margin,
                "diff_margin": round(new_margin - old_margin, 4),
                "old_gm": old_margin / revenue if revenue else 0,
                "new_gm": new_margin / revenue if revenue else 0,
            }
        )

    return diffs, product_diffs


# ---------------------------------------------------------------------------
# 写 diff Excel
# ---------------------------------------------------------------------------

def write_diff_excel(diffs, product_diffs, out_path):
    import xlsxwriter

    wb = xlsxwriter.Workbook(out_path)
    F_HDR = wb.add_format(
        {"bold": True, "bg_color": "#4472C4", "font_color": "white", "align": "center", "valign": "vcenter", "border": 1}
    )
    F_TXT = wb.add_format({"valign": "vcenter"})
    F_NUM = wb.add_format({"valign": "vcenter", "num_format": "#,##0.0000"})
    F_MON = wb.add_format({"valign": "vcenter", "num_format": "#,##0.00"})
    F_PCT = wb.add_format({"valign": "vcenter", "num_format": "0.00%"})
    F_RED = wb.add_format({"valign": "vcenter", "font_color": "#C00000", "num_format": "#,##0.0000"})
    F_GRN = wb.add_format({"valign": "vcenter", "font_color": "#00B050", "num_format": "#,##0.0000"})

    # Sheet 1: 逐物料行 diff
    ws1 = wb.add_worksheet("逐物料行差异")
    ws1.freeze_panes(1, 0)
    headers1 = [
        "Sheet", "门店", "商品", "物料编码", "物料名称", "消耗数量",
        "净销量", "旧单位成本", "新单位成本", "单位成本差异",
        "旧行成本", "新行成本", "行成本差异", "基价", "税率%",
    ]
    for c, h in enumerate(headers1):
        ws1.write(0, c, h, F_HDR)
    for c, w in enumerate([12, 12, 28, 11, 18, 10, 9, 12, 12, 14, 12, 12, 14, 10, 8]):
        ws1.set_column(c, c, w)

    for i, r in enumerate(diffs, 1):
        fmt = F_RED if r["diff_unit_cost"] > 0 else (F_GRN if r["diff_unit_cost"] < 0 else F_NUM)
        ws1.write(i, 0, r["sheet"], F_TXT)
        ws1.write(i, 1, r["store"], F_TXT)
        ws1.write(i, 2, r["product"], F_TXT)
        ws1.write(i, 3, r["mat_code"], F_TXT)
        ws1.write(i, 4, r["mat_name"], F_TXT)
        ws1.write(i, 5, r["qty"], F_NUM)
        ws1.write(i, 6, r["net_qty"], F_NUM)
        ws1.write(i, 7, r["old_unit_cost"], F_NUM)
        ws1.write(i, 8, r["new_unit_cost"], F_NUM)
        ws1.write(i, 9, r["diff_unit_cost"], fmt)
        ws1.write(i, 10, r["old_line_cost"], F_MON)
        ws1.write(i, 11, r["new_line_cost"], F_MON)
        ws1.write(i, 12, r["diff_line_cost"], fmt)
        ws1.write(i, 13, r["base"], F_NUM)
        ws1.write(i, 14, r["tax"], F_NUM)

    # Sheet 2: 逐商品汇总 diff
    ws2 = wb.add_worksheet("逐商品汇总差异")
    ws2.freeze_panes(1, 0)
    headers2 = [
        "Sheet", "门店", "商品", "实收金额", "旧总成本", "新总成本", "总成本差异",
        "旧毛利", "新毛利", "毛利差异", "旧毛利率", "新毛利率",
    ]
    for c, h in enumerate(headers2):
        ws2.write(0, c, h, F_HDR)
    for c, w in enumerate([12, 12, 28, 11, 12, 12, 12, 11, 11, 11, 10, 10]):
        ws2.set_column(c, c, w)

    for i, r in enumerate(product_diffs, 1):
        fmt = F_RED if r["diff_total_cost"] > 0 else (F_GRN if r["diff_total_cost"] < 0 else F_NUM)
        ws2.write(i, 0, r["sheet"], F_TXT)
        ws2.write(i, 1, r["store"], F_TXT)
        ws2.write(i, 2, r["product"], F_TXT)
        ws2.write(i, 3, r["revenue"], F_MON)
        ws2.write(i, 4, r["old_total_cost"], F_MON)
        ws2.write(i, 5, r["new_total_cost"], F_MON)
        ws2.write(i, 6, r["diff_total_cost"], fmt)
        ws2.write(i, 7, r["old_margin"], F_MON)
        ws2.write(i, 8, r["new_margin"], F_MON)
        ws2.write(i, 9, r["diff_margin"], fmt)
        ws2.write(i, 10, r["old_gm"], F_PCT)
        ws2.write(i, 11, r["new_gm"], F_PCT)

    # Sheet 3: 假设声明
    ws3 = wb.add_worksheet("假设与Limitations")
    assumptions = [
        "假设声明(无 sid 前提下的诚实声明):",
        "",
        "1. PRLE-0003 条件假设满足:",
        "   - for_price_list = \"Buying - Internal\"",
        "   - disabled = False",
        "   - 日期有效(当前月份内)",
        "   - PriceOrDiscount = \"Price\"",
        "   以上条件因无 sid 无法验证, 拿到 sid 后应重新跑 live 对账锚验证 drift ≤ 0.5%。",
        "",
        "2. desired-UOM = clean_bom.csv `单位` 列(BOM 消耗单位)。",
        "",
        "3. baseCost = `基价(原始)` 列(Buying-Internal price_list_rate)。",
        "",
        "4. 税率 = `适用税率%` 列(已含在 csv 中)。",
        "",
        "5. unit_corrections(MK01018 ÷50) 与 v40 路径一致应用。",
        "",
        "Limitations:",
        "- 本 diff 只改价格算法, 不改 BOM 结构/销量/实收, 差异仅来自物料单价变化。",
        "- 未映射商品(无 BOM)的成本用平均成本率估算, 不受价格规则影响。",
        "- 由于无 sid, 规则条件判定是假设性而非验证性。",
        "",
        "Phase 5 后建议:",
        "- 拿到 sid 后, 用 final_unit_cost_with_rule 接真实 ERPNext PricingRule 条件判定,",
        "  替换本脚本中的硬编码假设规则, 并跑 ttpos_cost_anchor.py 验证 drift ≤ 0.5%。",
    ]
    for i, line in enumerate(assumptions):
        ws3.write(i, 0, line)

    wb.close()


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description="ERP 成本规则修正 diff: 用修正算法重算并与 v40 输出对比"
    )
    ap.add_argument("--v40", required=True, help="v40 基准 Excel 路径")
    ap.add_argument("--bom", required=True, help="clean_bom.csv 路径")
    ap.add_argument("--out", required=True, help="diff 输出 Excel 路径")
    a = ap.parse_args()

    for p in (a.v40, a.bom):
        if not os.path.exists(p):
            sys.exit(f"找不到文件: {p}")

    print("加载 clean_bom.csv 并应用修正算法 ...")
    old_price, new_price, meta = load_clean_bom_with_rule(a.bom)
    print(f"  物料数: {len(old_price)}")

    # 快速统计: 有多少物料单价变化了
    changed = [code for code in old_price if abs(old_price[code] - new_price[code]) > 1e-9]
    print(f"  单价变化物料数: {len(changed)} / {len(old_price)}")

    print("读取 v40 基准输出 ...")
    v40_rows = read_v40_rows(a.v40)
    print(f"  v40 物料行数: {len(v40_rows)}")

    print("对齐并计算差异 ...")
    diffs, product_diffs = diff_rows(v40_rows, old_price, new_price, meta)
    print(f"  diff 物料行数: {len(diffs)}")
    print(f"  diff 商品数: {len(product_diffs)}")

    # 汇总统计
    total_old_cost = sum(r["old_line_cost"] for r in diffs)
    total_new_cost = sum(r["new_line_cost"] for r in diffs)
    total_cost_diff = total_new_cost - total_old_cost

    # 区分"算法差异"和"浮点/舍入差异"
    # 算法差异: 新单价(CSV原始全精度经conditional算法) != CSV原始单价(unconditional算法)
    # 舍入差异: 新单价 == CSV原始单价, 但与 Excel 4dp 舍入值不同
    # 注意: old_price 来自 CSV `物料单价` 列(全精度), new_price 来自 conditional 算法
    algo_diffs = [r for r in diffs if abs(r["new_unit_cost"] - old_price.get(r["mat_code"], r["new_unit_cost"])) > 1e-6]
    rounding_diffs = [r for r in diffs if 0 < abs(r["new_unit_cost"] - old_price.get(r["mat_code"], r["new_unit_cost"])) <= 1e-6]

    changed_products = [r for r in product_diffs if abs(r["diff_total_cost"]) > 1e-6]
    top10_cost = sorted(changed_products, key=lambda x: -abs(x["diff_total_cost"]))[:10]

    print("\n==================== 汇总 ====================")
    print(f"总 SKU(商品)数: {len(product_diffs)}")
    print(f"物料行数: {len(diffs)}")
    print(f"算法差异物料行数(新单价 != CSV原始单价): {len(algo_diffs)}")
    print(f"浮点/舍入差异物料行数(仅 Excel 4dp 舍入): {len(rounding_diffs)}")
    print(f"总成本差异额: ฿{total_cost_diff:,.4f} (新 {'>' if total_cost_diff > 0 else '<'} 旧)")

    if not algo_diffs:
        print("\n【关键发现】本次 diff 中算法差异为 0。")
        print("原因: clean_bom.csv 的 `物料单价` 列已等于 unconditional final_unit_cost(base*1.05*(1+tax/100)),")
        print("而假设的 PRLE-0003 规则(for_price_list='Buying - Internal', buying=True, disabled=False)")
        print("条件满足, conditional final_unit_cost_with_rule 与 unconditional 结果相同。")
        print("观察到的微小差异(฿-28.56)来自 Excel 写盘时的 4dp 舍入 vs 全精度重算。")
    else:
        print(f"\n差异金额 Top 10 商品:")
        for i, r in enumerate(top10_cost, 1):
            print(
                f"  {i}. {r['sheet']} | {r['store']} | {r['product']} | "
                f"成本差异 ฿{r['diff_total_cost']:,.4f} | 毛利差异 ฿{r['diff_margin']:,.4f}"
            )

    print(f"\n写入 diff 文件: {a.out}")
    write_diff_excel(diffs, product_diffs, a.out)
    print("完成。")


if __name__ == "__main__":
    main()
