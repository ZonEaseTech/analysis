"""把 5 个外卖平台对账单 + 1 个 RBH 汇总成标准 CSV。

订单号没法跟 BQ 1:1 对上 → 用 (store_num, year_month, platform) 粒度的聚合事实表.
报表层按 SKU takeout revenue 比例分摊到 SKU.

输入: resources/wallace.20260515/*.xlsx (6 个原始文件)
输出: resources/wallace.20260515/platform_fees_202603.csv
列:  store_num, year_month, platform, gross_revenue, commission_total, source_file

口径:
  - LINE MAN: REPORT_MAR26 sheet, Store# 列 (收001→001), total_revenue, 佣金
  - Shopee Food (4 法人): {prefix}_SHP_INCOME_MAR26 sheet, 应收, 手续费+税
  - RBH: RBH_INCOME_MAR2026 sheet, 应收, 佣金合计 (实收/应收差额未解释, 暂只用佣金列)
"""
from __future__ import annotations
import csv
from pathlib import Path
from openpyxl import load_workbook

BASE = Path("/home/weifashi/hwt/analysis/resources/wallace.20260515")
OUT = BASE / "platform_fees_202603.csv"
YEAR_MONTH = "2026-03"


def _normalize_store(raw) -> str | None:
    """收001 / 收 001 / 001 → '001'. 空/无效 → None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s in ("-", "0"):
        return None
    # 剥 '收' 前缀
    if s.startswith("收"):
        s = s[1:].strip()
    # 剥前缀引号 ('001 → 001)
    s = s.lstrip("'").strip()
    if not s.isdigit():
        return None
    return s.zfill(3)


def ingest_lineman(path: Path) -> list[dict]:
    """LINE MAN: REPORT_MAR26 sheet, daily granularity → group by store."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["REPORT_MAR26"]
    # r3 是 header (r1-r2 是 title/summary)
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    header = rows[0]
    print(f"  LINEMAN header: {header[:10]}")
    # 字段位置 (基于 preview):
    # 0=wongnai_id, 1=Store #, 2=name, 3=branch, 4=Date,
    # 5=total_revenue, 6=cash, 7=应收, 8=佣金, ...
    by_store: dict[str, dict] = {}
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        store = _normalize_store(row[1])
        if not store:
            continue
        gross = float(row[7] or 0)       # 应收
        commission = float(row[8] or 0)  # 佣金
        slot = by_store.setdefault(store, {"gross": 0.0, "commission": 0.0})
        slot["gross"] += gross
        slot["commission"] += commission
    wb.close()
    return [
        {"store_num": s, "year_month": YEAR_MONTH, "platform": "LINEMAN",
         "gross_revenue": round(d["gross"], 2),
         "commission_total": round(d["commission"], 2),
         "source_file": path.name}
        for s, d in by_store.items()
    ]


def ingest_shp(path: Path, legal_entity: str) -> list[dict]:
    """Shopee Food: {prefix}_SHP_INCOME_MAR26 sheet.
    Header r6: Date | Store# | 应收 | 实收 | 手续费 | 税费 | 佣金合计 | INV#
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = None
    for sn in wb.sheetnames:
        if "_SHP_INCOME_" in sn:
            sheet_name = sn
            break
    if not sheet_name:
        print(f"  ⚠️ {path.name}: no *_SHP_INCOME_* sheet, skip")
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=6, values_only=True))  # header at r6
    header = rows[0]
    print(f"  SHP/{legal_entity} sheet={sheet_name} header: {header[:8]}")
    by_store: dict[str, dict] = {}
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        store = _normalize_store(row[1])
        if not store:
            continue
        gross = float(row[2] or 0)       # 应收
        commission = float(row[6] or 0)  # 佣金合计 (手续费+税)
        slot = by_store.setdefault(store, {"gross": 0.0, "commission": 0.0})
        slot["gross"] += gross
        slot["commission"] += commission
    wb.close()
    return [
        {"store_num": s, "year_month": YEAR_MONTH, "platform": f"SHP-{legal_entity}",
         "gross_revenue": round(d["gross"], 2),
         "commission_total": round(d["commission"], 2),
         "source_file": path.name}
        for s, d in by_store.items()
    ]


def ingest_rbh(path: Path) -> list[dict]:
    """Robinhood: RBH_INCOME_MAR2026 sheet.
    Header r5: Date | Store # | 应收 | 实收 | 手续费 | 税费 | 佣金合计
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["RBH_INCOME_MAR2026"]
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    header = rows[0]
    print(f"  RBH header: {header[:7]}")
    by_store: dict[str, dict] = {}
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        store = _normalize_store(row[1])
        if not store:
            continue
        gross = float(row[2] or 0)       # 应收
        commission = float(row[6] or 0)  # 佣金合计
        slot = by_store.setdefault(store, {"gross": 0.0, "commission": 0.0})
        slot["gross"] += gross
        slot["commission"] += commission
    wb.close()
    return [
        {"store_num": s, "year_month": YEAR_MONTH, "platform": "RBH",
         "gross_revenue": round(d["gross"], 2),
         "commission_total": round(d["commission"], 2),
         "source_file": path.name}
        for s, d in by_store.items()
    ]


def main():
    print("=== 接入外卖平台抽佣对账单 → 标准 CSV ===\n")
    all_rows: list[dict] = []

    print("[1/3] LINE MAN")
    all_rows += ingest_lineman(BASE / "LINEMAN_MARCH_2026_原始.xlsx")
    print(f"  → {len([r for r in all_rows if r['platform']=='LINEMAN'])} stores")

    print("\n[2/3] Shopee Food (4 法人)")
    for legal in ["TR", "FF", "JJFM", "FBF"]:
        path = BASE / f"SHP_{legal}_MAR26_原始.xlsx"
        rows = ingest_shp(path, legal)
        all_rows += rows
        print(f"  SHP-{legal}: {len(rows)} stores")

    print("\n[3/3] Robinhood")
    rbh = ingest_rbh(BASE / "RBH_MAR2026_原始.xlsx")
    all_rows += rbh
    print(f"  RBH: {len(rbh)} stores")

    # 写 CSV
    with OUT.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "store_num", "year_month", "platform",
            "gross_revenue", "commission_total", "source_file"])
        writer.writeheader()
        for r in sorted(all_rows, key=lambda x: (x["store_num"], x["platform"])):
            writer.writerow(r)

    print(f"\n→ 输出: {OUT}")
    print(f"  总行数: {len(all_rows)}")

    # 概览: 每个平台的总应收/总佣金/综合费率
    print("\n=== 平台费率汇总 ===")
    by_plat: dict[str, dict] = {}
    for r in all_rows:
        slot = by_plat.setdefault(r["platform"], {"gross": 0.0, "commission": 0.0, "stores": set()})
        slot["gross"] += r["gross_revenue"]
        slot["commission"] += r["commission_total"]
        slot["stores"].add(r["store_num"])
    total_gross = sum(d["gross"] for d in by_plat.values())
    total_commission = sum(d["commission"] for d in by_plat.values())
    print(f"{'platform':12s} {'stores':>7s} {'应收':>14s} {'佣金':>12s} {'实际费率':>10s}")
    print("-" * 60)
    for plat, d in sorted(by_plat.items()):
        rate = d["commission"] / d["gross"] * 100 if d["gross"] else 0
        print(f"{plat:12s} {len(d['stores']):>7d} {d['gross']:>14,.2f} {d['commission']:>12,.2f} {rate:>9.2f}%")
    print("-" * 60)
    blended = total_commission / total_gross * 100 if total_gross else 0
    print(f"{'TOTAL':12s} {'':>7s} {total_gross:>14,.2f} {total_commission:>12,.2f} {blended:>9.2f}%")

    # 按店汇总: 一店多平台合并
    by_store_month: dict[tuple, dict] = {}
    for r in all_rows:
        key = (r["store_num"], r["year_month"])
        slot = by_store_month.setdefault(key, {"gross": 0.0, "commission": 0.0, "plats": []})
        slot["gross"] += r["gross_revenue"]
        slot["commission"] += r["commission_total"]
        slot["plats"].append(r["platform"])
    print(f"\n=== 按 (店,月) 汇总 ===  共 {len(by_store_month)} 个店-月")
    print(f"前 10 店:")
    for (store, ym), d in sorted(by_store_month.items())[:10]:
        plats = ",".join(sorted(set(d["plats"])))
        rate = d["commission"] / d["gross"] * 100 if d["gross"] else 0
        print(f"  店{store} {ym}: 应收 {d['gross']:>10,.0f}  佣金 {d['commission']:>9,.0f}  ({rate:>5.2f}%)  [{plats}]")


if __name__ == "__main__":
    main()
