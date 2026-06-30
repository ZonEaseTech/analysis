#!/usr/bin/env python3
"""弧酷 BOM 补充包清洗 (Sheet1/2/3 一次性接入).

市场给的 xlsx 包含三张表:
  Sheet1  套餐: 简单 alias (21XXX→现有 / 吃爽→超值套餐) + 完整 BOM 展开 (~52 套餐)
  Sheet2  单品: 简单 alias (中可乐→百事 / 半只→脆皮半鸡 / 冰激凌→冰淇凌) + 完整 BOM 展开 (~7 单品)
  Sheet3  成分: 16 个成分级 BOM (汉堡/可乐/薯条/上校鸡块/烤肠/鸡肉卷/鸡米花/鸡肉棒/...)

输出 (进 git):
  resources/wallace.20260514/huku_补充BOM_202601.csv     新 BOM 商品 (跟现有重名的跳过)
  resources/wallace.20260514/huku_补充alias_202601.csv   弧酷名 → BOM 已有名映射

跟现有 最终BOM_202605.csv 同名的商品 — 按 A 方案跳过, 不覆盖 (现有 BOM 比较全).

一次性脚本, 跑完归档. xlsx 原文件不进 git.
"""
from __future__ import annotations

import argparse
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(
    os.path.abspath(__file__)))))

import openpyxl


RAW_DEFAULT = "/workspace/data/uploads/01KRZY6N45EV61Z2EJA9PMJJ2H.xlsx"
EXISTING_BOM = "resources/wallace.20260514/最终BOM_202605.csv"
OUT_BOM = "resources/wallace.20260514/huku_补充BOM_202601.csv"
OUT_ALIAS = "resources/wallace.20260514/huku_补充alias_202601.csv"

# ── 手工补充 BOM (市场后续在报表 xlsx 里手填的, 不在 Sheet1/2/3) ────────────
# 2026-05-19: THD华莱士庆祝套餐 1/2 — 市场在 profit_by_price_202601_v22(1).xlsx
#   套餐 sheet 里手工补的 BOM, 抽出来固化.
MANUAL_BOM = {
    "THD华莱士庆祝套餐 1": [
        ("PA99020", "单杯袋", 1, "pc"), ("PA99001", "16A 冷杯", 1, "pc"),
        ("BE01004", "百事可乐原味", 25, "g"), ("PA99024", "PLA尖头吸管", 1, "pc"),
        ("PA99027", "哑光塑料封口膜", 1, "pc"), ("PA99008", "4#全鸡袋", 1, "pc"),
        ("PA99025", "一次性餐包", 1, "pc"), ("CK01002", "鸡块", 5, "pc"),
        ("SA01003", "甜辣酱（小包）", 1, "pc"), ("DR02001", "起酥油", 10, "g"),
        ("PA99012", "小号方底袋", 1, "pc"), ("PA99018", "打包袋 (小号)", 1, "pc"),
        ("SA01004", "番茄酱 （小包）", 1, "pc"), ("SE05002", "辣味粉", 10, "g"),
        ("PA99023", "手套", 1, "pc"), ("FR01004", "翅根", 8, "pc"),
        ("DR02002", "炸鸡炸粉", 10, "g"),
    ],
    # 2026-05-19: 鸡翅桶 — 市场图片提供的 8 物料 BOM
    "鸡翅桶": [
        ("DR02001", "起酥油", 10, "g"), ("PA99004", "65oz 桶(含盖)", 1, "pc"),
        ("FR01004", "翅根", 8, "pc"), ("DR02002", "炸鸡炸粉", 10, "g"),
        ("PA99016", "托盘纸", 1, "pc"), ("PA99023", "手套", 1, "pc"),
        ("SA01008", "甜辣酱（大包）", 10, "g"), ("SA01009", "番茄酱（大包）", 10, "g"),
    ],
    "THD华莱士庆祝套餐2": [
        ("PA99020", "单杯袋", 1, "pc"), ("PA99001", "16A 冷杯", 1, "pc"),
        ("BE01004", "百事可乐原味", 25, "g"), ("PA99024", "PLA尖头吸管", 1, "pc"),
        ("PA99027", "哑光塑料封口膜", 1, "pc"), ("PA99019", "打包袋 (大号)", 1, "pc"),
        ("SA02001", "沙拉酱", 15, "g"), ("PA99025", "一次性餐包", 1, "pc"),
        ("PA99034", "食安封签", 1, "pc"), ("SE05002", "辣味粉", 10, "g"),
        ("PA99017", "五合一汉堡纸", 1, "pc"), ("DR02001", "起酥油", 10, "g"),
        ("FR01001", "无骨方腿腿肉汉堡用", 1, "pc"), ("SA01013", "番茄酱（小包）", 1, "pc"),
        ("DR02002", "炸鸡炸粉", 10, "g"), ("MK01019", "生菜", 15, "g"),
        ("DR01001", "汉堡面包", 1, "pc"), ("PA99008", "4#全鸡袋", 1, "pc"),
        ("FR01004", "翅根", 8, "pc"), ("PA99004", "65oz 桶(含盖)", 1, "pc"),
        ("PA99023", "手套", 1, "pc"), ("PA99016", "托盘纸", 1, "pc"),
    ],
}

# ── 调味衍生 BOM ────────────────────────────────────────────────────────
# 2026-05-19 用户拍板: 辣番茄系列 = 基础商品 + SE05004 辣番茄调味料 10g
#                      蒜香柠檬系列 = 基础商品 + DE01010 柠檬大蒜粉 10g
# 这里给 X炸鸡桶+饮料 拆解用的成分商品建 BOM (基础=炸鸡桶).
SEASONING_DERIVED = {
    # 衍生商品名: (基础商品名, 调味物料)
    "辣番茄炸鸡桶": ("炸鸡桶", ("SE05004", "辣番茄调味料", 10, "g")),
    "蒜香柠檬炸鸡桶": ("炸鸡桶", ("DE01010", "柠檬大蒜粉", 10, "g")),
}


def load_existing_bom_keys():
    keys = set()
    with open(EXISTING_BOM, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            keys.add(r["商品名称"])
    return keys


def load_existing_bom(name: str):
    """读现有 csv 里某商品的物料列表 → [(code, mname, qty, unit), ...]."""
    recs = []
    with open(EXISTING_BOM, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if r["商品名称"] == name:
                recs.append((r["物品编号"], r["物品名称"],
                             float(r["单耗"] or 0), r["单位"]))
    return recs


def parse_sheet1(ws):
    """Sheet1 套餐: 混合 alias + BOM 展开.

    每行 6 列: 套餐名 / (空) / 物料名 / 物料编码 / 消耗数量 / 物料单价

    pattern:
      - col 0 有 + col 3 无  → alias 行 (col 2 = 目标名)
      - col 0 有 + col 3 有  → 新套餐 BOM 起始行 (col 0 carry-forward 给后续)
      - col 0 无 + col 3 有  → 同上一套餐继续

    Returns: (alias_dict, bom_dict)
    """
    alias = {}
    bom = {}
    cur = None
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        a, _, c, code, qty, _price = (r[0], r[1], r[2], r[3], r[4], r[5])
        if a and not code:
            if c and str(c).strip():
                alias[str(a).strip()] = str(c).strip()
            cur = None
        elif a and code:
            cur = str(a).strip()
            bom.setdefault(cur, []).append(
                (str(code).strip(), str(c or "").strip(), float(qty or 0), "pc"))
        elif not a and code and cur:
            bom[cur].append(
                (str(code).strip(), str(c or "").strip(), float(qty or 0), "pc"))
    return alias, bom


def parse_sheet2(ws):
    """Sheet2 单品: 混合 alias + BOM 展开.

    每行: 商品名 / 物料名 / 物料编码 / 消耗 / 单价 / 单位

    pattern:
      - col 0 有 + col 2 无  → alias 行 (col 1 = 目标名, 跳过 col 1 为 None 的)
      - col 0 有 + col 2 有  → 新商品 BOM 起始
      - col 0 无 + col 2 有  → 继续
    """
    alias = {}
    bom = {}
    cur = None
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        a, b, code, qty, _price, unit = (r[0], r[1], r[2], r[3], r[4], r[5])
        if a and not code:
            if b and str(b).strip():
                alias[str(a).strip()] = str(b).strip()
            cur = None
        elif a and code:
            cur = str(a).strip()
            bom.setdefault(cur, []).append(
                (str(code).strip(), str(b or "").strip(),
                 float(qty or 0), str(unit or "").strip() or "pc"))
        elif not a and code and cur:
            bom[cur].append(
                (str(code).strip(), str(b or "").strip(),
                 float(qty or 0), str(unit or "").strip() or "pc"))
    return alias, bom


def parse_sheet3(ws):
    """Sheet3 成分: 纯 BOM 展开.

    每行: 商品名 / 物料名 / 物料编码 / 消耗 / 单价 / 单位
    第 1 列 carry-forward, 第 8-11 列 (堂食/外卖额外包装) 暂不读 — 主行作默认.

    用户已确认: 汉堡/薯条/可乐 默认加 4 项堂食包装 (托盘纸/手套/甜辣酱大包/番茄酱大包).
    """
    DINE_EXTRAS = [
        ("PA99016", "托盘纸", 1.0, "pc"),
        ("PA99023", "手套", 1.0, "pc"),
        ("SA01008", "甜辣酱（大包）", 10.0, "g"),
        ("SA01009", "番茄酱（大包）", 10.0, "g"),
    ]
    APPEND_DINE_EXTRAS = {"汉堡", "薯条", "可乐"}
    bom = {}
    cur = None
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        a, b, code, qty, _price, unit = (r[0], r[1], r[2], r[3], r[4], r[5])
        if a:
            cur = str(a).strip()
            if code:
                bom.setdefault(cur, []).append(
                    (str(code).strip(), str(b or "").strip(),
                     float(qty or 0), str(unit or "").strip() or "pc"))
        elif code and cur:
            bom.setdefault(cur, []).append(
                (str(code).strip(), str(b or "").strip(),
                 float(qty or 0), str(unit or "").strip() or "pc"))
    # 汉堡/薯条/可乐 加堂食额外包装
    for name in APPEND_DINE_EXTRAS:
        if name in bom:
            existing_codes = {x[0] for x in bom[name]}
            for code, mname, q, u in DINE_EXTRAS:
                if code not in existing_codes:
                    bom[name].append((code, mname, q, u))
    return bom


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", default=RAW_DEFAULT)
    args = ap.parse_args()

    print(f"[1/4] 加载 {args.raw}")
    wb = openpyxl.load_workbook(args.raw, data_only=True, read_only=True)
    s1_alias, s1_bom = parse_sheet1(wb["Sheet1"])
    s2_alias, s2_bom = parse_sheet2(wb["Sheet2"])
    s3_bom = parse_sheet3(wb["Sheet3"])
    print(f"  Sheet1: alias {len(s1_alias)}, BOM {len(s1_bom)}")
    print(f"  Sheet2: alias {len(s2_alias)}, BOM {len(s2_bom)}")
    print(f"  Sheet3: BOM {len(s3_bom)}")

    print(f"[2/4] 合并 + 去重 (跟现有 {EXISTING_BOM} 重名跳过)")
    existing_keys = load_existing_bom_keys()
    all_bom = {}
    skipped_dup = []
    for src_label, src_bom in [("sheet1", s1_bom), ("sheet2", s2_bom), ("sheet3", s3_bom)]:
        for name, recs in src_bom.items():
            if name in existing_keys:
                skipped_dup.append((name, src_label, len(recs)))
                continue
            if name in all_bom:
                # 后到的源覆盖 (sheet3 > sheet2 > sheet1)? 实际同一商品只应出现一份
                continue
            all_bom[name] = (recs, src_label)
    # 手工补充 BOM (THD套餐 等)
    for name, recs in MANUAL_BOM.items():
        if name in existing_keys:
            skipped_dup.append((name, "manual", len(recs)))
        else:
            all_bom[name] = (recs, "manual")

    # 调味衍生 BOM (辣番茄炸鸡桶 = 炸鸡桶 + 辣番茄调味料 等)
    for derived, (base, seasoning) in SEASONING_DERIVED.items():
        if derived in existing_keys:
            skipped_dup.append((derived, "seasoning", 0))
            continue
        base_recs = load_existing_bom(base)
        if not base_recs:
            print(f"  ⚠️ 调味衍生 {derived}: 基础商品 {base!r} 在现有 csv 找不到, 跳过")
            continue
        all_bom[derived] = (base_recs + [seasoning], "seasoning")

    print(f"  新加 BOM 商品: {len(all_bom)} ({sum(len(v[0]) for v in all_bom.values())} 物料行)")
    print(f"  跟现有 csv 重名跳过: {len(skipped_dup)}")
    for name, src, n in skipped_dup:
        print(f"    [{src}] {name} ({n} 物料) — 现有 csv 已有, 不覆盖")

    print(f"[3/4] 写 BOM csv → {OUT_BOM}")
    with open(OUT_BOM, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["商品名称", "物品编号", "物品名称", "单耗", "单位"])
        for name in sorted(all_bom):
            recs, _src = all_bom[name]
            for code, mname, qty, unit in recs:
                w.writerow([name, code, mname, qty, unit])

    print(f"[4/4] 写 alias csv → {OUT_ALIAS}")
    all_alias = {}
    for src_label, src_alias in [("sheet1", s1_alias), ("sheet2", s2_alias)]:
        for k, v in src_alias.items():
            all_alias[k] = v
    with open(OUT_ALIAS, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["弧酷名", "BOM目标名"])
        for k in sorted(all_alias):
            w.writerow([k, all_alias[k]])
    print(f"  共 {len(all_alias)} 条 alias")

    print("\n完成. 接下来:")
    print(f"  1. config.yaml 加 bom_sources 新条目 (path={OUT_BOM}, priority=90)")
    print(f"  2. external_sales/huku/aliases.py 读 {OUT_ALIAS} 进 SPELLING_ALIASES")
    print(f"  3. 删 match.py 的 + 拆解 + INGREDIENT_ALIAS 分支")


if __name__ == "__main__":
    main()
