#!/usr/bin/env python3
"""报表交付前 audit — 一条命令出摘要 + 交付说明.

固化 adhoc-export skill 步骤 6 的检查项, 替代"每次现写脚本"(会漏项,
上次就漏了 BOM 抽查). 跑:

    venv/bin/python scripts/audit_report.py exports/profit_by_price_202602_v1.xlsx
    venv/bin/python scripts/audit_report.py exports/profit_by_price_2026{01,02,04}_v*.xlsx

检查项 (profit_by_price / profit_margin 通用):
  1. BOM来源 / 价来源 分布          — 确认事实表来源符合预期
  2. 物料单价 <= 0 的 unique code   — strict 模式缺价物料 (zero_yellow 会标黄)
  3. BOM 抽查                      — 抽几个商品打印 BOM 物料明细 + COGS + 毛利率
  4. "无 BOM" 商品数               — 新事实表没覆盖的, 成本会算 0

输出末尾给一段「交付说明草稿」: 把异常项摘成给市场的话术.

只读, 不改任何文件.
"""
from __future__ import annotations

import sys
from collections import Counter

import openpyxl


def _safe_float(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def audit_one(path: str) -> dict:
    """audit 单个 xlsx, 返回结构化结果 dict."""
    wb = openpyxl.load_workbook(path, read_only=True)
    result = {"path": path, "sheets": {}}

    for sn in wb.sheetnames:
        ws = wb[sn]
        hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
        if "BOM来源" not in hdr:
            continue
        name_col = next((hdr.index(c) for c in
                         ("套餐名称", "单品名称", "商品名称") if c in hdr), None)
        store_col = hdr.index("门店编号") if "门店编号" in hdr else None
        bom_src_col = hdr.index("BOM来源")
        price_src_col = hdr.index("价来源") if "价来源" in hdr else None
        price_col = hdr.index("物料单价") if "物料单价" in hdr else None
        code_col = hdr.index("BOM物品编码") if "BOM物品编码" in hdr else None

        bom_src = Counter()
        price_src = Counter()
        zero_price_codes = Counter()       # code -> 累计行数
        nonzero_codes = set()
        items = set()

        for r in ws.iter_rows(min_row=2, values_only=True):
            if name_col is not None and store_col is not None:
                items.add((r[store_col], r[name_col]))
            if r[bom_src_col]:
                bom_src[str(r[bom_src_col])] += 1
            if price_src_col is not None and r[price_src_col]:
                price_src[str(r[price_src_col])[:40]] += 1
            if price_col is not None and code_col is not None:
                pv = _safe_float(r[price_col])
                code = r[code_col]
                if pv is not None and code and code != "-":
                    if pv <= 0:
                        zero_price_codes[code] += 1
                    else:
                        nonzero_codes.add(code)

        zero_unique = {c: n for c, n in zero_price_codes.items()
                       if c not in nonzero_codes}
        result["sheets"][sn] = {
            "items": len(items),
            "bom_src": dict(bom_src.most_common()),
            "price_src": dict(price_src.most_common()),
            "zero_price_unique": zero_unique,
            "no_bom": bom_src.get("无", 0),
        }
    wb.close()
    return result


def sample_bom(path: str, sheet: str, n: int = 3) -> list:
    """从 sheet 抽 n 个商品, 取每个商品第一个门店块的 BOM 明细 + COGS."""
    wb = openpyxl.load_workbook(path, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    name_col = next((hdr.index(c) for c in
                     ("套餐名称", "单品名称", "商品名称") if c in hdr), None)
    store_col = hdr.index("门店编号")
    bn = hdr.index("BOM物品名称")
    bc = hdr.index("BOM物品编码")
    bq = hdr.index("消耗数量")
    bp = hdr.index("物料单价")
    actual_col = hdr.index("实收金额") if "实收金额" in hdr else None
    netq_col = hdr.index("净销量") if "净销量" in hdr else None

    grabbed = {}
    cur = cur_store = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[name_col]:
            cur, cur_store = r[name_col], r[store_col]
            if len(grabbed) < n and cur not in grabbed:
                grabbed[cur] = {
                    "store": cur_store,
                    "actual": r[actual_col] if actual_col is not None else None,
                    "netq": r[netq_col] if netq_col is not None else None,
                    "rows": [],
                }
        if cur in grabbed and grabbed[cur]["store"] == cur_store \
                and r[bc] and r[bc] != "-":
            grabbed[cur]["rows"].append((r[bn], r[bc], r[bq], r[bp]))
        if len(grabbed) >= n and all(g["rows"] for g in grabbed.values()):
            # 已抓够且每个都有料 (但还要等当前块结束, 简单起见抓到就停下一个块)
            pass
    wb.close()
    return list(grabbed.items())


def margin_bands(path: str, sheet: str) -> dict:
    """扫全部 SKU 算毛利率, 统计落在异常波段的 — 固化"人肉扫报表看毛利率离谱没"。

    毛利率 = (客单实收 - 单份成本) / 客单实收
      客单实收 = 实收金额 / 净销量    (merge 列, SKU 块首行有值)
      单份成本 = Σ(BOM 消耗 × 物料单价)  (SKU 块内累加)

    波段 (餐饮经验): 正常 0-90%; >90% 高度可疑 (成本被严重低估, 多半口径错);
                     <0% 也可疑 (成本 > 售价). 这次"肉类成本错 120 倍"就是
                     靠 99% 毛利率暴露的 — 把那个判断固化成扫描.
    返回: {正常, 高于90, 低于0, 无法算(成本0/无净销量)} 计数 + 异常 SKU 样例.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    name_col = next((hdr.index(c) for c in
                     ("套餐名称", "单品名称", "商品名称") if c in hdr), None)
    store_col = hdr.index("门店编号")
    bc = hdr.index("BOM物品编码")
    bq = hdr.index("消耗数量")
    bp = hdr.index("物料单价")
    actual_col = hdr.index("实收金额")
    netq_col = hdr.index("净销量")

    bands = {"正常": 0, "高于90%": 0, "低于0%": 0, "无法算": 0}
    suspect = []  # (店, 商品, 毛利率, 成本, 客单实收)

    def _flush(blk):
        if blk is None:
            return
        actual = _safe_float(blk["actual"])
        netq = _safe_float(blk["netq"])
        if not actual or not netq:
            bands["无法算"] += 1
            return
        pu = actual / netq
        cost = blk["cost"]
        if cost <= 0 or pu <= 0:
            bands["无法算"] += 1
            return
        margin = (pu - cost) / pu
        if margin > 0.90:
            bands["高于90%"] += 1
            suspect.append((blk["store"], blk["name"], margin, cost, pu))
        elif margin < 0:
            bands["低于0%"] += 1
            suspect.append((blk["store"], blk["name"], margin, cost, pu))
        else:
            bands["正常"] += 1

    blk = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[name_col]:                       # 新 SKU 块
            _flush(blk)
            blk = {"store": r[store_col], "name": r[name_col],
                   "actual": r[actual_col], "netq": r[netq_col], "cost": 0.0}
        if blk is not None and r[bc] and r[bc] != "-":
            blk["cost"] += (_safe_float(r[bq]) or 0) * (_safe_float(r[bp]) or 0)
    _flush(blk)
    wb.close()
    # 异常样例按 |偏离| 排序, 取 top 10
    suspect.sort(key=lambda x: -abs(x[2]))
    return {"bands": bands, "suspect": suspect[:10]}


def print_report(result: dict):
    print(f"\n{'=' * 60}")
    print(f"AUDIT: {result['path']}")
    print('=' * 60)
    for sn, s in result["sheets"].items():
        print(f"\n[{sn}] {s['items']} 个 (店,商品)")
        print(f"  BOM来源: {s['bom_src']}")
        if s["price_src"]:
            print(f"  价来源:  {s['price_src']}")
        if s["no_bom"]:
            print(f"  ⚠️ 无 BOM: {s['no_bom']} 行 (新事实表未覆盖, 成本算 0)")
        zu = s["zero_price_unique"]
        if zu:
            top = sorted(zu.items(), key=lambda x: -x[1])[:8]
            print(f"  ⚠️ 物料单价<=0 的 unique code: {len(zu)} 个 "
                  f"(累计行数 top: {top})")

    # BOM 抽查
    for sn in result["sheets"]:
        samples = sample_bom(result["path"], sn, n=3)
        if not samples:
            continue
        print(f"\n  [BOM 抽查 — {sn}]")
        for name, d in samples:
            cost = sum((_safe_float(q) or 0) * (_safe_float(p) or 0)
                       for _, _, q, p in d["rows"])
            actual = _safe_float(d["actual"])
            netq = _safe_float(d["netq"])
            pu = actual / netq if actual and netq else 0
            margin = (pu - cost) / pu * 100 if pu else 0
            print(f"    {str(name)[:24]:24} 店{d['store']} "
                  f"客单实收{pu:>7.1f} 成本{cost:>7.2f} 毛利率{margin:>6.1f}% "
                  f"({len(d['rows'])} 料)")

    # 毛利率波段 — 扫全部 SKU, 固化"人肉扫报表看毛利率离谱没"
    for sn in result["sheets"]:
        mb = margin_bands(result["path"], sn)
        if not mb:
            continue
        b = mb["bands"]
        flag = "⚠️" if (b["高于90%"] or b["低于0%"]) else "  "
        print(f"\n  {flag}[毛利率波段 — {sn}] {b}")
        for store, name, m, cost, pu in mb["suspect"]:
            print(f"      店{store} {str(name)[:24]:24} "
                  f"毛利率{m*100:>6.1f}% (客单实收{pu:.1f} 成本{cost:.2f})")
        # 把波段结果挂到 result, 供 delivery_notes 用
        result["sheets"][sn]["margin_bands"] = b


def delivery_notes(results: list):
    """把所有报表的异常项摘成「交付说明草稿」."""
    print(f"\n{'=' * 60}")
    print("交付说明草稿 (复制给市场前自己再核一遍)")
    print('=' * 60)
    lines = []
    for res in results:
        month = res["path"].split("_")[-2] if "_" in res["path"] else res["path"]
        for sn, s in res["sheets"].items():
            if s["no_bom"] > 50:
                lines.append(
                    f"- {res['path']} [{sn}]: {s['no_bom']} 行无 BOM "
                    f"(事实表未覆盖, 这些商品成本显示 0 / 毛利率虚高, 别当真)")
            zu = s["zero_price_unique"]
            if zu:
                lines.append(
                    f"- {res['path']} [{sn}]: {len(zu)} 个物料无价 "
                    f"(strict 模式标黄, 含这些料的商品成本偏低)")
            mb = s.get("margin_bands")
            if mb and (mb["高于90%"] or mb["低于0%"]):
                lines.append(
                    f"- {res['path']} [{sn}]: 毛利率异常 "
                    f">90% 有 {mb['高于90%']} 个 / <0% 有 {mb['低于0%']} 个 "
                    f"(>90% 多半是成本口径错, 这次肉类 120× 就是这么暴露的, 必查)")
    if lines:
        print("\n".join(lines))
    else:
        print("  (无显著异常项)")
    print("\n⚠️ 通用提醒: 若本次换过事实表 (BOM/价格), 检查毛利率变化是否因"
          "口径修正 — 是修正不是变差, 要跟市场说清楚.")


def main():
    paths = sys.argv[1:]
    if not paths:
        print("用法: venv/bin/python scripts/audit_report.py <xlsx> [<xlsx> ...]")
        return 1
    results = []
    for p in paths:
        try:
            res = audit_one(p)
            print_report(res)
            results.append(res)
        except Exception as e:
            print(f"\n❌ {p}: audit 失败 — {e}")
    if results:
        delivery_notes(results)
    return 0


if __name__ == "__main__":
    sys.exit(main())
