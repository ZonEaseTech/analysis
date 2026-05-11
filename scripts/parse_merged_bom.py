#!/usr/bin/env python3
"""
解析手工维护的"合并 BOM.xlsx" → 输出 fallback BOM 兼容格式。

输入：resources/wallace.20260506/合并BOM_原始.xlsx
  Sheet1 列：Product / 活动内容 / 原材料 / 包装 / (外卖) / 备注
  Sheet2 列：Product / 开始 / 结束 / 活动内容 / 原材料 / 包装 / 价格 / 备注

输出：resources/wallace.20260506/合并BOM_结构化.xlsx
  Sheet "BOM配方(解析结果)"：商品名称 / 物品编号 / 物品名称 / 单耗 / 单位
  Sheet "_来源"：商品名称 / 来源 sheet / 行号 / 时间起 / 时间止 / 是否纳入 / 排除原因
  Sheet "_解析失败"：商品名称 / 来源 / 原始行 / 备注

时间过滤：
  Sheet2 行只保留与 2026-03 任一交集的（默认 month=2026-03 可改）。
  起止为空或非日期文本（如 "จนกว่าสินค้าจะหมด" 直至售完）→ 视为永久有效，纳入。

合并策略：
  原材料 + 包装两段合并解析（与现有 fallback 一致，不区分堂食/外卖）。
  "外卖" 列内容大量与"包装"重复，本版本忽略以避免双倍计数。
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl import Workbook


# ─────────────── 解析正则 ───────────────

# 物品编号：两字母 + 4-8 位数字，允许中间空格（PA 99034 → PA99034）
CODE_RE = re.compile(r"\b([A-Z]{2})\s?(\d{4,8})\b")

# 数量 + 单位（覆盖手写常见写法）
QTY_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*"
    r"(GM|gm|Gm|g|克|個|个|片|Pcs|Pc|pcs|pc|PCS|ml|ML|盒|張|张|个、|片、)",
)
# 兼容括号写法 "(1Pc)" "（1个）" "(30 GM)"
QTY_PAREN_RE = re.compile(
    r"[（(]\s*(\d+(?:\.\d+)?)\s*"
    r"(GM|gm|g|克|個|个|片|Pcs|Pc|pcs|pc|ml|ML|盒|張|张)\s*[）)]",
    re.IGNORECASE,
)

# 单位归一化：克类 → 克；个类 → 个；其他保留
UNIT_MAP = {
    "GM": "克", "gm": "克", "g": "克", "Gm": "克", "克": "克",
    "個": "个", "个": "个", "Pcs": "个", "Pc": "个",
    "pcs": "个", "pc": "个", "PCS": "个", "片": "个", "張": "个", "张": "个", "盒": "个",
    "ml": "ml", "ML": "ml",
}

# 月份过滤区间（默认 3 月）
DEFAULT_MONTH = "2026-03"


def normalize_unit(unit: str) -> str:
    if not unit:
        return ""
    return UNIT_MAP.get(unit, unit)


def parse_line(line: str):
    """从单行文本提取 (code, name, qty, unit)。无 code 返回 None。"""
    if not line:
        return None
    s = line.replace("\t", " ").replace("　", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return None
    cm = CODE_RE.search(s)
    if not cm:
        return None
    code = cm.group(1) + cm.group(2)
    rest = (s[: cm.start()] + " " + s[cm.end():]).strip()

    qm = QTY_RE.search(rest) or QTY_PAREN_RE.search(rest)
    if qm:
        qty = float(qm.group(1))
        unit = normalize_unit(qm.group(2))
        name = (rest[: qm.start()] + " " + rest[qm.end():]).strip()
    else:
        # 兜底 1：行尾或行首孤立整数（如 'Tomato sauce/take away 1'） → 默认"个"
        m = re.search(r"(?:^|\s)(\d+(?:\.\d+)?)(?:\s|$)", rest)
        if m:
            qty = float(m.group(1))
            unit = "个"
            name = (rest[: m.start()] + " " + rest[m.end():]).strip()
        else:
            qty = None
            unit = ""
            name = rest

    # 去掉名称头尾的标点 / 多余空格 / 引号
    name = re.sub(r"[\s,.\-、:：\"']+$", "", name)
    name = re.sub(r"^[\s,.\-、:：\"']+", "", name)

    return (code, name, qty, unit)


def parse_blob(blob: str):
    """解析多行文本，返回 (rows, fail_lines)。
    rows: List[(code, name, qty, unit)]; 完全失败/无 qty 的行进入 fail_lines。
    """
    rows = []
    fails = []
    if not blob:
        return rows, fails
    for raw in blob.split("\n"):
        s = raw.strip()
        if not s:
            continue
        # 跳过显然是换行符残留的孤立片段（如 'P', 'PA', '"', '）'）
        if len(s) <= 2 and not re.search(r"\d", s):
            continue
        p = parse_line(raw)
        if not p:
            fails.append(s)
            continue
        code, name, qty, unit = p
        if qty is None:
            fails.append(f"[无数量] {raw.strip()}")
            continue
        rows.append((code, name, qty, unit))
    return rows, fails


# ─────────────── 时间过滤 ───────────────

def in_month(start_val, end_val, year: int, month: int) -> tuple[bool, str]:
    """返回 (是否纳入, 原因)。"""
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1)
    else:
        month_end = date(year, month + 1, 1)
    # 把 datetime/date/None/文本统一处理
    def to_d(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        # 文本（如 "จนกว่าสินค้าจะหมด" / "until sold out"）→ 视为无界
        return "TEXT"

    s = to_d(start_val)
    e = to_d(end_val)
    # 起止都不是日期 → 视为永久有效
    if (s is None or s == "TEXT") and (e is None or e == "TEXT"):
        return True, "永久有效"
    # 起为日期，结束为文字"直至售完" → 起 ≤ 月末则纳入
    if isinstance(s, date) and (e is None or e == "TEXT"):
        return (s < month_end), f"开始 {s} 起，无截止"
    if (s is None or s == "TEXT") and isinstance(e, date):
        return (e >= month_start), f"无开始，结束 {e}"
    # 两个都是日期
    if isinstance(s, date) and isinstance(e, date):
        ok = (s < month_end) and (e >= month_start)
        return ok, f"{s} ~ {e}"
    return True, "未知，默认纳入"


# ─────────────── 主流程 ───────────────

def normalize_product_name(name: str) -> str:
    """商品名清洗：去多余空白、去末尾空白行。保留多语言/特殊符号。"""
    if not name:
        return ""
    n = name.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    n = re.sub(r"\s+", " ", n).strip()
    return n


def merge_bom_rows(rows):
    """同一商品下 (code, unit) 相同的行合并：累加 qty，取首次出现的 name。"""
    agg = {}
    order = []
    for code, name, qty, unit in rows:
        key = (code, unit)
        if key not in agg:
            agg[key] = {"code": code, "name": name, "qty": qty, "unit": unit}
            order.append(key)
        else:
            agg[key]["qty"] += qty
            # 名称缺失时补上
            if not agg[key]["name"] and name:
                agg[key]["name"] = name
    return [agg[k] for k in order]


def main():
    ap = argparse.ArgumentParser(description="解析手工 BOM 表 → fallback 兼容格式")
    ap.add_argument("--input", default="resources/wallace.20260506/合并BOM_原始.xlsx")
    ap.add_argument("--output", default="resources/wallace.20260506/合并BOM_结构化.xlsx")
    ap.add_argument("--month", default=DEFAULT_MONTH, help="月份过滤 YYYY-MM")
    ap.add_argument("--include-takeout-col", action="store_true",
                    help="是否合并'外卖'列（默认忽略，避免双倍计数）")
    args = ap.parse_args()

    year, month = int(args.month[:4]), int(args.month[5:7])

    print(f"[*] 读取 {args.input}")
    wb = openpyxl.load_workbook(args.input, read_only=True)

    bom_rows = []        # 最终 [(产品名, code, name, qty, unit)]
    source_rows = []     # 来源记录
    fail_rows = []       # 失败行

    # ─── Sheet1：无时间限制 ───
    ws = wb["Sheet1"]
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = row
            continue
        product = normalize_product_name(row[0]) if row[0] else ""
        if not product:
            continue
        materials = row[2] or ""
        packaging = row[3] or ""
        takeout = row[4] if (len(row) > 4 and args.include_takeout_col) else ""
        blob = "\n".join([str(materials), str(packaging), str(takeout or "")])
        rows, fails = parse_blob(blob)
        rows = merge_bom_rows(rows)
        for r in rows:
            bom_rows.append((product, r["code"], r["name"], r["qty"], r["unit"]))
        source_rows.append({
            "product": product, "sheet": "Sheet1", "row": i + 1,
            "start": "", "end": "", "include": "是", "reason": "Sheet1 无期限",
            "bom_count": len(rows), "fail_count": len(fails),
        })
        for f in fails:
            fail_rows.append({"product": product, "source": f"Sheet1!{i+1}", "raw": f})

    # ─── Sheet2：带时间，过滤 ───
    ws = wb["Sheet2"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= 1:  # 前两行表头
            continue
        product = normalize_product_name(row[0]) if row[0] else ""
        if not product:
            continue
        start_v = row[1]
        end_v = row[2]
        materials = row[4] or ""
        packaging = row[5] or ""
        # Sheet2 没有独立的"外卖"列（第 7 列是价格）

        ok, reason = in_month(start_v, end_v, year, month)
        if not ok:
            source_rows.append({
                "product": product, "sheet": "Sheet2", "row": i + 1,
                "start": str(start_v or ""), "end": str(end_v or ""),
                "include": "否", "reason": reason, "bom_count": 0, "fail_count": 0,
            })
            continue

        blob = "\n".join([str(materials), str(packaging)])
        rows, fails = parse_blob(blob)
        rows = merge_bom_rows(rows)
        for r in rows:
            bom_rows.append((product, r["code"], r["name"], r["qty"], r["unit"]))
        source_rows.append({
            "product": product, "sheet": "Sheet2", "row": i + 1,
            "start": str(start_v or ""), "end": str(end_v or ""),
            "include": "是", "reason": reason,
            "bom_count": len(rows), "fail_count": len(fails),
        })
        for f in fails:
            fail_rows.append({"product": product, "source": f"Sheet2!{i+1}", "raw": f})

    # ─── 同名商品合并：相同 (product, code, unit) 累加 qty ───
    # 但 Sheet1 + Sheet2 同名（套餐重复）情况少，先按 (product, code, unit) dedup 即可
    dedup = {}
    order = []
    for product, code, name, qty, unit in bom_rows:
        key = (product, code, unit)
        if key not in dedup:
            dedup[key] = {"product": product, "code": code, "name": name, "qty": qty, "unit": unit}
            order.append(key)
        else:
            dedup[key]["qty"] += qty
            if not dedup[key]["name"]:
                dedup[key]["name"] = name
    final_rows = [dedup[k] for k in order]

    print(f"[*] 解析完成:")
    print(f"    商品总数:        {len(set(r[0] for r in bom_rows))}")
    print(f"    BOM 明细行(去重后): {len(final_rows)}")
    print(f"    解析失败行:      {len(fail_rows)}")
    in_count = sum(1 for s in source_rows if s["include"] == "是")
    ex_count = sum(1 for s in source_rows if s["include"] == "否")
    print(f"    Sheet 来源: 纳入 {in_count} / 排除 {ex_count}")

    # ─── 写出 ───
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb_out = Workbook()

    ws1 = wb_out.active
    ws1.title = "BOM配方(解析结果)"
    ws1.append(["商品名称", "物品编号", "物品名称", "单耗", "单位"])
    # 按商品名 → 物品编号 排序，方便 review
    final_rows.sort(key=lambda r: (r["product"], r["code"]))
    for r in final_rows:
        ws1.append([r["product"], r["code"], r["name"], r["qty"], r["unit"]])

    ws2 = wb_out.create_sheet("_来源")
    ws2.append(["商品名称", "来源 sheet", "行号", "时间起", "时间止",
                "是否纳入", "原因/说明", "BOM 行数", "解析失败数"])
    for s in source_rows:
        ws2.append([s["product"], s["sheet"], s["row"], s["start"], s["end"],
                    s["include"], s["reason"], s["bom_count"], s["fail_count"]])

    ws3 = wb_out.create_sheet("_解析失败")
    ws3.append(["商品名称", "来源", "原始行"])
    for f in fail_rows:
        ws3.append([f["product"], f["source"], f["raw"]])

    wb_out.save(str(out))
    print(f"[*] 已写出: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
