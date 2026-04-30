#!/usr/bin/env python3
"""
Grab 对账单与 TTPOS 订单匹配对账 —— 外部数据注入版

功能：
  1. 读取 Grab 平台对账单（Excel/CSV）
  2. 自动检测列名（门店/日期/金额/Grab订单号）
  3. 按门店+日期批量查询 TTPOS
  4. 按金额匹配 TTPOS 订单（容差 ±1）
  5. 在原表基础上加列：TTPOS订单号、TTPOS订单总价、差异金额、差异状态
  6. 输出对比报表

新增列说明：
  - 订单号一致       ：Grab 订单号 vs TTPOS 订单号 是否一致（是=绿 / 否=红）——用于审计异常
  - TTPOS订单号      ：匹配到的 TTPOS 订单号（POS=serial_number，外卖=order_number）
  - TTPOS订单总价    ：TTPOS 侧订单金额
  - Grab平台金额      ：对账单原始金额（清洗后）
  - 差异金额         ：Grab平台金额 - TTPOS订单总价
  - 差异状态         ：一致(绿) / 微小差异(黄) / 差异(红) / 未匹配(灰)
  - 匹配时间         ：TTPOS 订单的支付/完成时间

列顺序：Grab 订单号（原始列）→ 订单号一致 → TTPOS订单号 → TTPOS订单总价 → ...
       两边订单号相邻排列，方便一眼对比审计。

Usage:
    # 自动检测列名
    venv/bin/python scripts/match_grab_statement.py --input grab_statement.xlsx --output grab_reconciled.xlsx

    # 手动指定列名
    venv/bin/python scripts/match_grab_statement.py \
        --input grab_statement.xlsx \
        --store-col "Store Name" \
        --date-col "Transaction Date" \
        --amount-col "Gross Amount" \
        --order-col "Order ID" \
        --output grab_reconciled.xlsx

    # CSV 文件 + 指定日期格式
    venv/bin/python scripts/match_grab_statement.py \
        --input grab_statement.csv \
        --date-format "%d/%m/%Y" \
        --output grab_reconciled.xlsx
"""

import argparse
import re
import sys
from datetime import datetime, timezone
from datetime import date as Date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, '/home/weifashi/hwt/analysis')
sys.path.insert(0, '/home/weifashi/hwt/analysis/scripts')
from bq_reports.utils.bq_client import get_bq_client, setup_proxy
from trace_order_by_payment import (
    query_pos_orders, query_takeout_orders, _ts_to_bkk_str, _calc_diff
)

PROJECT_ID = "diyl-407103"
BQ_LOCATION = "asia-southeast1"
THAILAND_TZ_OFFSET = 7 * 3600

# ===== 列名自动检测候选 =====
STORE_CANDIDATES = ['store', 'merchant', 'outlet', 'branch', 'restaurant', '门店', '商家', '店铺', '分店', 'store name', 'outlet name', 'branch name']
DATE_CANDIDATES = ['date', 'transaction date', 'order date', 'settlement date', '日期', '交易日期', '订单日期', '结算日期']
AMOUNT_CANDIDATES = ['amount', 'total', 'price', 'gross amount', 'net amount', 'total amount', 'order amount', '金额', '总价', '订单金额', 'gross']
ORDER_CANDIDATES = ['order number', 'order id', 'order no', 'transaction id', 'reference', 'ref', '订单号', '交易号', '订单编号']
TIME_CANDIDATES = ['time', 'transaction time', 'order time', '时间']


def _detect_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """自动检测列名，返回匹配到的原始列名"""
    col_map = {c.lower().strip(): c for c in df.columns}
    for c in candidates:
        if c in col_map:
            return col_map[c]
    # 模糊匹配
    for col in df.columns:
        col_lower = col.lower().strip()
        for c in candidates:
            if c in col_lower or col_lower in c:
                return col
    return None


def _parse_amount(val) -> float:
    """清洗金额字符串，去除货币符号和千分位"""
    if pd.isna(val):
        return 0.0
    s = str(val).strip()
    # 去除货币符号
    s = re.sub(r'[฿$€¥£]', '', s)
    # 去除空白
    s = s.replace(' ', '')
    # 处理千分位逗号
    if ',' in s and '.' in s:
        # 有逗号有点，判断哪个是小数点
        last_comma = s.rfind(',')
        last_dot = s.rfind('.')
        if last_comma > last_dot:
            # 逗号是小数点（欧洲格式）
            s = s.replace('.', '').replace(',', '.')
        else:
            # 点是小数点（美式格式）
            s = s.replace(',', '')
    elif ',' in s:
        # 只有逗号，可能是小数点（欧洲）或千分位
        parts = s.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_date(val, fmt: str | None = None) -> Date | None:
    """解析日期字符串"""
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if fmt:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    # 自动尝试常见格式
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y%m%d']:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # 尝试 pandas 解析
    try:
        return pd.to_datetime(s).date()
    except (ValueError, pd.errors.OutOfBoundsDatetime):
        return None


def _date_to_ts_range(d: Date) -> tuple[int, int]:
    """日期转 Unix 时间戳范围（泰国时间当日 00:00~24:00）"""
    dt_start = datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
    start_ts = int(dt_start.timestamp()) - THAILAND_TZ_OFFSET
    end_ts = start_ts + 86400
    return start_ts, end_ts


def find_all_stores(client) -> list[dict]:
    """获取所有华莱士泰国门店"""
    job = client.query('''
        SELECT c.uuid, c.name, cs.erpnext_company_abbr
        FROM `diyl-407103`.`saas`.`ttpos_company` c
        LEFT JOIN `diyl-407103`.`saas`.`ttpos_company_setting` cs
          ON cs.company_uuid = c.uuid AND cs.delete_time = 0
        WHERE c.delete_time = 0
          AND cs.headquarter_uuid = 5080409448448000
          AND cs.erpnext_company_abbr LIKE 'TH%%'
        ORDER BY cs.erpnext_company_abbr
    ''', location=BQ_LOCATION)

    stores = []
    datasets = list(client.list_datasets())
    shop_dataset_ids = set(d.dataset_id for d in datasets if d.dataset_id.startswith('shop'))

    for row in job.result():
        dataset = f"shop{row.uuid}"
        if dataset in shop_dataset_ids:
            stores.append({
                'uuid': str(row.uuid),
                'name': row.name or '',
                'abbr': row.erpnext_company_abbr or '',
                'dataset': dataset,
            })
    return stores


def match_store(store_name: str, all_stores: list[dict]) -> dict | None:
    """模糊匹配门店名称"""
    if not store_name or pd.isna(store_name):
        return None
    s_lower = str(store_name).lower().strip()

    # 精确匹配
    for s in all_stores:
        if s_lower == s['name'].lower() or s_lower == s['abbr'].lower():
            return s

    # 包含匹配
    for s in all_stores:
        if s_lower in s['name'].lower() or s['name'].lower() in s_lower:
            return s
        if s['abbr'] and (s_lower in s['abbr'].lower() or s['abbr'].lower() in s_lower):
            return s

    # 关键词匹配（去掉常见后缀）
    s_clean = re.sub(r'\s*(wallace|burger|restaurant|store|branch|\d+)$', '', s_lower).strip()
    if s_clean and s_clean != s_lower:
        for s in all_stores:
            name_clean = re.sub(r'\s*(wallace|burger|restaurant|store|branch|\d+)$', '', s['name'].lower()).strip()
            if s_clean in name_clean or name_clean in s_clean:
                return s

    return None


def fetch_orders_for_date(client, store: dict, date_val: datetime.date) -> list[dict]:
    """查询指定门店指定日期的所有 Grab 相关订单"""
    start_ts, end_ts = _date_to_ts_range(date_val)

    # POS 侧
    pos_orders = query_pos_orders(client, store, 'grab', start_ts, end_ts, None, None, None)
    for o in pos_orders:
        o['source'] = 'pos'
        o['match_time'] = o.get('complete_time', 0)

    # 外卖侧
    to_orders = query_takeout_orders(client, store, 'grab', start_ts, end_ts, None, None, None)
    for o in to_orders:
        o['source'] = 'takeout'
        o['match_time'] = o.get('pay_time', 0)

    return pos_orders + to_orders


def match_order(grab_amount: float, orders: list[dict], grab_order_id: str = "",
                tolerance: float = 1.0) -> tuple[dict | None, int]:
    """按订单号优先、金额次之的顺序匹配最佳订单。

    Returns:
        (matched_order, matched_index) — matched_index 为 -1 表示未匹配
    """
    if not orders:
        return None, -1

    grab_oid = str(grab_order_id).strip() if grab_order_id else ""

    # 1. 优先用 Grab 订单号匹配（TTPOS 外卖订单的 order_number 通常就是 Grab 订单号）
    if grab_oid:
        for i, o in enumerate(orders):
            if o['source'] == 'takeout':
                ttpos_oid = str(o.get('order_number', '')).strip()
                if ttpos_oid and grab_oid.lower() == ttpos_oid.lower():
                    return o, i

    # 2. 金额匹配
    best = None
    best_idx = -1
    best_diff = float('inf')

    for i, o in enumerate(orders):
        if o['source'] == 'pos':
            order_amount = o.get('payment_amount', 0)
        else:
            order_amount = o.get('platform_total', 0)

        diff = abs(order_amount - grab_amount)
        if diff < best_diff:
            best_diff = diff
            best = o
            best_idx = i

    if best and best_diff <= tolerance:
        return best, best_idx
    return None, -1


def export_matched(df: pd.DataFrame, output_path: str, unmatched_count: int, order_col: str | None = None):
    """导出带匹配结果的对账报表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "对账结果"

    # 样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    diff_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    diff_red_font = Font(color='9C0006')
    diff_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    diff_green_font = Font(color='006100')
    diff_yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    diff_yellow_font = Font(color='9C5700')
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    gray_font = Font(color='666666')

    # 写入 DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(vertical='center')

    # 对新增列做条件格式高亮
    diff_status_col = None
    order_match_col = None
    for c_idx, col_name in enumerate(df.columns, 1):
        if col_name == '差异状态':
            diff_status_col = c_idx
        if col_name == '订单号一致':
            order_match_col = c_idx

    if diff_status_col:
        for r_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=r_idx, column=diff_status_col).value
            cell = ws.cell(row=r_idx, column=diff_status_col)
            if status == '一致':
                cell.fill = diff_green_fill
                cell.font = diff_green_font
            elif status == '微小差异':
                cell.fill = diff_yellow_fill
                cell.font = diff_yellow_font
            elif status == '差异':
                cell.fill = diff_red_fill
                cell.font = diff_red_font
            elif status == '未匹配':
                cell.fill = gray_fill
                cell.font = gray_font

    # 订单号一致列高亮：否=红色，是=绿色
    if order_match_col:
        for r_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=r_idx, column=order_match_col).value
            cell = ws.cell(row=r_idx, column=order_match_col)
            if val == '否':
                cell.fill = diff_red_fill
                cell.font = diff_red_font
            elif val == '是':
                cell.fill = diff_green_fill
                cell.font = diff_green_font

    # 自动调整列宽
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    matched = len(df) - unmatched_count
    diff_count = len(df[df['差异状态'].isin(['差异', '微小差异'])])
    order_mismatch = len(df[df['订单号一致'] == '否']) if '订单号一致' in df.columns else 0
    print(f"\n报表已导出: {output_path}")
    print(f"  总记录: {len(df)} 行")
    print(f"  匹配成功: {matched} 行 (含 {diff_count} 行金额差异)")
    if order_mismatch > 0:
        print(f"  订单号不一致: {order_mismatch} 行 (红色高亮)")
    print(f"  未匹配: {unmatched_count} 行")


def main():
    parser = argparse.ArgumentParser(description="Grab 对账单与 TTPOS 订单匹配对账")
    parser.add_argument("--input", required=True, help="Grab 对账单文件路径（Excel/CSV）")
    parser.add_argument("--store-col", help="门店名称列名（默认自动检测）")
    parser.add_argument("--date-col", help="日期列名（默认自动检测）")
    parser.add_argument("--amount-col", help="金额列名（默认自动检测）")
    parser.add_argument("--order-col", help="Grab 订单号列名（可选，默认自动检测）")
    parser.add_argument("--date-format", help="日期格式（如 %%d/%%m/%%Y），默认自动解析")
    parser.add_argument("--tolerance", type=float, default=1.0, help="金额匹配容差（默认 ±1.0）")
    parser.add_argument("--output", default="exports/grab_reconciled.xlsx", help="输出路径")
    args = parser.parse_args()

    # 1. 读取文件
    print(f"[1/5] 读取对账单: {args.input}")
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"错误: 文件不存在: {args.input}")
        return 1

    if input_path.suffix.lower() == '.csv':
        df = pd.read_csv(args.input)
    else:
        df = pd.read_excel(args.input)

    print(f"  共 {len(df)} 行, {len(df.columns)} 列: {list(df.columns)}")

    # 2. 检测/确认列名
    print("\n[2/5] 检测列名...")
    store_col = args.store_col or _detect_column(df, STORE_CANDIDATES)
    date_col = args.date_col or _detect_column(df, DATE_CANDIDATES)
    amount_col = args.amount_col or _detect_column(df, AMOUNT_CANDIDATES)
    order_col = args.order_col or _detect_column(df, ORDER_CANDIDATES)

    if not store_col:
        print("错误: 无法检测门店列，请通过 --store-col 指定")
        print(f"  可用列: {list(df.columns)}")
        return 1
    if not date_col:
        print("错误: 无法检测日期列，请通过 --date-col 指定")
        return 1
    if not amount_col:
        print("错误: 无法检测金额列，请通过 --amount-col 指定")
        return 1

    print(f"  门店列: '{store_col}'")
    print(f"  日期列: '{date_col}'")
    print(f"  金额列: '{amount_col}'")
    if order_col:
        print(f"  Grab订单号列: '{order_col}'")

    # 3. 解析金额和日期
    print("\n[3/5] 解析数据...")
    df['_grab_amount'] = df[amount_col].apply(_parse_amount)
    df['_date'] = df[date_col].apply(lambda x: _parse_date(x, args.date_format))

    # 过滤无效数据
    valid_mask = (df['_grab_amount'] > 0) & (df['_date'].notna())
    invalid = len(df) - valid_mask.sum()
    if invalid > 0:
        print(f"  跳过 {invalid} 行无效数据（金额≤0 或日期无效）")
    df_valid = df[valid_mask].copy()

    if len(df_valid) == 0:
        print("错误: 没有有效数据可处理")
        return 1

    # 4. 连接 BQ 并查询 TTPOS
    print("\n[4/5] 连接 BigQuery 查询 TTPOS...")
    setup_proxy()
    client = get_bq_client(project_id=PROJECT_ID)

    all_stores = find_all_stores(client)
    print(f"  共 {len(all_stores)} 家 TTPOS 门店")

    # 按门店匹配
    df_valid['_matched_store'] = df_valid[store_col].apply(lambda x: match_store(x, all_stores))
    unmatched_store = df_valid['_matched_store'].isna().sum()
    if unmatched_store > 0:
        print(f"  警告: {unmatched_store} 行无法匹配到 TTPOS 门店")
        # 列出未匹配的唯一门店名
        unmatched_names = df_valid[df_valid['_matched_store'].isna()][store_col].unique()
        for name in unmatched_names[:5]:
            print(f"    - '{name}'")
        if len(unmatched_names) > 5:
            print(f"    ... 还有 {len(unmatched_names) - 5} 个")

    # 按 (门店, 日期) 分组批量查询
    query_groups = {}
    for idx, row in df_valid.iterrows():
        store = row['_matched_store']
        if store is None:
            continue
        date_val = row['_date']
        key = (store['uuid'], date_val)
        if key not in query_groups:
            query_groups[key] = {'store': store, 'date': date_val, 'rows': []}
        query_groups[key]['rows'].append(idx)

    print(f"  需要查询 {len(query_groups)} 个门店/日期组合...")

    # 批量查询
    orders_cache = {}
    for i, (key, group) in enumerate(query_groups.items(), 1):
        store = group['store']
        date_val = group['date']
        print(f"  [{i}/{len(query_groups)}] {store['abbr']} | {store['name']} | {date_val}...", end='')
        orders = fetch_orders_for_date(client, store, date_val)
        orders_cache[key] = orders
        print(f" {len(orders)} 笔订单")

    # 5. 逐行匹配
    print("\n[5/5] 匹配订单...")
    results = []
    unmatched_count = 0

    for idx, row in df_valid.iterrows():
        store = row['_matched_store']
        date_val = row['_date']
        grab_amount = row['_grab_amount']
        grab_oid = str(row[order_col]).strip() if order_col and not pd.isna(row.get(order_col, '')) else ''

        if store is None or date_val is None:
            results.append({
                'TTPOS订单号': '',
                'TTPOS订单总价': '',
                'Grab平台金额': grab_amount,
                '差异金额': '',
                '差异状态': '未匹配',
                '匹配时间': '',
                '订单号一致': '',
            })
            unmatched_count += 1
            continue

        key = (store['uuid'], date_val)
        orders = orders_cache.get(key, [])
        matched, matched_idx = match_order(grab_amount, orders, grab_oid, tolerance=args.tolerance)

        if matched:
            # 防止同一笔 TTPOS 订单被多行匹配
            if matched_idx >= 0:
                orders.pop(matched_idx)

            grab_price, ttpos_price, diff, diff_status, order_no = _calc_diff(matched)

            # 判断订单号是否一致（外卖订单 TTPOS order_number 通常就是 Grab 订单号）
            order_match = ''
            if grab_oid and order_no:
                if grab_oid.lower() == str(order_no).strip().lower():
                    order_match = '是'
                else:
                    order_match = '否'

            results.append({
                'TTPOS订单号': order_no,
                'TTPOS订单总价': ttpos_price,
                'Grab平台金额': grab_amount,
                '差异金额': diff,
                '差异状态': diff_status,
                '匹配时间': _ts_to_bkk_str(matched.get('match_time', 0)),
                '订单号一致': order_match,
            })
            if diff_status != '一致':
                unmatched_count += 1
        else:
            results.append({
                'TTPOS订单号': '',
                'TTPOS订单总价': '',
                'Grab平台金额': grab_amount,
                '差异金额': '',
                '差异状态': '未匹配',
                '匹配时间': '',
                '订单号一致': '',
            })
            unmatched_count += 1

    # 合并结果到原 DataFrame
    result_df = pd.DataFrame(results)
    for col in result_df.columns:
        df_valid[col] = result_df[col].values

    # 清理临时列
    df_valid = df_valid.drop(columns=['_grab_amount', '_date', '_matched_store'], errors='ignore')

    # 调整列顺序：Grab 订单号 和 TTPOS 订单号 放一起
    cols = list(df_valid.columns)
    new_cols = ['TTPOS订单号', 'TTPOS订单总价', 'Grab平台金额', '差异金额', '差异状态', '匹配时间', '订单号一致']
    for col in new_cols:
        if col in cols:
            cols.remove(col)

    # 找到 Grab 订单号列的位置，把 TTPOS 订单号放在它旁边
    if order_col and order_col in cols:
        order_idx = cols.index(order_col)
        # Grab 订单号后面依次插入：订单号一致、TTPOS订单号、TTPOS订单总价、Grab平台金额、差异金额、差异状态、匹配时间
        insert_cols = ['订单号一致', 'TTPOS订单号', 'TTPOS订单总价', 'Grab平台金额', '差异金额', '差异状态', '匹配时间']
        for col in reversed(insert_cols):
            if col in new_cols:  # 确保是新增列
                cols.insert(order_idx + 1, col)
    elif amount_col in cols:
        #  fallback：放在金额列后面
        amount_idx = cols.index(amount_col)
        for col in reversed(new_cols):
            cols.insert(amount_idx + 1, col)
    else:
        cols = new_cols + cols

    df_valid = df_valid[[c for c in cols if c in df_valid.columns]]

    # 导出
    export_matched(df_valid, args.output, unmatched_count, order_col)

    return 0


if __name__ == "__main__":
    sys.exit(main())
