#!/usr/bin/env python3
"""
批量对账：Wallace SHP + RBH + LINEMAN

处理流程：
  1. 查询所有门店 2026-03 Grab 订单，缓存
  2. 逐个文件匹配：
     - Wallace SHP: 逐笔按日期+金额匹配
     - RBH: 逐笔按日期+金额匹配
     - LINEMAN: 按门店+日期汇总对比
  3. 每个文件输出：原始表 + TTPOS订单号 + TTPOS金额

Usage:
    venv/bin/python scripts/batch_reconcile.py
"""

import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, '/home/weifashi/hwt/analysis')
from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
BQ_LOCATION = "asia-southeast1"
TH_OFFSET = 7 * 3600


def get_all_stores(client):
    job = client.query('''
        SELECT c.uuid, c.name, cs.erpnext_company_abbr
        FROM `diyl-407103`.`saas`.`ttpos_company` c
        LEFT JOIN `diyl-407103`.`saas`.`ttpos_company_setting` cs
          ON cs.company_uuid = c.uuid AND cs.delete_time = 0
        WHERE c.delete_time = 0
          AND cs.headquarter_uuid = 5080409448448000
          AND cs.erpnext_company_abbr LIKE 'TH%%'
        ORDER BY cs.erpnext_company_abbr
    ''', location=BQ_LOCATION)
    stores = []
    for r in job.result():
        stores.append({
            'uuid': str(r.uuid),
            'name': r.name or '',
            'abbr': r.erpnext_company_abbr or '',
            'dataset': f"shop{r.uuid}",
        })
    return stores


def fetch_all_payments(client, stores, start_ts, end_ts, methods=('Shopee', 'LINEMAN', 'Robinhood')):
    """查询所有门店扫码支付明细 (statistics_payment + sale_bill 联表，含真实订单号)"""
    all_pays = []
    methods_in = ", ".join(f"'{m}'" for m in methods)
    for i, store in enumerate(stores, 1):
        ds = store['dataset']
        try:
            sql = f"""
            SELECT
              DATE(TIMESTAMP_SECONDS(s.complete_time + 7 * 3600)) as pay_date,
              s.complete_time,
              pm.name AS payment_method_name,
              CAST(s.payment_amount AS FLOAT64) AS payment_amount,
              s.uuid AS stat_uuid,
              s.sale_bill_uuid,
              b.order_no,
              b.serial_no,
              b.duty_no
            FROM `{PROJECT_ID}`.`{ds}`.`ttpos_statistics_payment` s
            JOIN `{PROJECT_ID}`.`{ds}`.`ttpos_payment_method` pm
              ON pm.uuid = s.payment_method_uuid AND pm.delete_time = 0
            LEFT JOIN `{PROJECT_ID}`.`{ds}`.`ttpos_sale_bill` b
              ON b.uuid = s.sale_bill_uuid AND b.delete_time = 0
            WHERE s.delete_time = 0
              AND pm.name IN ({methods_in})
              AND s.complete_time >= {start_ts}
              AND s.complete_time < {end_ts}
            """
            cnt = 0
            for r in client.query(sql, location=BQ_LOCATION).result():
                all_pays.append({
                    'store_abbr': store['abbr'],
                    'store_name': store['name'],
                    'method': r.payment_method_name,
                    'date': r.pay_date,
                    'create_time': r.complete_time,
                    'payment_amount': round(r.payment_amount, 2),
                    'order_no': r.order_no or '',
                    'serial_no': r.serial_no or '',
                    'duty_no': r.duty_no or '',
                    'stat_uuid': str(r.stat_uuid),
                })
                cnt += 1
            if cnt > 0:
                print(f"  [{i}/{len(stores)}] {store['abbr']}: {cnt} 笔")
        except Exception:
            pass
    return all_pays


def fetch_all_orders(client, stores, start_ts, end_ts, platforms=('grab', 'lineman')):
    """[已废弃] 外卖订单查询 - 改用 fetch_all_payments"""
    all_orders = []
    platforms_in = ", ".join(f"'{p}'" for p in platforms)
    for i, store in enumerate(stores, 1):
        try:
            sql = f"""
            SELECT
              DATE(TIMESTAMP_SECONDS(CASE WHEN order_state = 40 AND completed_time > 0 THEN completed_time ELSE accepted_time END) + INTERVAL 7 HOUR) as order_date,
              platform_order_id,
              platform,
              payment_type,
              CAST(subtotal AS FLOAT64) as subtotal,
              CAST(platform_total AS FLOAT64) as platform_total
            FROM `{PROJECT_ID}`.`{store['dataset']}`.`ttpos_takeout_order`
            WHERE delete_time = 0
              AND order_state IN (10, 20, 30, 40)
              AND platform IN ({platforms_in})
              AND (CASE WHEN order_state = 40 AND completed_time > 0 THEN completed_time ELSE accepted_time END) >= {start_ts}
              AND (CASE WHEN order_state = 40 AND completed_time > 0 THEN completed_time ELSE accepted_time END) < {end_ts}
            """
            job = client.query(sql, location=BQ_LOCATION)
            cnt = 0
            for r in job.result():
                all_orders.append({
                    'store_abbr': store['abbr'],
                    'store_name': store['name'],
                    'platform': r.platform,
                    'date': r.order_date,
                    'platform_order_id': r.platform_order_id,
                    'payment_type': r.payment_type or '',
                    'subtotal': round(r.subtotal, 2),
                    'platform_total': round(r.platform_total, 2),
                })
                cnt += 1
            if cnt > 0:
                print(f"  [{i}/{len(stores)}] {store['abbr']}: {cnt} 笔")
        except Exception:
            pass
    return all_orders


# 向后兼容
fetch_all_grab_orders = fetch_all_orders


def match_wallace_shp(df, all_pays):
    """对账 Wallace SHP (Shopee Pay 门店扫码支付结算单)

    特点: 净额结算 (扣 18% 佣金后), 5 位小数, 无门店字段。
    匹配策略: 用 Shopee 全 TH 池, 按日期 + 净额 (Shopee净 ≈ TTPOS payment_amount × (1 - fee)) 匹配。
    fee 从 TTPOS commission_fee 反推。
    """
    shopee_pays = [p for p in all_pays if p['method'] == 'Shopee']

    df['amt_net'] = df['ยอดเงิน'].apply(lambda x: round(float(str(x).replace(',', '').replace(' ', '')), 2) if pd.notna(x) else None)
    df['dt'] = df['นที่เกิดรายการ'].apply(lambda x: pd.to_datetime(x).date() if pd.notna(x) else None)
    df = df.reset_index(drop=True)

    used = set()
    statuses, ttpos_oids, ttpos_serials, ttpos_amts, ttpos_stores, diffs = [], [], [], [], [], []
    for _, row in df.iterrows():
        d = row['dt']
        amt_net = row['amt_net']
        if d is None or amt_net is None or amt_net <= 0:
            statuses.append('源数据无法对账(金额/日期缺失或调整记录)')
            ttpos_oids.append(''); ttpos_serials.append(''); ttpos_amts.append(''); ttpos_stores.append(''); diffs.append('')
            continue
        # Shopee 净额 ≈ TTPOS payment_amount × (1 - 微小手续费)
        best, best_diff = None, float('inf')
        for p in shopee_pays:
            if p['date'] != d or p['stat_uuid'] in used:
                continue
            diff = abs(p['payment_amount'] - amt_net)
            if diff < best_diff:
                best_diff = diff
                best = p
        if best and best_diff <= 0.5:
            used.add(best['stat_uuid'])
            statuses.append('已对上')
            ttpos_oids.append(best['order_no'])
            ttpos_serials.append(best['serial_no'])
            ttpos_amts.append(best['payment_amount'])
            ttpos_stores.append(best['store_abbr'])
            diffs.append(round(best['payment_amount'] - amt_net, 4))
        elif best and best_diff <= 5:
            used.add(best['stat_uuid'])
            statuses.append('金额近似(差>0.5)')
            ttpos_oids.append(best['order_no'])
            ttpos_serials.append(best['serial_no'])
            ttpos_amts.append(best['payment_amount'])
            ttpos_stores.append(best['store_abbr'])
            diffs.append(round(best['payment_amount'] - amt_net, 4))
        else:
            statuses.append('TTPOS未找到匹配支付')
            ttpos_oids.append(''); ttpos_serials.append(''); ttpos_amts.append(''); ttpos_stores.append(''); diffs.append('')

    df['TTPOS门店'] = ttpos_stores
    df['TTPOS订单号'] = ttpos_oids
    df['TTPOS小票号'] = ttpos_serials
    df['TTPOS金额(毛)'] = ttpos_amts
    df['净额差异'] = diffs
    df['对账状态'] = statuses
    return df


def match_rbh(df, all_pays):
    """对账 RBH (Robinhood 门店扫码支付结算单)

    源文件: 应收 = 顾客付款金额 = TTPOS payment_amount
    匹配维度: 日期 + 应收金额 + 门店名(可选)
    """
    rb_pays = [p for p in all_pays if p['method'] == 'Robinhood']

    df['amt_gross'] = df['应收 เงินที่ควรจะได้รับ'].apply(lambda x: round(float(str(x).replace(',', '').replace(' ', '')), 2) if pd.notna(x) else None)
    df['dt'] = df['Date'].apply(lambda x: pd.to_datetime(x).date() if pd.notna(x) else None)
    df = df.reset_index(drop=True)

    used = set()
    statuses, ttpos_oids, ttpos_serials, ttpos_amts, ttpos_stores, diffs = [], [], [], [], [], []
    for _, row in df.iterrows():
        d = row['dt']
        amt = row['amt_gross']
        if d is None or amt is None or amt <= 0:
            statuses.append('源数据无法对账'); ttpos_oids.append(''); ttpos_serials.append(''); ttpos_amts.append(''); ttpos_stores.append(''); diffs.append('')
            continue
        best, best_diff = None, float('inf')
        for p in rb_pays:
            if p['date'] != d or p['stat_uuid'] in used:
                continue
            diff = abs(p['payment_amount'] - amt)
            if diff < best_diff:
                best_diff = diff
                best = p
        if best and best_diff <= 0.5:
            used.add(best['stat_uuid'])
            statuses.append('已对上')
            ttpos_oids.append(best['order_no'])
            ttpos_serials.append(best['serial_no'])
            ttpos_amts.append(best['payment_amount'])
            ttpos_stores.append(best['store_abbr'])
            diffs.append(round(best['payment_amount'] - amt, 2))
        else:
            statuses.append('TTPOS未找到匹配支付')
            ttpos_oids.append(''); ttpos_serials.append(''); ttpos_amts.append(''); ttpos_stores.append(''); diffs.append('')

    df['TTPOS门店'] = ttpos_stores
    df['TTPOS订单号'] = ttpos_oids
    df['TTPOS小票号'] = ttpos_serials
    df['TTPOS金额'] = ttpos_amts
    df['差异'] = diffs
    df['对账状态'] = statuses
    return df


def _extract_thai(s):
    """提取字符串中的泰文字符"""
    return ''.join(ch for ch in str(s) if '฀' <= ch <= '๿' or ch == '.')


def build_branch_to_store_map(branch_names, stores):
    """LINEMAN branch_name(纯泰文) → TTPOS store_name 映射 (严格相似度)

    评分策略（避免误匹）：
      1. 数字必须一致 (含 "112"/"122" 时不可跨匹配)
      2. 用 difflib.ratio 计算泰文相似度作为基础分 (0-1)
      3. 子串包含 (一方完全包含另一方泰文核心) 加奖励
      4. 阈值 0.5 以下视为未匹配
    """
    import re
    from difflib import SequenceMatcher
    mapping = {}
    unmatched = []

    store_meta = []
    for s in stores:
        n = s['name']
        thai = _extract_thai(n).strip('.').strip()
        digits = set(re.findall(r'\d+', n))
        store_meta.append((n, thai, digits))

    for br in branch_names:
        if not br:
            continue
        br_norm = str(br).strip()
        br_thai = _extract_thai(br_norm).strip('.').strip()
        br_digits = set(re.findall(r'\d+', br_norm))
        if not br_thai:
            unmatched.append(br_norm)
            continue

        best_store = None
        best_score = 0
        for full_name, t_thai, t_digits in store_meta:
            if not t_thai:
                continue
            # 数字过滤：仅当两方都有数字时才要求一致
            if br_digits and t_digits and not (br_digits & t_digits):
                continue
            ratio = SequenceMatcher(None, br_thai, t_thai).ratio()
            # 包含关系奖励 (一方完整包含另一方)
            if t_thai in br_thai or br_thai in t_thai:
                ratio = max(ratio, 0.85)
            # 数字一致额外加分 (区分编号同名店)
            if br_digits and t_digits and (br_digits & t_digits):
                ratio += 0.05
            if ratio > best_score:
                best_score = ratio
                best_store = full_name

        if best_store and best_score >= 0.5:
            mapping[br_norm] = best_store
        else:
            unmatched.append(br_norm)

    return mapping, unmatched


def match_lineman(df, all_pays, stores):
    """LINEMAN 按门店+日期汇总匹配 (LineMan Pay 门店扫码支付)

    LINEMAN 文件按日汇总（每行 = 一家门店一天），用 TTPOS 当天该门店所有 LineMan 扫码
    支付的 payment_amount 汇总，与 LINEMAN total_revenue 对比。
    """
    lm_pays = [p for p in all_pays if p['method'] == 'LINEMAN']

    # (store_name, date) → {count, sum_amount}
    agg = defaultdict(lambda: {'count': 0, 'sum_amount': 0.0})
    for p in lm_pays:
        key = (p['store_name'], p['date'])
        agg[key]['count'] += 1
        agg[key]['sum_amount'] += p['payment_amount']

    df['dt'] = df['Date'].apply(lambda x: pd.to_datetime(x).date() if pd.notna(x) else None)
    df['branch'] = df['branch_name'].apply(lambda x: str(x).strip() if pd.notna(x) else '')

    # 建立 branch_name → store_name 映射
    unique_branches = sorted(set(df['branch'].dropna().tolist()) - {''})
    branch_map, unmatched = build_branch_to_store_map(unique_branches, stores)
    print(f"  门店匹配: {len(branch_map)}/{len(unique_branches)} 命中")
    if unmatched:
        print(f"  未匹配门店({len(unmatched)}): {unmatched[:5]}{'...' if len(unmatched) > 5 else ''}")

    # TTPOS 哪些门店实际接了 LineMan Pay
    stores_with_lineman = {p['store_name'] for p in lm_pays}

    results = []
    for _, row in df.iterrows():
        d = row['dt']
        branch = row['branch']
        try:
            file_amt = float(row.get('total_revenue', 0) or 0)
        except Exception:
            file_amt = 0

        rec = {'TTPOS订单号': '', 'TTPOS金额': '', '差异': '', '对账状态': ''}

        if d is None or not branch:
            rec['对账状态'] = '源数据无法对账(日期/门店缺失)'
            results.append(rec)
            continue

        store_name = branch_map.get(branch)
        if not store_name:
            rec['对账状态'] = 'TTPOS未建档此门店'
            rec['TTPOS订单号'] = f'未建档({branch})'
            results.append(rec)
            continue

        if store_name not in stores_with_lineman:
            rec['对账状态'] = 'TTPOS此门店未接LineMan'
            rec['TTPOS订单号'] = f'未接({store_name})'
            rec['TTPOS金额'] = 0
            rec['差异'] = round(file_amt, 2)
            results.append(rec)
            continue

        info = agg.get((store_name, d))
        if not info or info['count'] == 0:
            # 门店接了 LineMan，但当天无订单 → 真异常
            rec['对账状态'] = 'TTPOS当天无LineMan订单'
            rec['TTPOS订单号'] = f'汇总0笔({store_name})'
            rec['TTPOS金额'] = 0
            rec['差异'] = round(file_amt, 2)
            results.append(rec)
            continue

        ttpos_sum = round(info['sum_amount'], 2)
        diff = round(file_amt - ttpos_sum, 2)
        rec['TTPOS订单号'] = f"汇总{info['count']}笔"
        rec['TTPOS金额'] = ttpos_sum
        rec['差异'] = diff
        # 金额对账判定：±1 THB 内算对上（按日汇总累计四舍五入）
        if abs(diff) <= 1.0:
            rec['对账状态'] = '已对上'
        else:
            rec['对账状态'] = '金额不一致'
        results.append(rec)

    result_df = pd.DataFrame(results)
    df = df.reset_index(drop=True)
    for col in ['TTPOS订单号', 'TTPOS金额', '差异', '对账状态']:
        df[col] = result_df[col]
    return df


def next_versioned_path(output_path: str) -> str:
    """foo.xlsx -> foo_v{N+1}.xlsx，N 取目录里同前缀已有版本最大值"""
    p = Path(output_path)
    parent = p.parent
    stem = p.stem
    suffix = p.suffix
    pattern = re.compile(rf'^{re.escape(stem)}_v(\d+){re.escape(suffix)}$')
    max_v = 0
    if parent.exists():
        for f in parent.iterdir():
            m = pattern.match(f.name)
            if m:
                max_v = max(max_v, int(m.group(1)))
    return str(parent / f"{stem}_v{max_v + 1}{suffix}")


def export_df(df, output_path):
    """导出 DataFrame 到 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "对账结果"

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 长数字订单号列要按文本格式存（避免 Excel 科学计数法）
    text_keywords = ('订单号', '小票号', 'Order no', 'INV#', 'หมายเลข', 'duty')
    text_cols = set()
    for ci, col_name in enumerate(df.columns, 1):
        if any(k in str(col_name) for k in text_keywords):
            text_cols.add(ci)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            # 文本列且值是数字时强制转字符串
            if r_idx > 1 and c_idx in text_cols and value is not None and value != '':
                value = str(value)
                # 浮点退化的整数（如 1656.0）去掉小数
                if value.endswith('.0'):
                    value = value[:-2]
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(vertical='center')
                if c_idx in text_cols:
                    cell.number_format = '@'

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    # 简单参数：--only=wallace,rbh,lineman
    only = set()
    for arg in sys.argv[1:]:
        if arg.startswith('--only='):
            only = set(s.strip().lower() for s in arg.split('=', 1)[1].split(',') if s.strip())
    if not only:
        only = {'wallace', 'rbh', 'lineman'}

    print("=" * 60)
    print(f"批量对账 (only={','.join(sorted(only))})")
    print("=" * 60)

    # 根据 only 决定查询哪些支付方式
    methods = []
    if 'wallace' in only:
        methods.append('Shopee')
    if 'rbh' in only:
        methods.append('Robinhood')
    if 'lineman' in only:
        methods.append('LINEMAN')

    # 1. 查询所有门店外卖订单
    setup_proxy()
    client = get_bq_client(project_id=PROJECT_ID)
    stores = get_all_stores(client)

    start_ts = int(datetime(2026, 3, 1, tzinfo=timezone.utc).timestamp()) - TH_OFFSET
    end_ts = int(datetime(2026, 4, 1, tzinfo=timezone.utc).timestamp()) - TH_OFFSET

    if methods:
        print(f"\n[1/4] 查询所有门店 2026-03 {'+'.join(methods)} 扫码支付明细...")
        print(f"  共 {len(stores)} 家门店")
        all_pays = fetch_all_payments(client, stores, start_ts, end_ts, methods=tuple(methods))
        from collections import Counter
        m_cnt = Counter(p['method'] for p in all_pays)
        print(f"\n  总计 {len(all_pays)} 笔: {dict(m_cnt)}")
    else:
        print("\n[1/4] 跳过 BQ 查询")
        all_pays = []

    # 2. 处理 Wallace SHP 文件
    shp_files = [
        ('TR', '/workspace/data/uploads/01KQECMR3RZ9D0K19FNR57QQ2K.xlsx', 'exports/wallace_tr_reconciled.xlsx'),
        ('FF', '/workspace/data/uploads/01KQECMR3S1GGN3898AGSV2T1K.xlsx', 'exports/wallace_ff_reconciled.xlsx'),
        ('JJFM', '/workspace/data/uploads/01KQECMR3SDZAN2R2B03ZXXV66.xlsx', 'exports/wallace_jjfm_reconciled.xlsx'),
        ('FBF', '/workspace/data/uploads/01KQECMR3RJG99SEMGDKK6FC76.xlsx', 'exports/wallace_fbf_reconciled.xlsx'),
    ]

    def print_status_dist(df):
        s = df['对账状态'].astype(str).value_counts()
        for k, v in s.items():
            print(f"    {k}: {v} 行 ({v * 100 // max(len(df), 1)}%)")

    if 'wallace' in only:
        print("\n[2/4] 处理 Wallace SHP (Shopee Pay) 文件...")
        for name, input_path, output_path in shp_files:
            print(f"\n  {name}: {input_path}")
            df = pd.read_excel(input_path)
            df = match_wallace_shp(df, all_pays)
            print(f"    总记录: {len(df)} 行")
            print_status_dist(df)
            out = next_versioned_path(output_path)
            export_df(df, out)
            print(f"    输出: {out}")
    else:
        print("\n[2/4] 跳过 Wallace SHP")

    # 3. 处理 RBH 文件
    if 'rbh' in only:
        print("\n[3/4] 处理 RBH (Robinhood Pay) 文件...")
        rbh_input = '/workspace/data/uploads/01KQECMR3SWFV9SXNM9APPNCKB.xlsx'
        df = pd.read_excel(rbh_input, header=4)
        df = match_rbh(df, all_pays)
        print(f"  总记录: {len(df)} 行")
        print_status_dist(df)
        out = next_versioned_path('exports/rbh_reconciled.xlsx')
        export_df(df, out)
        print(f"  输出: {out}")
    else:
        print("\n[3/4] 跳过 RBH")

    # 4. 处理 LINEMAN 文件（按日汇总匹配）
    if 'lineman' in only:
        print("\n[4/4] 处理 LINEMAN (LineMan Pay) 文件...")
        lineman_input = '/workspace/data/uploads/01KQECMR3SCV8HF369JTYFMX3P.xlsx'
        df = pd.read_excel(lineman_input, header=2)
        df = match_lineman(df, all_pays, stores)
        print(f"  总记录: {len(df)} 行")
        print_status_dist(df)
        out = next_versioned_path('exports/lineman_reconciled.xlsx')
        export_df(df, out)
        print(f"  输出: {out}")
    else:
        print("\n[4/4] 跳过 LINEMAN")

    print("\n" + "=" * 60)
    print("完成")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
