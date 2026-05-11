"""把 profit_margin 报表的「标价应收」列删除，并修正所有受影响的公式。

输入: exports/profit_202603_v38.xlsx
输出: exports/profit_202603_v39.xlsx

变更：
  - P (异常损失):   =O-H            →  =F-L-H
  - AC (单份总成本): =SUMPRODUCT(Z,AA) →  =SUMPRODUCT(Y,Z)   (列字母左移 1)
  - AD (单品毛利):   =H/E-AC          →  =H/E-AB
  - AE (总毛利):     =AD*E            →  =AC*E
  - AF (毛利率):     =AD/(H/E)        →  =AC/(H/E)
  - 然后 delete_cols(15)  即删除 O 列
"""

import re
import sys
from openpyxl import load_workbook

SRC = 'exports/profit_202603_v38.xlsx'
DST = 'exports/profit_202603_v39.xlsx'

SUMPROD_RE = re.compile(r'=SUMPRODUCT\(Z(\d+):Z(\d+),AA(\d+):AA(\d+)\)')


def patch_sheet(ws):
    rewrites = {'P': 0, 'AC': 0, 'AD': 0, 'AE': 0, 'AF': 0}
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        cell = row[0]
        if cell.value is None:
            continue
        r = cell.row

        # P (16) 异常损失
        ws.cell(r, 16).value = f'=F{r}-L{r}-H{r}'
        rewrites['P'] += 1

        # AC (29) 单份总成本：解析出原来的 BOM 行范围，整体左移一列
        v_ac = ws.cell(r, 29).value
        if isinstance(v_ac, str):
            m = SUMPROD_RE.match(v_ac)
            if m:
                s, e = m.group(1), m.group(2)
                ws.cell(r, 29).value = f'=SUMPRODUCT(Y{s}:Y{e},Z{s}:Z{e})'
                rewrites['AC'] += 1
            else:
                print(f'  ! AC{r} 公式不符合预期: {v_ac!r}', file=sys.stderr)

        # AD (30) 单品毛利
        ws.cell(r, 30).value = f'=H{r}/E{r}-AB{r}'
        rewrites['AD'] += 1

        # AE (31) 总毛利
        ws.cell(r, 31).value = f'=AC{r}*E{r}'
        rewrites['AE'] += 1

        # AF (32) 毛利率
        ws.cell(r, 32).value = f'=AC{r}/(H{r}/E{r})'
        rewrites['AF'] += 1

    return rewrites


def main():
    print(f'load {SRC} ...')
    wb = load_workbook(SRC)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f'== sheet: {name} ==')
        stats = patch_sheet(ws)
        print('  rewrites:', stats)
        print(f'  delete_cols(15) on {name} ...')
        ws.delete_cols(15, 1)
        print(f'  done. new max_col = {ws.max_column}')
    print(f'save -> {DST}')
    wb.save(DST)
    print('OK')


if __name__ == '__main__':
    main()
