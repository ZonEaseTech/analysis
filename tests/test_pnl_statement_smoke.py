"""P3.5b — pnl_statement.py 主入口 smoke test (mock 数据).

不实际跑 BQ — 用 mock sales_rows 走完 build_pnl_artifact + write_pnl_excel
全流程, 验证:
  - 聚合层 → P&L → KPI → MoM 链路通
  - Excel 文件能写出来 (xlsxwriter 不抛错)
  - 财务化格式被正确应用 (健康度颜色 / 千分位等通过 xlsxwriter 内部 dictionary
    pass; smoke 验证不抛错即 OK, 完整视觉验证留实测)

跑法: venv/bin/python -m unittest tests.test_pnl_statement_smoke -v
"""
from __future__ import annotations

import os
import tempfile
import unittest

import openpyxl

from bq_reports.pnl_statement import (
    aggregate_sales_by_channel,
    build_pnl_artifact,
    _per_store_artifacts,
    write_pnl_excel,
)
from semantic.aggregations.kpi_ratios import HealthStatus
from semantic.resolvers import DictProvider, Resolver


# ═══════════════════════════════════════════════════════════════════
# aggregate_sales_by_channel
# ═══════════════════════════════════════════════════════════════════

class AggregateByChannelTests(unittest.TestCase):
    def test_dine_and_takeout_summed(self):
        # 金额是萨当 (×100); aggregate 在边界 /100 转元
        rows = [
            {"channel": "dine", "qty": 10, "sales_price": 10000,
             "actual_amount": 9000, "refund_amount": 500, "free_amount": 300,
             "give_amount": 200, "discount_amount": 0, "cancelled_amount": 0},
            {"channel": "takeout", "qty": 5, "sales_price": 6000,
             "actual_amount": 6000, "refund_amount": 0, "free_amount": 0,
             "give_amount": 0, "discount_amount": 0, "cancelled_amount": 1000},
        ]
        out = aggregate_sales_by_channel(rows)

        # 总数 = 两 channel 之和 (元)
        self.assertEqual(out["qty"], 15)
        self.assertEqual(out["sales_price"], 160)
        self.assertEqual(out["actual_amount"], 150)
        self.assertEqual(out["refund_amount"], 5)
        self.assertEqual(out["cancelled_amount"], 10)

        # 渠道拆分 (元)
        self.assertEqual(out["dine_qty"], 10)
        self.assertEqual(out["dine_sales_price"], 100)
        self.assertEqual(out["takeout_qty"], 5)
        self.assertEqual(out["takeout_sales_price"], 60)

    def test_empty_rows(self):
        out = aggregate_sales_by_channel([])
        # 空输入返回空 dict (defaultdict 没触发 += 的 key 不会出现).
        # 调用方 (build_pnl_artifact) 把它喂给 aggregate_sales 会安全 fallback 0
        self.assertEqual(out.get("qty", 0), 0)
        self.assertEqual(out.get("sales_price", 0), 0)

    def test_unknown_channel_only_contributes_to_totals(self):
        """channel 不是 dine/takeout 的 row 只算总数, 不拆渠道."""
        rows = [
            {"channel": "unknown", "qty": 100, "sales_price": 100000,
             "actual_amount": 100000},   # 萨当 → 1000 元
        ]
        out = aggregate_sales_by_channel(rows)
        self.assertEqual(out["qty"], 100)
        self.assertEqual(out["sales_price"], 1000)
        # 拆分字段都是 0 (默认 dict 没设置 → defaultdict 给 0)
        self.assertEqual(out.get("dine_sales_price", 0), 0)
        self.assertEqual(out.get("takeout_sales_price", 0), 0)


# ═══════════════════════════════════════════════════════════════════
# build_pnl_artifact — 端到端编排
# ═══════════════════════════════════════════════════════════════════

def _make_realistic_rows():
    """模拟一份真实月度 sale_event 数据 (跨 channel).

    交易金额是萨当整数 (PR-B 7b): aggregate_sales_by_channel 在边界 /100 转元,
    故 fixture 值 = 期望元值 × 100 (e.g. 50000 元 → 5_000_000 萨当).
    """
    return [
        {"channel": "dine", "qty": 1000, "sales_price": 5_000_000,
         "actual_amount": 4_500_000, "refund_amount": 150_000, "cancelled_amount": 0,
         "free_amount": 80_000, "give_amount": 50_000, "discount_amount": 220_000,
         "order_count": 600},
        {"channel": "takeout", "qty": 400, "sales_price": 2_400_000,
         "actual_amount": 2_400_000, "refund_amount": 0, "cancelled_amount": 100_000,
         "free_amount": 0, "give_amount": 0, "discount_amount": 0,
         "order_count": 200},
    ]


class BuildPnlArtifactTests(unittest.TestCase):
    def test_basic_artifact_shape(self):
        artifact = build_pnl_artifact(
            period="2026-04", scope="Test",
            sales_rows=_make_realistic_rows(),
        )
        self.assertIn("pnl", artifact)
        self.assertIn("kpis", artifact)
        self.assertIn("mom", artifact)
        self.assertIn("yoy", artifact)
        self.assertIn("sales_totals", artifact)

    def test_pnl_values_correct(self):
        rows = _make_realistic_rows()
        artifact = build_pnl_artifact(
            period="2026-04", scope="Test", sales_rows=rows,
        )
        pnl = artifact["pnl"]
        self.assertEqual(pnl.by_code("gmv").amount, 50000 + 24000)
        self.assertEqual(pnl.by_code("dine_gmv").amount, 50000)
        self.assertEqual(pnl.by_code("takeout_gmv").amount, 24000)
        # Net Sales = ttpos actual_sale_amount = sum actual_amount
        self.assertEqual(pnl.by_code("net_sales").amount, 45000 + 24000)

    def test_with_cogs_data(self):
        artifact = build_pnl_artifact(
            period="2026-04", scope="Test",
            sales_rows=_make_realistic_rows(),
            cogs_data={"total": 20000, "dine": 14000, "takeout": 6000},
        )
        pnl = artifact["pnl"]
        # Gross Profit = Net Sales (69000) − COGS (20000) = 49000
        self.assertEqual(pnl.by_code("gross_profit").amount, 49000)

    def test_with_commission_resolver(self):
        resolver = Resolver([
            DictProvider(name="defaults", priority=0,
                        data={"default": 0.25}),
        ])
        artifact = build_pnl_artifact(
            period="2026-04", scope="Test",
            sales_rows=_make_realistic_rows(),
            cogs_data={"total": 20000, "dine": 14000, "takeout": 6000},
            commission_rate_resolver=resolver,
        )
        pnl = artifact["pnl"]
        # platform_commission = takeout_gmv (24000) × 0.25 = 6000
        self.assertEqual(pnl.by_code("platform_commission").amount, -6000)
        # Contribution Margin = Gross Profit (49000) − commission (6000) = 43000
        self.assertEqual(pnl.by_code("contribution_margin").amount, 43000)

    def test_mom_with_previous_pnl(self):
        previous = build_pnl_artifact(
            period="2026-03", scope="Test", sales_rows=[
                {"channel": "dine", "qty": 800, "sales_price": 4_000_000,
                 "actual_amount": 3_600_000},   # 萨当 → 40000/36000 元
            ],
        )["pnl"]
        current = build_pnl_artifact(
            period="2026-04", scope="Test",
            sales_rows=_make_realistic_rows(),
            previous_pnl=previous,
        )
        # GMV: 40000 → 74000 = +85%
        gmv_change = current["mom"]["gmv"]
        self.assertAlmostEqual(gmv_change.pct_delta, (74000 - 40000) / 40000)

    def test_kpis_emitted(self):
        artifact = build_pnl_artifact(
            period="2026-04", scope="Test",
            sales_rows=_make_realistic_rows(),
            cogs_data={"total": 20000, "dine": 14000, "takeout": 6000},
        )
        kpi_codes = {k.code for k in artifact["kpis"]}
        # 至少有 gross_margin, food_cost, aov, dine_mix, takeout_mix
        for required in ["gross_margin", "food_cost", "aov",
                         "dine_mix", "takeout_mix"]:
            self.assertIn(required, kpi_codes)


# ═══════════════════════════════════════════════════════════════════
# write_pnl_excel — 端到端 Excel 输出
# ═══════════════════════════════════════════════════════════════════

class WritePnlExcelTests(unittest.TestCase):
    def _build_default_artifact(self):
        resolver = Resolver([
            DictProvider(name="defaults", priority=0,
                        data={"default": 0.25}),
        ])
        previous = build_pnl_artifact(
            period="2026-03", scope="Test",
            sales_rows=[{"channel": "dine", "qty": 800,
                         "sales_price": 4_000_000, "actual_amount": 3_600_000}],
        )["pnl"]   # 萨当 → 40000/36000 元
        return build_pnl_artifact(
            period="2026-04", scope="全集团",
            sales_rows=_make_realistic_rows(),
            cogs_data={"total": 20000, "dine": 14000, "takeout": 6000},
            commission_rate_resolver=resolver,
            previous_pnl=previous,
        )

    def test_excel_file_is_created(self):
        artifact = self._build_default_artifact()
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            write_pnl_excel(artifact, tmp.name)
            self.assertTrue(os.path.exists(tmp.name))
            self.assertGreater(os.path.getsize(tmp.name), 0)
        finally:
            os.unlink(tmp.name)

    def test_excel_has_two_sheets(self):
        artifact = self._build_default_artifact()
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            write_pnl_excel(artifact, tmp.name)
            wb = openpyxl.load_workbook(tmp.name, data_only=False)
            self.assertIn("集团损益表", wb.sheetnames)
            self.assertIn("KPI Dashboard", wb.sheetnames)
            wb.close()
        finally:
            os.unlink(tmp.name)

    def test_pnl_sheet_contains_subtotal_labels(self):
        artifact = self._build_default_artifact()
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            write_pnl_excel(artifact, tmp.name)
            wb = openpyxl.load_workbook(tmp.name, data_only=True)
            ws = wb["集团损益表"]

            # 找几个 subtotal 标签
            cell_values = [row[0].value for row in ws.iter_rows()]
            joined = " ".join(str(v) for v in cell_values if v)
            self.assertIn("GMV", joined)
            self.assertIn("净销售额", joined)
            self.assertIn("销售毛利", joined)
            wb.close()
        finally:
            os.unlink(tmp.name)

    def test_kpi_sheet_contains_health_labels(self):
        artifact = self._build_default_artifact()
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            write_pnl_excel(artifact, tmp.name)
            wb = openpyxl.load_workbook(tmp.name, data_only=True)
            ws = wb["KPI Dashboard"]

            # 评级列应该有 ✅/🟡/🔴/N/A 之一
            ratings_col = [row[3].value for row in ws.iter_rows(min_row=2)]
            ratings_str = " ".join(str(v) for v in ratings_col if v)
            has_some_rating = any(
                marker in ratings_str
                for marker in ["✅", "🟡", "🔴", "⚪", "N/A"]
            )
            self.assertTrue(has_some_rating,
                          f"KPI Dashboard 应该有健康度标记, got: {ratings_str!r}")
            wb.close()
        finally:
            os.unlink(tmp.name)

    def test_kpi_without_benchmark_has_empty_rating(self):
        """没 benchmark 的 KPI (AOV / Channel Mix) 评级列应空白, 不是 'N/A'.

        防止 UX 误解: AOV 数值是 OK 的, 只是没行业 universal 基准做评级.
        """
        artifact = self._build_default_artifact()
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            write_pnl_excel(artifact, tmp.name)
            wb = openpyxl.load_workbook(tmp.name, data_only=True)
            ws = wb["KPI Dashboard"]

            # 找 AOV / 堂食占比 行
            for row in ws.iter_rows(min_row=2):
                name = str(row[0].value or "")
                rating = row[3].value  # 第 4 列是评级
                if "客单价" in name or "占比" in name:
                    self.assertIsNone(
                        rating,
                        msg=f"{name} 评级列应该空 (没 benchmark), 实际: {rating!r}",
                    )
            wb.close()
        finally:
            os.unlink(tmp.name)

    def test_extended_sheets_render(self):
        """Sheet 3 (按店) / Sheet 4 (按渠道) / Sheet 6 (来源审计) 能正常写出."""
        artifact = self._build_default_artifact()

        # Mock per_store_artifacts (用 store_num attribute 兼容 BQ Row 形式)
        from types import SimpleNamespace
        # 交易金额萨当 (×100); _per_store_artifacts → aggregate 在边界 /100 转元
        store_rows = [
            SimpleNamespace(
                store_num="001", store_name="Test Store A",
                channel="dine", qty=100, sales_price=500000,
                actual_amount=450000, refund_amount=20000, free_amount=5000,
                give_amount=3000, discount_amount=22000, cancelled_amount=0,
                order_count=80,
            ),
            SimpleNamespace(
                store_num="001", store_name="Test Store A",
                channel="takeout", qty=40, sales_price=240000,
                actual_amount=240000, refund_amount=0, free_amount=0,
                give_amount=0, discount_amount=0, cancelled_amount=10000,
                order_count=20,
            ),
        ]
        per_store = _per_store_artifacts(
            store_rows, {"001": {"dine": 1500, "takeout": 500, "total": 2000}},
            None, "2026-04",
        )
        channel_data = {
            "dine":    {"cogs": 1500},
            "takeout": {"cogs": 500},
        }

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            write_pnl_excel(
                artifact, tmp.name,
                per_store_artifacts=per_store,
                channel_data=channel_data,
            )
            wb = openpyxl.load_workbook(tmp.name, data_only=True)
            # 5 个 sheets
            self.assertIn("集团损益表", wb.sheetnames)
            self.assertIn("KPI Dashboard", wb.sheetnames)
            self.assertIn("按店损益", wb.sheetnames)
            self.assertIn("按渠道对比", wb.sheetnames)
            self.assertIn("数据来源审计", wb.sheetnames)

            # Sheet 3 应该至少 1 行 + 表头
            ws3 = wb["按店损益"]
            self.assertGreaterEqual(ws3.max_row, 2)

            # Sheet 4 包含"堂食"/"外卖"列名
            ws4 = wb["按渠道对比"]
            header_row = [c.value for c in ws4[1]]
            self.assertIn("堂食", header_row)
            self.assertIn("外卖", header_row)

            # Sheet 6 每行带 confidence
            ws6 = wb["数据来源审计"]
            confidence_col_vals = {
                row[2].value for row in ws6.iter_rows(min_row=2)
            }
            # 应该至少出现 actual / derived / n/a 中的两种
            self.assertTrue(
                len(confidence_col_vals & {"actual", "derived", "estimated", "n/a"}) >= 2,
                msg=f"confidence 列应有多种值, got {confidence_col_vals}",
            )
            wb.close()
        finally:
            os.unlink(tmp.name)

    def test_na_layers_show_as_na(self):
        """没接 cogs/resolver 时 Gross Profit / Contribution / Operating
        应该显示 N/A."""
        artifact = build_pnl_artifact(
            period="2026-04", scope="Test",
            sales_rows=_make_realistic_rows(),
            # 故意不传 cogs_data / resolver
        )
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            write_pnl_excel(artifact, tmp.name)
            wb = openpyxl.load_workbook(tmp.name, data_only=True)
            ws = wb["集团损益表"]

            # 找 "经营利润" 行的金额列, 应该是 "N/A"
            for row in ws.iter_rows():
                if row[0].value and "经营利润" in str(row[0].value):
                    self.assertEqual(row[1].value, "N/A")
                    break
            else:
                self.fail("没找到经营利润行")
            wb.close()
        finally:
            os.unlink(tmp.name)


if __name__ == "__main__":
    unittest.main()
