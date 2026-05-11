"""`_load_uploaded_prices` parses customer-supplied Excel.

Sheets:
  '盘点单位匹配分析'  → col A=code, col I=销售换算系数
  '干冻货' / '设备材料' → col A=code, col H=单价
Result: unit_price = price / conversion_rate (defaults to /1).
"""
import os
import tempfile
import unittest

from tests import _setup  # noqa: F401

import openpyxl

from bq_reports.profit_margin_report import _load_uploaded_prices


def make_excel(path: str, *, conv_rows=(), dry_rows=(), equip_rows=(),
               include_conv_sheet=True, include_dry=True, include_equip=True):
    wb = openpyxl.Workbook()
    # Remove default sheet so we control all sheets explicitly.
    wb.remove(wb.active)

    if include_conv_sheet:
        ws = wb.create_sheet("盘点单位匹配分析")
        # Column index: A=1 (code), I=9 (conversion). Header row 1.
        ws.append(["code"] + [None] * 7 + ["销售换算系数"])
        for code, conv in conv_rows:
            row = [code] + [None] * 7 + [conv]
            ws.append(row)

    def _add_price_sheet(name, rows):
        ws = wb.create_sheet(name)
        ws.append(["code"] + [None] * 6 + ["单价"])
        for code, price in rows:
            ws.append([code] + [None] * 6 + [price])

    if include_dry:
        _add_price_sheet("干冻货", dry_rows)
    if include_equip:
        _add_price_sheet("设备材料", equip_rows)

    wb.save(path)


class LoadUploadedPricesTests(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.path = os.path.join(self.tmpdir, "prices.xlsx")

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_returns_empty_when_path_missing(self):
        prices, convs = _load_uploaded_prices("/nonexistent/path.xlsx")
        self.assertEqual(prices, {})
        self.assertEqual(convs, {})

    def test_returns_empty_when_path_empty(self):
        prices, convs = _load_uploaded_prices("")
        self.assertEqual((prices, convs), ({}, {}))

    def test_price_divided_by_conversion(self):
        make_excel(self.path,
                   conv_rows=[("M001", 50), ("M002", 1)],
                   dry_rows=[("M001", 500), ("M002", 7.5)])
        prices, convs = _load_uploaded_prices(self.path)
        self.assertAlmostEqual(prices["M001"], 10.0)   # 500/50
        self.assertAlmostEqual(prices["M002"], 7.5)
        self.assertEqual(convs, {"M001": 50.0, "M002": 1.0})

    def test_no_conversion_defaults_to_one(self):
        make_excel(self.path, conv_rows=[], dry_rows=[("M003", 4.2)])
        prices, _ = _load_uploaded_prices(self.path)
        self.assertAlmostEqual(prices["M003"], 4.2)

    def test_na_sentinel_skipped(self):
        make_excel(self.path, dry_rows=[("BAD", "#N/A"), ("GOOD", 3)])
        prices, _ = _load_uploaded_prices(self.path)
        self.assertIn("GOOD", prices)
        self.assertNotIn("BAD", prices)

    def test_code_stripped(self):
        make_excel(self.path, dry_rows=[("  M004  ", 2.0)])
        prices, _ = _load_uploaded_prices(self.path)
        self.assertIn("M004", prices)

    def test_equipment_sheet_also_loaded(self):
        make_excel(self.path, equip_rows=[("E001", 9.0)], include_dry=False)
        prices, _ = _load_uploaded_prices(self.path)
        self.assertEqual(prices["E001"], 9.0)

    def test_missing_conversion_sheet_yields_no_conv_but_still_prices(self):
        make_excel(self.path, dry_rows=[("M005", 6)], include_conv_sheet=False)
        prices, convs = _load_uploaded_prices(self.path)
        self.assertEqual(convs, {})
        self.assertEqual(prices["M005"], 6.0)


if __name__ == "__main__":
    unittest.main()
