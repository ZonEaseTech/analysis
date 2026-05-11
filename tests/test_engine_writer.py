"""`write_configured_sheet` round-trip: write via xlsxwriter, read via openpyxl.

Locked behaviours:
  - Header cells contain `name`; columns with `comment` emit an openpyxl Comment.
  - `merge: true` columns produce a merged range across each detected block.
  - `block_formula` writes the formula at the block's first row (in the merged cell).
  - `negative_red` colours the cell font red when the precomputed value is < 0.
  - `positive_red` registers a `>0` conditional format on the column.
  - Freeze panes set to A2 (first row frozen).
"""
import os
import tempfile
import unittest

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from tests import _setup  # noqa: F401

from utils.report_engine import (
    ColumnConfig,
    SheetConfig,
    write_configured_sheet,
)


def _sheet_cfg() -> SheetConfig:
    """Tiny three-column sheet: merge-id, value, formula (negative-red + positive-red sibling)."""
    return SheetConfig(
        name="测试",
        columns=[
            ColumnConfig(name="商品", field_index=0, merge=True, comment="字段语义说明"),
            ColumnConfig(name="销量", field_index=1, col_type="value",
                         format_str="0.00", merge=True),
            ColumnConfig(name="单品毛利", field_index=2, col_type="block_formula",
                         format_str="0.00", merge=True, negative_red=True,
                         formula_template="=B{block_start}-1"),
            ColumnConfig(name="异常损失", field_index=3, col_type="value",
                         format_str="0.00", positive_red=True),
        ],
        merge_key_indices=[0],  # block = unique product
    )


def _write_to_tmp(rows):
    tmpdir = tempfile.mkdtemp()
    path = os.path.join(tmpdir, "out.xlsx")
    wb = xlsxwriter.Workbook(path)
    write_configured_sheet(wb, "测试", _sheet_cfg(), rows)
    wb.close()
    return path, tmpdir


class EngineWriterRoundTrip(unittest.TestCase):
    def setUp(self):
        # Two products, each spanning 2 rows (so merge kicks in).
        # `field_index=2` carries pre-computed value (negative for product B → font red).
        # `field_index=3` carries cancellation amount (positive triggers conditional format).
        self.rows = [
            ["商品A", 10, 5.0, 0],
            ["商品A", 10, 5.0, 0],
            ["商品B", 8, -2.5, 7.5],
            ["商品B", 8, -2.5, 7.5],
        ]
        self.path, self.tmpdir = _write_to_tmp(self.rows)

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _load(self):
        return load_workbook(self.path)["测试"]

    def test_header_row_and_comment(self):
        ws = self._load()
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        self.assertEqual(headers, ["商品", "销量", "单品毛利", "异常损失"])
        # First header cell has a comment (字段语义说明)
        c = ws.cell(row=1, column=1).comment
        self.assertIsNotNone(c, "comment missing on 商品 header")
        self.assertEqual(c.text, "字段语义说明")

    def test_block_merges(self):
        ws = self._load()
        merged = {str(r) for r in ws.merged_cells.ranges}
        # Two products × three merge columns (商品, 销量, 单品毛利) × rows 2-3 / 4-5.
        # Each merged range spans 2 rows.
        expected_pairs = [
            ("A2:A3", "B2:B3", "C2:C3"),
            ("A4:A5", "B4:B5", "C4:C5"),
        ]
        for trio in expected_pairs:
            for rng in trio:
                self.assertIn(rng, merged, f"missing merged range {rng}")

    def test_block_formula_written_at_first_row(self):
        ws = self._load()
        # block 1 first row = row 2 ; formula =B2-1
        f1 = ws.cell(row=2, column=3).value
        f2 = ws.cell(row=4, column=3).value
        self.assertEqual(f1, "=B2-1")
        self.assertEqual(f2, "=B4-1")

    def test_negative_red_applied_only_when_value_negative(self):
        ws = self._load()
        # openpyxl Font.color.rgb can be an ARGB string OR an RGB object — normalise.
        def _rgb(cell):
            col = cell.font.color
            if col is None:
                return ""
            raw = getattr(col, "rgb", None) or ""
            return str(raw).upper()

        red = "FFFF0000"
        c_pos = ws.cell(row=2, column=3)  # product A: precomputed 5.0 → no red
        c_neg = ws.cell(row=4, column=3)  # product B: precomputed -2.5 → red
        self.assertNotEqual(_rgb(c_pos), red,
                            "non-negative block must not be red")
        self.assertEqual(_rgb(c_neg), red,
                         "negative_red did not paint negative block")

    def test_positive_red_conditional_format_registered(self):
        ws = self._load()
        # openpyxl exposes conditional_formatting as a dict-like
        ranges = [str(rng) for rng in ws.conditional_formatting]
        # Column D = index 3; data rows 2..5 → "D2:D5"
        self.assertTrue(
            any("D2:D5" in r for r in ranges),
            f"positive_red did not register conditional format; got ranges={ranges}",
        )

    def test_freeze_panes_first_row(self):
        ws = self._load()
        self.assertEqual(ws.freeze_panes, "A2")


class EmptyRowsTest(unittest.TestCase):
    """Empty input must still produce a sheet with header and freeze pane."""

    def test_empty_rows_no_crash(self):
        path, tmpdir = _write_to_tmp(rows=[])
        try:
            ws = load_workbook(path)["测试"]
            self.assertEqual(ws.cell(row=1, column=1).value, "商品")
            self.assertEqual(ws.freeze_panes, "A2")
            # No conditional formatting necessary; no merged ranges.
            self.assertEqual(list(ws.merged_cells.ranges), [])
        finally:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)


class SingleRowBlockTest(unittest.TestCase):
    """Block of size 1 must not produce a merged range; formula written directly."""

    def test_single_row_block(self):
        rows = [["商品X", 1, 9.0, 0]]
        path, tmpdir = _write_to_tmp(rows)
        try:
            ws = load_workbook(path)["测试"]
            merged = [str(r) for r in ws.merged_cells.ranges]
            # No range should span A2:A2 (single-row blocks bypass merge_range).
            for rng in merged:
                a, b, c, d = range_boundaries(rng)
                self.assertFalse(b == d, f"single-row block got merged: {rng}")
            # Block formula still placed at row 2 col C.
            self.assertEqual(ws.cell(row=2, column=3).value, "=B2-1")
        finally:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
