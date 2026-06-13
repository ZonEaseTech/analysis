"""bq_exporter 的闸门钩子: set_gate 后写盘前必过闸。"""
import os
import tempfile
import unittest

import tests._setup  # noqa: F401

from bq_reports.utils.bq_exporter import BaseExporter, ReportExporter
from semantic.validators.gate import GateSpec
from semantic.validators.identities import DEFAULT_IDENTITIES


class _FakeQueryJob:
    def result(self):
        return []


class _FakeClient:
    """最小 BQ stub: query().result() → 空, 让无 merchants 的导出跑到写盘段。"""

    def query(self, sql):
        return _FakeQueryJob()


def _make_sales_exporter(tmpdir, *, force):
    """构造一个不连 BQ、merchants 为空的 ReportExporter。

    空 merchants → 空 sales_results → 闸门 min_rows=1 触发 → 复现写盘路径的
    "闸门 sys.exit 时是否漏文件" (无需真实数据/BQ)。
    """
    exporter = ReportExporter.__new__(ReportExporter)  # 不触发 BQ 连接
    exporter.project_id = "test-proj"
    exporter.client = _FakeClient()
    exporter.gate_spec = None
    exporter.merchants = []
    exporter.output_path = os.path.join(tmpdir, "out.xlsx")
    exporter.load_merchants = lambda *a, **k: None  # 保持 merchants=[]
    exporter.set_gate(GateSpec(
        identities=[], force=force, report_name="sales_consumption",
        build_check_rows=lambda rows: rows))
    return exporter


class TestExporterGate(unittest.TestCase):
    def test_gate_spec_holds_config(self):
        spec = GateSpec(identities=DEFAULT_IDENTITIES, force=False,
                        report_name="daily_sales",
                        build_check_rows=lambda rows: rows)
        self.assertEqual(spec.report_name, "daily_sales")

    def test_exporter_accepts_gate(self):
        exporter = BaseExporter.__new__(BaseExporter)  # 不触发 BQ 连接
        exporter.gate_spec = None
        spec = GateSpec(identities=DEFAULT_IDENTITIES, force=False,
                        report_name="t", build_check_rows=lambda rows: rows)
        exporter.set_gate(spec)
        self.assertIs(exporter.gate_spec, spec)

    def test_gate_spec_run_blocks_on_empty(self):
        spec = GateSpec(identities=DEFAULT_IDENTITIES, force=False,
                        report_name="t", build_check_rows=lambda rows: rows)
        with self.assertRaises(SystemExit) as cm:
            spec.run([])
        self.assertEqual(cm.exception.code, 2)


class TestXlsxwriterGateNoFileLeak(unittest.TestCase):
    """回归: xlsxwriter 写盘路径在闸门阻断时不得漏文件 (PR-A #24 review)。

    根因: export_sales_and_consumption 把 gate.run()(含 sys.exit(2))放 try、
    wb.close() 放 finally — SystemExit 传播前 finally 仍 flush 一个无水印文件到盘,
    违反闸门契约「🔴 且无 force → 不产文件」。openpyxl 路径安全是因 wb.save() 在
    gate 之后且不在 finally。修复: gate 在 xlsxwriter.Workbook() 构造之前跑。
    """

    def test_blocked_export_leaves_no_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            exporter = _make_sales_exporter(tmp, force=False)
            out = exporter.output_path
            with self.assertRaises(SystemExit) as cm:
                exporter.export_sales_and_consumption(
                    month="2026-06", merchant_xlsx="(ignored)", output_path=out)
            self.assertEqual(cm.exception.code, 2)
            self.assertFalse(
                os.path.exists(out),
                "闸门阻断后不得有文件落盘 (xlsxwriter finally: wb.close() 漏)")

    def test_forced_export_still_writes_watermarked_file(self):
        # --force 逃生口必须仍产文件且带水印 — 守护修复没把它改坏
        with tempfile.TemporaryDirectory() as tmp:
            exporter = _make_sales_exporter(tmp, force=True)
            out = exporter.output_path
            exporter.export_sales_and_consumption(
                month="2026-06", merchant_xlsx="(ignored)", output_path=out)
            self.assertTrue(os.path.exists(out), "--force 应仍产出文件")
            import openpyxl
            wb = openpyxl.load_workbook(out)
            self.assertIn("⚠️校验未通过", wb.sheetnames,
                          "--force 强制导出的文件必须带水印 sheet")


if __name__ == "__main__":
    unittest.main()
