"""End-to-end smoke for profit_by_price — 客户交付报表（per-SKU + 横向价格档 + BOM 展开）。

Per-SKU rollup pattern:
  fine-grain (店, SKU, price) → _rollup_per_sku → (店, SKU) + price_tiers
  → _build_by_price_rows → 每 SKU 多 BOM 物料行（SKU/价格档列跨行 merge）

Anchors:
  - SKU 多档价格 → 横向 售价1-5/净销量1-5 + 其它销量
  - 售价按净销量降序排
  - BOM 展开 = 同 SKU 多行
  - 利润 4 件套用 Excel block formula（成本=0 时利润留空）
"""
import os
import tempfile
import unittest
from types import SimpleNamespace

import xlsxwriter
from openpyxl import load_workbook

from tests._setup import REPO_ROOT  # noqa: F401

from bq_reports.profit_by_price_report import (
    FINE_GRAIN_KEYS,
    GRAIN_KEYS,
    METRIC_KEYS,
    TOP_N_PRICES,
    _build_by_price_rows,
    _rollup_per_sku,
)
from semantic.aggregations.by_grain import aggregate_by_grain
from utils.report_engine import load_sheet_config, write_configured_sheet


YAML_PATH = REPO_ROOT / "resources/reports/profit_by_price.yaml"


def event(store_num="1", store_name="店A",
          item_uuid="X", item_name="商品X",
          price=10.0, channel="dine",
          qty=0, sales_price=0, actual_amount=0, original_amount=0,
          refund_qty=0, refund_amount=0, free_qty=0, give_qty=0,
          free_amount=0, give_amount=0, discount_amount=0,
          cancelled_qty=0, cancelled_amount=0):
    return SimpleNamespace(
        store_num=store_num, store_name=store_name,
        item_uuid=item_uuid, item_name=item_name,
        price=price, channel=channel,
        qty=qty, sales_price=sales_price, actual_amount=actual_amount,
        original_amount=original_amount,
        refund_qty=refund_qty, refund_amount=refund_amount,
        free_qty=free_qty, give_qty=give_qty,
        free_amount=free_amount, give_amount=give_amount,
        discount_amount=discount_amount,
        cancelled_qty=cancelled_qty, cancelled_amount=cancelled_amount,
    )


class RollupPerSkuTests(unittest.TestCase):
    """fine-grain (店, SKU, price) → per-SKU + price_tiers."""

    def test_collapses_to_one_sku_with_two_tiers(self):
        events = [
            event(item_uuid="A", item_name="X",
                  price=10.0, qty=8, sales_price=80, actual_amount=80),
            event(item_uuid="A", item_name="X",
                  price=12.0, qty=3, sales_price=36, actual_amount=36),
        ]
        fine = aggregate_by_grain(events, FINE_GRAIN_KEYS, METRIC_KEYS)
        by_sku = _rollup_per_sku(fine)
        self.assertEqual(len(by_sku), 1)
        sku_key = next(iter(by_sku))
        data = by_sku[sku_key]
        # 总和
        self.assertEqual(data["qty"], 11)
        self.assertEqual(data["actual_amount"], 116)
        # 价格档列表（净销量降序）
        self.assertEqual(data["price_tiers"], [(10.0, 8), (12.0, 3)])

    def test_price_tiers_sorted_by_net_qty_desc(self):
        events = [
            event(item_uuid="A", price=10.0, qty=5, sales_price=50, actual_amount=50),
            event(item_uuid="A", price=12.0, qty=20, sales_price=240, actual_amount=240),
            event(item_uuid="A", price=8.0,  qty=2, sales_price=16, actual_amount=16),
        ]
        fine = aggregate_by_grain(events, FINE_GRAIN_KEYS, METRIC_KEYS)
        by_sku = _rollup_per_sku(fine)
        tiers = next(iter(by_sku.values()))["price_tiers"]
        prices_only = [p for p, _ in tiers]
        self.assertEqual(prices_only, [12.0, 10.0, 8.0],
                         "价格档应按净销量降序，¥12 销量 20 最高")

    def test_dine_takeout_same_price_collapse_into_same_tier(self):
        events = [
            event(item_uuid="A", price=10.0, channel="dine",
                  qty=5, sales_price=50, actual_amount=50),
            event(item_uuid="A", price=10.0, channel="takeout",
                  qty=7, sales_price=70, actual_amount=70),
        ]
        fine = aggregate_by_grain(events, FINE_GRAIN_KEYS, METRIC_KEYS)
        by_sku = _rollup_per_sku(fine)
        tiers = next(iter(by_sku.values()))["price_tiers"]
        # ¥10 这一档合并 dine + takeout 净销量 = 12
        self.assertEqual(tiers, [(10.0, 12)])


class BuildRowsShapeTests(unittest.TestCase):
    """每行 = (SKU, BOM 物料)；SKU 维度跨 BOM 行 merge。"""

    def setUp(self):
        self.events = [
            event(item_uuid="A", item_name="鸡腿堡",
                  price=10.0, qty=100, sales_price=1000, actual_amount=1000),
            event(item_uuid="A", item_name="鸡腿堡",
                  price=5.0,  qty=50,  sales_price=250,  actual_amount=250),
        ]
        # 2 个 BOM 物料 → 2 行
        self.bom_data = {"1": {"A": [
            ("M1", "面包", 1.0, "pc", 1.0, 1.0),
            ("M2", "鸡腿", 1.0, "pc", 1.0, 5.0),
        ]}}

    def test_one_sku_collapses_horizontal_prices(self):
        fine = aggregate_by_grain(self.events, FINE_GRAIN_KEYS, METRIC_KEYS)
        rows = _build_by_price_rows(
            fine, self.bom_data, combo_structure={}, mode="single",
        )
        # 1 SKU × 2 BOM = 2 行 (不是 2 价 × 2 BOM = 4)
        self.assertEqual(len(rows), 2)
        # 47 列 (37 visible + 1 hidden + 2 audit + 5 净利润口径 + 2 实收审计: 应收金额(AT)/订单级折扣(AU))
        self.assertEqual(len(rows[0]), 47)

    def test_row_columns_match_yaml_layout(self):
        fine = aggregate_by_grain(self.events, FINE_GRAIN_KEYS, METRIC_KEYS)
        rows = _build_by_price_rows(
            fine, self.bom_data, combo_structure={}, mode="single",
        )
        r = rows[0]
        # SKU 维度
        self.assertEqual(r[0], "1")            # 门店编号
        self.assertEqual(r[2], "鸡腿堡")        # SKU名
        self.assertEqual(r[3], 150)            # 销量 (gross)
        self.assertEqual(r[4], 150)            # 净销量
        self.assertEqual(r[5], 1250)           # 营业额
        self.assertEqual(r[7], 1250)           # 实收
        # 价格档 1-5 (按净销量降序：¥10 销 100 > ¥5 销 50)
        self.assertEqual(r[11], 10.0)          # 售价1
        self.assertEqual(r[12], 100)           # 净销量1
        self.assertEqual(r[13], 5.0)           # 售价2
        self.assertEqual(r[14], 50)            # 净销量2
        self.assertIsNone(r[15])               # 售价3 (空，只有 2 档)
        self.assertEqual(r[21], 0)             # 其它销量
        # BOM 第一行
        self.assertEqual(r[28], "面包")        # BOM物品名
        self.assertEqual(r[29], "M1")          # BOM物品编码
        # 利润 4 件套 formula 占位
        for idx in (33, 34, 35, 36):
            self.assertIsNone(r[idx])
        self.assertEqual(r[37], "A")           # hidden uuid

    def test_seven_price_tiers_top5_plus_others(self):
        events = [
            event(item_uuid="X", price=p, qty=q,
                  sales_price=p*q, actual_amount=p*q)
            for p, q in [(119, 850), (109, 180), (129, 90), (99, 35),
                          (139, 18), (149, 5), (89, 2)]
        ]
        fine = aggregate_by_grain(events, FINE_GRAIN_KEYS, METRIC_KEYS)
        rows = _build_by_price_rows(
            fine, bom_data={}, combo_structure={}, mode="single",
        )
        self.assertEqual(len(rows), 1, "1 SKU = 1 row (无 BOM 用 '-' 占位)")
        r = rows[0]
        # Top 5 by qty desc: 119/109/129/99/139
        self.assertEqual(r[11], 119)
        self.assertEqual(r[13], 109)
        self.assertEqual(r[15], 129)
        self.assertEqual(r[17], 99)
        self.assertEqual(r[19], 139)
        # 其它销量 = 5 + 2 = 7
        self.assertEqual(r[21], 7)


class ExcelRoundTrip(unittest.TestCase):
    def setUp(self):
        self.events = [
            event(item_uuid="A", item_name="鸡腿堡",
                  price=10.0, qty=100, sales_price=1000, actual_amount=1000),
            event(item_uuid="A", item_name="鸡腿堡",
                  price=5.0, qty=50, sales_price=250, actual_amount=250),
        ]
        self.bom_data = {"1": {"A": [
            ("M1", "面包", 1.0, "pc", 1.0, 1.0),
            ("M2", "鸡腿", 1.0, "pc", 1.0, 5.0),
        ]}}
        self.tmpdir = tempfile.mkdtemp()
        self.xlsx = os.path.join(self.tmpdir, "test.xlsx")

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_excel_writes_with_per_sku_layout(self):
        fine = aggregate_by_grain(self.events, FINE_GRAIN_KEYS, METRIC_KEYS)
        flat = _build_by_price_rows(
            fine, self.bom_data, combo_structure={}, mode="single",
        )
        sheet_cfg = load_sheet_config(str(YAML_PATH), "单品")
        wb = xlsxwriter.Workbook(self.xlsx)
        write_configured_sheet(wb, "单品", sheet_cfg, flat)
        wb.close()

        ws = load_workbook(self.xlsx)["单品"]
        # Header row + 2 data rows (2 BOM materials)
        self.assertEqual(ws.cell(row=1, column=1).value, "门店编号")
        # 售价1 column (L)
        headers = {ws.cell(row=1, column=c).value: c
                   for c in range(1, ws.max_column + 1)}
        self.assertIn("售价1", headers)
        self.assertEqual(ws.cell(row=2, column=headers["售价1"]).value, 10.0)
        self.assertEqual(ws.cell(row=2, column=headers["净销量1"]).value, 100)

    def test_excel_formulas_use_correct_letters(self):
        """新布局列字母: E=净销量, H=实收, AH=单份成本, AJ=总毛利。"""
        fine = aggregate_by_grain(self.events, FINE_GRAIN_KEYS, METRIC_KEYS)
        flat = _build_by_price_rows(
            fine, self.bom_data, combo_structure={}, mode="single",
        )
        sheet_cfg = load_sheet_config(str(YAML_PATH), "单品")
        wb = xlsxwriter.Workbook(self.xlsx)
        write_configured_sheet(wb, "单品", sheet_cfg, flat)
        wb.close()

        ws = load_workbook(self.xlsx)["单品"]
        headers = {ws.cell(row=1, column=c).value: c
                   for c in range(1, ws.max_column + 1)}
        self.assertEqual(ws.cell(row=2, column=headers["折损金额"]).value, "=F2-H2")
        self.assertEqual(ws.cell(row=2, column=headers["客单实收"]).value, "=IF(E2=0,0,H2/E2)")
        self.assertEqual(ws.cell(row=2, column=headers["实收占比"]).value, "=IF(F2=0,0,H2/F2)")
        # SUMPRODUCT 跨 block (row 2-3 because 2 BOM materials)
        self.assertEqual(ws.cell(row=2, column=headers["单份总成本"]).value,
                         "=SUMPRODUCT(AE2:AE3,AF2:AF3)")
        self.assertEqual(ws.cell(row=2, column=headers["单品毛利"]).value,
                         '=IF(OR(E2=0,AH2=0),"",H2/E2-AH2)')
        self.assertEqual(ws.cell(row=2, column=headers["总毛利"]).value,
                         '=IF(AH2=0,"",H2-E2*AH2)')
        self.assertEqual(ws.cell(row=2, column=headers["毛利率"]).value,
                         '=IF(OR(H2=0,AH2=0),"",AJ2/H2)')


class YamlFieldIndexCrossCheck(unittest.TestCase):
    """yaml field_index 必须都落在 row 长度内（防 yaml 跟 Python 漂移）。"""

    def test_all_yaml_indices_within_row(self):
        import yaml
        cfg = yaml.safe_load(YAML_PATH.read_text(encoding="utf-8"))

        events = [event(item_uuid="A", item_name="x",
                        price=10.0, qty=1, sales_price=10, actual_amount=10)]
        fine = aggregate_by_grain(events, FINE_GRAIN_KEYS, METRIC_KEYS)
        rows = _build_by_price_rows(
            fine, bom_data={"1": {"A": []}}, combo_structure={}, mode="single",
        )
        row_len = len(rows[0])
        for sheet_name, sheet in cfg["sheets"].items():
            for col in sheet["columns"]:
                idx = col.get("field_index", 0)
                self.assertLess(
                    idx, row_len,
                    f"yaml '{sheet_name}.{col['name']}' field_index={idx} "
                    f"out of bounds (row len {row_len})",
                )


if __name__ == "__main__":
    unittest.main()
