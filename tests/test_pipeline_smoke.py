"""End-to-end smoke: fake BQ rows → aggregate_with_bom → _build_rows →
write_configured_sheet → openpyxl read back → assert ON SPECIFIC CELLS.

This is the ultimate safety net. Even if every unit test passes, this asserts
that *the wiring* between layers still produces a recognisable Excel file.

If a refactor:
  - changes a field name on the BQ row contract
  - re-orders columns
  - breaks block detection
  - mis-aligns yaml field_index against _build_rows output
… one of these assertions fires.
"""
import os
import tempfile
import unittest

import xlsxwriter
from openpyxl import load_workbook

from tests._setup import REPO_ROOT, order_row  # noqa: F401 — also wires sys.path

from bq_reports.profit_margin_report import (
    _build_rows,
    aggregate_with_bom,
)
from utils.report_engine import load_sheet_config, write_configured_sheet


YAML_PATH = REPO_ROOT / "resources/reports/profit_margin.yaml"


class ProfitMarginPipelineSmoke(unittest.TestCase):
    """Single-product, single-BOM happy path through every layer."""

    def setUp(self):
        # Two raw BQ rows: same store/item, different price tiers — so the
        # weighted-discount + qty-accumulation paths execute.
        self.bq_rows = [
            order_row(
                store_num="001", store_name="店A",
                item_uuid="ITEM_X", item_name="单品X",
                qty=8, revenue=240,
                sales_price=320, original_amount=320,
                avg_member_discount=1.0,
                refund_qty=0, refund_amount=0,
                free_qty=0, give_qty=0,
                cancelled_qty=0, cancelled_amount=0,
                list_price=40,
                price_1=40, qty_1=8,
                price_2=None, qty_2=None,
                price_3=None, qty_3=None,
                other_price_qty=None,
            ),
            order_row(
                store_num="001", store_name="店A",
                item_uuid="ITEM_X", item_name="单品X",
                qty=2, revenue=70,
                sales_price=80, original_amount=80,
                avg_member_discount=0.875,    # discounted 2 units
                refund_qty=0, refund_amount=0,
                cancelled_qty=1, cancelled_amount=35,   # 1 cancelled order
                list_price=40,
            ),
        ]
        self.bom_data = {
            "001": {"ITEM_X": [("M1", "盐", 2.0, "g", 1.0, 0.5)]}
        }
        # Uploaded price overrides BQ default 0.5 → 0.8
        self.uploaded = {"M1": 0.8}

        self.tmpdir = tempfile.mkdtemp()
        self.xlsx = os.path.join(self.tmpdir, "smoke.xlsx")

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _run_pipeline(self) -> dict:
        agg = aggregate_with_bom(
            self.bq_rows, self.bom_data, combo_structure={},
            uploaded_prices=self.uploaded, mode="single",
        )
        flat = _build_rows(agg, mode="single",
                          uploaded_prices=self.uploaded)
        sheet_cfg = load_sheet_config(str(YAML_PATH), "单品")
        wb = xlsxwriter.Workbook(self.xlsx)
        write_configured_sheet(wb, "单品", sheet_cfg, flat)
        wb.close()
        return {"agg": agg, "flat": flat}

    # ------------------------------------------------------------------
    # Layer-by-layer assertions on the SAME pipeline run.
    # ------------------------------------------------------------------

    def test_aggregation_rolls_two_rows_into_one(self):
        result = self._run_pipeline()
        agg = result["agg"]
        self.assertEqual(len(agg), 1, "two BQ rows for same item must merge")
        v = list(agg.values())[0]
        self.assertEqual(v["qty"], 10)
        self.assertEqual(v["revenue"], 310)
        self.assertEqual(v["cancelled_qty"], 1)
        self.assertEqual(v["cancelled_amount"], 35)
        # Weighted discount: (8*1.0 + 2*0.875)/10 = 0.975
        self.assertAlmostEqual(v["avg_member_discount"], 0.975)

    def test_unit_price_comes_from_uploaded_list(self):
        result = self._run_pipeline()
        bom = list(result["agg"].values())[0]["bom"]
        # (code, name, num, unit_price, uom)
        self.assertEqual(bom[0][3], 0.8,
                         "uploaded_prices must override BQ default of 0.5")

    def test_flat_row_columns_match_yaml_indices(self):
        result = self._run_pipeline()
        row = result["flat"][0]
        # Spot-check the indices most likely to drift in a refactor.
        self.assertEqual(row[0], "001")       # 门店编号
        self.assertEqual(row[2], "单品X")      # 商品名
        self.assertEqual(row[4], 10.0)        # 销量
        self.assertEqual(row[21], "M1")       # BOM code
        self.assertEqual(row[32], 1.0)        # 取消数量
        self.assertEqual(row[33], 35.0)       # 取消金额
        self.assertEqual(row[25], "ITEM_X")   # 隐藏 merge key

    def test_excel_file_renders_expected_cells(self):
        """End-to-end: file actually opens; the cells match the source rows."""
        self._run_pipeline()
        wb = load_workbook(self.xlsx)
        self.assertIn("单品", wb.sheetnames)
        ws = wb["单品"]

        # Header row + 1 data row (single BOM).
        # Column A in Excel = field_index 0 = 门店编号.
        self.assertEqual(ws.cell(row=2, column=1).value, "001")
        # 门店名称 (B) — merged column, value at top of block (=row 2).
        self.assertEqual(ws.cell(row=2, column=2).value, "店A")
        # 商品名称 (C) for sheet '单品' the yaml uses '单品名称' but field index 2.
        self.assertEqual(ws.cell(row=2, column=3).value, "单品X")
        # 销量 — field_index 4 in row; column position depends on yaml order
        # so we look up by header name to be robust.
        headers = {ws.cell(row=1, column=c).value: c
                   for c in range(1, ws.max_column + 1)}
        sales_col = headers["销量"]
        self.assertEqual(ws.cell(row=2, column=sales_col).value, 10)

        bom_code_col = headers["BOM物品编码"]
        self.assertEqual(ws.cell(row=2, column=bom_code_col).value, "M1")

        cancelled_qty_col = headers["取消数量"]
        self.assertEqual(ws.cell(row=2, column=cancelled_qty_col).value, 1)

    def test_freeze_panes_and_header_present(self):
        self._run_pipeline()
        ws = load_workbook(self.xlsx)["单品"]
        self.assertEqual(ws.freeze_panes, "A2",
                         "freeze_panes regressed — header must stay visible on scroll")
        # Header row is row 1; every column must have a non-empty name.
        names = [ws.cell(row=1, column=c).value
                 for c in range(1, ws.max_column + 1)]
        self.assertTrue(all(n for n in names), f"empty header detected: {names}")


class MultiBomBlockSmoke(unittest.TestCase):
    """Two BOM lines for the same item must produce a merged block in Excel."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.xlsx = os.path.join(self.tmpdir, "block.xlsx")

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_two_bom_lines_share_merge_block(self):
        rows = [order_row(item_uuid="P", item_name="商品P", qty=5,
                          revenue=100, sales_price=100, list_price=20)]
        bom_data = {"001": {"P": [
            ("M1", "盐", 1.0, "g", 1.0, 0.0),
            ("M2", "糖", 2.0, "g", 1.0, 0.0),
        ]}}
        agg = aggregate_with_bom(rows, bom_data, combo_structure={},
                                  uploaded_prices={"M1": 1.0, "M2": 2.0}, mode="single")
        flat = _build_rows(agg, mode="single",
                            uploaded_prices={"M1": 1.0, "M2": 2.0})
        self.assertEqual(len(flat), 2)

        sheet_cfg = load_sheet_config(str(YAML_PATH), "单品")
        wb = xlsxwriter.Workbook(self.xlsx)
        write_configured_sheet(wb, "单品", sheet_cfg, flat)
        wb.close()

        ws = load_workbook(self.xlsx)["单品"]
        merged = {str(r) for r in ws.merged_cells.ranges}
        # 门店编号 col = A; data rows 2 and 3 must be merged.
        self.assertIn("A2:A3", merged,
                      f"two-BOM block didn't merge column A; ranges={merged}")


if __name__ == "__main__":
    unittest.main()
