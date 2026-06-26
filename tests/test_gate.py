"""导出闸门: 阻断 / 放行 / --force 水印 三态行为 (spec §5 A4)。"""
import io
import unittest
from contextlib import redirect_stdout

import tests._setup  # noqa: F401

from semantic.validators.gate import validate_and_gate, GateOutcome
from semantic.validators.identities import DEFAULT_IDENTITIES


def _good_row():
    # 金额是萨当整数 (PR-B 7c: = 元 × 100)
    return {
        "qty": 100.0, "net_qty": 80.0, "free_qty": 5.0, "give_qty": 5.0,
        "refund_qty": 6.0, "cancelled_qty": 4.0,
        "sales_price": 100000, "revenue": 80000, "refund_amount": 6000,
        "free_amount": 5000, "give_amount": 5000, "discount_amount": 4000,
        "cancelled_amount": 3000, "gross_amount": 103000,
    }


def _bad_row():
    row = _good_row()
    row["revenue"] += 50000  # +500 元 = 50000 萨当 → 破坏金额恒等式 → MUST_FIX
    return row


class TestGate(unittest.TestCase):
    def test_clean_rows_pass_through(self):
        buf = io.StringIO()
        with redirect_stdout(buf):
            out = validate_and_gate([_good_row()], DEFAULT_IDENTITIES,
                                    force=False, report_name="t")
        self.assertIsInstance(out, GateOutcome)
        self.assertFalse(out.needs_watermark)

    def test_must_fix_blocks_with_exit_2(self):
        buf = io.StringIO()
        with self.assertRaises(SystemExit) as cm, redirect_stdout(buf):
            validate_and_gate([_bad_row()], DEFAULT_IDENTITIES,
                              force=False, report_name="t")
        self.assertEqual(cm.exception.code, 2)
        self.assertIn("零容差", buf.getvalue())

    def test_force_returns_with_watermark(self):
        buf = io.StringIO()
        with redirect_stdout(buf):
            out = validate_and_gate([_bad_row()], DEFAULT_IDENTITIES,
                                    force=True, report_name="t")
        self.assertTrue(out.needs_watermark)
        lines = out.watermark_lines()
        self.assertTrue(any("未通过" in ln for ln in lines))

    def test_empty_export_blocks(self):
        """0 行也是无声错误 — '成功导出空表' 不许发生."""
        buf = io.StringIO()
        with self.assertRaises(SystemExit) as cm, redirect_stdout(buf):
            validate_and_gate([], DEFAULT_IDENTITIES, force=False, report_name="t")
        self.assertEqual(cm.exception.code, 2)

    def test_empty_export_forced_gets_watermark(self):
        buf = io.StringIO()
        with redirect_stdout(buf):
            out = validate_and_gate([], DEFAULT_IDENTITIES, force=True, report_name="t")
        self.assertTrue(out.needs_watermark)
        self.assertTrue(any("数据为空" in ln for ln in out.watermark_lines()))


class TestWatermarkHelpers(unittest.TestCase):
    def test_openpyxl_watermark_first_sheet(self):
        import openpyxl
        from semantic.validators.gate import add_watermark_sheet_openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "数据"
        add_watermark_sheet_openpyxl(wb, ["⚠️ 本表未通过零容差校验", "强制导出"])
        self.assertEqual(wb.sheetnames[0], "⚠️校验未通过")

    def test_xlsxwriter_watermark_activated(self):
        import os
        import tempfile
        import xlsxwriter
        from semantic.validators.gate import add_watermark_sheet_xlsxwriter
        path = os.path.join(tempfile.mkdtemp(), "t.xlsx")
        wb = xlsxwriter.Workbook(path)
        wb.add_worksheet("数据")
        ws = add_watermark_sheet_xlsxwriter(wb, ["⚠️ 本表未通过零容差校验"])
        self.assertEqual(ws.name, "⚠️校验未通过")
        wb.close()
        self.assertTrue(os.path.exists(path))


if __name__ == "__main__":
    unittest.main()
