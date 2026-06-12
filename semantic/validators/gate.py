"""导出闸门 — 零容差校验失败时硬阻断, --force 留逃生口但水印可见 (spec §2 决策 5)。

任何把数据写到 Excel/CSV 的报表脚本, 写盘前必须调 validate_and_gate():

    from semantic.validators.gate import validate_and_gate
    outcome = validate_and_gate(check_rows, FULL_IDENTITIES,
                                force=args.force, report_name="profit_margin",
                                row_label=lambda r: f"店 {r['store_num']}")
    # ... 写盘 ...
    if outcome.needs_watermark:
        add_watermark_sheet_xlsxwriter(wb, outcome.watermark_lines())  # 或 openpyxl 版

行为:
  - 全绿 / 仅 🟡         → 返回 GateOutcome, 正常写盘
  - 有 🔴 且无 force     → 打印违反清单, sys.exit(2), 不产出文件
  - 有 🔴 且 force=True  → 返回 needs_watermark=True, 脚本写盘后必须打水印
  - check_rows 为空      → 按 🔴 处理 ("成功导出空表"是无声错误)

结构性测试 (tests/test_validator_coverage.py) 强制 19 个报表脚本全部接入.
"""
from __future__ import annotations

import sys
from dataclasses import dataclass
from typing import Callable

from .core import Identity, Result, Severity, check, print_result

EXIT_GATE_BLOCKED = 2

WATERMARK_SHEET_NAME = "⚠️校验未通过"


@dataclass
class GateOutcome:
    result: Result
    forced: bool
    report_name: str
    empty_blocked: bool = False  # 空导出被 --force 放行 — 也要水印

    @property
    def needs_watermark(self) -> bool:
        return self.forced and (self.result.has_must_fix or self.empty_blocked)

    def watermark_lines(self) -> list[str]:
        must = [v for v in self.result.violations
                if v.severity == Severity.MUST_FIX]
        detail = ("数据为空 (0 行)" if self.empty_blocked
                  else f"🔴 离谱违反: {len(must)} 条")
        return [
            "⚠️ 本表未通过零容差校验 (--force 强制导出)",
            f"报表: {self.report_name}   {detail}",
            "本文件中的数字未经数学校验背书, 不得作为对外交付口径.",
            "违反明细见导出时的 console 输出.",
        ]


def validate_and_gate(
    check_rows: list[dict],
    identities: list[Identity],
    *,
    force: bool,
    report_name: str,
    row_label: Callable[[dict], str] = lambda r: str(r),
    min_rows: int = 1,
) -> GateOutcome:
    if len(check_rows) < min_rows:
        print(f"🔴 [{report_name}] 零容差闸门: check_rows 仅 {len(check_rows)} 行 "
              f"(< {min_rows}) — 空导出按无声错误阻断.")
        if not force:
            sys.exit(EXIT_GATE_BLOCKED)
        return GateOutcome(result=check([], identities), forced=True,
                           report_name=report_name, empty_blocked=True)

    result = check(check_rows, identities)
    print(f"\n[{report_name}] 零容差闸门校验 ({len(check_rows)} 行 × {len(identities)} 条恒等式)")
    print_result(result, row_label=row_label)

    if result.has_must_fix and not force:
        print(f"🔴 [{report_name}] 零容差闸门: 有离谱违反, 已阻断导出 (exit {EXIT_GATE_BLOCKED}).")
        print("    修复数据/口径后重跑; 或 --force 强制导出 (文件将带水印, 不得对外交付).")
        sys.exit(EXIT_GATE_BLOCKED)
    if result.has_must_fix:
        print(f"⚠️  [{report_name}] --force 强制导出: 文件将打水印.")
    return GateOutcome(result=result, forced=force, report_name=report_name)


def add_watermark_sheet_xlsxwriter(workbook, lines: list[str]):
    """xlsxwriter 无法事后调 sheet 顺序, 用 activate+set_first_sheet 让水印页
    成为打开文件的第一眼."""
    ws = workbook.add_worksheet(WATERMARK_SHEET_NAME)
    fmt = workbook.add_format(
        {"bold": True, "font_color": "white", "bg_color": "#C00000", "font_size": 14})
    for i, line in enumerate(lines):
        ws.write(i, 0, line, fmt)
    ws.set_column(0, 0, 90)
    ws.activate()
    ws.set_first_sheet()
    return ws


def add_watermark_sheet_openpyxl(workbook, lines: list[str]):
    from openpyxl.styles import Font, PatternFill
    ws = workbook.create_sheet(WATERMARK_SHEET_NAME, 0)
    fill = PatternFill("solid", fgColor="C00000")
    font = Font(bold=True, color="FFFFFF", size=14)
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.fill = fill
        cell.font = font
    ws.column_dimensions["A"].width = 90
    workbook.active = 0
    return ws
