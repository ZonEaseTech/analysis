#!/usr/bin/env python3
"""
利润报表交叉验证脚本
"""

import os
import sys

sys.path.insert(0, os.path.dirname(__file__))

from collections import defaultdict
from bq_reports.utils.bq_client import get_bq_client, setup_proxy
from bq_reports.utils.erpnext_api import load_erpnext_prices

setup_proxy()
client = get_bq_client('diyl-407103')
erp = load_erpnext_prices()


def query(sql):
    return list(client.query(sql).result())


REPORT = []

def log(section, detail):
    REPORT.append(f"## {section}\n{detail}\n")
    print(f"\n=== {section} ===")
    print(detail)


# ============================================================================
# 验证1: BOM结构跨门店一致性
# ============================================================================
stores = ['1958987436032000', '2269470793728000', '2598648160256000']
items = ['秘制烤鸡腿堡', '脆皮炸鸡', '香辣鸡翅']

details = []
for item_name in items:
    details.append(f"\n{item_name}:")
    for store in stores:
        sql = f"""
        SELECT
          JSON_EXTRACT_SCALAR(pp.name, '$.zh') as product_name,
          m.code,
          rm.num as bom_num,
          rm_base.max_rate,
          m.price as bq_price
        FROM `diyl-407103`.shop{store}.ttpos_product_package pp
        JOIN `diyl-407103`.shop{store}.ttpos_product_bom pb
          ON pb.product_package_uuid = pp.uuid AND pb.delete_time = 0
        LEFT JOIN `diyl-407103`.shop{store}.ttpos_related_material rm
          ON (
            (pb.product_bom_card_uuid > 0 AND rm.related_uuid = pb.product_bom_card_uuid)
            OR (pb.product_bom_card_uuid = 0 AND rm.related_uuid = pb.uuid)
          )
          AND rm.delete_time = 0
        LEFT JOIN `diyl-407103`.shop{store}.ttpos_material m
          ON m.uuid = rm.material_uuid AND m.delete_time = 0
        LEFT JOIN (
          SELECT material_uuid, MAX(base_unit_conversion_rate) as max_rate
          FROM `diyl-407103`.shop{store}.ttpos_related_material
          WHERE delete_time = 0
          GROUP BY material_uuid
        ) rm_base ON rm_base.material_uuid = m.uuid
        WHERE JSON_EXTRACT_SCALAR(pp.name, '$.zh') = '{item_name}'
          AND pp.delete_time = 0
        ORDER BY m.code
        """
        rows = query(sql)
        total = 0
        for r in rows:
            rate = float(r.max_rate or 1)
            if rate <= 0: rate = 1
            bq_price = float(r.bq_price or 0)
            erp_price = erp.get(r.code, erp.get(r.code.upper(), erp.get(r.code.lower(), bq_price))) if r.code else bq_price
            consumption = float(r.bom_num or 0) / rate
            cost = consumption * erp_price
            total += cost
        details.append(f"  门店{store[-4:]}: {len(rows)}个BOM, 单份成本={total:.2f}")

log("验证1: 同一单品在不同门店的BOM一致性", "\n".join(details))


# ============================================================================
# 验证2: ERPNext价格覆盖率
# ============================================================================
sql = """
SELECT DISTINCT m.code
FROM `diyl-407103`.shop1958987436032000.ttpos_material m
WHERE m.delete_time = 0
  AND m.code IS NOT NULL
"""
all_codes = {r.code for r in query(sql)}
covered = sum(1 for c in all_codes if c in erp or c.upper() in erp or c.lower() in erp)
details = [
    f"TTPOS 物料总数: {len(all_codes)}",
    f"ERPNext 有价格: {covered}",
    f"覆盖率: {covered/len(all_codes)*100:.1f}%",
    f"缺失价格物料数: {len(all_codes) - covered}",
]
log("验证2: ERPNext价格覆盖率", "\n".join(details))


# ============================================================================
# 验证3: 毛利率分布合理性
# ============================================================================
from openpyxl import load_workbook

wb = load_workbook('exports/combo_profit_202603.xlsx', data_only=False)
ws = wb.active

margins = []
for i in range(2, ws.max_row + 1):
    m = ws.cell(row=i, column=13).value
    if m is not None:
        margins.append(m)

neg = sum(1 for m in margins if m < 0)
pos = sum(1 for m in margins if m > 0)
zero = sum(1 for m in margins if m == 0)

details = [
    f"独立商品数: {len(margins)}",
    f"负毛利率: {neg} ({neg/len(margins)*100:.1f}%)",
    f"正毛利率: {pos} ({pos/len(margins)*100:.1f}%)",
    f"零毛利率: {zero}",
    f"最低: {min(margins)*100:.1f}%",
    f"最高: {max(margins)*100:.1f}%",
    f"中位数: {sorted(margins)[len(margins)//2]*100:.1f}%",
]
log("验证3: 套餐报表毛利率分布", "\n".join(details))

# 单品
wb2 = load_workbook('exports/single_profit_202603.xlsx', data_only=False)
ws2 = wb2.active
margins2 = []
for i in range(2, ws2.max_row + 1):
    m = ws2.cell(row=i, column=13).value
    if m is not None:
        margins2.append(m)

neg2 = sum(1 for m in margins2 if m < 0)
pos2 = sum(1 for m in margins2 if m > 0)
details = [
    f"独立商品数: {len(margins2)}",
    f"负毛利率: {neg2} ({neg2/len(margins2)*100:.1f}%)",
    f"正毛利率: {pos2} ({pos2/len(margins2)*100:.1f}%)",
    f"最低: {min(margins2)*100:.1f}%",
    f"最高: {max(margins2)*100:.1f}%",
]
log("验证4: 单品报表毛利率分布", "\n".join(details))


# ============================================================================
# 验证5: 销量 vs BOM行数关系（检查是否有销量为0但BOM很多的异常）
# ============================================================================
sql = """
SELECT
  JSON_EXTRACT_SCALAR(pp.name, '$.zh') as combo_name,
  COUNT(*) as bom_count
FROM `diyl-407103`.shop1958987436032000.ttpos_product_package pp
JOIN `diyl-407103`.shop1958987436032000.ttpos_product_bom pb
  ON pb.product_package_uuid = pp.uuid AND pb.delete_time = 0
LEFT JOIN `diyl-407103`.shop1958987436032000.ttpos_related_material rm
  ON (
    (pb.product_bom_card_uuid > 0 AND rm.related_uuid = pb.product_bom_card_uuid)
    OR (pb.product_bom_card_uuid = 0 AND rm.related_uuid = pb.uuid)
  )
  AND rm.delete_time = 0
WHERE pp.delete_time = 0
  AND pp.product_type = 2
GROUP BY pp.uuid, combo_name
ORDER BY bom_count DESC
LIMIT 20
"""
rows = query(sql)
details = [f"{r.combo_name}: {r.bom_count}个BOM" for r in rows]
log("验证5: 套餐BOM复杂度Top20", "\n".join(details))


# ============================================================================
# 验证6: 高负毛利率商品的共同特征
# ============================================================================
wb = load_workbook('exports/combo_profit_202603.xlsx', data_only=False)
ws = wb.active

# 收集负毛利率商品
neg_items = defaultdict(lambda: {"count": 0, "costs": [], "prices": []})
for i in range(2, ws.max_row + 1):
    margin = ws.cell(row=i, column=13).value
    if margin is not None and margin < -1.0:  # 低于-100%
        name = ws.cell(row=i, column=2).value
        price = ws.cell(row=i, column=4).value
        cost = ws.cell(row=i, column=11).value
        neg_items[name]["count"] += 1
        neg_items[name]["prices"].append(price)
        neg_items[name]["costs"].append(cost)

details = ["毛利率 < -100% 的商品:"]
for name, data in sorted(neg_items.items(), key=lambda x: x[1]["count"], reverse=True)[:15]:
    avg_price = sum(data["prices"]) / len(data["prices"])
    avg_cost = sum(data["costs"]) / len(data["costs"])
    details.append(f"  {name}: 出现{data['count']}次, 均价={avg_price:.0f}, 均成本={avg_cost:.0f}")
log("验证6: 高亏损商品特征", "\n".join(details))


# ============================================================================
# 生成报告
# ============================================================================
report_path = 'exports/validation_report.md'
with open(report_path, 'w') as f:
    f.write("# 利润报表交叉验证报告\n\n")
    f.write(f"生成时间: 2026-04-22\n\n")
    f.write("\n".join(REPORT))

print(f"\n\n报告已保存: {report_path}")
