"""HukuLoader — 弧酷 xlsx → ExternalSKURow 列表.

完整流程:
  1. 读 xlsx, 按 (店, 菜品, 业务日期) 聚合
  2. 店名 → BQ 店编 (clean rule + 手工 override)
  3. 日期切割: 业务日期 < BQ 该店切换日 → 保留, 否则丢 (BQ 接手)
  4. 跨日聚合到 (店, 菜品) → qty + revenue
  5. BOM 匹配 (strict + alias + split) → 物料明细
  6. 算成本 / 毛利 / 包装成 ExternalSKURow
"""
from __future__ import annotations

import datetime
import os
from typing import Dict, List, Optional

import openpyxl

from external_sales.base import ExternalSalesSource, ExternalSKURow
from external_sales.cutover import query_cutover_dates
from external_sales.huku.match import match_item
from external_sales.huku.normalize import (
    MANUAL_STORE,
    clean_bq_store_name,
    clean_ext_store_name,
)


class HukuLoader(ExternalSalesSource):

    def __init__(self, path: str):
        if not os.path.exists(path):
            raise FileNotFoundError(f"huku xlsx not found: {path}")
        self.path = path

    def load(self, *, month: str, bq_client=None, merchants=None,
             config=None) -> List[ExternalSKURow]:
        # 1. 读 xlsx + 按日期聚合
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        # 列序: 4 门店编码 / 5 门店名称 / 7 业务日期 / 9 大类 / 11 菜品编码 / 12 菜品名 / 14 净收数量 / 15 净收金额
        store_to_bq, ext_stores_unique = self._build_store_map(wb, config)

        # 2. 查 BQ 切换日 (用于日期切割)
        cutover_dates: Dict[str, datetime.date] = {}
        if bq_client is not None and merchants is not None:
            bq_codes = sorted(set(store_to_bq.values()))
            print(f"  [huku] 查 BQ 切换日 ({len(bq_codes)} 店)...")
            cutover_dates = query_cutover_dates(bq_client, merchants, bq_codes)
            for bc, d in sorted(cutover_dates.items()):
                print(f"    {bc} 切换日: {d}")

        # 3. 按 (店, 菜品, 大类) 聚合, 日期切割
        ws = wb["菜品销售报表"]
        agg = {}      # (bq_code, bq_name, item, category) → {qty, rev}
        dropped_rev = 0.0; dropped_rows = 0
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i < 2 or not r[5] or not r[12]: continue
            ext_store = r[5].strip()
            bq_code = store_to_bq.get(ext_store)
            if not bq_code: continue
            biz_date_raw = r[7]
            biz_date = self._parse_date(biz_date_raw)
            cutover = cutover_dates.get(bq_code)
            if cutover and biz_date and biz_date >= cutover:
                dropped_rev += float(r[15] or 0); dropped_rows += 1
                continue
            bq_name = self._bq_name_for(bq_code, config)
            key = (bq_code, bq_name, r[12].strip(), r[9] or "")
            qty = float(r[14] or 0); rev = float(r[15] or 0)
            if key not in agg: agg[key] = {"qty": 0.0, "rev": 0.0}
            agg[key]["qty"] += qty; agg[key]["rev"] += rev
        if cutover_dates:
            print(f"  [huku] 日期切割丢弃: {dropped_rows} 行 / {dropped_rev:,.0f} 泰铢 (BQ 接手部分)")

        # 4. BOM 匹配 + 写成 SKU rows
        from bq_reports.profit_margin_report import (
            _load_bom_layers, _load_material_price_layers,
            _build_material_price_resolver, _resolve_unit_price_with_source,
            load_config as _load_config,
        )
        cfg = config or _load_config()
        bom_layers = _load_bom_layers(cfg)
        price_layers = _load_material_price_layers(cfg)
        resolver = _build_material_price_resolver({}, {}, price_layers, strict=True)

        out: List[ExternalSKURow] = []
        match_stats = {}
        for (bq_code, bq_name, item, category), m in sorted(agg.items()):
            bom_records, match_type, bom_src = match_item(item, bom_layers)
            match_stats[match_type] = match_stats.get(match_type, 0) + 1
            materials = []
            unit_cost = 0.0
            srcs = set()
            if bom_records:
                for code, mname, q, unit in bom_records:
                    p, src = _resolve_unit_price_with_source(
                        code, 0, {}, {}, price_layers=price_layers,
                        strict=True, material_name=mname, resolver=resolver,
                    )
                    srcs.add(src)
                    materials.append((code, mname, q, p, unit))
                    unit_cost += q * p
            price_src = " + ".join(sorted(srcs)) if srcs else "无"

            out.append(ExternalSKURow(
                store_num=bq_code,
                store_name=bq_name or "",
                item_name=item,
                category=("套餐" if category == "套餐类" else "单品"),
                qty_net=m["qty"],
                revenue=m["rev"],
                materials=materials,
                unit_cost=unit_cost,
                bom_source=bom_src,
                price_source=price_src,
                match_type=match_type,
                item_uuid=f"ext:huku:{bq_code}:{hash(item) & 0xFFFFFFFF:08x}",
            ))
        print(f"  [huku] 匹配统计: {match_stats}")
        return out

    # ── 内部 ───────────────────────────────────────────────

    def _build_store_map(self, wb, config) -> tuple[dict, set]:
        """弧酷店名 → BQ 店编 (3 位).

        BQ ttpos_setting 是店名权威源, 覆盖所有 56 店.
        旧版从 Excel 加载会缺 030 等漏录的店, 已切换.
        """
        if config is None:
            from bq_reports.profit_margin_report import load_config
            config = load_config()
        # 从 BQ 加载 (需要 bq_client, 通过 self._bq_client_for_names 临时取)
        from bq_reports.utils.bq_client import setup_proxy, get_bq_client
        from bq_reports.profit_margin_report import _load_store_names
        setup_proxy()
        client = get_bq_client()
        names = _load_store_names(config, client=client)
        # names dict 已含 "001" 和 "1" 两种形式; 取 zfill(3) 版本作 BQ_all
        bq_all = {k: v for k, v in names.items() if len(k) == 3}
        self._bq_all = bq_all
        bqc = {c: clean_bq_store_name(n) for c, n in bq_all.items()}

        ext_stores = set()
        ws = wb["菜品销售报表"]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i < 2: continue
            if r[5]: ext_stores.add(r[5].strip())

        mapping = {}
        for en in ext_stores:
            if en in MANUAL_STORE:
                mapping[en] = MANUAL_STORE[en]; continue
            ec = clean_ext_store_name(en); hit = None
            for bc, bn in bqc.items():
                if bn == ec: hit = bc; break
            if not hit:
                for bc, bn in bqc.items():
                    if ec and (ec in bn or bn in ec): hit = bc; break
            if hit: mapping[en] = hit
        return mapping, ext_stores

    def _bq_name_for(self, bq_code: str, config) -> str:
        return getattr(self, "_bq_all", {}).get(bq_code, "")

    @staticmethod
    def _parse_date(v) -> Optional[datetime.date]:
        if v is None: return None
        if isinstance(v, datetime.datetime): return v.date()
        if isinstance(v, datetime.date): return v
        s = str(v).strip()[:10]
        try:
            return datetime.date.fromisoformat(s)
        except ValueError:
            return None
