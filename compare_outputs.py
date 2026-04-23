#!/usr/bin/env python3
"""比较旧版和新版输出差异。"""
import sys
from openpyxl import load_workbook

old_path = "exports/profit_20260401_0415_test.xlsx"
new_path = "exports/profit_20260401_0415_opt.xlsx"

wb_old = load_workbook(old_path, data_only=True)
wb_new = load_workbook(new_path, data_only=True)

for sheet_name in ["套餐", "单品"]:
    ws_old = wb_old[sheet_name]
    ws_new = wb_new[sheet_name]

    old_rows = []
    for row in ws_old.iter_rows(min_row=2, values_only=True):
        old_rows.append(tuple(row))

    new_rows = []
    for row in ws_new.iter_rows(min_row=2, values_only=True):
        new_rows.append(tuple(row))

    old_set = set(old_rows)
    new_set = set(new_rows)

    only_old = old_set - new_set
    only_new = new_set - old_set

    print(f"\n=== {sheet_name} ===")
    print(f"  旧版: {len(old_rows)} 行")
    print(f"  新版: {len(new_rows)} 行")
    print(f"  只在旧版: {len(only_old)} 行")
    print(f"  只在新版: {len(only_new)} 行")

    if only_old:
        print(f"\n  旧版独有的前 5 行:")
        for row in list(only_old)[:5]:
            print(f"    {row}")

    if only_new:
        print(f"\n  新版独有的前 5 行:")
        for row in list(only_new)[:5]:
            print(f"    {row}")

wb_old.close()
wb_new.close()
