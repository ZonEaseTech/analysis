#!/usr/bin/env python3
"""
修复后的利润报表交叉验证（15列格式 + 双sheet）
"""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import load_workbook

REPORT = []

def section(title):
    print(f"\n=== {title} ===")
    REPORT.append(f"\n## {title}\n")

def detail(text):
    print(text)
    REPORT.append(text + "\n")

FILE = "exports/profit_202603.xlsx"

# 列索引（15列格式）
COL_STORE_NUM = 1
COL_STORE_NAME = 2
COL_NAME = 3
COL_QTY = 4
COL_PRICE = 5
COL_REVENUE = 6
COL_BOM_NAME = 7
COL_BOM_CODE = 8
COL_MAT_PRICE = 9
COL_BOM_NUM = 10
COL_UOM = 11
COL_MAT_COST = 12
COL_UNIT_COST = 13
COL_PROFIT = 14
COL_MARGIN = 15

# ============================================================================
# 1. 验证: 同一商品在不同门店的成本一致性
# ============================================================================
section("验证1: 跨门店成本一致性")

for mode, sheet_name in [("套餐", "套餐"), ("单品", "单品")]:
    wb = load_workbook(FILE, data_only=False)
    ws = wb[sheet_name]

    from collections import defaultdict
    item_costs = defaultdict(dict)
    for i in range(2, ws.max_row + 1):
        name = ws.cell(row=i, column=COL_NAME).value
        store_num = ws.cell(row=i, column=COL_STORE_NUM).value
        cost = ws.cell(row=i, column=COL_UNIT_COST).value
        if cost is not None and name:
            item_costs[name][store_num] = cost

    inconsistent = []
    for name, stores in item_costs.items():
        if len(stores) >= 3:
            costs = list(stores.values())
            if max(costs) - min(costs) > 1.0:
                inconsistent.append((name, min(costs), max(costs), len(stores)))

    detail(f"{mode}: 检查了 {len(item_costs)} 个商品")
    detail(f"  跨门店成本不一致(差异>1泰铢): {len(inconsistent)} 个")
    for name, mn, mx, cnt in sorted(inconsistent, key=lambda x: x[2]-x[1], reverse=True)[:5]:
        detail(f"    {name}: {cnt}个门店, 成本范围 {mn:.2f} ~ {mx:.2f}")

    wb.close()

# ============================================================================
# 2. 验证: 极端毛利率检查
# ============================================================================
section("验证2: 极端毛利率案例抽查")

for mode, sheet_name in [("套餐", "套餐"), ("单品", "单品")]:
    wb = load_workbook(FILE, data_only=False)
    ws = wb[sheet_name]

    all_rows = []
    for i in range(2, ws.max_row + 1):
        margin = ws.cell(row=i, column=COL_MARGIN).value
        if margin is not None:
            row = [ws.cell(row=i, column=c).value for c in range(1, 16)]
            all_rows.append((margin, row))

    all_rows.sort()
    detail(f"\n{mode} - 毛利率最低的5个:")
    for margin, row in all_rows[:5]:
        detail(f"  {row[0]}/{row[1]}/{row[2]}: 售价{row[4]}, 成本{row[12]}, 毛利{margin:.1%}")
        detail(f"    BOM: {row[6]}({row[7]}), 消耗{row[9]}, 单位{row[10]}, 单价{row[8]}")

    detail(f"\n{mode} - 毛利率最高的5个:")
    for margin, row in all_rows[-5:]:
        detail(f"  {row[0]}/{row[1]}/{row[2]}: 售价{row[4]}, 成本{row[12]}, 毛利{margin:.1%}")

    wb.close()

# ============================================================================
# 3. 验证: 零成本商品占比
# ============================================================================
section("验证3: 零成本商品占比")

for mode, sheet_name in [("套餐", "套餐"), ("单品", "单品")]:
    wb = load_workbook(FILE, data_only=False)
    ws = wb[sheet_name]

    zero_cost = 0
    low_cost = 0
    normal_cost = 0
    high_cost = 0

    for i in range(2, ws.max_row + 1):
        cost = ws.cell(row=i, column=COL_UNIT_COST).value
        if cost is None:
            continue
        if cost == 0:
            zero_cost += 1
        elif cost < 1:
            low_cost += 1
        elif cost > 200:
            high_cost += 1
        else:
            normal_cost += 1

    total = zero_cost + low_cost + normal_cost + high_cost
    detail(f"{mode} ({total}个商品):")
    detail(f"  成本=0: {zero_cost} ({zero_cost/total*100:.1f}%)")
    detail(f"  0<成本<1: {low_cost} ({low_cost/total*100:.1f}%)")
    detail(f"  1<=成本<=200: {normal_cost} ({normal_cost/total*100:.1f}%)")
    detail(f"  成本>200: {high_cost} ({high_cost/total*100:.1f}%)")

    wb.close()

# ============================================================================
# 4. 验证: 成本 > 售价的案例（真实亏损）
# ============================================================================
section("验证4: 成本 > 售价的案例（真实亏损商品）")

for mode, sheet_name in [("套餐", "套餐"), ("单品", "单品")]:
    wb = load_workbook(FILE, data_only=False)
    ws = wb[sheet_name]

    loss_items = []
    for i in range(2, ws.max_row + 1):
        price = ws.cell(row=i, column=COL_PRICE).value
        cost = ws.cell(row=i, column=COL_UNIT_COST).value
        margin = ws.cell(row=i, column=COL_MARGIN).value
        name = ws.cell(row=i, column=COL_NAME).value
        if cost is not None and price is not None and cost > price and margin is not None and margin < 0:
            loss_items.append((name, price, cost, margin))

    from collections import Counter
    counts = Counter(name for name, _, _, _ in loss_items)
    detail(f"{mode}: 真实亏损的商品种类: {len(counts)} 个")
    for name, cnt in counts.most_common(10):
        avg_price = sum(p for n, p, _, _ in loss_items if n == name) / cnt
        avg_cost = sum(c for n, _, c, _ in loss_items if n == name) / cnt
        detail(f"  {name}: {cnt}个门店, 均价={avg_price:.0f}, 均成本={avg_cost:.0f}")

    wb.close()

# ============================================================================
# 5. 验证: 关键商品成本合理性（人工判断）
# ============================================================================
section("验证5: 关键商品成本合理性抽查")

wb = load_workbook(FILE, data_only=False)
ws = wb["套餐"]

for i in range(2, ws.max_row + 1):
    name = ws.cell(row=i, column=COL_NAME).value
    if name == "脆皮全鸡+2杯可乐":
        row = [ws.cell(row=i, column=c).value for c in range(1, 16)]
        detail(f"脆皮全鸡+2杯可乐 @ {row[0]} - {row[1]}:")
        detail(f"  销量={row[3]}, 单价={row[4]}, 销售额={row[5]}")
        detail(f"  单份总成本={row[12]}, 毛利={row[13]}, 毛利率={row[14]:.1%}")
        j = i
        while j < ws.max_row and ws.cell(row=j+1, column=COL_MARGIN).value is None:
            j += 1
            r = [ws.cell(row=j, column=c).value for c in range(1, 16)]
            detail(f"    {r[6]}({r[7]}): 单价={r[8]}, 消耗={r[9]}, 单位={r[10]}, 总成本={r[11]}")
        break

wb.close()

# ============================================================================
# 保存报告
# ============================================================================
report_path = "exports/validation_report_v4.md"
with open(report_path, "w") as f:
    f.write("# 利润报表交叉验证报告 V4（15列门店编号+名称格式）\n\n")
    f.write("".join(REPORT))

print(f"\n\n报告已保存: {report_path}")
