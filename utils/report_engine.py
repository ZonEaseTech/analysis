#!/usr/bin/env python3
"""
报表引擎 —— 封装 BigQuery 并发查询、外部资源加载、配置驱动 Excel 写入。

将"容易踩坑"的公共逻辑集中管理：
  - 代理设置 + BQ 连接
  - 多门店并发查询
  - 外部资源加载（adapter + cache）
  - Excel 样式（表头、边框、列宽、冻结首行）
  - 合并单元格
  - 公式列（行级 + block 级）
  - 条件格式（负值标红）

报表脚本只需关心：
  1. SQL 模板
  2. 聚合逻辑（去重、分组、BOM 展开）
  3. 列配置（YAML 或代码）
  4. 调用引擎执行

Usage:
    from utils.report_engine import ReportEngine, ColumnConfig, SheetConfig, query_all_shops

    # 1. 初始化
    engine = ReportEngine(project_id="diyl-407103")

    # 2. 并发查询
    raw_rows, errors = query_all_shops(
        engine.client, sql_template, merchants, start_ts, end_ts, workers=10,
        row_proxy_factory=lambda row, acc, num, name: RowProxy(row, acc, num, name)
    )

    # 3. 聚合（报表脚本自定义）
    agg_data = my_aggregate(raw_rows)
    rows = build_flat_rows(agg_data)   # 扁平化为 list[list]

    # 4. 加载列配置（YAML）
    sheet_cfg = engine.load_sheet_config("resources/reports/profit_margin.yaml", "套餐")

    # 5. 写入 Excel
    wb = Workbook()
    ws = wb.create_sheet("套餐")
    engine.write_sheet(ws, sheet_cfg, rows)
    wb.save("output.xlsx")
"""

import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.cache import get_cache, set_cache, cache_key
from utils.resource_adapter import get_adapter

# BQ 相关 import 延迟到 ReportEngine 初始化时，避免无 google-cloud-bigquery 环境报错


# ───────────────────────────────────────────────────────────────
# 数据模型
# ───────────────────────────────────────────────────────────────

@dataclass
class ColumnConfig:
    """Excel 列配置"""
    name: str
    field_index: int                    # 在行数据中的索引（0-based）
    col_type: str = "value"             # value | formula | block_formula
    format_str: str = ""                # 数字/百分比格式，如 "0.00", "0.00%"
    width: int = 15
    merge: bool = False                 # 是否合并（block 级别）
    formula_template: str = ""          # 公式模板，支持 {row} {block_start} {block_end} {col}
    negative_red: bool = False          # 负值标红（基于预计算值）


@dataclass
class SheetConfig:
    """Sheet 配置"""
    name: str
    columns: List[ColumnConfig]
    merge_key_indices: List[int] = field(default_factory=list)  # block 判定字段索引
    freeze_panes: str = "A2"


# ───────────────────────────────────────────────────────────────
# Excel 样式常量
# ───────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
RED_FONT = Font(color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# ───────────────────────────────────────────────────────────────
# 查询层
# ───────────────────────────────────────────────────────────────

def query_all_shops(
    client,
    sql_template: str,
    merchants: List[Tuple],
    start_ts: int,
    end_ts: int,
    workers: int = 10,
    row_proxy_factory: Optional[Callable] = None,
    label: str = "门店",
) -> Tuple[List[Any], List[Dict]]:
    """
    并发查询所有门店。

    Args:
        client: BigQuery client
        sql_template: SQL 模板，含 {project} {dataset} {start_ts} {end_ts}
        merchants: [(account, uuid_str, store_num, store_name), ...]
        start_ts, end_ts: 时间范围（Unix 秒）
        workers: 并发线程数
        row_proxy_factory: 可选，(row, account, store_num, store_name) -> proxy_row
        label: 日志标签

    Returns:
        (all_raw_rows, errors)
    """
    all_rows = []
    errors = []

    def _query_one(account_uuid):
        account, uuid_str, store_num, store_name = account_uuid
        dataset = f"shop{uuid_str}"
        sql = sql_template.format(
            project=client.project,
            dataset=dataset,
            start_ts=start_ts,
            end_ts=end_ts,
        )
        try:
            rows = list(client.query(sql).result())
            if row_proxy_factory:
                rows = [row_proxy_factory(r, account, store_num, store_name) for r in rows]
            return {"account": account, "rows": rows, "error": None}
        except Exception as e:
            return {"account": account, "rows": [], "error": str(e)}

    max_workers = min(workers, len(merchants))
    print(f"[{label}] 并发查询: {max_workers} 线程")

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_query_one, m): m[0] for m in merchants}
        for future in as_completed(futures):
            result = future.result()
            if result["error"]:
                errors.append({"account": result["account"], "error": result["error"]})
                print(f"[错误] {result['account']}: {result['error']}")
            else:
                all_rows.extend(result["rows"])
                print(f"[OK] {result['account']}: {len(result['rows'])} 行")

    print(f"\n[{label}] 查询完成: {len(merchants) - len(errors)}/{len(merchants)} 个门店成功")
    if errors:
        print(f"[{label}] 错误数: {len(errors)}")

    return all_rows, errors


# ───────────────────────────────────────────────────────────────
# 资源加载层
# ───────────────────────────────────────────────────────────────

def load_resources(resource_configs: List[Dict]) -> Dict[str, Any]:
    """
    加载多个外部资源，返回 name -> data 字典。

    resource_configs: [
        {"name": "store_names", "adapter": "excel", "path": "...", "mapping": {...}, "cache_ttl": 604800},
        ...
    ]
    """
    results = {}
    for cfg in resource_configs:
        name = cfg["name"]
        adapter_name = cfg["adapter"]
        cache_ttl = cfg.get("cache_ttl", 3600)

        cache_key_str = cache_key(name, {"path": cfg.get("path", "")})
        cached = get_cache(cache_key_str, ttl_seconds=cache_ttl)
        if cached is not None:
            print(f"[Resource] 缓存命中: {name}")
            results[name] = cached
            continue

        adapter = get_adapter(adapter_name)
        # 去掉 engine 专用字段再传给 adapter
        adapter_cfg = {k: v for k, v in cfg.items() if k not in ("name", "cache_ttl")}
        data = adapter.load(adapter_cfg)
        set_cache(cache_key_str, data)
        count = len(data) if hasattr(data, "__len__") else "?"
        print(f"[Resource] 加载: {name}, {count} 条")
        results[name] = data

    return results


# ───────────────────────────────────────────────────────────────
# Excel 写入层
# ───────────────────────────────────────────────────────────────

def _detect_blocks(rows: List[List], key_indices: List[int]) -> List[Tuple[int, int]]:
    """检测数据 block 边界。返回 [(start_idx, end_idx), ...]"""
    if not rows:
        return []
    if not key_indices:
        return [(0, len(rows) - 1)]

    blocks = []
    start = 0
    for i in range(1, len(rows)):
        if any(rows[i][k] != rows[i - 1][k] for k in key_indices):
            blocks.append((start, i - 1))
            start = i
    blocks.append((start, len(rows) - 1))
    return blocks


def _resolve_formula(template: str, row: int, block_start: int, block_end: int, col_letter: str) -> str:
    """替换公式模板变量。"""
    return (
        template
        .replace("{row}", str(row))
        .replace("{block_start}", str(block_start))
        .replace("{block_end}", str(block_end))
        .replace("{col}", col_letter)
    )


def write_configured_sheet(ws, sheet_config: SheetConfig, rows: List[List]):
    """
    按配置写入 Sheet。

    Args:
        ws: openpyxl Worksheet
        sheet_config: SheetConfig（列定义、合并规则、公式规则）
        rows: 扁平数据行，每个内层列表长度 ≥ max(field_index) + 1
    """
    columns = sheet_config.columns
    merge_key_indices = sheet_config.merge_key_indices

    # 1. 表头
    for col_idx, col_cfg in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_cfg.name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    if not rows:
        return

    # 2. 检测 block
    blocks = _detect_blocks(rows, merge_key_indices)

    # 3. 写入 value / formula 列
    for block_start_idx, block_end_idx in blocks:
        excel_start = block_start_idx + 2
        excel_end = block_end_idx + 2

        for data_idx in range(block_start_idx, block_end_idx + 1):
            excel_row = data_idx + 2

            for col_idx, col_cfg in enumerate(columns, 1):
                field_idx = col_cfg.field_index

                if col_cfg.col_type == "value":
                    value = rows[data_idx][field_idx] if field_idx < len(rows[data_idx]) else None
                    cell = ws.cell(row=excel_row, column=col_idx, value=value)
                    cell.border = THIN_BORDER

                    if isinstance(value, (int, float)) and value is not None:
                        cell.alignment = Alignment(horizontal="right")
                        if col_cfg.format_str:
                            cell.number_format = col_cfg.format_str
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                elif col_cfg.col_type == "formula":
                    col_letter = get_column_letter(col_idx)
                    formula = _resolve_formula(
                        col_cfg.formula_template, excel_row, excel_start, excel_end, col_letter
                    )
                    cell = ws.cell(row=excel_row, column=col_idx, value=formula)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="right")
                    if col_cfg.format_str:
                        cell.number_format = col_cfg.format_str

        # 4. block_formula 列（只在 block 第一行写）
        for col_idx, col_cfg in enumerate(columns, 1):
            if col_cfg.col_type != "block_formula":
                continue

            field_idx = col_cfg.field_index
            col_letter = get_column_letter(col_idx)
            formula = _resolve_formula(
                col_cfg.formula_template, excel_start, excel_start, excel_end, col_letter
            )
            cell = ws.cell(row=excel_start, column=col_idx, value=formula)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_cfg.format_str:
                cell.number_format = col_cfg.format_str

            # 负值标红（基于预计算值）
            if col_cfg.negative_red:
                precomputed = rows[block_start_idx][field_idx] if field_idx < len(rows[block_start_idx]) else None
                if isinstance(precomputed, (int, float)) and precomputed < 0:
                    cell.font = RED_FONT

    # 5. 合并单元格
    if merge_key_indices:
        merge_col_indices = [i + 1 for i, c in enumerate(columns) if c.merge]
        for block_start_idx, block_end_idx in blocks:
            if block_start_idx == block_end_idx:
                continue
            excel_start = block_start_idx + 2
            excel_end = block_end_idx + 2
            for col_idx in merge_col_indices:
                ws.merge_cells(start_row=excel_start, start_column=col_idx,
                               end_row=excel_end, end_column=col_idx)
                ws.cell(row=excel_start, column=col_idx).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    # 6. 自动列宽
    for col_idx, col_cfg in enumerate(columns, 1):
        max_len = max(
            (len(str(rows[r][col_cfg.field_index] or "")) for r in range(len(rows)) if col_cfg.field_index < len(rows[r])),
            default=0
        )
        header_len = len(col_cfg.name)
        max_len = max(max_len, header_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # 7. 冻结首行
    ws.freeze_panes = sheet_config.freeze_panes


# ───────────────────────────────────────────────────────────────
# 配置加载
# ───────────────────────────────────────────────────────────────

def load_sheet_config(yaml_path: str, sheet_name: str) -> SheetConfig:
    """从 YAML 加载 Sheet 配置。"""
    import yaml
    with open(yaml_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    sheet_data = config["sheets"][sheet_name]
    columns = []
    for c in sheet_data.get("columns", []):
        columns.append(ColumnConfig(
            name=c["name"],
            field_index=c.get("field_index", 0),
            col_type=c.get("type", "value"),
            format_str=c.get("format", ""),
            width=c.get("width", 15),
            merge=c.get("merge", False),
            formula_template=c.get("formula", ""),
            negative_red=c.get("negative_red", False),
        ))

    return SheetConfig(
        name=sheet_name,
        columns=columns,
        merge_key_indices=sheet_data.get("merge_key_indices", []),
        freeze_panes=sheet_data.get("freeze_panes", "A2"),
    )


# ───────────────────────────────────────────────────────────────
# 引擎类（便利封装）
# ───────────────────────────────────────────────────────────────

class ReportEngine:
    """报表引擎便利封装。"""

    def __init__(self, project_id: str = "diyl-407103"):
        # 延迟 import，避免无 BQ 环境时模块加载失败
        from bq_reports.utils.bq_client import get_bq_client, setup_proxy
        setup_proxy()
        self.client = get_bq_client(project_id)

    def query(self, sql_template, merchants, start_ts, end_ts, workers=10, row_proxy_factory=None, label="门店"):
        """并发查询。"""
        return query_all_shops(
            self.client, sql_template, merchants, start_ts, end_ts,
            workers=workers, row_proxy_factory=row_proxy_factory, label=label
        )

    def load_resources(self, configs: List[Dict]) -> Dict[str, Any]:
        """加载外部资源。"""
        return load_resources(configs)

    def write_sheet(self, ws, sheet_config: SheetConfig, rows: List[List]):
        """按配置写入 Sheet。"""
        return write_configured_sheet(ws, sheet_config, rows)

    def load_sheet_config(self, yaml_path: str, sheet_name: str) -> SheetConfig:
        """从 YAML 加载配置。"""
        return load_sheet_config(yaml_path, sheet_name)
