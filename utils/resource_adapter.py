#!/usr/bin/env python3
"""
灵活资源适配器 —— 解耦报表代码与外部文件格式。

通过配置（而非硬编码）描述如何从 Excel/CSV/JSON 映射到标准字段。
支持：单 sheet 读取、多 sheet 合并、列回退、适配器回退链、条件路由。

Usage:
    # 1. 基础用法（单 sheet）
    config = {
        "adapter": "excel",
        "path": "data.xlsx",
        "sheet": "Sheet1",
        "mapping": {"store_number": "门店编号", "store_name": "门店名称"}
    }
    records = get_adapter("excel").load(config)

    # 2. 列回退（主列为空时从备用列取）
    mapping = {
        "store_number": {"primary": "门店编号", "fallback": "编号"}
    }

    # 3. 多 sheet 合并（同一文件不同 sheet）
    config = {
        "adapter": "multi_sheet",
        "path": "data.xlsx",
        "sheets": [
            {"name": "门店列表", "mapping": {"num": "编号", "name": "名称"}},
            {"name": "BOM配方", "mapping": {"product": "商品", "code": "编码"}}
        ]
    }

    # 4. 适配器回退链（A 失败回退到 B）
    config = {
        "adapter": "fallback",
        "chain": [
            {"adapter": "excel", "path": "new.xlsx", "sheet": "BOM", "mapping": {...}},
            {"adapter": "excel", "path": "old.xlsx", "sheet": "配方", "mapping": {...}}
        ]
    }

    # 5. 条件路由（按文件存在性选择策略）
    config = {
        "adapter": "conditional",
        "rules": [
            {"if": {"file_exists": "new.xlsx"}, "then": {"adapter": "excel", "path": "new.xlsx", ...}},
            {"else": {"adapter": "excel", "path": "old.xlsx", ...}}
        ]
    }
"""

from abc import ABC, abstractmethod
from typing import Any, Dict, List, Optional, Union


class ResourceAdapter(ABC):
    """外置资源读取接口。"""

    @abstractmethod
    def load(self, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        返回标准化记录列表，每条记录是 dict。
        """
        pass


# ───────────────────────────────────────────────────────────────
# 基础适配器
# ───────────────────────────────────────────────────────────────

class ExcelAdapter(ResourceAdapter):
    """Excel 适配器：通过配置指定 sheet、列名→字段映射。"""

    def load(self, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        from openpyxl import load_workbook

        filepath = config["path"]
        sheet_ref = config.get("sheet", 0)
        mapping = config["mapping"]
        header_row = config.get("header_row", 1)
        skip_empty = config.get("skip_empty", True)

        wb = load_workbook(filepath, data_only=True)
        try:
            ws = wb[sheet_ref] if isinstance(sheet_ref, str) else wb.worksheets[sheet_ref]
        except (KeyError, IndexError) as e:
            available = [s.title for s in wb.worksheets]
            wb.close()
            raise ValueError(f"Sheet '{sheet_ref}' 不存在。可用: {available}") from e

        # 读取表头（1-based）
        headers = [cell.value for cell in ws[header_row]]

        # 构建 标准字段 → 列索引（0-based）映射
        # 支持两种格式：
        #   "field": "列名"               # 直接映射
        #   "field": {"primary": "A", "fallback": "B"}  # 主列+备用列
        col_index = {}
        fallback_map = {}  # std_field -> fallback_col_idx

        for std_field, excel_ref in mapping.items():
            if isinstance(excel_ref, dict):
                # 列回退策略
                primary_col = self._resolve_column(headers, excel_ref["primary"])
                col_index[std_field] = primary_col
                if "fallback" in excel_ref:
                    fallback_col = self._resolve_column(headers, excel_ref["fallback"])
                    fallback_map[std_field] = fallback_col
            else:
                col_index[std_field] = self._resolve_column(headers, excel_ref)

        # 读取数据行
        results = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if skip_empty and all(v is None or str(v).strip() == "" for v in row):
                continue

            record = {}
            for std_field, idx in col_index.items():
                val = row[idx] if idx is not None and idx < len(row) else None
                # 列回退：主列为空时从备用列取
                if val is None or str(val).strip() == "":
                    if std_field in fallback_map:
                        fb_idx = fallback_map[std_field]
                        if fb_idx is not None and fb_idx < len(row):
                            val = row[fb_idx]
                record[std_field] = val
            results.append(record)

        wb.close()
        return results

    @staticmethod
    def _resolve_column(headers: List[str], ref: Union[str, int]) -> Optional[int]:
        """将列引用（名称或索引）解析为 0-based 列索引。"""
        if isinstance(ref, int):
            return ref
        ref_str = str(ref).strip()
        for idx, h in enumerate(headers):
            if h and str(h).strip() == ref_str:
                return idx
        return None


class CSVAdapter(ResourceAdapter):
    """CSV 适配器。"""

    def load(self, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        import csv

        filepath = config["path"]
        mapping = config["mapping"]
        encoding = config.get("encoding", "utf-8-sig")
        skip_empty = config.get("skip_empty", True)

        results = []
        with open(filepath, newline="", encoding=encoding) as f:
            reader = csv.DictReader(f)
            for row in reader:
                if skip_empty and all(v is None or str(v).strip() == "" for v in row.values()):
                    continue
                record = {}
                for std_field, csv_col in mapping.items():
                    if isinstance(csv_col, dict):
                        # 列回退
                        val = row.get(csv_col["primary"])
                        if val is None or str(val).strip() == "":
                            val = row.get(csv_col.get("fallback"))
                        record[std_field] = val
                    else:
                        record[std_field] = row.get(csv_col)
                results.append(record)
        return results


class JSONAdapter(ResourceAdapter):
    """JSON 适配器：从 JSON 数组中按路径提取字段。"""

    def load(self, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        import json

        filepath = config["path"]
        mapping = config["mapping"]
        data_key = config.get("data_key", None)

        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        if data_key:
            records = data.get(data_key, [])
        else:
            records = data if isinstance(data, list) else []

        results = []
        for row in records:
            record = {}
            for std_field, json_key in mapping.items():
                keys = json_key.split(".")
                val = row
                for k in keys:
                    val = val.get(k) if isinstance(val, dict) else None
                record[std_field] = val
            results.append(record)
        return results


# ───────────────────────────────────────────────────────────────
# 组合/策略适配器
# ───────────────────────────────────────────────────────────────

class MultiSheetAdapter(ResourceAdapter):
    """
    多 sheet 适配器：从同一个 Excel 的多个 sheet 读取数据。

    支持两种合并模式：
    - merge_mode="concat"（默认）：所有 sheet 的记录合并成一个列表
    - merge_mode="group"：返回 dict，key 为 sheet 名
    - merge_mode="join"：按 merge_key 字段 join（实验性）
    """

    def load(self, config: Dict[str, Any]) -> Union[List[Dict], Dict[str, List[Dict]]]:
        from openpyxl import load_workbook

        filepath = config["path"]
        sheets_config = config["sheets"]  # list of {"name": ..., "mapping": ..., ...}
        merge_mode = config.get("merge_mode", "concat")  # concat | group | join
        merge_key = config.get("merge_key", None)

        wb = load_workbook(filepath, data_only=True)
        all_data = {}

        for sheet_cfg in sheets_config:
            sheet_name = sheet_cfg["name"]
            try:
                ws = wb[sheet_name]
            except KeyError:
                available = [s.title for s in wb.worksheets]
                print(f"[警告] Sheet '{sheet_name}' 不存在，跳过。可用: {available}")
                continue

            # 复用 ExcelAdapter 的单 sheet 读取逻辑
            single_config = {
                "path": filepath,
                "sheet": sheet_name,
                "mapping": sheet_cfg["mapping"],
                "header_row": sheet_cfg.get("header_row", 1),
                "skip_empty": sheet_cfg.get("skip_empty", True),
            }
            adapter = ExcelAdapter()
            records = adapter.load(single_config)

            # 可选：添加 sheet 来源标记
            if sheet_cfg.get("tag_source", False):
                for r in records:
                    r["_source_sheet"] = sheet_name

            # 可选：字段前缀
            prefix = sheet_cfg.get("prefix", "")
            if prefix:
                records = [
                    {f"{prefix}{k}": v for k, v in r.items()}
                    for r in records
                ]

            all_data[sheet_name] = records

        wb.close()

        if merge_mode == "group":
            return all_data

        if merge_mode == "concat":
            result = []
            for records in all_data.values():
                result.extend(records)
            return result

        if merge_mode == "join" and merge_key:
            # 按 merge_key 合并所有 sheet 的记录
            # 假设第一个 sheet 是主表，其余是补充
            master_sheet = list(all_data.keys())[0]
            master = {r.get(merge_key): r for r in all_data[master_sheet]}
            for sheet_name, records in list(all_data.items())[1:]:
                for r in records:
                    key = r.get(merge_key)
                    if key in master:
                        master[key].update(r)
            return list(master.values())

        return all_data


class FallbackAdapter(ResourceAdapter):
    """
    回退适配器：按顺序尝试多个适配器，第一个成功的结果返回。

    适用于：文件格式升级，新旧版本并存，优先读新版，失败回退旧版。
    """

    def load(self, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        chain = config["chain"]  # list of adapter configs
        last_error = None

        for i, item_config in enumerate(chain):
            adapter_name = item_config["adapter"]
            try:
                adapter = get_adapter(adapter_name)
                records = adapter.load(item_config)
                if i > 0:
                    print(f"[Fallback] 回退到第 {i+1} 个适配器成功: {adapter_name}")
                return records
            except Exception as e:
                last_error = e
                print(f"[Fallback] 第 {i+1} 个适配器失败 ({adapter_name}): {e}")
                continue

        raise RuntimeError(f"所有回退适配器均失败。最后一个错误: {last_error}")


class ConditionalAdapter(ResourceAdapter):
    """
    条件适配器：根据条件动态选择适配器配置。

    支持的条件：
    - file_exists: "path"        — 文件是否存在
    - sheet_exists: {"path": "x.xlsx", "sheet": "Sheet1"} — sheet 是否存在
    - env: "VAR_NAME"            — 环境变量是否存在且非空
    """

    def load(self, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        rules = config["rules"]  # list of {"if": condition, "then": config} or {"else": config}

        for rule in rules:
            if "if" in rule:
                condition = rule["if"]
                if self._evaluate_condition(condition):
                    then_config = rule["then"]
                    adapter = get_adapter(then_config["adapter"])
                    return adapter.load(then_config)
            elif "else" in rule:
                else_config = rule["else"]
                adapter = get_adapter(else_config["adapter"])
                return adapter.load(else_config)

        raise RuntimeError("没有匹配的条件规则，且未提供 else 分支")

    @staticmethod
    def _evaluate_condition(condition: Dict[str, Any]) -> bool:
        if "file_exists" in condition:
            import os
            return os.path.exists(condition["file_exists"])

        if "sheet_exists" in condition:
            from openpyxl import load_workbook
            info = condition["sheet_exists"]
            try:
                wb = load_workbook(info["path"], read_only=True)
                exists = info["sheet"] in wb.sheetnames
                wb.close()
                return exists
            except Exception:
                return False

        if "env" in condition:
            import os
            return bool(os.environ.get(condition["env"], "").strip())

        return False


# ───────────────────────────────────────────────────────────────
# 成本价专用适配器 (业务语义编码: ✅ 直读 vs 算法路径 / 规格 parser)
# ───────────────────────────────────────────────────────────────

import re


def _parse_spec_to_min_unit_count(spec):
    """规格字符串 → 每箱最小单位数 (提取所有数字相乘).
    示例:
        '100只/包；100包/袋/箱'  → 10000
        '52个/条*25条/箱'       → 1300
        '500张/包*4包/箱'       → 2000
        '3043杯/卷*6卷/箱'      → 18258
    """
    if not spec:
        return None
    nums = re.findall(r'\d+(?:\.\d+)?', str(spec))
    if not nums:
        return None
    result = 1.0
    for n in nums:
        result *= float(n)
    return int(result) if result == int(result) else result


def _normalize_overseas_suffix(name):
    """去 (海外版) / （海外版）/ ( 海外版 ) 尾缀, trim 空白"""
    if not name:
        return ''
    s = str(name).strip()
    s = re.sub(r'[\(（]\s*海外版\s*[\)）]\s*$', '', s).strip()
    return s


class CostPriceTaixiAdapter(ResourceAdapter):
    """读 '所有成本价_含ItemCode.xlsx' 的【泰国采】sheet, 输出 {编码: {price, unit, source_tag}}.

    取价规则:
      - col 14「系统匹配」== '✅'  → 直读 col 10「单价(最小单位)」
      - 否则                    → 算 col 8「销售单价」÷ col 6「转换系数」
    单位: col 11「最小单位」 (克/个 等).
    """

    def load(self, config):
        from openpyxl import load_workbook
        filepath = config["path"]
        sheet_name = config.get("sheet", "泰国采")

        wb = load_workbook(filepath, data_only=True)
        try:
            ws = wb[sheet_name]
        except KeyError as e:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' 不存在: {[s.title for s in wb.worksheets]}") from e

        result = {}
        for r in range(2, ws.max_row + 1):
            code = ws.cell(r, 1).value
            sales_price = ws.cell(r, 8).value
            conv = ws.cell(r, 6).value
            min_unit_price = ws.cell(r, 10).value
            min_unit = ws.cell(r, 11).value
            sys_match = ws.cell(r, 14).value
            if not code:
                continue
            code_str = str(code).strip()
            if sys_match == '✅' and isinstance(min_unit_price, (int, float)):
                price = float(min_unit_price)
                tag = '泰国采[✅直读]'
            elif isinstance(sales_price, (int, float)) and isinstance(conv, (int, float)) and conv > 0:
                price = sales_price / conv
                tag = '泰国采[销售单价/转换系数]'
            else:
                continue
            if code_str not in result:
                result[code_str] = {
                    "price": price,
                    "unit": str(min_unit or '').strip(),
                    "source_tag": tag,
                }
        wb.close()
        return result


class CostPriceImportAdapter(ResourceAdapter):
    """读 '2026进口货物(2).xlsx' 的【2026.1】sheet (报关单), 输出 {归一化名: {price, unit, ...}}.

    取价规则: col 17「销售无税单价」÷ parse_spec(col 6「规格」) = per-最小单位.
    匹配键: col 5「NC 名称」去「(海外版)」/「（海外版）」尾缀.
    """

    def load(self, config):
        from openpyxl import load_workbook
        filepath = config["path"]
        sheet_name = config.get("sheet", "2026.1")

        wb = load_workbook(filepath, data_only=True)
        try:
            ws = wb[sheet_name]
        except KeyError as e:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' 不存在: {[s.title for s in wb.worksheets]}") from e

        result = {}
        for r in range(2, ws.max_row + 1):
            nc_name = ws.cell(r, 5).value
            spec = ws.cell(r, 6).value
            sales_no_tax = ws.cell(r, 17).value
            if not nc_name or not isinstance(sales_no_tax, (int, float)):
                continue
            normalized = _normalize_overseas_suffix(nc_name)
            if not normalized or normalized in result:
                continue
            per_box_total = _parse_spec_to_min_unit_count(spec)
            if not per_box_total or per_box_total <= 0:
                continue
            price = float(sales_no_tax) / per_box_total
            # 单位: 规格里第一个 数字+单位 token
            m = re.match(r'(\d+(?:\.\d+)?)\s*([^\d\s/*；,]+)', str(spec))
            unit = m.group(2) if m else ''
            result[normalized] = {
                "price": price,
                "unit": unit,
                "source_tag": '2026进口[规格×销售无税]',
                "raw_nc_name": nc_name,
                "spec": spec,
            }
        wb.close()
        return result


# ───────────────────────────────────────────────────────────────
# 适配器注册表
# ───────────────────────────────────────────────────────────────

_ADAPTERS = {
    "excel": ExcelAdapter,
    "csv": CSVAdapter,
    "json": JSONAdapter,
    "multi_sheet": MultiSheetAdapter,
    "fallback": FallbackAdapter,
    "conditional": ConditionalAdapter,
    "cost_price_taixi": CostPriceTaixiAdapter,
    "cost_price_import": CostPriceImportAdapter,
}


def get_adapter(name: str) -> ResourceAdapter:
    """获取指定名称的适配器实例。"""
    if name not in _ADAPTERS:
        raise ValueError(f"未知适配器: {name}。可用: {list(_ADAPTERS.keys())}")
    return _ADAPTERS[name]()


def register_adapter(name: str, cls: type):
    """注册自定义适配器。"""
    if not issubclass(cls, ResourceAdapter):
        raise TypeError("适配器必须继承 ResourceAdapter")
    _ADAPTERS[name] = cls
