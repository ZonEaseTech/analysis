#!/usr/bin/env python3
"""
Wallace 商品成本毛利分析生成器
================================
用【0514 总部 BOM + 物料价格】(模板店级 BOM 补缺)套到【我生成的销量/实收表】,
重算成本毛利,输出与模板相同的 16 列结构,拆成 4 个 sheet:
堂食-单品 / 堂食-套餐 / 外卖-单品 / 外卖-套餐。

读用 openpyxl,写用 xlsxwriter(合并单元格 + 公式,大表性能好)。

BOM / 成本来源(优先级)
-----------------------
- 单品 BOM : 0514「单品」-> 缺则模板店级 BOM 补
- 套餐 BOM : 0514「套餐」(已展开)-> 缺则「明细组分 × 单品 BOM」拼 -> 再缺模板补
- 物料单价 : 0514「价格」(按编码)-> 缺则模板 BOM 行自带单价补
- 单份总成本 = Σ(消耗数量 × 物料单价)

销量 / 实收
-----------
- 净销量 : 销量表(堂食 / 外卖 分开;只取独立单品 + 套餐,排除套餐内单品)
- 实收金额 = round(原价 × 该店该渠道实收率)  [原价−优惠,取整,无小数]
- 净利润 / 单份总成本 / 总毛利 / 净总毛利 / 毛利率 : Excel 公式

用法
----
    python3 wallace_bom_margin.py --bom "0514最终 BOM和价格 (1).xlsx" \
        --tmpl "商品成本毛利分析....xlsx" \
        --sales "Wallace全店_套餐单品销量_堂食+外卖_2026-05.xlsx" \
        --recon "Wallace门店实收明细_2026-05.json" \
        --out  "Wallace商品成本毛利分析_2026-05.xlsx"
"""
import argparse, json, os, re, sys
import openpyxl.styles.colors as _colors
_colors.aRGB_REGEX = re.compile(r".*")   # 放宽非法颜色校验(0514 文件 styles 含非法 rgb)
from openpyxl import load_workbook
import xlsxwriter

OUT_COLS = ["门店", "毛利率", "净毛利率", "商品", "单价", "BOM物品名称", "BOM物品编码",
            "消耗数量", "物品单价", "单位", "净销量", "实收金额", "净利润",
            "单份总成本", "总毛利", "净总毛利"]
# 0-based 列号:A门店0 B毛利率1 C净毛利率2 D商品3 E单价4 F名5 G码6 H消耗7 I物品单价8
#              J单位9 K净销量10 L实收11 M净利润12 N单份成本13 O总毛利14 P净总毛利15

# 注: clean_bom 时代前的 drop_* 渠道过滤死代码已删(2026-06-26)，
# 渠道/物料删除规则已迁入 bom_rules.py + load_clean_bom()。

# ── 多语言 (--lang en/th); zh 时全部 identity, 导出代码零改动 ──────────
import json as _json
_LANG = "zh"
_NM = {}
_HDR_T = {
 "en": {"门店":"Store","毛利率":"Gross Margin %","净毛利率":"Net Margin %","商品":"Product","单价":"Unit Price","BOM物品名称":"BOM Material","BOM物品编码":"Material Code","消耗数量":"Qty Used","物品单价":"Material Unit Price","单位":"Unit","净销量":"Net Qty Sold","实收金额":"Net Revenue","净利润":"Net Profit","单份总成本":"Unit Cost","总毛利":"Gross Profit","净总毛利":"Net Gross Profit","来源sheet":"Source Sheet","店号":"Store No.","说明":"Note","组分单品":"Component Item","涉及店数":"Stores Affected","被套餐引用次数":"Combo Refs","影响套餐销量合计":"Affected Combo Qty"},
 "th": {"门店":"ร้าน","毛利率":"อัตรากำไรขั้นต้น %","净毛利率":"อัตรากำไรสุทธิ %","商品":"สินค้า","单价":"ราคาต่อหน่วย","BOM物品名称":"วัตถุดิบ (BOM)","BOM物品编码":"รหัสวัตถุดิบ","消耗数量":"ปริมาณที่ใช้","物品单价":"ราคาวัตถุดิบ/หน่วย","单位":"หน่วย","净销量":"ยอดขายสุทธิ","实收金额":"รายรับสุทธิ","净利润":"กำไรสุทธิ","单份总成本":"ต้นทุน/หน่วย","总毛利":"กำไรขั้นต้น","净总毛利":"กำไรขั้นต้นสุทธิ","来源sheet":"ชีตต้นทาง","店号":"รหัสร้าน","说明":"หมายเหตุ","组分单品":"ส่วนประกอบ","涉及店数":"จำนวนร้าน","被套餐引用次数":"ครั้งที่อ้างอิง","影响套餐销量合计":"ยอดขายชุดที่กระทบ"},
}
_SHEET_T = {
 "en": {"堂食-单品":"Dine-in Single","堂食-套餐":"Dine-in Combo","外卖-单品":"Delivery Single","外卖-套餐":"Delivery Combo","未匹配清单":"Unmatched","缺失组分单品":"Missing Components"},
 "th": {"堂食-单品":"ทานในร้าน เดี่ยว","堂食-套餐":"ทานในร้าน ชุด","外卖-单品":"เดลิเวอรี เดี่ยว","外卖-套餐":"เดลิเวอรี ชุด","未匹配清单":"ไม่จับคู่","缺失组分单品":"ส่วนประกอบที่ขาด"},
}
_UNIT_T = {"en": {"份":"portion","个":"pc","包":"pack","张":"sheet","只":"pc","块":"pc"},
           "th": {"份":"ที่","个":"ชิ้น","包":"แพ็ค","张":"แผ่น","只":"ชิ้น","块":"ชิ้น"}}
_MISC_T = {
 "en": {"(估算·未映射)":"(Estimated·Unmapped)","玉米杯无 BOM":"Corn cup no BOM","别名目标无 BOM":"Alias target no BOM","无单品 BOM":"No single-item BOM","无 BOM 且组分无法拼":"No BOM & components unresolvable"},
 "th": {"(估算·未映射)":"(ประมาณ·ไม่จับคู่)","玉米杯无 BOM":"ถ้วยข้าวโพดไม่มี BOM","别名目标无 BOM":"เป้าหมาย alias ไม่มี BOM","无单品 BOM":"ไม่มี BOM เดี่ยว","无 BOM 且组分无法拼":"ไม่มี BOM และประกอบไม่ได้"},
}

def t_hdr(x):
    return _HDR_T.get(_LANG, {}).get(x, x) if _LANG != "zh" else x

def t_sheet(x):
    return _SHEET_T.get(_LANG, {}).get(x, x) if _LANG != "zh" else x

def t_prod(x):
    if _LANG == "zh" or not isinstance(x, str):
        return x
    return _NM.get("prod_" + _LANG, {}).get(x.strip(), x)

def t_mat(name, code=None):
    if _LANG == "zh" or not isinstance(name, str):
        return name
    if code:
        v = _NM.get("matc_" + _LANG, {}).get(str(code).strip())
        if v:
            return v
    return _NM.get("mat_" + _LANG, {}).get(name.strip(), name)

def t_unit(x):
    if _LANG == "zh" or not isinstance(x, str):
        return x
    return _UNIT_T.get(_LANG, {}).get(x.strip(), x)

def t_misc(x):
    if _LANG == "zh" or not isinstance(x, str):
        return x
    m = _MISC_T.get(_LANG, {})
    if x in m:
        return m[x]
    if x.startswith("缺组分: "):
        return ("Missing: " if _LANG == "en" else "ขาด: ") + x[len("缺组分: "):]
    return x
# ────────────────────────────────────────────────────────────────────


def norm(name):
    if name is None:
        return ""
    s = str(name).split('/')[0]
    s = re.sub(r'\s+', '', s).replace('（', '(').replace('）', ')')
    return s.strip().lower()


def fnum(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def month_from_sales_path(path):
    m = re.search(r"(20\d{2}-\d{2})", os.path.basename(path))
    return m.group(1) if m else None


def resolve_recon_path(sales_path, recon_path=None):
    if recon_path:
        return recon_path
    # recon 与 sales 同目录查找(原 BASE 未定义 bug，2026-06-26 修)。
    # 兼容旧命名 Wallace门店实收明细_<月>.json 和新命名 recon.json；
    # recon.json 与月份无关，故不受 sales 文件名是否含月份影响。
    base = os.path.dirname(os.path.abspath(sales_path))
    names = ["recon.json"]
    month = month_from_sales_path(sales_path)
    if month:
        names.insert(0, f"Wallace门店实收明细_{month}.json")
    for name in names:
        candidate = os.path.join(base, name)
        if os.path.exists(candidate):
            return candidate
    return None


# ---------------- 读 0514 BOM + 价格 ----------------
def load_0514(path):
    wb = load_workbook(path, data_only=True)
    price = {}
    for r in wb["价格"].iter_rows(min_row=2, values_only=True):
        if r and r[1]:
            price[str(r[1]).strip()] = fnum(r[2])

    def build_d(sheet):
        # 商品名只在块首行,后续物料行该列为空 -> 用 cur 延续商品名(否则每个商品只收到首行 1 个物料)。
        # 同一商品名按 (物料编码|物料名) 去重:0514 sheet 是店×商品展开(同名重复 ~60 次),
        # 不去重会把 BOM 累加几十倍导致成本爆炸。
        d, seen = {}, {}
        cur = None
        for r in sheet.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            if r[0] and str(r[0]).strip():
                cur = norm(r[0])
            if cur is None:
                continue
            mname = str(r[1]).strip() if r[1] else ""
            code = str(r[2]).strip() if r[2] else ""
            if not (mname or code):       # 空物料行跳过
                continue
            dk = code or mname
            s = seen.setdefault(cur, set())
            if dk in s:
                continue
            s.add(dk)
            d.setdefault(cur, []).append((mname, code, fnum(r[3]), str(r[4]).strip() if r[4] else ""))
        return d
    single = build_d(wb["单品"])
    combo = build_d(wb["套餐"])
    wb.close()
    return price, single, combo


# ---------------- 读模板(补缺 BOM + 补价) ----------------
def load_tmpl_supp(path, price):
    wb = load_workbook(path, read_only=True, data_only=True)
    fb, seen = {}, {}
    it = wb["单品"].iter_rows(values_only=True)
    next(it)
    cur = None
    for row in it:
        prod = row[3]
        bname, bcode, qty, uprice, unit = row[5], row[6], row[7], row[8], row[9]
        if prod and str(prod).strip():
            cur = norm(prod)
            fb.setdefault(cur, [])
            seen.setdefault(cur, set())
        if bname and str(bname).strip() and cur is not None:
            code = str(bcode).strip() if bcode else ""
            mname = str(bname).strip()
            dk = code or mname
            if dk not in seen[cur]:        # 模板是店×商品,同名跨店去重避免 BOM concat 翻倍
                seen[cur].add(dk)
                fb[cur].append((mname, code, fnum(qty), str(unit).strip() if unit else ""))
            if code and code not in price:
                price[code] = fnum(uprice)
    wb.close()
    return {k: v for k, v in fb.items() if v}


# ---------------- 读人工补充 BOM(缺失BOM.xlsx)----------------
def load_manual(path):
    """两段混合:套餐组分明细(col1) + 单品 BOM(col2 有编码;列序 商品/物料名/编码/单位/消耗)。"""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    msingle, mdetail = {}, {}
    cur, curtype = None, None
    for r in ws.iter_rows(min_row=2, values_only=True):
        c0 = str(r[0]).strip() if r[0] else ""
        c1 = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        c2 = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        c3 = str(r[3]).strip() if len(r) > 3 and r[3] else ""
        c4 = str(r[4]).strip() if len(r) > 4 and r[4] else ""
        if c0:
            if c2:                       # 单品 BOM(有物料编码)
                curtype, cur = "B", norm(c0)
                msingle.setdefault(cur, []).append((c1, c2, fnum(c4), c3))
            else:                        # 套餐组分明细
                curtype, cur = "A", norm(c0)
                if c1:
                    mdetail[cur] = c1
        elif curtype == "B" and (c1 or c2):   # 单品 BOM 续行
            msingle.setdefault(cur, []).append((c1, c2, fnum(c4), c3))
    wb.close()
    return {k: v for k, v in msingle.items() if v}, mdetail


# ---------------- 读销量表 ----------------
def _sheet(wb, prefix):
    for s in wb.sheetnames:
        if s.startswith(prefix):
            return wb[s]
    raise SystemExit(f"销量表缺少 sheet: {prefix}*")


def chan_big(chan):
    c = str(chan or "")
    return "堂食" if c.startswith("堂食") else ("外卖" if c.startswith("外卖") else "其它")


def load_sales(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    din_rate, to_orig, to_settle = {}, {}, {}
    for r in _sheet(wb, "堂食实收").iter_rows(min_row=2, values_only=True):
        if not r[0] or str(r[0]).strip() in ("合计", ""):
            continue
        s3 = str(r[0]).strip()[:3]
        orig = fnum(r[3])
        din_rate[s3] = (fnum(r[7]) / orig) if orig else 1.0
    for r in _sheet(wb, "外卖促销").iter_rows(min_row=2, values_only=True):
        if not r[0] or str(r[0]).strip() in ("合计", ""):
            continue
        s3 = str(r[0]).strip()[:3]
        to_orig[s3] = to_orig.get(s3, 0.0) + fnum(r[4])
        to_settle[s3] = to_settle.get(s3, 0.0) + fnum(r[10])
    to_rate = {s3: (to_settle[s3] / to_orig[s3]) if to_orig.get(s3) else 1.0 for s3 in to_orig}

    sales = {}

    def add(s3, cb, kind, disp, nm, qty, unit, amt, detail=None):
        k = (s3, cb, kind, norm(nm))
        e = sales.get(k)
        if e is None:
            sales[k] = {"qty": qty, "amt": amt, "unit": unit, "disp": disp, "detail": detail}
        else:
            e["qty"] += qty
            e["amt"] += amt
            if detail and not e["detail"]:
                e["detail"] = detail

    for r in _sheet(wb, "套餐").iter_rows(min_row=2, values_only=True):
        if not r[0] or str(r[0]).strip() in ("合计", "") or not r[3]:
            continue
        cb = chan_big(r[2])
        if cb != "其它":
            add(str(r[0]).strip()[:3], cb, "combo", str(r[3]), r[3],
                fnum(r[4]), fnum(r[5]), fnum(r[6]), detail=r[7])
    for r in _sheet(wb, "单品").iter_rows(min_row=2, values_only=True):
        if not r[0] or str(r[0]).strip() in ("合计", "") or not r[4]:
            continue
        if str(r[3] or "") != "独立":
            continue
        cb = chan_big(r[2])
        if cb != "其它":
            add(str(r[0]).strip()[:3], cb, "single", str(r[4]), r[4],
                fnum(r[5]), fnum(r[6]), fnum(r[7]))
    wb.close()
    return sales, din_rate, to_rate


# ---------------- BOM 解析 ----------------
def parse_detail(detail):
    # 组分格式 "名*数量",组分间用顿号或空格分隔(名本身可含空格,如"经典炸鸡 5 块")
    out = []
    if not detail or str(detail).strip() in ("", "(无明细)"):
        return out
    for nm, q in re.findall(r'(.+?)\*\s*(\d+(?:\.\d+)?)(?:[、\s]+|$)', str(detail)):
        nm = nm.strip().lstrip('、').strip()
        if nm:
            out.append((nm, fnum(q) or 1.0))
    return out


# 别名映射(用户确认):缺失商品名 -> 已有 BOM 的商品名。键值都会被 norm 规范化。
ALIAS_RAW = {
    "华莱士脆皮鸡": "脆皮全鸡",       # 用户:=炸全鸡(0514 无"炸全鸡",取"脆皮全鸡")
    "华莱士烤鸡": "烤全鸡",
    "汉堡": "辣味鸡肉汉堡",
    "抹茶冰淇淋": "抹茶冰淇凌",        # 0514 名为"抹茶冰淇凌"
    "抹茶冰激凌": "抹茶冰淇凌",
    "冰淇淋圣代": "巧克力圣代",        # 取巧克力圣代(另有草莓圣代)
    "冰激凌圣代": "巧克力圣代",
    "周四促销": "香辣脆皮鸡 1 份",     # 用户:=特价大脆鸡1块
    "周四优惠": "香辣脆皮鸡 1 份",
    # 以下为剩余缺失商品按"已有同类"匹配(模糊对应,可调整)
    "新价格的辣味鸡翅": "炸鸡翅",         # 另有 香脆鸡翅14.55 / 烤鸡翅15.09
    "16、20、27号抹茶冰淇淋店": "抹茶冰淇凌",
    "脆皮鸡腿排饭(旧价格)": "脆皮鸡腿饭",
    "巨无霸烤鸡饭": "脆皮鸡腿饭",
    "新店开业促销，满意十足，10%特别折扣": "实惠满足 3",
    # 命名差异、0514 有 BOM 的(用户确认)
    "快乐套餐8": "快乐套餐 8（旧）",
    "超值10件套餐": "超值套餐 10",
    "[促销] 辣味番茄双人套餐 1 特价！": "【促销】辣味番茄人套餐特惠！",
    "饱足感2": "饱足满心 2 经典 旧",
}
ALIAS = {norm(k): norm(v) for k, v in ALIAS_RAW.items()}

# 内置组分明细(用户提供,商品=多份某单品):合并进 manual_detail,走组分拼
EXTRA_DETAIL_RAW = {
    "[热卖] 买一份薯条，再加20泰铢得另一份": "薯条*2",
}
EXTRA_DETAIL = {norm(k): v for k, v in EXTRA_DETAIL_RAW.items()}

# 人工内置 BOM(用户提供,0514/模板都没有的商品)。格式: (物料名, 编码, 消耗, 单位)
EXTRA_BOM_RAW = {
    "特别套装迎接新学期 1": [
        ("单杯袋", "PA99020", 1, "pc"), ("16A 冷杯", "PA99001", 1, "pc"),
        ("百事可乐原味", "BE01004", 25, "g"), ("PLA尖头吸管", "PA99024", 1, "pc"),
        ("哑光塑料封口膜", "PA99027", 1, "pc"), ("8寸面饼", "FR03001", 1, "pc"),
        ("五合一汉堡纸", "PA99017", 1, "pc"), ("打包袋(小号)", "PA99018", 2, "pc"),
        ("食安封签", "PA99034", 1, "pc"), ("沙拉酱", "SA02001", 15, "g"),
        ("番茄酱(大包)", "SA01009", 25, "g"), ("生菜", "VE00004", 20, "g"),
        ("中薯盒", "PA99006", 1, "pc"), ("薯条细款7mm", "FR02001", 150, "g"),
        ("起酥油", "DR02001", 10, "g"), ("甜辣酱(大包)", "SA01008", 10, "g"),
        ("盐", "DR06002", 1, "g"), ("番茄酱(小包)", "SA01013", 1, "pc"),
        ("辣椒酱(小包)", "SA01010", 1, "pc"), ("外卖筒形薯盒", "PA99005", 1, "pc"),
        ("托盘纸", "PA99016", 1, "pc"),
    ],
}
EXTRA_BOM = {norm(k): v for k, v in EXTRA_BOM_RAW.items()}


def _lookup_bom(nm, source, cb):
    """渠道感知BOM查找"""
    r = None
    if cb:
        r = source.get((nm, cb)) or source.get((nm, '通用'))
    if not r:
        r = source.get(nm)
    return r

def get_single_bom(nm, single0, combo0=None, cb=None):
    """渠道感知单品BOM查找: 先查 ALIAS 名, 未命中回退原名"""
    aliased = ALIAS.get(nm, nm)
    r = _lookup_bom(aliased, single0, cb)
    if not r and aliased != nm:
        r = _lookup_bom(nm, single0, cb)  # ALIAS未命中, 用原名查
    if not r and combo0:
        r = _lookup_bom(aliased, combo0, cb)
        if not r and aliased != nm:
            r = _lookup_bom(nm, combo0, cb)
    return r


def get_combo_bom(nm, detail, single0, combo0, cb=None):
    """渠道感知套餐BOM查找: ALIAS未命中回退原名"""
    aliased = ALIAS.get(nm, nm)
    direct = _lookup_bom(aliased, combo0, cb)
    if not direct and aliased != nm:
        direct = _lookup_bom(nm, combo0, cb)
    if direct:
        return direct, []
    merged, missing = {}, []
    for cname, cqty in parse_detail(detail):
        b = get_single_bom(norm(cname), single0, combo0, cb)
        if not b:
            missing.append(cname)
            continue
        for mname, code, qty, unit in b:
            key = code or mname
            if key in merged:
                merged[key][2] += qty * cqty
            else:
                merged[key] = [mname, code, qty * cqty, unit]
    items = [(v[0], v[1], round(v[2], 4), v[3]) for v in merged.values()]
    return items, missing


# ---------------- 读清洁BOM CSV (单源) ----------------
# 渠道过滤 / 物料删除规则已抽到 bom_rules.py（2026-06-26），口径在那里收口
from bom_rules import DINEIN_DEL, TAKEOUT_SINGLE_DEL, PRODUCT_DEL


def load_clean_bom(csv_path):
    """从 clean_bom.csv 加载, 落地所有渠道/删除规则, 返回 price + channel-aware BOM"""
    import csv as _csv
    price = {}
    single0, combo0 = {}, {}
    with open(csv_path, encoding='utf-8') as f:
        for r in _csv.DictReader(f):
            pname = r['商品']
            ptype = r['类型']
            mat = r['物料名称']
            code = r['物料编码'] = r['物料编码'].strip()
            qty = float(r['消耗数量'])
            unit = r['单位']
            ch = r['适用渠道']

            price[code] = float(r['物料单价']) if r['物料单价'] else 0.0

            # === 应用删除规则 ===
            # 1. 堂食全局删除 (适用渠道=堂食/通用 的行中去掉) — 通过调整渠道实现
            if mat in DINEIN_DEL and ch in ('通用', '堂食'):
                if ch == '通用': ch = '外卖'  # 改为仅外卖可用
                else: continue  # 堂食专属行直接丢弃
            # 2. 外卖单品删除
            if ptype == 'single' and mat in TAKEOUT_SINGLE_DEL and ch in ('通用', '外卖'):
                if ch == '通用': ch = '堂食'
                else: continue
            # 3. 商品专属删除
            if pname in PRODUCT_DEL and mat in PRODUCT_DEL[pname]:
                continue

            entry = (mat, code, round(float(qty), 4), unit)
            target = single0 if ptype == 'single' else combo0

            # 渠道感知: 同时存(channel)和(product)两种key
            for c in ([ch] if ch != '通用' else ['堂食', '外卖']):
                target.setdefault((norm(pname), c), [])
                dup = target[(norm(pname), c)]
                if not any(e[1] == code for e in dup):
                    dup.append(entry)
            # backward compat key
            target.setdefault(norm(pname), [])
            if not any(e[1] == code for e in target[norm(pname)]):
                target[norm(pname)].append(entry)

    return price, single0, combo0


# ---------------- 生成 ----------------
def build(bom_csv_path, sales_path, out_path, recon_path=None):
    print("读取清洁 BOM CSV ...", flush=True)
    price, single0, combo0 = load_clean_bom(bom_csv_path)
    print(f"  物料价 {len(price)} | 单品 {len(single0)} | 套餐 {len(combo0)}")
    manual_detail = {}
    fb = {}  # no fallback needed
    print("读取销量表 ...", flush=True)
    sales, din_rate, to_rate = load_sales(sales_path)
    print(f"  销量聚合行 {len(sales)}")

    # 实收对齐:把每店每渠道的实收率分子换成门店汇总口径实收(堂食 payment−退款−挂账 / 外卖 platform_total),
    # 分母用毛利表该店该渠道实际原价和。这样 Σ(原价×rate)=门店汇总实收,4 sheet 合计与 ttpos shop 报表一致。
    if recon_path:
        with open(recon_path, encoding="utf-8") as f:
            raw_recon = json.load(f)
        recon = {str(sid).strip()[:3]: rv for sid, rv in raw_recon.items()}
        orig = {}   # (s3, 渠道): Σ原价(qty>0 的行,与写表口径一致)
        for (s3, cb, kind, _), e in sales.items():
            if e["qty"] > 0:
                orig[(s3, cb)] = orig.get((s3, cb), 0.0) + e["amt"]
        din_rate, to_rate = {}, {}
        missing_recon = []
        for (s3, cb), amount in sorted(orig.items()):
            if amount <= 0:
                continue
            rv = recon.get(s3)
            if rv is None:
                missing_recon.append(f"{s3}-{cb}")
                continue
            recv_key = "din_recv" if cb == "堂食" else "to_recv"
            rate = fnum(rv.get(recv_key)) / amount
            if cb == "堂食":
                din_rate[s3] = rate
            else:
                to_rate[s3] = rate
        if missing_recon:
            raise SystemExit("recon 缺少这些门店/渠道的实收明细: " + "、".join(missing_recon[:20]))
        print(f"  实收已对齐门店汇总口径(recon {len(recon)} 店):堂食率 {len(din_rate)} | 外卖率 {len(to_rate)}")
    else:
        raise SystemExit(
            "缺少 --recon，不能生成正确实收金额。请先运行 wallace_shop_summary.py 生成 "
            "Wallace门店实收明细_<月份>.json，或通过 --recon 指定。"
        )

    wb = xlsxwriter.Workbook(out_path)
    F_HDR = wb.add_format({"bold": True, "bg_color": "#4472C4", "font_color": "white",
                           "align": "center", "valign": "vcenter", "border": 1})
    F_TXT = wb.add_format({"valign": "vcenter"})
    F_NUM = wb.add_format({"valign": "vcenter", "num_format": "#,##0"})
    F_MON = wb.add_format({"valign": "vcenter", "num_format": "#,##0.00"})
    F_PCT = wb.add_format({"valign": "vcenter", "num_format": "0.00%"})

    unmatched, missing_comp = [], {}
    miss_price_codes = set()
    stat = {}

    def rec_missing(s3, comps, qty):
        for m in comps:
            r = missing_comp.setdefault(norm(m), {"name": m, "stores": set(), "combos": 0, "qty": 0})
            r["stores"].add(s3)
            r["combos"] += 1
            r["qty"] += qty

    def item_cost(items):
        c = 0.0
        for _, code, qty, _ in items:
            if code and code not in price:
                miss_price_codes.add(code)
            c += qty * price.get(code, 0.0)
        return c

    def make_sheet(cb, kind, title):
        ws = wb.add_worksheet(t_sheet(title))
        ws.freeze_panes(1, 0)
        for c, w in enumerate([12, 9, 9, 26, 8, 18, 11, 9, 10, 6, 9, 11, 11, 12, 12, 12]):
            ws.set_column(c, c, w)
        for c, name in enumerate(OUT_COLS):
            ws.write(0, c, t_hdr(name), F_HDR)
        rate_map = din_rate if cb == "堂食" else to_rate
        keys = sorted([k for k in sales if k[1] == cb and k[2] == kind],
                      key=lambda k: (k[0], k[3]))
        # ---- 第一遍:取 BOM、算该 sheet 已匹配商品的平均成本率(供未映射估算)----
        recs = []
        bom_cost = bom_rev = 0.0
        for k in keys:
            s3 = k[0]
            e = sales[k]
            qty = e["qty"]
            if qty <= 0:
                continue
            disp = e["disp"]
            if s3 not in rate_map:
                raise SystemExit(f"recon 缺少 {s3}-{cb} 的实收率，无法计算正确实收金额")
            rev = round(e["amt"] * rate_map[s3])
            nmd = norm(disp)
            md = manual_detail.get(nmd)
            if "烤玉米" in nmd and nmd not in combo0 and nmd not in single0:  # 直配方优先;否则烤玉米=玉米杯×(送?2:1)
                cup = get_single_bom("玉米杯", single0, combo0, cb)
                miss = []
                if cup:
                    mult = 2 if "送" in nmd else 1
                    items = [(n, c, round(q * mult, 4), u) for n, c, q, u in cup]
                else:
                    items = []
                    unmatched.append([title, s3, disp, "玉米杯无 BOM"])
            elif nmd in ALIAS and not md:    # market 说明(manual_detail) 优先于写死 ALIAS
                items = get_single_bom(nmd, single0, combo0, cb)
                miss = []
                if not items:
                    unmatched.append([title, s3, disp, "别名目标无 BOM"])
            elif kind == "single" and not md:
                items = get_single_bom(nmd, single0, combo0, cb)
                miss = []
                if not items:
                    unmatched.append([title, s3, disp, "无单品 BOM"])
            else:
                detail = md or e["detail"]
                items, miss = get_combo_bom(nmd, detail, single0, combo0, cb)
                if not items:
                    unmatched.append([title, s3, disp, "无 BOM 且组分无法拼"])
                elif miss:
                    unmatched.append([title, s3, disp, "缺组分: " + "、".join(miss[:6])])
                if miss:
                    rec_missing(s3, miss, qty)
            # 所有渠道/商品过滤规则已在 load_clean_bom() 中落地, 此处不再需要
            ucost = round(item_cost(items), 4) if items else None
            recs.append({"s3": s3, "disp": disp, "qty": qty, "rev": rev,
                         "unit": round(e["unit"], 2), "items": items, "ucost": ucost})
            if items:
                bom_cost += round(ucost * qty)
                bom_rev += rev
        cr = (bom_cost / bom_rev) if bom_rev else 0.0   # 已匹配商品平均成本率

        # ---- 第二遍:写入(无 BOM 的用 cr 估算成本)----
        srow = scost = srev = 0
        ridx = 1
        for rc in recs:
            s3, disp, qty, rev, unit_price = rc["s3"], rc["disp"], rc["qty"], rc["rev"], rc["unit"]
            items, ucost = rc["items"], rc["ucost"]
            store = f"{s3} {s3}"
            r0 = ridx
            n = len(items) if items else 1
            r1 = r0 + n - 1
            er0 = r0 + 1

            def put(col, value, fmt, formula=False, result=None):
                if formula:
                    if n > 1:
                        ws.merge_range(r0, col, r1, col, "", fmt)
                    ws.write_formula(r0, col, value, fmt, result if result is not None else 0)
                else:
                    if n > 1:
                        ws.merge_range(r0, col, r1, col, value, fmt)
                    else:
                        ws.write(r0, col, value, fmt)

            if items:
                total_cost = round(ucost * qty)
                margin = rev - total_cost
                put(0, store, F_TXT)
                put(3, t_prod(disp), F_TXT)
                put(4, unit_price, F_MON)
                put(10, int(qty), F_NUM)
                put(11, rev, F_NUM)
                put(12, f"=L{er0}", F_NUM, formula=True, result=rev)
                put(13, f"=SUMPRODUCT(H{er0}:H{r1+1},I{er0}:I{r1+1})", F_MON,
                    formula=True, result=round(ucost, 2))
                put(14, f"=L{er0}-N{er0}*K{er0}", F_NUM, formula=True, result=margin)
                put(15, f"=O{er0}", F_NUM, formula=True, result=margin)
                pr = (margin / rev) if rev else 0
                put(1, f"=IF(L{er0}=0,0,O{er0}/L{er0})", F_PCT, formula=True, result=pr)
                put(2, f"=B{er0}", F_PCT, formula=True, result=pr)
                for i, it in enumerate(items):
                    rr = r0 + i
                    ws.write(rr, 5, t_mat(it[0], it[1]), F_TXT)
                    ws.write(rr, 6, it[1], F_TXT)
                    ws.write(rr, 7, it[2], F_MON)
                    ws.write(rr, 8, round(price.get(it[1], 0.0), 4), F_MON)
                    ws.write(rr, 9, t_unit(it[3]), F_TXT)
                scost += total_cost
            else:
                # 无 BOM(未映射):用该渠道平均成本率 cr 估算成本,标注"(估算·未映射)"
                est_cost = round(rev * cr)
                est_margin = rev - est_cost
                pr = (est_margin / rev) if rev else 0
                ws.write(r0, 0, store, F_TXT)
                ws.write(r0, 1, pr, F_PCT)
                ws.write(r0, 2, pr, F_PCT)
                ws.write(r0, 3, t_prod(disp), F_TXT)
                ws.write(r0, 4, unit_price, F_MON)
                ws.write(r0, 5, t_misc("(估算·未映射)"), F_TXT)
                ws.write(r0, 10, int(qty), F_NUM)
                ws.write(r0, 11, rev, F_NUM)
                ws.write(r0, 12, rev, F_NUM)
                ws.write(r0, 13, round(est_cost / qty, 2) if qty else 0, F_MON)
                ws.write(r0, 14, est_margin, F_NUM)
                ws.write(r0, 15, est_margin, F_NUM)
                scost += est_cost
            srow += 1
            srev += rev
            ridx = r1 + 1
        stat[title] = {"row": srow, "rev": srev, "cost": scost, "cr": round(cr * 100, 1)}

    make_sheet("堂食", "single", "堂食-单品")
    make_sheet("堂食", "combo", "堂食-套餐")
    make_sheet("外卖", "single", "外卖-单品")
    make_sheet("外卖", "combo", "外卖-套餐")

    # 未匹配清单
    wsx = wb.add_worksheet(t_sheet("未匹配清单"))
    for c, name in enumerate(["来源sheet", "店号", "商品", "说明"]):
        wsx.write(0, c, t_hdr(name), F_HDR)
    for c, w in enumerate([14, 8, 30, 44]):
        wsx.set_column(c, c, w)
    for i, r in enumerate(unmatched, 1):
        for c, v in enumerate(r):
            wsx.write(i, c, ([t_sheet, str, t_prod, t_misc][c](v) if c < 4 else v))

    # 缺失组分单品(补这些可让套餐算全),按影响销量降序
    wsm = wb.add_worksheet(t_sheet("缺失组分单品"))
    for c, name in enumerate(["组分单品", "涉及店数", "被套餐引用次数", "影响套餐销量合计"]):
        wsm.write(0, c, t_hdr(name), F_HDR)
    for c, w in enumerate([34, 10, 16, 18]):
        wsm.set_column(c, c, w)
    for i, rec in enumerate(sorted(missing_comp.values(), key=lambda x: -x["qty"]), 1):
        wsm.write(i, 0, t_prod(rec["name"]))
        wsm.write(i, 1, len(rec["stores"]))
        wsm.write(i, 2, rec["combos"])
        wsm.write(i, 3, int(rec["qty"]))

    wb.close()
    print("\n==================== 完成 ====================")
    print(f"输出: {out_path}")
    for t in ("堂食-单品", "堂食-套餐", "外卖-单品", "外卖-套餐"):
        s = stat[t]
        gm = (s["rev"] - s["cost"]) / s["rev"] * 100 if s["rev"] else 0
        print(f"{t}: 商品 {s['row']:5} | 实收 ฿{s['rev']:>12,} | 成本 ฿{s['cost']:>11,} "
              f"| 毛利 ฿{s['rev']-s['cost']:>12,} | 毛利率 {gm:5.1f}%")
    print(f"未匹配/缺BOM 条目: {len(unmatched)} | 缺失组分单品: {len(missing_comp)} "
          f"| 缺价物料编码: {len(miss_price_codes)}")


def main():
    ap = argparse.ArgumentParser(description="用 0514 BOM 给销量表生成成本毛利分析")
    ap.add_argument("--bom", required=True, help="清洁 BOM CSV 文件路径")
    ap.add_argument("--sales", required=True)
    ap.add_argument("--recon", default=None,
                    help="门店实收明细 JSON(wallace_shop_summary 产出)。未传时自动查找同目录 Wallace门店实收明细_<月份>.json")
    ap.add_argument("--out", required=True)
    ap.add_argument("--lang", default="zh", choices=["zh", "en", "th"])
    a = ap.parse_args()
    global _LANG, _NM
    _LANG = a.lang
    if a.lang != "zh":
        _mp = os.path.join(os.path.dirname(os.path.abspath(__file__)), "名称映射_中英泰.json")
        if not os.path.exists(_mp):
            _mp = "/tmp/name_maps.json"
        _NM = _json.load(open(_mp, encoding="utf-8")) if os.path.exists(_mp) else {}
        if not _NM:
            print("⚠️ 名称映射缺失, 商品/物料名保持中文 (改 zh 或重建映射可恢复)")
    for p in (a.bom, a.sales):
        if not os.path.exists(p):
            sys.exit(f"找不到文件: {p}")
    recon = resolve_recon_path(a.sales, a.recon)
    if not recon:
        sys.exit("找不到门店实收明细 JSON，请先运行 wallace_shop_summary.py，或通过 --recon 指定")
    if not os.path.exists(recon):
        sys.exit(f"找不到文件: {recon}")
    build(a.bom, a.sales, a.out, recon)


if __name__ == "__main__":
    main()
