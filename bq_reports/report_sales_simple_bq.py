#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
销售业绩明细表 (BigQuery 版本) - 简化版

导出原始订单商品明细（不去重）

使用方法:
    python3 report_sales_simple_bq.py [YYYY-MM]
"""

import os
import sys
import csv
import calendar
import subprocess
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from google.cloud import bigquery
from google.oauth2.credentials import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PROJECT_ID = "diyl-407103"
OUTPUT_DIR = Path(__file__).parent / "output"

STORE_LIST = [
    ("1", "1958987436032000"), ("2", "2269470793728000"), ("3", "2598648160256000"),
    ("4", "2876210421760000"), ("5", "3446618988544000"), ("6", "3870122057728000"),
    ("7", "4149605310464000"), ("8", "4358842359808000"), ("9", "4912616316928000"),
    ("10", "5567347171328000"), ("11", "5999171739648000"), ("12", "6542950670336000"),
    ("13", "6789240201216000"), ("14", "6977459593216000"), ("15", "7191251656704000"),
    ("16", "7400123801600000"), ("17", "7648065888256000"), ("18", "7863653113856000"),
    ("19", "8100551598080000"), ("20", "8501761941504000"), ("21", "8722592047104000"),
    ("22", "2947521978368000"), ("23", "3448951017472000"), ("24", "3782477877248000"),
    ("25", "4024875094016000"), ("26", "4229506797568000"), ("27", "4418766376960000"),
    ("28", "4613872816128000"), ("29", "4805464432640000"), ("30", "5001267122176000"),
    ("31", "5250979205120000"), ("32", "5444022046720000"), ("33", "7600687026176000"),
    ("34", "7813128523776000"), ("35", "8051063001088000"), ("36", "8535580610560000"),
    ("37", "8723170856960000"), ("38", "1515821506560000"), ("39", "3631470387200000"),
    ("40", "5498438983680000"), ("42", "9231705128960000"), ("44", "1379607252992000"),
    ("46", "1745354756096000"), ("47", "1919875551232000"), ("49", "2277263806464000"),
    ("51", "2618629820416000"), ("52", "2788834676736000"), ("53", "2992442970112000"),
    ("54", "3169446793216000"), ("55", "3367514411008000"), ("56", "3662462062592000"),
    ("59", "4197894328320000"), ("61", "3087884357632000"),
]

UUID_SET = {uuid for _, uuid in STORE_LIST}


def log(msg: str):
    if os.getenv("QUIET", "0") != "1":
        print(msg, flush=True)


def setup_proxy():
    proxy_url = os.environ.get("BQ_PROXY")
    if not proxy_url:
        return
    for k in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"]:
        os.environ[k] = proxy_url


def get_creds():
    r = subprocess.run(["gcloud", "auth", "print-access-token"], capture_output=True, text=True, check=True)
    return Credentials(token=r.stdout.strip(), scopes=["https://www.googleapis.com/auth/cloud-platform"])


def get_time_range():
    if len(sys.argv) > 1:
        ym = sys.argv[1]
    else:
        d = date.today().replace(day=1) - timedelta(days=1)
        ym = d.strftime("%Y-%m")
    y, m = map(int, ym.split('-'))
    ld = calendar.monthrange(y, m)[1]
    st = int(datetime(y, m, 1, 0, 0, 0).timestamp())
    et = int(datetime(y, m, ld, 23, 59, 59).timestamp())
    return ym, st, et


def query_shop(client, dataset_id, start_ts, end_ts):
    """查询单个门店的销售明细"""
    query = f"""
    SELECT
        sb.order_no,
        FORMAT_TIMESTAMP('%Y-%m-%d %H:%M:%S', TIMESTAMP_SECONDS(sb.finish_time)) as finish_time,
        IF(sop.name LIKE '{{%', 
            (SELECT STRING_AGG(val, ' / ') FROM UNNEST([
                NULLIF(JSON_EXTRACT_SCALAR(sop.name, '$.zh'), 'null'),
                NULLIF(JSON_EXTRACT_SCALAR(sop.name, '$.en'), 'null'),
                NULLIF(JSON_EXTRACT_SCALAR(sop.name, '$.th'), 'null')
            ]) AS val WHERE val IS NOT NULL),
            sop.name
        ) as product_name,
        sop.product_type,
        sop.num,
        sop.total_price as unit_price,
        ROUND(sop.total_price * sop.num, 2) as total_price
    FROM `{PROJECT_ID}.{dataset_id}.ttpos_sale_order_product` sop
    JOIN `{PROJECT_ID}.{dataset_id}.ttpos_sale_bill` sb
        ON sb.uuid = sop.sale_bill_uuid AND sb.delete_time = 0
    WHERE sop.delete_time = 0
      AND sop.cancel_time = 0
      AND sb.status = 1
      AND sb.finish_time >= {start_ts}
      AND sb.finish_time <= {end_ts}
      AND sop.product_type IN (0, 2)
    ORDER BY sb.finish_time, sb.order_no
    """
    try:
        result = client.query(query).result()
        return [(r.order_no, r.finish_time, r.product_name, r.product_type,
                 r.num, r.unit_price, r.total_price) for r in result]
    except Exception as e:
        log(f"  {dataset_id} 查询失败: {e}")
        return []


def get_store_name(client, dataset_id):
    try:
        r = list(client.query(f"""
            SELECT JSON_EXTRACT_SCALAR(values, '$.store_code') as no,
                   JSON_EXTRACT_SCALAR(values, '$.store_name') as name
            FROM `{PROJECT_ID}.{dataset_id}.ttpos_setting`
            WHERE `key` = 'store' AND delete_time = 0 LIMIT 1
        """).result())[0]
        return r.no or "", r.name or ""
    except:
        return "", ""


def process_one(args):
    idx, total, uuid, start_ts, end_ts = args
    dataset_id = f"shop{uuid}"
    
    try:
        client = bigquery.Client(project=PROJECT_ID, credentials=get_creds())
    except Exception as e:
        return {"error": str(e), "idx": idx}
    
    store_no, store_name = get_store_name(client, dataset_id)
    if not store_name:
        store_name = f"门店_{uuid[:12]}"
    
    rows = query_shop(client, dataset_id, start_ts, end_ts)
    
    return {
        "idx": idx,
        "store_no": store_no,
        "store_name": store_name,
        "rows": rows,
        "count": len(rows)
    }


def _parse_force_flag():
    """最小 argparse — 仅支持 --force, 其余 sys.argv 行为不变."""
    import argparse
    p = argparse.ArgumentParser(add_help=False)
    p.add_argument("--force", action="store_true",
                   help="强制导出 (有 🔴 违反时打水印而非阻断)")
    args, _ = p.parse_known_args()
    return args.force


def main():
    # 技术债: 本脚本导出订单行明细, 行结构为 order_no×product_name, 缺少聚合销售桶字段,
    # 无法跑 DEFAULT_IDENTITIES. 当前用基线恒等式 (必填+唯一) 替代.
    # 升级路径: 若需对账, 改用 profit_margin_report 或在此加销售桶聚合.
    force = _parse_force_flag()

    t0 = time.time()
    log("=" * 60)
    log("销售业绩明细表 (BigQuery 简化版)")
    log("=" * 60)

    setup_proxy()

    ym, start_ts, end_ts = get_time_range()
    log(f"统计月份: {ym}")
    log(f"时间戳: {start_ts} ~ {end_ts}\n")
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # 获取所有门店数据集
    client = bigquery.Client(project=PROJECT_ID, credentials=get_creds())
    all_datasets = [d.dataset_id for d in client.list_datasets() if d.dataset_id.startswith("shop")]
    shop_datasets = [(i+1, uuid) for i, (_, uuid) in enumerate(STORE_LIST) 
                     if f"shop{uuid}" in all_datasets]
    
    log(f"找到 {len(shop_datasets)} 个门店数据集")
    log(f"开始并发查询 (5并发)...\n")
    
    all_data = []
    total_rows = 0
    
    with ThreadPoolExecutor(max_workers=5) as ex:
        futures = {ex.submit(process_one, (idx, len(shop_datasets), uuid, start_ts, end_ts)): idx 
                   for idx, uuid in shop_datasets}
        
        for fut in as_completed(futures):
            r = fut.result()
            if "error" in r:
                log(f"[{r['idx']}] 错误: {r['error']}")
                continue
            
            log(f"[{r['idx']}/{len(shop_datasets)}] {r['store_name']}: {r['count']} 行")
            
            for row in r["rows"]:
                all_data.append((r["store_no"], r["store_name"]) + row)
            total_rows += r["count"]
    
    log(f"\n总共 {total_rows} 行数据，正在生成文件...")
    
    # 生成 CSV
    csv_path = OUTPUT_DIR / f"销售业绩明细-{ym}.csv"
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(['门店编号', '门店名称', '订单号', '完成时间', '商品名称', 
                    '商品类型', '数量', '单价', '金额'])
        for row in all_data:
            w.writerow([row[0], row[1], row[2], row[3], row[4],
                       '普通商品' if row[5] == 0 else '套餐子商品',
                       f"{row[6]:.2f}", f"{row[7]:.2f}", f"{row[8]:.2f}"])
    
    # 生成 Excel
    xlsx_path = OUTPUT_DIR / f"销售业绩明细-{ym}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = f"销售业绩-{ym}"
    
    headers = ['门店编号', '门店名称', '订单号', '完成时间', '商品名称', 
               '商品类型', '数量', '单价', '金额']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin = Border(left=Side(style='thin', color='D9D9D9'),
                  right=Side(style='thin', color='D9D9D9'),
                  top=Side(style='thin', color='D9D9D9'),
                  bottom=Side(style='thin', color='D9D9D9'))
    
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin
        c.alignment = Alignment(horizontal='center')
    
    for ri, row in enumerate(all_data, 2):
        ws.cell(row=ri, column=1, value=row[0]).border = thin
        ws.cell(row=ri, column=2, value=row[1]).border = thin
        ws.cell(row=ri, column=3, value=row[2]).border = thin
        ws.cell(row=ri, column=4, value=row[3]).border = thin
        ws.cell(row=ri, column=5, value=row[4]).border = thin
        ws.cell(row=ri, column=6, value='普通商品' if row[5] == 0 else '套餐子商品').border = thin
        ws.cell(row=ri, column=7, value=float(row[6])).border = thin
        ws.cell(row=ri, column=8, value=float(row[7])).border = thin
        ws.cell(row=ri, column=9, value=float(row[8])).border = thin
        
        if ri % 10000 == 0:
            log(f"  已写入 {ri-1} 行...")
    
    for i, w in enumerate([12, 28, 18, 18, 45, 12, 10, 10, 12], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A2'

    from semantic.validators.gate import (
        add_watermark_sheet_openpyxl, validate_and_gate)
    from semantic.validators.identities import (
        make_required_fields_identity, make_unique_key_identity)
    uniq_ident, _prepare = make_unique_key_identity(
        ("store_no", "order_no", "product_name"), name="店×订单×商品主键唯一")
    _check_rows = _prepare([
        {"store_no": str(r[0]), "order_no": str(r[2]), "product_name": r[4] or ""}
        for r in all_data
    ])
    _outcome = validate_and_gate(
        _check_rows,
        [make_required_fields_identity(
            ("store_no", "order_no", "product_name"), name="销售明细必填字段"),
         uniq_ident],
        force=force, report_name="report_sales_simple_bq",
        row_label=lambda r: f"{r.get('store_no', '')} {r.get('order_no', '')} {r.get('product_name', '')}",
    )
    if _outcome.needs_watermark:
        add_watermark_sheet_openpyxl(wb, _outcome.watermark_lines())

    wb.save(xlsx_path)
    
    elapsed = time.time() - t0
    log(f"\n{'='*60}")
    log(f"完成! 总行数: {total_rows}")
    log(f"CSV: {csv_path}")
    log(f"Excel: {xlsx_path}")
    log(f"总耗时: {elapsed:.1f} 秒 ({elapsed/60:.1f} 分钟)")
    log(f"{'='*60}")


if __name__ == "__main__":
    main()
