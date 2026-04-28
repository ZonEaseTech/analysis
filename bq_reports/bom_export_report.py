#!/usr/bin/env python3
"""
BOM 配方导出 —— 市场需求版

输出：三 sheet Excel
  Sheet 1「BOM清单」：5 列(门店名称 | 商品名称 | BOM物品名称 | BOM消耗 | BOM单位)
                     已配置 BOM 的单品/加料
  Sheet 2「未配置BOM」：3 列(门店名称 | 商品类型 | 商品名称)
                       未配置 BOM 的单品 + 加料,各商品/加料 1 行
  Sheet 3「套餐组成」：7 列(门店名称 | 套餐名称 | 分组名 | 子商品 | 数量 | 是否必选 | 加价)
                     套餐(product_type=1)的组与子商品明细

范围：
  - 商户真实名称取自 ttpos_company.name(优先 JSON zh,回退 en,再回退原始值)
  - 商品类型:单品 + 加料 + 套餐
    - 单品:ttpos_product_package.product_type = 0
    - 加料:ttpos_product_sauce 表(独立菜单,前端 /product/sauce/*)
    - 套餐:ttpos_product_package.product_type = 1
  - 53 家分店产品配置完全一致(总部下发)

Usage:
    venv/bin/python -m bq_reports.bom_export_report --output exports/bom_export.xlsx
"""

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
# 任意一家分店 dataset(总部下发,配置一致)
DEFAULT_DATASET = "shop1958987436032000"

# 报表版本号 —— 每次正式交付前递增
REPORT_VERSION = "v8"

# 门店名称从 BQ ttpos_company 动态加载
STORE_NAME_SQL = """
SELECT
  COALESCE(
    JSON_EXTRACT_SCALAR(name, '$.zh'),
    JSON_EXTRACT_SCALAR(name, '$.en'),
    name
  ) AS company_name
FROM `{project}`.`{dataset}`.`ttpos_company`
WHERE delete_time = 0
LIMIT 1
"""

# 单品 BOM —— LEFT JOIN 保留无 BOM 的单品
# 商品名后缀规则(优先级从高到低):
#   1. 同 pkg_uuid 有≥2 张不同 BOM 卡(多规格不同配方) → 追加 "(规格名)"   例: 矿泉水 (1pc)
#   2. 商品名在多个 pkg 中重复(同名不同 SKU,多渠道销售) → 追加 "(分类名)" 例: 鸡肉棒 (Snacks)
#   3. 否则保留原商品名
SINGLE_BOM_SQL = """
WITH bom_dedup AS (
  SELECT
    product_package_uuid,
    product_bom_card_uuid,
    MIN(product_flavor_uuid) AS flavor_uuid
  FROM `{project}`.`{dataset}`.`ttpos_product_bom`
  WHERE delete_time = 0 AND product_bom_card_uuid > 0
  GROUP BY product_package_uuid, product_bom_card_uuid
),
pkg_card_count AS (
  SELECT product_package_uuid, COUNT(*) AS card_cnt
  FROM bom_dedup
  GROUP BY product_package_uuid
),
pkg_meta AS (
  -- 每个 pkg 的去空白基础名 + 分类名(用于同名 SKU 区分)
  SELECT
    pp.uuid AS pkg_uuid,
    REGEXP_REPLACE(COALESCE(
      JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
      JSON_EXTRACT_SCALAR(pp.name, '$.en'),
      ''
    ), r'^\\s+|\\s+$', '') AS base_name,
    COALESCE(
      JSON_EXTRACT_SCALAR(c.name, '$.zh'),
      JSON_EXTRACT_SCALAR(c.name, '$.en'),
      ''
    ) AS category_name
  FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
  LEFT JOIN `{project}`.`{dataset}`.`ttpos_product_category` c
    ON c.uuid = pp.category_uuid AND c.delete_time = 0
  WHERE pp.delete_time = 0 AND pp.product_type = 0
),
duplicate_names AS (
  -- 哪些 base_name 在多个 pkg 中重复(同名不同 SKU)
  SELECT base_name
  FROM pkg_meta
  WHERE base_name != ''
  GROUP BY base_name
  HAVING COUNT(*) > 1
)
SELECT
  0 AS sort_type,
  pp.uuid AS product_uuid,
  CASE
    -- 优先级 1: 多卡 → 追加规格名
    WHEN IFNULL(pcc.card_cnt, 0) > 1 AND pf.uuid IS NOT NULL THEN CONCAT(
      pm.base_name,
      ' (',
      COALESCE(JSON_EXTRACT_SCALAR(pf.name, '$.zh'),
               JSON_EXTRACT_SCALAR(pf.name, '$.en'),
               ''),
      ')'
    )
    -- 优先级 2: 同名不同 SKU → 追加分类名
    WHEN dn.base_name IS NOT NULL AND pm.category_name != '' THEN CONCAT(
      pm.base_name,
      ' (',
      pm.category_name,
      ')'
    )
    -- 默认: 原商品名
    ELSE pm.base_name
  END AS product_name,
  pm.category_name AS category_name,
  COALESCE(JSON_EXTRACT_SCALAR(m.name, '$.zh'),
           JSON_EXTRACT_SCALAR(m.name, '$.en')) AS material_name,
  rm.num AS bom_num,
  COALESCE(JSON_EXTRACT_SCALAR(rm.unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.unit_name, '$.en'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.en')) AS bom_unit
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
LEFT JOIN bom_dedup pb ON pb.product_package_uuid = pp.uuid
LEFT JOIN pkg_card_count pcc ON pcc.product_package_uuid = pp.uuid
LEFT JOIN pkg_meta pm ON pm.pkg_uuid = pp.uuid
LEFT JOIN duplicate_names dn ON dn.base_name = pm.base_name
LEFT JOIN `{project}`.`{dataset}`.`ttpos_product_flavor` pf
  ON pf.uuid = pb.flavor_uuid AND pf.delete_time = 0
LEFT JOIN `{project}`.`{dataset}`.`ttpos_related_material` rm
  ON rm.related_uuid = pb.product_bom_card_uuid AND rm.delete_time = 0
LEFT JOIN `{project}`.`{dataset}`.`ttpos_material` m
  ON m.uuid = rm.material_uuid AND m.delete_time = 0
WHERE pp.delete_time = 0 AND pp.product_type = 0
"""

# 加料 BOM —— LEFT JOIN 保留无 BOM 的加料
SAUCE_BOM_SQL = """
SELECT
  1 AS sort_type,
  ps.uuid AS product_uuid,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(ps.name, '$.zh'),
    JSON_EXTRACT_SCALAR(ps.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS product_name,
  '' AS category_name,
  COALESCE(JSON_EXTRACT_SCALAR(m.name, '$.zh'),
           JSON_EXTRACT_SCALAR(m.name, '$.en')) AS material_name,
  rm.num AS bom_num,
  COALESCE(JSON_EXTRACT_SCALAR(rm.unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.unit_name, '$.en'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.en')) AS bom_unit
FROM `{project}`.`{dataset}`.`ttpos_product_sauce` ps
LEFT JOIN `{project}`.`{dataset}`.`ttpos_related_material` rm
  ON rm.related_uuid = ps.product_bom_card_uuid
  AND ps.product_bom_card_uuid > 0
  AND rm.delete_time = 0
LEFT JOIN `{project}`.`{dataset}`.`ttpos_material` m
  ON m.uuid = rm.material_uuid AND m.delete_time = 0
WHERE ps.delete_time = 0
"""

# 套餐组成(product_package_group + group_item)
# group_type: 0=固定(套餐必含), 1=可选(顾客可选择)
COMBO_SQL = """
SELECT
  pp.uuid AS combo_uuid,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pp.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS combo_name,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pgrp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pgrp.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS group_name,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(child.name, '$.zh'),
    JSON_EXTRACT_SCALAR(child.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS child_name,
  gpi.num AS child_num,
  pgrp.group_type AS group_type,  -- 0=固定, 1=可选
  gpi.is_required,
  gpi.add_price
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
JOIN `{project}`.`{dataset}`.`ttpos_product_package_group` pgrp
  ON pgrp.product_package_uuid = pp.uuid AND pgrp.delete_time = 0
JOIN `{project}`.`{dataset}`.`ttpos_product_package_group_item` gpi
  ON gpi.product_package_group_uuid = pgrp.uuid AND gpi.delete_time = 0
JOIN `{project}`.`{dataset}`.`ttpos_product_package` child
  ON child.uuid = gpi.related_uuid AND child.delete_time = 0
WHERE pp.delete_time = 0 AND pp.product_type = 1
ORDER BY pp.uuid, pgrp.uuid, gpi.sort
"""

TYPE_LABEL = {0: "单品", 1: "加料"}


def query_rows(client, dataset: str):
    sql = f"""
    {SINGLE_BOM_SQL.format(project=client.project, dataset=dataset)}
    UNION ALL
    {SAUCE_BOM_SQL.format(project=client.project, dataset=dataset)}
    ORDER BY sort_type, product_name, material_name
    """
    return list(client.query(sql).result())


def query_combo_rows(client, dataset: str):
    sql = COMBO_SQL.format(project=client.project, dataset=dataset)
    return list(client.query(sql).result())


def get_store_name(client, dataset: str) -> str:
    sql = STORE_NAME_SQL.format(project=client.project, dataset=dataset)
    for row in client.query(sql).result():
        name = row.company_name or ""
        return name.strip() if name else ""
    return ""


def split_rows(rows):
    """拆分为「有 BOM」明细行 vs「无 BOM」商品列表(每商品/加料 1 条)"""
    with_bom = []
    no_bom = []
    seen_no_bom = set()
    for r in rows:
        if r.material_name:
            with_bom.append(r)
        else:
            key = (r.sort_type, r.product_uuid)
            if key in seen_no_bom:
                continue
            seen_no_bom.add(key)
            no_bom.append(r)
    return with_bom, no_bom


def _style_header(ws, headers):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, size=11, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align


def _merge_runs(ws, col_letter: str, start_row: int, end_row: int, value_of):
    """合并 col_letter 列上 [start_row, end_row] 区间内相邻相同值的单元格"""
    if end_row <= start_row:
        return
    run_start = start_row
    run_value = value_of(start_row)
    for r in range(start_row + 1, end_row + 1):
        v = value_of(r)
        if v != run_value:
            if r - 1 > run_start:
                ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{r - 1}")
            run_start = r
            run_value = v
    if end_row > run_start:
        ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{end_row}")


def _write_bom_sheet(ws, with_bom, store_name: str):
    headers = ["门店名称", "商品名称", "分类", "BOM物品名称", "BOM消耗", "BOM单位"]
    _style_header(ws, headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(with_bom, start=2):
        ws.cell(row=r_idx, column=1, value=store_name).alignment = center
        ws.cell(row=r_idx, column=2, value=row.product_name or "").alignment = center
        ws.cell(row=r_idx, column=3, value=row.category_name or "").alignment = center
        ws.cell(row=r_idx, column=4, value=row.material_name or "")
        if row.bom_num is not None:
            ws.cell(row=r_idx, column=5, value=float(row.bom_num))
        ws.cell(row=r_idx, column=6, value=row.bom_unit or "")

    if with_bom:
        last_row = 1 + len(with_bom)
        if last_row > 2:
            ws.merge_cells(f"A2:A{last_row}")
            _merge_runs(ws, "B", 2, last_row, lambda r: with_bom[r - 2].product_name)
            _merge_runs(ws, "C", 2, last_row,
                        lambda r: f"{with_bom[r - 2].product_name}|{with_bom[r - 2].category_name or ''}")

    for col_idx, w in enumerate([14, 30, 18, 28, 10, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


def _write_no_bom_sheet(ws, no_bom, store_name: str):
    headers = ["门店名称", "商品类型", "商品名称"]
    _style_header(ws, headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(no_bom, start=2):
        ws.cell(row=r_idx, column=1, value=store_name).alignment = center
        ws.cell(row=r_idx, column=2, value=TYPE_LABEL.get(row.sort_type, "")).alignment = center
        ws.cell(row=r_idx, column=3, value=row.product_name or "")

    if no_bom:
        last_row = 1 + len(no_bom)
        if last_row > 2:
            ws.merge_cells(f"A2:A{last_row}")
            _merge_runs(ws, "B", 2, last_row,
                        lambda r: TYPE_LABEL.get(no_bom[r - 2].sort_type, ""))

    for col_idx, w in enumerate([14, 12, 30], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


def _write_combo_sheet(ws, combo_rows, store_name: str):
    headers = ["门店名称", "套餐名称", "分组", "类型", "单品名称", "数量", "是否必选", "加价"]
    _style_header(ws, headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(combo_rows, start=2):
        ws.cell(row=r_idx, column=1, value=store_name).alignment = center
        ws.cell(row=r_idx, column=2, value=row.combo_name or "").alignment = center
        ws.cell(row=r_idx, column=3, value=row.group_name or "").alignment = center
        ws.cell(row=r_idx, column=4, value="固定" if row.group_type == 0 else "可选").alignment = center
        ws.cell(row=r_idx, column=5, value=row.child_name or "").alignment = center
        ws.cell(row=r_idx, column=6, value=float(row.child_num) if row.child_num is not None else 0)
        ws.cell(row=r_idx, column=7, value="是" if row.is_required else "否").alignment = center
        ws.cell(row=r_idx, column=8, value=float(row.add_price) if row.add_price else 0)

    if combo_rows:
        last_row = 1 + len(combo_rows)
        if last_row > 2:
            # A 列(门店名称)整列合并
            ws.merge_cells(f"A2:A{last_row}")
            # B 列(套餐名称)相邻同名合并
            _merge_runs(ws, "B", 2, last_row, lambda r: combo_rows[r - 2].combo_name)
            # C 列(分组)相邻同组合并(同一套餐内的同名组合并)
            _merge_runs(ws, "C", 2, last_row,
                        lambda r: f"{combo_rows[r - 2].combo_name}|{combo_rows[r - 2].group_name}")
            # D 列(类型)相邻同类型合并(同一组内的同类型合并)
            _merge_runs(ws, "D", 2, last_row,
                        lambda r: f"{combo_rows[r - 2].combo_name}|{combo_rows[r - 2].group_name}|{'固定' if combo_rows[r - 2].group_type == 0 else '可选'}")

    for col_idx, w in enumerate([14, 25, 12, 10, 25, 8, 10, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


def write_excel(with_bom, no_bom, combo_rows, store_name: str, output_path: str):
    wb = Workbook()
    ws_bom = wb.active
    ws_bom.title = "BOM清单"
    _write_bom_sheet(ws_bom, with_bom, store_name)

    ws_no = wb.create_sheet("未配置BOM")
    _write_no_bom_sheet(ws_no, no_bom, store_name)

    ws_combo = wb.create_sheet("套餐组成")
    _write_combo_sheet(ws_combo, combo_rows, store_name)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _inject_version(path: str) -> str:
    """在文件名后缀前插入版本号: foo.xlsx → foo_v1.xlsx (已带版本号则不重复加)"""
    p = Path(path)
    suffix = f"_{REPORT_VERSION}"
    if p.stem.endswith(suffix):
        return str(p)
    return str(p.with_name(f"{p.stem}{suffix}{p.suffix}"))


def parse_args():
    p = argparse.ArgumentParser(description="BOM 配方导出")
    p.add_argument("--project", default=PROJECT_ID, help="GCP 项目 ID")
    p.add_argument("--dataset", default=DEFAULT_DATASET,
                   help=f"BQ dataset(任意一家分店即可,默认 {DEFAULT_DATASET})")
    p.add_argument("--output", required=True,
                   help=f"输出 Excel 文件路径(自动追加版本号 _{REPORT_VERSION})")
    return p.parse_args()


def main():
    args = parse_args()
    setup_proxy()

    output_path = _inject_version(args.output)

    print(f"[BQ] project={args.project} dataset={args.dataset}")
    print(f"[版本] {REPORT_VERSION} → 输出路径: {output_path}")
    client = get_bq_client(project_id=args.project)

    store_name = get_store_name(client, args.dataset)
    if not store_name:
        store_name = ""
    print(f"[门店] {store_name}")

    print("[BQ] 查询单品 + 加料 BOM...")
    rows = query_rows(client, args.dataset)
    print(f"[BQ] 单品+加料共 {len(rows)} 行")

    print("[BQ] 查询套餐组成...")
    combo_rows = query_combo_rows(client, args.dataset)
    print(f"[BQ] 套餐组成共 {len(combo_rows)} 行")

    with_bom, no_bom = split_rows(rows)
    write_excel(with_bom, no_bom, combo_rows, store_name, output_path)

    no_bom_singles = sum(1 for r in no_bom if r.sort_type == 0)
    no_bom_sauces = sum(1 for r in no_bom if r.sort_type == 1)
    print(f"[完成] 输出 {output_path}")
    print(f"  Sheet 1「BOM清单」: {len(with_bom)} 行")
    print(f"  Sheet 2「未配置BOM」: {len(no_bom)} 个 (单品 {no_bom_singles} 个, 加料 {no_bom_sauces} 个)")
    print(f"  Sheet 3「套餐组成」: {len(combo_rows)} 行")
    return 0


if __name__ == "__main__":
    sys.exit(main())
