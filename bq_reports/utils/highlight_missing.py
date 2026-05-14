"""报表后处理: 物料单价为空 (或 0) 的 BOM 行标黄.

为啥独立后处理: 报表写入用 xlsxwriter (流式 write_only), 写完才能精确知道哪些
行 的 物料单价 列是空; 此外样式需要按数据条件分支决定. 直接在写入时判断成本高
(每 cell 都要重判定); 写完一次性扫描 yellow-fill 更简单, 也跟下游格式无关。

只 yellow 那 5 列 (BOM 物品名 / 编码 / 消耗 / 单价 / 单位) — 不染 SKU 维度列 (那些
跨 BOM 行 merge 着, 染整个会很乱).
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


def highlight_missing_unit_prices(xlsx_path):
    """打开 xlsx, 把所有 '物料单价' 列为空 / 0 / None 的 BOM 行标黄.

    判定:
      - 物料单价 cell 值 == None | 0 | '' | '-'   → 是空
      - 但 BOM 物品编码 cell != '-' / None         → 是真 BOM 行 (排除"无 BOM" 占位行)
      - 同行 5 列 (BOM 物品名 / 编码 / 消耗 / 单价 / 单位) 标黄
    """
    wb = load_workbook(xlsx_path)
    total_highlighted = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        # 找表头
        hdr = [c.value for c in ws[1]]

        def col_idx(name):
            try: return hdr.index(name) + 1
            except ValueError: return None

        price_col = col_idx('物料单价')
        code_col = col_idx('BOM物品编码')
        name_col = col_idx('BOM物品名称')
        qty_col = col_idx('消耗数量')
        unit_col = col_idx('单位')
        if price_col is None or code_col is None:
            continue   # 不是 BOM 展开报表

        for r in range(2, ws.max_row + 1):
            code = ws.cell(r, code_col).value
            price = ws.cell(r, price_col).value
            # 跳过"无 BOM"占位行 (编码 = '-')
            if not code or str(code).strip() == '-':
                continue
            # 物料单价空判定
            if price is None or price == 0 or (isinstance(price, str) and price.strip() in ('', '-')):
                for c in (name_col, code_col, qty_col, price_col, unit_col):
                    if c is not None:
                        ws.cell(r, c).fill = YELLOW
                total_highlighted += 1
    wb.save(xlsx_path)
    print(f"[Yellow] 已标黄 {total_highlighted} 个 BOM 行 (物料单价为空)")
    return total_highlighted
