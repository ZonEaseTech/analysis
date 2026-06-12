#!/usr/bin/env python3
"""
BigQuery 导出框架 (BQ Export Framework)

提供可复用的多门店数据导出能力，支持：
- 批量查询多个门店
- 统一的 Excel 输出格式
- 内置校验机制
- 进度追踪和报告

Usage:
    from utils.bq_exporter import MultiShopExporter
    from utils.validators import create_default_validators
    
    exporter = MultiShopExporter(
        project_id="diyl-407103",
        output_path="exports/report.xlsx"
    )
    
    # 添加校验器
    exporter.set_validators(create_default_validators())
    
    # 设置商家列表
    exporter.load_merchants("resources/merchants.xlsx")
    
    # 执行导出
    result = exporter.export(
        sql_template="SELECT ... FROM `{project}.{dataset}.table` ...",
        start_ts=1772323200,
        end_ts=1775001600
    )
    
    print(f"导出完成: {result.success_count}/{result.total_count}")
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .bq_client import get_bq_client, setup_proxy
from .validators import DataValidator, ValidationChain, ValidationResult
from .sql_templates import get_template
from semantic.validators.gate import add_watermark_sheet_openpyxl, add_watermark_sheet_xlsxwriter


@dataclass
class ExportResult:
    """导出结果"""
    total_count: int
    success_count: int
    failed_count: int
    output_path: str
    validation_result: Optional[ValidationResult] = None
    errors: List[Dict[str, Any]] = field(default_factory=list)
    
    @property
    def success_rate(self) -> float:
        """成功率"""
        if self.total_count == 0:
            return 0.0
        return self.success_count / self.total_count
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            "total_count": self.total_count,
            "success_count": self.success_count,
            "failed_count": self.failed_count,
            "success_rate": f"{self.success_rate:.1%}",
            "output_path": self.output_path,
            "validation_passed": self.validation_result.is_valid if self.validation_result else None,
            "error_count": len(self.errors)
        }


@dataclass
class ExcelConfig:
    """Excel 配置"""
    sheet_name: str = "Sheet1"
    headers: List[str] = field(default_factory=list)
    header_fill_color: str = "4472C4"
    header_font_color: str = "FFFFFF"
    number_format: str = "0.00"
    auto_column_width: bool = True
    freeze_header: bool = True
    border_style: str = "thin"
    border_color: str = "D9D9D9"


class BaseExporter:
    """
    BigQuery 导出器基类
    
    提供基础的导出功能，单门店查询
    """
    
    def __init__(
        self,
        project_id: str,
        output_path: Union[str, Path],
        client: Optional[bigquery.Client] = None
    ):
        self.project_id = project_id
        self.output_path = Path(output_path)
        self.client = client or get_bq_client()
        self.validators: List[DataValidator] = []
        self.excel_config = ExcelConfig()
        self.gate_spec = None
    
    def set_validators(self, validators: Union[List[DataValidator], ValidationChain]):
        """设置校验器"""
        if isinstance(validators, ValidationChain):
            self.validators = validators.validators
        else:
            self.validators = validators
    
    def set_gate(self, gate_spec):
        """零容差导出闸门 (semantic.validators.gate.GateSpec).
        设置后写盘前强制过闸; 未设置 = 脚本必须以其他方式直调 validate_and_gate
        (tests/test_validator_coverage.py 结构性强制)."""
        self.gate_spec = gate_spec

    def set_excel_config(self, config: ExcelConfig):
        """设置 Excel 配置"""
        self.excel_config = config
    
    def query(
        self, 
        sql: str, 
        dataset: Optional[str] = None
    ) -> List[bigquery.Row]:
        """
        执行 BigQuery 查询
        
        Args:
            sql: SQL 查询语句
            dataset: 可选，dataset 名称用于格式化
        """
        if dataset:
            sql = sql.format(project=self.project_id, dataset=dataset)
        
        return list(self.client.query(sql).result())
    
    def validate(self, data: Any) -> ValidationResult:
        """执行所有校验"""
        if not self.validators:
            return ValidationResult.success()
        
        chain = ValidationChain(self.validators)
        return chain.validate(data)
    
    def write_excel(
        self,
        data: List[Dict[str, Any]],
        output_path: Optional[Path] = None
    ) -> Path:
        """
        写入 Excel 文件
        
        Args:
            data: 数据列表，每项为字典
            output_path: 输出路径，默认使用 self.output_path
        
        Returns:
            输出文件路径
        """
        output_path = output_path or self.output_path
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        cfg = self.excel_config
        wb = Workbook()
        ws = wb.active
        ws.title = cfg.sheet_name
        
        # 样式
        header_fill = PatternFill(
            start_color=cfg.header_fill_color,
            end_color=cfg.header_fill_color,
            fill_type="solid"
        )
        header_font = Font(bold=True, size=11, color=cfg.header_font_color)
        thin_border = Border(
            left=Side(style=cfg.border_style, color=cfg.border_color),
            right=Side(style=cfg.border_style, color=cfg.border_color),
            top=Side(style=cfg.border_style, color=cfg.border_color),
            bottom=Side(style=cfg.border_style, color=cfg.border_color),
        )
        
        # 写入表头
        headers = cfg.headers or (list(data[0].keys()) if data else [])
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "") if isinstance(row_data, dict) else row_data[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # 数字格式
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = cfg.number_format
        
        # 自动列宽
        if cfg.auto_column_width:
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, len(headers) + 1):
                max_len = 0
                for r in range(1, len(data) + 2):
                    val = ws.cell(row=r, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
        
        # 冻结首行
        if cfg.freeze_header:
            ws.freeze_panes = 'A2'

        # 零容差闸门 (set_gate 设置后执行)
        if self.gate_spec is not None:
            outcome = self.gate_spec.run(data)
            if outcome.needs_watermark:
                add_watermark_sheet_openpyxl(wb, outcome.watermark_lines())

        wb.save(output_path)
        return output_path


class MultiShopExporter(BaseExporter):
    """
    多门店批量导出器
    
    支持批量查询多个门店，统一输出到 Excel
    """
    
    def __init__(
        self,
        project_id: str,
        output_path: Union[str, Path],
        client: Optional[bigquery.Client] = None
    ):
        super().__init__(project_id, output_path, client)
        self.merchants: List[Tuple[str, str, str]] = []  # [(account, uuid, store_num), ...]
        self.progress_callback: Optional[Callable[[int, int, str, Dict], None]] = None
    
    def load_merchants(self, merchant_xlsx: Union[str, Path]):
        """
        从 Excel 加载商家列表

        Excel 格式：每行包含 [序号, 账号, UUID]
        门店编号从账号 (admin-XXX@wallace.com) 解析，用于外部销售源的日期切割。
        """
        import re
        from openpyxl import load_workbook

        wb = load_workbook(merchant_xlsx, data_only=True)
        ws = wb.active

        self.merchants = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 3 and row[1] and row[2]:
                account = str(row[1]).strip()
                uuid_str = str(row[2]).strip()
                m = re.search(r'admin-(\d+)@', account)
                store_num = m.group(1) if m else account
                self.merchants.append((account, uuid_str, store_num))

        wb.close()
        return self
    
    def set_merchants(self, merchants: List[Tuple[str, str]]):
        """直接设置商家列表"""
        self.merchants = merchants
        return self
    
    def set_progress_callback(self, callback: Callable[[int, int, str, Dict], None]):
        """
        设置进度回调函数
        
        Args:
            callback: 函数签名 (current, total, account, data) -> None
        """
        self.progress_callback = callback
        return self
    
    def export(
        self,
        sql_template: str,
        start_ts: int,
        end_ts: int,
        row_processor: Optional[Callable[[bigquery.Row, str], Dict]] = None
    ) -> ExportResult:
        """
        执行批量导出
        
        Args:
            sql_template: SQL 模板，使用 {project}, {dataset}, {start_ts}, {end_ts} 占位符
            start_ts: 开始时间戳
            end_ts: 结束时间戳
            row_processor: 可选的行处理器，将查询结果转换为字典
        
        Returns:
            ExportResult
        """
        if not self.merchants:
            raise ValueError("商家列表为空，请先调用 load_merchants() 或 set_merchants()")
        
        results = []
        errors = []
        
        for idx, (account, uuid_str, _) in enumerate(self.merchants, 1):
            dataset = f"shop{uuid_str}"
            
            # 格式化 SQL
            sql = sql_template.format(
                project=self.project_id,
                dataset=dataset,
                start_ts=start_ts,
                end_ts=end_ts
            )
            
            try:
                rows = self.query(sql)
                
                if rows:
                    if row_processor:
                        data = row_processor(rows[0], account)
                    else:
                        data = dict(rows[0])
                        data["_account"] = account
                        data["_dataset"] = dataset
                    
                    results.append(data)
                    
                    # 进度回调
                    if self.progress_callback:
                        self.progress_callback(idx, len(self.merchants), account, data)
                else:
                    errors.append({
                        "index": idx,
                        "account": account,
                        "dataset": dataset,
                        "error": "无数据"
                    })
                    
            except Exception as e:
                errors.append({
                    "index": idx,
                    "account": account,
                    "dataset": dataset,
                    "error": str(e)
                })
        
        # 执行校验
        validation_result = None
        if self.validators and results:
            validation_result = self.validate(results)

        # 写入 Excel (gate 在 write_excel 内部执行)
        output_file = self.write_excel(results)
        
        return ExportResult(
            total_count=len(self.merchants),
            success_count=len(results),
            failed_count=len(errors),
            output_path=str(output_file),
            validation_result=validation_result,
            errors=errors
        )


class TakeoutRevenueExporter(MultiShopExporter):
    """
    外卖营业额导出器（专门用于华莱士外卖报表）
    
    预置外卖统计逻辑和校验规则
    """
    
    # 外卖营业额统计 SQL 模板
    SQL_TEMPLATE = """
WITH 
-- 1. POS侧已完成账单
finished_bills AS (
  SELECT
    sb.uuid AS bill_uuid,
    sb.amount AS bill_amount,
    sb.payment_amount AS bill_payment_amount,
    sb.bill_type,
    sb.order_source_uuid,
    sb.order_source_name
  FROM `{project}`.`{dataset}`.`ttpos_sale_bill` AS sb
  WHERE sb.delete_time = 0
    AND sb.status = 1
    AND sb.finish_time >= {start_ts}
    AND sb.finish_time < {end_ts}
),

-- 2. Grab/Lineman渠道的order_source
grab_lineman_sources AS (
  SELECT os.uuid AS source_uuid
  FROM `{project}`.`{dataset}`.`ttpos_order_source` AS os
  LEFT JOIN `{project}`.`{dataset}`.`ttpos_multi_language_name` AS mln
    ON mln.uuid = os.multi_language_name_uuid AND mln.delete_time = 0
  WHERE os.delete_time = 0
    AND REGEXP_CONTAINS(
      LOWER(REPLACE(CONCAT(IFNULL(mln.zh_name,''),IFNULL(mln.th_name,''),IFNULL(mln.en_name,'')), ' ', '')),
      r'grab|lineman'
    )
),

-- 3. 标记外卖订单（3个条件任一即计入）
bill_takeout AS (
  SELECT
    fb.bill_uuid,
    fb.bill_amount,
    fb.bill_payment_amount,
    (
      -- 条件1: 支付方式 Robinhood/Grab/Lineman/Shopee
      EXISTS (
        SELECT 1 FROM `{project}`.`{dataset}`.`ttpos_payment_order` po
        JOIN `{project}`.`{dataset}`.`ttpos_sale_order` so 
          ON so.uuid = po.related_uuid AND po.related_type = 0 AND so.delete_time = 0
        WHERE so.sale_bill_uuid = fb.bill_uuid AND po.delete_time = 0 AND po.status = 1
          AND REGEXP_CONTAINS(LOWER(REPLACE(po.payment_method_name, ' ', '')), r'robinhood|grab|lineman|shopee')
      )
      -- 条件2: 订单来源 Grab/Lineman
      OR (fb.order_source_uuid > 0 AND fb.order_source_uuid IN (SELECT source_uuid FROM grab_lineman_sources))
      OR REGEXP_CONTAINS(LOWER(REPLACE(CONCAT(
          IFNULL(JSON_EXTRACT_SCALAR(fb.order_source_name, '$.zh'), ''),
          IFNULL(JSON_EXTRACT_SCALAR(fb.order_source_name, '$.th'), ''),
          IFNULL(JSON_EXTRACT_SCALAR(fb.order_source_name, '$.en'), '')), ' ', '')), r'grab|lineman')
      -- 条件3: 会员外送
      OR fb.bill_type = 2
    ) AS is_takeout
  FROM finished_bills fb
),

-- 4. POS侧汇总
pos_summary AS (
  SELECT
    ROUND(SUM(bill_amount), 2) AS pos_turnover,
    ROUND(SUM(bill_payment_amount), 2) AS pos_received,
    ROUND(SUM(IF(is_takeout, bill_amount, 0)), 2) AS pos_takeout_turnover,
    ROUND(SUM(IF(is_takeout, bill_payment_amount, 0)), 2) AS pos_takeout_received
  FROM bill_takeout
),

-- 5. 外卖平台订单汇总
takeout_summary AS (
  SELECT
    ROUND(SUM(subtotal), 2) AS platform_turnover,
    ROUND(SUM(platform_total), 2) AS platform_received
  FROM `{project}`.`{dataset}`.`ttpos_takeout_order`
  WHERE delete_time = 0 AND order_state = 40
    AND platform IN ('grab', 'lineman', 'shopee')
    AND completed_time >= {start_ts} AND completed_time < {end_ts}
)

-- 6. 最终结果
SELECT
  c.name AS store_name,
  IFNULL(ps.pos_turnover, 0) AS total_turnover,
  IFNULL(ps.pos_received, 0) AS total_received,
  IFNULL(ps.pos_takeout_turnover, 0) + IFNULL(ts.platform_turnover, 0) AS takeout_turnover,
  IFNULL(ps.pos_takeout_received, 0) + IFNULL(ts.platform_received, 0) AS takeout_received
FROM pos_summary ps
CROSS JOIN takeout_summary ts
CROSS JOIN `{project}`.`{dataset}`.`ttpos_company` c
WHERE c.delete_time = 0
LIMIT 1
"""
    
    def __init__(
        self,
        project_id: str = "diyl-407103",
        output_path: Union[str, Path] = "exports/takeout_revenue.xlsx",
        client: Optional[bigquery.Client] = None
    ):
        super().__init__(project_id, output_path, client)
        
        # 设置默认 Excel 配置
        self.set_excel_config(ExcelConfig(
            sheet_name="外卖营业额统计",
            headers=["门店编号", "门店名称", "总营业额", "实收金额", 
                    "外卖营业额", "外卖实收金额", "非外卖营业额", "非外卖实收金额"],
            number_format="0.00"
        ))
    
    def export_takeout_revenue(
        self,
        start_date: str,
        end_date: str,
        merchant_xlsx: Union[str, Path]
    ) -> ExportResult:
        """
        导出外卖营业额报表
        
        Args:
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)
            merchant_xlsx: 商家列表 Excel 路径
        """
        from datetime import datetime, timezone
        from utils.validators import create_default_validators
        
        # 加载商家列表
        self.load_merchants(merchant_xlsx)
        
        # 转换时间戳
        start_ts = int(datetime.strptime(start_date, "%Y-%m-%d")
                      .replace(tzinfo=timezone.utc).timestamp())
        end_ts = int(datetime.strptime(end_date, "%Y-%m-%d")
                    .replace(tzinfo=timezone.utc).timestamp())
        
        # 设置默认校验器
        self.set_validators(create_default_validators())
        
        # 行处理器：添加非外卖计算和门店编号
        def process_row(row, account):
            total = float(row.total_turnover or 0)
            received = float(row.total_received or 0)
            takeout = float(row.takeout_turnover or 0)
            takeout_recv = float(row.takeout_received or 0)
            
            return {
                "门店名称": row.store_name or "",
                "总营业额": round(total, 2),
                "实收金额": round(received, 2),
                "外卖营业额": round(takeout, 2),
                "外卖实收金额": round(takeout_recv, 2),
                "非外卖营业额": round(total - takeout, 2),
                "非外卖实收金额": round(received - takeout_recv, 2),
                "_raw": row
            }
        
        # 执行导出
        result = self.export(
            sql_template=self.SQL_TEMPLATE,
            start_ts=start_ts,
            end_ts=end_ts,
            row_processor=process_row
        )
        
        # 添加门店编号
        if result.success_count > 0:
            from openpyxl import load_workbook
            wb = load_workbook(result.output_path)
            ws = wb.active
            
            for idx in range(2, ws.max_row + 1):
                ws.cell(row=idx, column=1, value=idx - 1)
            
            wb.save(result.output_path)
        
        return result


class ReportExporter(MultiShopExporter):
    """
    通用报表导出器（适配 n8n-scheduler 脚本需求）
    
    提供 n8n-scheduler 中常见报表类型的 BQ 版本实现
    """
    
    def __init__(
        self,
        project_id: str = "diyl-407103",
        output_path: Union[str, Path] = "exports/report.xlsx",
        client: Optional[bigquery.Client] = None
    ):
        super().__init__(project_id, output_path, client)
    
    def _to_timestamp(self, date_str: str, end_of_day: bool = False) -> int:
        """将日期字符串转换为 Unix 时间戳（UTC）"""
        from datetime import datetime, timezone, timedelta
        
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        if end_of_day:
            dt = dt.replace(hour=23, minute=59, second=59)
        else:
            dt = dt.replace(hour=0, minute=0, second=0)
        
        return int(dt.replace(tzinfo=timezone.utc).timestamp())
    
    def export_sales_and_consumption(
        self,
        month: str,
        merchant_xlsx: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None,
        external_spec: Optional[str] = None,
    ) -> ExportResult:
        """
        销售业绩 + 物品消耗报表（对应 report.sh）

        Args:
            month: 月份 (YYYY-MM)
            merchant_xlsx: 商家列表 Excel 路径
            output_path: 可选，自定义输出路径
            external_spec: 可选，外部销售源规格，如 'huku:path=/path/to/file.xlsx'

        Returns:
            ExportResult 包含两个 sheet：
            - 销售业绩: 总营业额、实收金额、订单数
            - 物品消耗: 各原料消耗明细
        """
        from .sql_templates import get_template

        # 计算月份时间范围（按 BKK 时区，跟 ttpos 业务对齐）
        from datetime import datetime, timezone, timedelta
        BKK = timezone(timedelta(hours=7))
        year, mon = int(month[:4]), int(month[5:7])
        start_ts = int(datetime(year, mon, 1, tzinfo=BKK).timestamp())

        # 计算下月1日
        if mon == 12:
            end_year, end_mon = year + 1, 1
        else:
            end_year, end_mon = year, mon + 1
        end_ts = int(datetime(end_year, end_mon, 1, tzinfo=BKK).timestamp())
        
        # 加载商家列表
        self.load_merchants(merchant_xlsx)
        
        sales_results = []
        # 物品消耗：(店编号, material_uuid) → {仓库出库, BOM 推算}
        consumption_map: Dict[Tuple[str, int], Dict[str, Any]] = {}
        # 商品销量原始行（按规格拆，含 package_uuid/bom_uuid，待 Python 层后处理决定是否显示规格）
        bom_sales_raw: List[Dict[str, Any]] = []
        # 支付方式明细（按门店 × 支付方式 × 渠道 拆开）
        payment_results: List[Dict[str, Any]] = []
        # 支付方式明细 - 按天（按门店 × 日期 × 支付方式 × 渠道 拆开，给客户做日报对账）
        payment_daily_results: List[Dict[str, Any]] = []
        errors = []

        sales_sql = get_template('comprehensive_sales')
        consumption_sql = get_template('material_consumption')
        bom_consumption_sql = get_template('material_bom_consumption')
        bom_sales_sql = get_template('bom_product_sales')
        payment_sql = get_template('payment_breakdown')
        payment_daily_sql = get_template('payment_breakdown_daily')

        # --- order_excludes: 加载报表层排除清单（与利润表对齐） ---
        dine_excludes: set[int] = set()
        takeout_excludes: set[int] = set()
        _order_excludes_cfg = getattr(self, '_config', None) or {}
        if not _order_excludes_cfg:
            import yaml
            cfg_path = Path("resources/config.yaml")
            if cfg_path.exists():
                with open(cfg_path, encoding="utf-8") as f:
                    _order_excludes_cfg = yaml.safe_load(f) or {}
        _oe_spec = _order_excludes_cfg.get("order_excludes") or {}
        _oe_path = _oe_spec.get("path")
        if _oe_path and Path(_oe_path).exists():
            import csv as _csv
            with open(_oe_path, encoding="utf-8-sig") as f:
                for r in _csv.DictReader(f):
                    ch = (r.get("channel") or "").strip()
                    key = (r.get("exclude_key") or "").strip()
                    if not key:
                        continue
                    try:
                        k = int(key)
                    except ValueError:
                        continue
                    if ch == "dine":
                        dine_excludes.add(k)
                    elif ch == "takeout":
                        takeout_excludes.add(k)
            print(f"[Order Excludes] 加载 {len(dine_excludes)} 单堂食排除 / {len(takeout_excludes)} 单外卖排除")

        def _build_exclude_orders_clause(uuids: set[int], field: str) -> str:
            if not uuids:
                return ""
            ids_str = ", ".join(str(u) for u in sorted(uuids))
            return f"AND {field} NOT IN ({ids_str})"

        # 探测哪些店有 ttpos_business_status_period 表（测试营业时段需要排除，对齐 ttpos UI）
        # 一次 UNION ALL 查 INFORMATION_SCHEMA，避免 53 次串行 RTT
        union_parts = [
            f"SELECT '{uuid_str}' AS uuid, COUNT(*) AS has_table "
            f"FROM `{self.project_id}.shop{uuid_str}.INFORMATION_SCHEMA.TABLES` "
            f"WHERE table_name = 'ttpos_business_status_period'"
            for _, uuid_str, _ in self.merchants
        ]
        check_sql = " UNION ALL ".join(union_parts)
        has_test_period_table = {
            r.uuid for r in self.client.query(check_sql).result() if r.has_table > 0
        }

        def _build_exclude_test_business(uuid_str: str, field: str) -> str:
            """ttpos ExcludeTestBusinessByBillSQL 的 BQ 等价：测试时段内创建的 bill 不算入统计。

            field: 调用方指定字段名（statistics_sale 用 'sale_bill_uuid'，sale_bill 直接用 'uuid'）
            """
            if uuid_str not in has_test_period_table:
                return ""
            return (
                f"AND {field} NOT IN ("
                f"  SELECT _sb.uuid "
                f"  FROM `{self.project_id}.shop{uuid_str}.ttpos_sale_bill` _sb "
                f"  JOIN `{self.project_id}.shop{uuid_str}.ttpos_business_status_period` _bsp "
                f"    ON _bsp.delete_time = 0 "
                f"    AND _sb.create_time >= _bsp.start_time "
                f"    AND (_bsp.end_time = 0 OR _sb.create_time <= _bsp.end_time) "
                f"  WHERE _sb.delete_time = 0"
                f")"
            )

        from concurrent.futures import ThreadPoolExecutor, as_completed
        from threading import Lock
        merge_lock = Lock()

        def _query_sales(account, uuid_str):
            dataset = f"shop{uuid_str}"
            _dine_clause = _build_exclude_orders_clause(dine_excludes, 'sale_order_uuid')
            _dine_bill_clause = ""
            if dine_excludes:
                ids_str = ", ".join(str(u) for u in sorted(dine_excludes))
                _dine_bill_clause = (
                    f"AND sb.uuid NOT IN ("
                    f"  SELECT _so.sale_bill_uuid "
                    f"  FROM `{self.project_id}.{dataset}.ttpos_sale_order` _so"
                    f"  WHERE _so.uuid IN ({ids_str}) AND _so.delete_time = 0"
                    f")"
                )
            sql = sales_sql.format(
                project=self.project_id, dataset=dataset,
                start_ts=start_ts, end_ts=end_ts,
                exclude_test_business_ss=_build_exclude_test_business(uuid_str, 'sale_bill_uuid'),
                exclude_test_business_sb=_build_exclude_test_business(uuid_str, 'uuid'),
                exclude_dine_orders=_dine_clause,
                exclude_takeout_orders=_build_exclude_orders_clause(takeout_excludes, 'uuid'),
                exclude_dine_sale_bills=_dine_bill_clause,
            )
            rows = list(self.client.query(sql).result())
            if not rows:
                return account, None
            row = rows[0]
            store_name = (row.store_name or "").replace("\r", "").replace("\n", " ").strip()
            return account, {
                "门店编号": row.store_code or "",
                "门店名称": store_name,
                "总营业额": float(row.total_turnover or 0),
                "实收金额": float(row.total_received or 0),
                "订单数": int(row.total_orders or 0),
            }

        def _clean_consumption(s):
            if not s:
                return ""
            return s.replace("\r", "").replace("\n", " ").strip()

        def _query_consumption_warehouse(account, uuid_str):
            dataset = f"shop{uuid_str}"
            sql = consumption_sql.format(project=self.project_id, dataset=dataset,
                                          start_ts=start_ts, end_ts=end_ts)
            rows = list(self.client.query(sql).result())
            with merge_lock:
                for row in rows:
                    key = (row.store_code or "", int(row.material_uuid))
                    rec = consumption_map.get(key)
                    if rec is None:
                        rec = {
                            "门店编号": row.store_code or "",
                            "门店名称": _clean_consumption(row.store_name),
                            "物品名称": _clean_consumption(row.material_name),
                            "消耗量(仓库出库)": 0.0,
                            "消耗量(BOM推算)": 0.0,
                            "单位": _clean_consumption(row.unit_name),
                        }
                        consumption_map[key] = rec
                    rec["消耗量(仓库出库)"] = float(row.total_num or 0)
                    if not rec["物品名称"]:
                        rec["物品名称"] = _clean_consumption(row.material_name)
                    if not rec["单位"]:
                        rec["单位"] = _clean_consumption(row.unit_name)
            return account

        def _query_consumption_bom(account, uuid_str):
            dataset = f"shop{uuid_str}"
            sql = bom_consumption_sql.format(project=self.project_id, dataset=dataset,
                                              start_ts=start_ts, end_ts=end_ts)
            rows = list(self.client.query(sql).result())
            with merge_lock:
                for row in rows:
                    key = (row.store_code or "", int(row.material_uuid))
                    rec = consumption_map.get(key)
                    if rec is None:
                        rec = {
                            "门店编号": row.store_code or "",
                            "门店名称": _clean_consumption(row.store_name),
                            "物品名称": _clean_consumption(row.material_name),
                            "消耗量(仓库出库)": 0.0,
                            "消耗量(BOM推算)": 0.0,
                            "单位": _clean_consumption(row.unit_name),
                        }
                        consumption_map[key] = rec
                    rec["消耗量(BOM推算)"] = float(row.total_num or 0)
                    if not rec["物品名称"]:
                        rec["物品名称"] = _clean_consumption(row.material_name)
                    if not rec["单位"]:
                        rec["单位"] = _clean_consumption(row.unit_name)
            return account

        def _clean(s: Optional[str]) -> str:
            if not s:
                return ""
            return s.replace("\r", "").replace("\n", " ").strip()

        def _query_bom_sales(account, uuid_str):
            dataset = f"shop{uuid_str}"
            sql = bom_sales_sql.format(project=self.project_id, dataset=dataset,
                                        start_ts=start_ts, end_ts=end_ts)
            rows = list(self.client.query(sql).result())
            return account, [{
                "门店编号": row.store_code or "",
                "门店名称": _clean(row.store_name),
                "package_uuid": int(row.product_package_uuid or 0),
                "bom_uuid": int(row.product_bom_uuid or 0),
                "package_name": _clean(row.package_name),
                "spec_name": _clean(row.spec_name),
                "bom_type": int(row.bom_type or 0),
                "销量": float(row.total_qty or 0),
                "has_bom": int(row.has_bom or 0),
            } for row in rows]

        def _query_payment(account, uuid_str):
            dataset = f"shop{uuid_str}"
            sql = payment_sql.format(
                project=self.project_id, dataset=dataset,
                start_ts=start_ts, end_ts=end_ts,
                exclude_test_business_ss=_build_exclude_test_business(uuid_str, 'sale_bill_uuid'),
            )
            rows = list(self.client.query(sql).result())
            return account, [{
                "门店编号": row.store_code or "",
                "门店名称": _clean(row.store_name),
                "支付方式": _clean(row.method_name),
                "渠道": row.channel or "",
                "笔数": int(row.bill_cnt or 0),
                "收款总额": float(row.total_amount or 0),
                "退款金额": float(row.total_refund or 0),
                "净收金额": float(row.net_amount or 0),
            } for row in rows]

        def _query_payment_daily(account, uuid_str):
            dataset = f"shop{uuid_str}"
            sql = payment_daily_sql.format(
                project=self.project_id, dataset=dataset,
                start_ts=start_ts, end_ts=end_ts,
                exclude_test_business_ss=_build_exclude_test_business(uuid_str, 'sale_bill_uuid'),
            )
            rows = list(self.client.query(sql).result())
            return account, [{
                "门店编号": row.store_code or "",
                "门店名称": _clean(row.store_name),
                "日期": row.pay_date or "",
                "支付方式": _clean(row.method_name),
                "渠道": row.channel or "",
                "笔数": int(row.bill_cnt or 0),
                "收款总额": float(row.total_amount or 0),
                "退款金额": float(row.total_refund or 0),
                "净收金额": float(row.net_amount or 0),
            } for row in rows]

        with ThreadPoolExecutor(max_workers=10) as ex:
            sales_futures = {
                ex.submit(_query_sales, account, uuid_str): account
                for account, uuid_str, _ in self.merchants
            }
            cons_wh_futures = {
                ex.submit(_query_consumption_warehouse, account, uuid_str): account
                for account, uuid_str, _ in self.merchants
            }
            cons_bom_futures = {
                ex.submit(_query_consumption_bom, account, uuid_str): account
                for account, uuid_str, _ in self.merchants
            }
            bom_sales_futures = {
                ex.submit(_query_bom_sales, account, uuid_str): account
                for account, uuid_str, _ in self.merchants
            }
            payment_futures = {
                ex.submit(_query_payment, account, uuid_str): account
                for account, uuid_str, _ in self.merchants
            }
            payment_daily_futures = {
                ex.submit(_query_payment_daily, account, uuid_str): account
                for account, uuid_str, _ in self.merchants
            }

            for fut in as_completed(sales_futures):
                account = sales_futures[fut]
                try:
                    _, data = fut.result()
                    if data:
                        sales_results.append(data)
                except Exception as e:
                    errors.append({"account": account, "type": "sales", "error": str(e)})

            for fut in as_completed(cons_wh_futures):
                account = cons_wh_futures[fut]
                try:
                    fut.result()
                except Exception as e:
                    errors.append({"account": account, "type": "consumption_wh", "error": str(e)})

            for fut in as_completed(cons_bom_futures):
                account = cons_bom_futures[fut]
                try:
                    fut.result()
                except Exception as e:
                    errors.append({"account": account, "type": "consumption_bom", "error": str(e)})

            for fut in as_completed(bom_sales_futures):
                account = bom_sales_futures[fut]
                try:
                    _, items = fut.result()
                    bom_sales_raw.extend(items)
                except Exception as e:
                    errors.append({"account": account, "type": "bom_sales", "error": str(e)})

            for fut in as_completed(payment_futures):
                account = payment_futures[fut]
                try:
                    _, items = fut.result()
                    payment_results.extend(items)
                except Exception as e:
                    errors.append({"account": account, "type": "payment", "error": str(e)})

            for fut in as_completed(payment_daily_futures):
                account = payment_daily_futures[fut]
                try:
                    _, items = fut.result()
                    payment_daily_results.extend(items)
                except Exception as e:
                    errors.append({"account": account, "type": "payment_daily", "error": str(e)})

        # ── 外部销售源（可选）：加载并合并到销售业绩 ──
        if external_spec:
            from external_sales import load_external
            print(f"\n[External] 加载外部源: {external_spec}")
            ext_rows = load_external(
                external_spec,
                bq_client=self.client,
                merchants=[(acc, uuid_str, store_num, "") for acc, uuid_str, store_num in self.merchants],
                month=month,
            )
            # 按门店聚合（外部源是 SKU 级，需要 rollup 到店级）
            ext_by_store: Dict[str, Dict[str, Any]] = {}
            for r in ext_rows:
                code = str(r.store_num).zfill(3)
                slot = ext_by_store.setdefault(code, {
                    "store_name": r.store_name,
                    "turnover": 0.0,
                    "received": 0.0,
                })
                # 外部源没有 gross_amount/标准金额概念，用 revenue 近似营业额
                slot["turnover"] += r.revenue
                slot["received"] += r.revenue
            print(f"[External] 合并 {len(ext_by_store)} 店, "
                  f"实收合计 {sum(d['received'] for d in ext_by_store.values()):,.2f}")
            # 合并到 sales_results
            sales_by_code = {r["门店编号"]: r for r in sales_results}
            for code, data in ext_by_store.items():
                if code in sales_by_code:
                    rec = sales_by_code[code]
                    rec["总营业额"] = round(rec.get("总营业额", 0) + data["turnover"], 2)
                    rec["实收金额"] = round(rec.get("实收金额", 0) + data["received"], 2)
                    # 订单数保持 BQ 值（外部源没有订单数维度）
                else:
                    sales_results.append({
                        "门店编号": code,
                        "门店名称": data["store_name"] or code,
                        "总营业额": round(data["turnover"], 2),
                        "实收金额": round(data["received"], 2),
                        "订单数": 0,  # 外部源无订单数
                    })

        # ===== Python 层后处理 =====
        # 1. 商品销量：决定每行是否显示规格后缀（同 store + 同 package 有多个 bom_uuid 才带规格）
        bom_yes_results, bom_no_results = self._build_bom_sales_rows(bom_sales_raw)

        # 2. 物品消耗：dict → list，并按 (店, 仓库出库量降序) 排序
        consumption_results = list(consumption_map.values())

        # 3. 支付方式明细：补"未分类"行（兜底 ttpos statistics_payment 异步生成失败的几单）
        # 销售业绩 sheet 实收用 sale_bill 兜底（事实数据源），但 sale_bill 没有 method_uuid → 那几单
        # 在支付明细 sheet 找不到归属。用 (sales 实收 - SUM(payment 净收)) 算差额，差 > 0.5 元就补 1 行。
        sales_received_by_store: Dict[str, Tuple[str, float]] = {
            r["门店编号"]: (r["门店名称"], r["实收金额"]) for r in sales_results
        }
        payment_net_by_store: Dict[str, float] = {}
        for r in payment_results:
            payment_net_by_store[r["门店编号"]] = (
                payment_net_by_store.get(r["门店编号"], 0) + r["净收金额"]
            )
        for code, (store_name, sales_recv) in sales_received_by_store.items():
            unclassified = round(sales_recv - payment_net_by_store.get(code, 0), 2)
            if unclassified > 0.5:
                payment_results.append({
                    "门店编号": code,
                    "门店名称": store_name,
                    "支付方式": "未分类(系统数据流异常)",
                    "渠道": "POS",
                    "笔数": 0,
                    "收款总额": unclassified,
                    "退款金额": 0.0,
                    "净收金额": unclassified,
                })

        # 写入 Excel（4 sheet，xlsxwriter 流式写）
        output_path = output_path or self.output_path
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        def _store_key(rec):
            code = rec.get("门店编号", "")
            try:
                return (0, int(code))
            except (TypeError, ValueError):
                return (1, str(code))

        sales_results.sort(key=_store_key)
        consumption_results.sort(key=lambda r: (_store_key(r), -float(r.get("消耗量(仓库出库)", 0) or 0)))
        bom_yes_results.sort(key=lambda r: (_store_key(r), -float(r.get("销量", 0) or 0)))
        bom_no_results.sort(key=lambda r: (_store_key(r), -float(r.get("销量", 0) or 0)))
        # 支付方式：店内按 净收金额 降序（财务最关心总收，大的在前）
        payment_results.sort(key=lambda r: (_store_key(r), -float(r.get("净收金额", 0) or 0)))
        # 日报：店内按 (日期升序, 净收降序) — 财务按时间线对账
        payment_daily_results.sort(
            key=lambda r: (_store_key(r), r.get("日期", ""), -float(r.get("净收金额", 0) or 0))
        )

        import xlsxwriter
        wb = xlsxwriter.Workbook(str(output_path))
        try:
            self._write_data_sheet_xw(wb, "销售业绩", sales_results,
                ["门店编号", "门店名称", "总营业额", "实收金额", "订单数"])
            self._write_data_sheet_xw(wb, "物品消耗", consumption_results,
                ["门店编号", "门店名称", "物品名称", "消耗量(仓库出库)", "消耗量(BOM推算)", "单位"])
            self._write_data_sheet_xw(wb, "已设置BOM商品销量", bom_yes_results,
                ["门店编号", "门店名称", "商品名称", "销量"])
            self._write_data_sheet_xw(wb, "未设置BOM商品销量", bom_no_results,
                ["门店编号", "门店名称", "商品名称", "销量"])
            self._write_data_sheet_xw(wb, "支付方式明细", payment_results,
                ["门店编号", "门店名称", "支付方式", "渠道", "笔数", "收款总额", "退款金额", "净收金额"])
            self._write_data_sheet_xw(wb, "支付方式明细(按天)", payment_daily_results,
                ["门店编号", "门店名称", "日期", "支付方式", "渠道", "笔数", "收款总额", "退款金额", "净收金额"])

            # 零容差闸门 (set_gate 设置后执行; 主表 sales_results 为闸门输入)
            if self.gate_spec is not None:
                outcome = self.gate_spec.run(sales_results)
                if outcome.needs_watermark:
                    add_watermark_sheet_xlsxwriter(wb, outcome.watermark_lines())
        finally:
            wb.close()
        
        return ExportResult(
            total_count=len(self.merchants),
            success_count=len(sales_results),
            failed_count=len(errors),
            output_path=str(output_path),
            errors=errors
        )
    
    def _write_data_to_sheet(self, ws, data: List[Dict], headers: List[str]):
        """辅助方法（保留 openpyxl 旧路径，给未迁移的 export 函数用）"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "0.00"
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(headers) + 1):
            max_len = 0
            for r in range(1, len(data) + 2):
                val = ws.cell(row=r, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    @staticmethod
    def _build_bom_sales_rows(raw_rows: List[Dict[str, Any]]) -> Tuple[List[Dict], List[Dict]]:
        """
        商品销量后处理：把按 (package_uuid, bom_uuid) 拆出的细行聚合到用户视角的"商品名"。

        - 同店 + 同 package_uuid 下若有多个 bom_uuid (规格行) → 显示 "包名(规格名)"
        - 否则只显示 "包名"
        - 同店 + 同显示名 → **合并销量**（覆盖以下 3 种情况）：
            (a) 同 package 的堂食(有 bom_uuid) + 外卖(bom_uuid=0)
            (b) 商家重复创建同名 package（不同 uuid，软删/活跃混合）
            (c) 同 package 同显示名的多个 bom_uuid（如规格名都是 "pc"）
        - has_bom 判定：组内任意一行 has_bom=1 → 整组归到"已设置"

        返回 (已设置 BOM 行, 未设置 BOM 行)
        """
        from collections import defaultdict

        # (店编号, package_uuid) → set(堂食 bom_uuid)
        # 外卖行 bom_uuid=0 不算"规格"维度，避免误判 multi_spec
        spec_groups: Dict[Tuple[str, int], set] = defaultdict(set)
        for r in raw_rows:
            if r["bom_uuid"] > 0:
                spec_groups[(r["门店编号"], r["package_uuid"])].add(r["bom_uuid"])

        # 第一遍：算每行的显示名 + 聚合
        # key: (店编号, 店名, 显示商品名) → {qty, has_bom}
        agg: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
        for r in raw_rows:
            multi_spec = len(spec_groups[(r["门店编号"], r["package_uuid"])]) > 1
            spec = r["spec_name"]
            if multi_spec and spec:
                name = f"{r['package_name']}({spec})"
            else:
                name = r["package_name"]
            key = (r["门店编号"], r["门店名称"], name)
            cur = agg.setdefault(key, {"qty": 0.0, "has_bom": 0})
            cur["qty"] += r["销量"]
            if r["has_bom"] == 1:
                cur["has_bom"] = 1

        yes_rows: List[Dict[str, Any]] = []
        no_rows: List[Dict[str, Any]] = []
        for (code, store_name, name), v in agg.items():
            rec = {
                "门店编号": code,
                "门店名称": store_name,
                "商品名称": name,
                "销量": v["qty"],
            }
            (yes_rows if v["has_bom"] == 1 else no_rows).append(rec)
        return yes_rows, no_rows

    def _write_data_sheet_xw(self, workbook, sheet_name: str,
                              data: List[Dict], headers: List[str]):
        """xlsxwriter 版本：流式写入，比 openpyxl 快 5-10×。"""
        ws = workbook.add_worksheet(sheet_name)

        # 预创建格式（缓存）
        hdr_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter',
            'border': 1, 'border_color': '#D9D9D9',
        })
        text_fmt = workbook.add_format({'border': 1, 'border_color': '#D9D9D9', 'valign': 'vcenter'})
        num_fmt = workbook.add_format({
            'border': 1, 'border_color': '#D9D9D9', 'valign': 'vcenter',
            'align': 'right', 'num_format': '0.00',
        })

        # 表头
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header, hdr_fmt)

        # 列宽（先算后 set_column；写完不能改）
        for col_idx, header in enumerate(headers):
            max_len = len(str(header))
            for row in data:
                v = row.get(header, "")
                if v is not None:
                    s = str(v)
                    if len(s) > max_len:
                        max_len = len(s)
            ws.set_column(col_idx, col_idx, min(max_len + 4, 50))

        # 数据
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, header in enumerate(headers):
                value = row_data.get(header, "")
                if isinstance(value, (int, float)):
                    ws.write_number(row_idx, col_idx, value, num_fmt)
                elif value is None or value == "":
                    ws.write_blank(row_idx, col_idx, None, text_fmt)
                else:
                    ws.write(row_idx, col_idx, value, text_fmt)
    
    def export_bom_sales(
        self,
        month: str,
        merchant_xlsx: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None
    ) -> ExportResult:
        """
        BOM商品销量报表（对应 report_bom_sales.sh）
        
        区分已设置BOM和未设置BOM的商品销量
        
        Returns:
            ExportResult 包含两个 sheet：
            - 已设置BOM
            - 未设置BOM
        """
        from .sql_templates import get_template
        from datetime import datetime, timezone
        
        # 计算月份时间范围
        year, mon = int(month[:4]), int(month[5:7])
        start_ts = int(datetime(year, mon, 1, tzinfo=timezone.utc).timestamp())
        if mon == 12:
            end_year, end_mon = year + 1, 1
        else:
            end_year, end_mon = year, mon + 1
        end_ts = int(datetime(end_year, end_mon, 1, tzinfo=timezone.utc).timestamp())
        
        self.load_merchants(merchant_xlsx)
        
        has_bom_data = []
        no_bom_data = []
        errors = []
        
        sql = get_template('bom_product_sales')
        
        for idx, (account, uuid_str, _) in enumerate(self.merchants, 1):
            dataset = f"shop{uuid_str}"
            
            try:
                query = sql.format(
                    project=self.project_id,
                    dataset=dataset,
                    start_ts=start_ts,
                    end_ts=end_ts
                )
                rows = list(self.client.query(query).result())
                
                for row in rows:
                    package_name = row.package_name or ""
                    spec_name = row.spec_name or ""
                    record = {
                        "账号": account,
                        "商品名称": f"{package_name}({spec_name})" if spec_name else package_name,
                        "分类": row.category_name or "",
                        "销量": float(row.total_qty or 0)
                    }
                    if row.has_bom:
                        has_bom_data.append(record)
                    else:
                        no_bom_data.append(record)
                        
            except Exception as e:
                errors.append({"account": account, "error": str(e)})
        
        # 写入 Excel
        output_path = output_path or self.output_path
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "已设置BOM"
        self._write_data_to_sheet(ws1, has_bom_data, ["账号", "商品名称", "分类", "销量"])
        
        ws2 = wb.create_sheet(title="未设置BOM")
        self._write_data_to_sheet(ws2, no_bom_data, ["账号", "商品名称", "分类", "销量"])

        # 零容差闸门 (set_gate 设置后执行; has_bom_data + no_bom_data 合并作为闸门输入)
        if self.gate_spec is not None:
            _all_bom_rows = has_bom_data + no_bom_data
            outcome = self.gate_spec.run(_all_bom_rows)
            if outcome.needs_watermark:
                add_watermark_sheet_openpyxl(wb, outcome.watermark_lines())

        wb.save(output_path)

        return ExportResult(
            total_count=len(self.merchants),
            success_count=len(self.merchants) - len(errors),
            failed_count=len(errors),
            output_path=str(output_path),
            errors=errors
        )
    
    def export_material_consumption_statistics(
        self,
        month: str,
        material_codes: List[str],
        merchant_xlsx: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None
    ) -> ExportResult:
        """
        原料经营明细（对应 report_item_consumption_statistics.sh）
        
        指定原料的：BOM消耗量、涉及该原料的销售金额、采购入库量
        
        Args:
            month: 月份 (YYYY-MM)
            material_codes: 原料 code 列表，如 ['flour', 'popcorn_chicken', 'whole_chicken']
            merchant_xlsx: 商家列表 Excel 路径
            output_path: 可选，自定义输出路径
            
        Returns:
            ExportResult
        """
        from .sql_templates import get_template
        from datetime import datetime, timezone
        
        # 计算月份时间范围
        year, mon = int(month[:4]), int(month[5:7])
        start_ts = int(datetime(year, mon, 1, tzinfo=timezone.utc).timestamp())
        if mon == 12:
            end_year, end_mon = year + 1, 1
        else:
            end_year, end_mon = year, mon + 1
        end_ts = int(datetime(end_year, end_mon, 1, tzinfo=timezone.utc).timestamp())
        
        self.load_merchants(merchant_xlsx)
        
        results = []
        errors = []
        
        consumption_sql = get_template('specific_material_consumption')
        sales_sql = get_template('material_related_sales')
        purchase_sql = get_template('purchase_in')
        
        # 构建 code 字符串
        codes_str = ", ".join(f"'{c}'" for c in material_codes)
        
        for idx, (account, uuid_str, _) in enumerate(self.merchants, 1):
            dataset = f"shop{uuid_str}"
            
            try:
                # BOM消耗量
                query = consumption_sql.format(
                    project=self.project_id,
                    dataset=dataset,
                    start_ts=start_ts,
                    end_ts=end_ts,
                    material_codes=codes_str
                )
                consumption_rows = list(self.client.query(query).result())
                consumption = float(consumption_rows[0].total_consumption or 0) if consumption_rows else 0
                
                # 销售金额
                query = sales_sql.format(
                    project=self.project_id,
                    dataset=dataset,
                    start_ts=start_ts,
                    end_ts=end_ts,
                    material_codes=codes_str
                )
                sales_rows = list(self.client.query(query).result())
                sales = float(sales_rows[0].total_sales or 0) if sales_rows else 0
                
                # 采购入库量
                query = purchase_sql.format(
                    project=self.project_id,
                    dataset=dataset,
                    start_ts=start_ts,
                    end_ts=end_ts,
                    material_codes=codes_str
                )
                purchase_rows = list(self.client.query(query).result())
                purchase = float(purchase_rows[0].total_in or 0) if purchase_rows else 0
                
                results.append({
                    "账号": account,
                    "BOM消耗量": round(consumption, 2),
                    "涉及销售金额": round(sales, 2),
                    "采购入库量": round(purchase, 2)
                })
                
            except Exception as e:
                errors.append({"account": account, "error": str(e)})
        
        # 写入 Excel
        output_path = output_path or self.output_path
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "原料经营明细"
        self._write_data_to_sheet(ws, results, ["账号", "BOM消耗量", "涉及销售金额", "采购入库量"])

        # 零容差闸门 (set_gate 设置后执行)
        if self.gate_spec is not None:
            outcome = self.gate_spec.run(results)
            if outcome.needs_watermark:
                add_watermark_sheet_openpyxl(wb, outcome.watermark_lines())

        wb.save(output_path)
        
        return ExportResult(
            total_count=len(self.merchants),
            success_count=len(results),
            failed_count=len(errors),
            output_path=str(output_path),
            errors=errors
        )
    
    def export_daily_item_sales(
        self,
        month: str,
        merchant_xlsx: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None
    ) -> ExportResult:
        """
        单品日销量报表（对应 report_wallace_daily_item_sales.sh）
        
        堂食 + 外卖的单品日销量明细
        
        Args:
            month: 月份 (YYYY-MM)
            merchant_xlsx: 商家列表 Excel 路径
            output_path: 可选，自定义输出路径
            
        Returns:
            ExportResult
        """
        from .sql_templates import get_template
        from datetime import datetime, timezone
        
        # 计算月份时间范围
        year, mon = int(month[:4]), int(month[5:7])
        start_ts = int(datetime(year, mon, 1, tzinfo=timezone.utc).timestamp())
        if mon == 12:
            end_year, end_mon = year + 1, 1
        else:
            end_year, end_mon = year, mon + 1
        end_ts = int(datetime(end_year, end_mon, 1, tzinfo=timezone.utc).timestamp())
        
        self.load_merchants(merchant_xlsx)
        
        results = []
        errors = []
        
        sql = get_template('daily_item_sales')
        
        for idx, (account, uuid_str, _) in enumerate(self.merchants, 1):
            dataset = f"shop{uuid_str}"
            
            try:
                query = sql.format(
                    project=self.project_id,
                    dataset=dataset,
                    start_ts=start_ts,
                    end_ts=end_ts
                )
                rows = list(self.client.query(query).result())
                
                for row in rows:
                    results.append({
                        "账号": account,
                        "日期": str(row.sale_date) if row.sale_date else "",
                        "商品名称": row.product_name or "",
                        "销量": float(row.total_qty or 0)
                    })
                    
            except Exception as e:
                errors.append({"account": account, "error": str(e)})
        
        # 写入 Excel
        output_path = output_path or self.output_path
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "单品日销量"
        self._write_data_to_sheet(ws, results, ["账号", "日期", "商品名称", "销量"])

        # 零容差闸门 (set_gate 设置后执行)
        if self.gate_spec is not None:
            outcome = self.gate_spec.run(results)
            if outcome.needs_watermark:
                add_watermark_sheet_openpyxl(wb, outcome.watermark_lines())

        wb.save(output_path)

        return ExportResult(
            total_count=len(self.merchants),
            success_count=len(self.merchants) - len(errors),
            failed_count=len(errors),
            output_path=str(output_path),
            errors=errors
        )

    def export_orders_by_nationality(
        self,
        start_date: str,
        end_date: str,
        output_path: Optional[Union[str, Path]] = None
    ) -> ExportResult:
        """
        订单国籍报表（单门店）

        Sheet1: 订单明细（订单号、时间、国籍、实收金额、订单明细）
        Sheet2: 按国籍汇总（国籍、总金额、订单数）

        Args:
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD，不包含)
            output_path: 可选，自定义输出路径

        Returns:
            ExportResult
        """
        from .sql_templates import get_template
        from datetime import datetime, timezone, timedelta
        from collections import defaultdict

        start_ts = self._to_timestamp(start_date)
        end_ts = self._to_timestamp(end_date)

        dataset = "shop1922991923200000"
        sql = get_template('order_nationality')

        query = sql.format(
            project=self.project_id,
            dataset=dataset,
            start_ts=start_ts,
            end_ts=end_ts
        )

        rows = list(self.client.query(query).result())

        # 格式化数据
        orders = []
        for row in rows:
            # finish_time 是 Unix 秒，转为 Asia/Bangkok 时间
            finish_ts = int(row.order_time or 0)
            if finish_ts > 0:
                dt = datetime.fromtimestamp(finish_ts, tz=timezone(timedelta(hours=7)))
                time_str = dt.strftime("%Y-%m-%d %H:%M:%S")
            else:
                time_str = ""

            orders.append({
                "订单号": row.order_number or "",
                "时间": time_str,
                "国籍": row.nationality or "未知",
                "实收金额": float(row.received_amount or 0),
                "订单明细": row.order_details or ""
            })

        # 按国籍汇总
        nationality_stats = defaultdict(lambda: {"总金额": 0.0, "订单数": 0})
        for order in orders:
            nat = order["国籍"]
            nationality_stats[nat]["总金额"] += order["实收金额"]
            nationality_stats[nat]["订单数"] += 1

        summary = []
        for nat, stats in sorted(nationality_stats.items(), key=lambda x: x[1]["总金额"], reverse=True):
            summary.append({
                "国籍": nat,
                "总金额": round(stats["总金额"], 2),
                "订单数": stats["订单数"]
            })

        # 写入 Excel（双 sheet）
        output_path = output_path or self.output_path
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "订单明细"
        self._write_data_to_sheet(ws1, orders, ["订单号", "时间", "国籍", "实收金额", "订单明细"])

        ws2 = wb.create_sheet(title="国籍汇总")
        self._write_data_to_sheet(ws2, summary, ["国籍", "总金额", "订单数"])

        # 零容差闸门 (set_gate 设置后执行; orders 作为闸门输入)
        if self.gate_spec is not None:
            outcome = self.gate_spec.run(orders)
            if outcome.needs_watermark:
                add_watermark_sheet_openpyxl(wb, outcome.watermark_lines())

        wb.save(output_path)

        # ========== 校验 ==========
        from .validators import RangeValidator, ValidationResult, ValidationError

        # 1. 范围校验：实收金额非负
        range_validator = RangeValidator([
            {"field": "实收金额", "min": 0, "name": "实收金额非负"}
        ])
        validation_result = range_validator.validate(orders)

        # 2. 自定义一致性校验：Sheet2 汇总 = Sheet1 聚合
        recalc_stats = defaultdict(lambda: {"总金额": 0.0, "订单数": 0})
        for order in orders:
            nat = order["国籍"]
            recalc_stats[nat]["总金额"] += order["实收金额"]
            recalc_stats[nat]["订单数"] += 1

        consistency_errors = []
        for row in summary:
            nat = row["国籍"]
            expected_amount = round(recalc_stats[nat]["总金额"], 2)
            expected_count = recalc_stats[nat]["订单数"]
            if abs(row["总金额"] - expected_amount) > 0.01 or row["订单数"] != expected_count:
                consistency_errors.append(ValidationError(
                    rule="nationality_consistency",
                    message=f"国籍 '{nat}' 汇总不一致: Sheet2(金额={row['总金额']}, 订单数={row['订单数']}) != Sheet1聚合(金额={expected_amount}, 订单数={expected_count})",
                    details={"国籍": nat, "sheet2_amount": row["总金额"], "sheet1_amount": expected_amount,
                             "sheet2_count": row["订单数"], "sheet1_count": expected_count}
                ))

        if consistency_errors:
            consistency_result = ValidationResult.failure(consistency_errors, {"checked": len(summary)})
            validation_result = validation_result.merge(consistency_result)
        else:
            validation_result = validation_result.merge(ValidationResult.success({"checked": len(summary)}))

        return ExportResult(
            total_count=1,
            success_count=1,
            failed_count=0,
            output_path=str(output_path),
            validation_result=validation_result,
            errors=[]
        )
