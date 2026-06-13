#!/usr/bin/env python3
"""
利润报表导出 —— 套餐/单品利润分析（配置驱动版，聚合优化版）

优化点：
1. 订单与 BOM 分两次查询，消除 JOIN 膨胀
2. 订单在 BQ 内聚合，大幅减少传输数据量
3. 套餐结构独立查询并缓存

Usage:
    python -m bq_reports.profit_margin_report --month 2026-03 --output exports/profit_202603.xlsx --use-erp-price
"""

import argparse
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone

# 跟 ttpos 业务时区对齐（曼谷 +07:00），月份边界以 BKK 时间为准。
# 真源在 semantic/dimensions/time.py；这里 re-export 保持现有 tests 和
# 报表脚本 from-import 兼容。
from semantic.dimensions.time import BKK_TZ  # noqa: E402
from pathlib import Path

from bq_reports.utils.bq_client import setup_proxy
from semantic.cogs import (
    BOM_UNIT_CORRECTIONS,
    MaterialPriceLayerProvider as _MaterialPriceLayerProvider,
    build_bom_resolver as _build_bom_resolver,
    build_material_price_resolver as _build_material_price_resolver,
    expand_item_bom,
    find_matched_bom_key as _find_matched_bom_key,
    match_bom_layered as _match_bom_layered,
    match_fallback_bom as _match_fallback_bom,
    resolve_unit_price as _resolve_unit_price_with_source,
)
from semantic.dimensions.time import month_to_ts_range as _month_to_ts_range, assert_month_not_frozen
from semantic.entities import bom, combo, price_breakdown, sale_line, takeout_line, total_line
from semantic.resolvers import (
    CallableProvider,
    Resolved,
    Resolver,
    YamlMatchProvider,
)
from utils.cache import get_cache, set_cache, cache_key
from utils.report_engine import ReportEngine, load_sheet_config
from utils.resource_adapter import get_adapter


# ============================================================================
# 门店名称映射加载
# ============================================================================

def _load_store_names_from_bq(client, merchant_records: list) -> dict:
    """并发查 BQ 每个店的 ttpos_setting (key='store' → values.name).

    BQ 是店名的权威真源 — 跟 POS 系统始终一致, 新店自动有名字.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _q(m):
        uuid = m["uuid"]; code = m["store_num"]
        sql = (
            f"SELECT JSON_VALUE(values, '$.name') AS name "
            f"FROM `{client.project}.shop{uuid}.ttpos_setting` "
            f"WHERE key = 'store' AND delete_time = 0 LIMIT 1"
        )
        try:
            for r in client.query(sql).result():
                if r.name:
                    return str(code).zfill(3), str(r.name).strip()
        except Exception:
            pass
        return str(code).zfill(3), None

    mapping = {}
    with ThreadPoolExecutor(max_workers=10) as ex:
        for f in as_completed([ex.submit(_q, m) for m in merchant_records]):
            code, name = f.result()
            if name:
                mapping[code] = name
                try:
                    mapping[str(int(code))] = name
                except ValueError:
                    pass
    return mapping


def _load_store_names(config: dict = None, client=None):
    """门店编号→名称映射. 优先 BQ ttpos_setting (权威源), Excel 作 fallback.

    Args:
        config: load_config() 返回值
        client: BigQuery client (可选). 传入则走 BQ; 否则 fallback Excel.
    """
    cfg = config or {}
    cache_ttl = cfg.get("cache", {}).get("store_names_ttl", 604800)

    # 路径 1: BQ ttpos_setting (主源)
    if client is not None:
        # 加载商户基础列表 (account + uuid + store_num) — 不依赖 store_name
        merchant_spec = cfg.get("merchant_list")
        if merchant_spec:
            try:
                from utils.resource_adapter import get_adapter as _ga
                ad = _ga(merchant_spec["adapter"])
                records_raw = ad.load(merchant_spec)
                merchants = []
                for r in records_raw:
                    acc = r.get("account") or ""
                    uuid = r.get("uuid") or ""
                    if not uuid: continue
                    # account 形如 admin-001@wallace.com → 抽 store_num
                    import re as _re
                    m = _re.search(r"admin-(\d+)", str(acc))
                    if not m: continue
                    merchants.append({"account": acc, "uuid": str(uuid),
                                      "store_num": m.group(1).zfill(3)})
                key = cache_key("store_names_bq_v1",
                                {"uuids": sorted([m["uuid"] for m in merchants])})
                cached = get_cache(key, ttl_seconds=cache_ttl)
                if cached is not None:
                    print(f"[Store Names] BQ 缓存命中: {len(set(cached.values()))} 个")
                    return cached
                mapping = _load_store_names_from_bq(client, merchants)
                set_cache(key, mapping)
                print(f"[Store Names] 从 BQ ttpos_setting 加载 "
                      f"{len(set(mapping.values()))} 个门店名称")
                return mapping
            except Exception as e:
                print(f"[警告] BQ 加载店名失败, fallback Excel: {e}")

    # 路径 2: Excel fallback (旧逻辑)
    mapping_config = cfg.get("store_name_mapping")
    if not mapping_config:
        return {}
    key = cache_key("store_names_v2", {"path": mapping_config.get("path", "")})
    cached = get_cache(key, ttl_seconds=cache_ttl)
    if cached is not None:
        print(f"[Store Names] Excel 缓存命中: {len(cached)} 个")
        return cached
    try:
        adapter = get_adapter(mapping_config["adapter"])
        records = adapter.load(mapping_config)
        mapping = {}
        for r in records:
            num = r.get("store_number"); name = r.get("store_name")
            if num is None or not name: continue
            s = str(num).strip()
            if not s: continue
            clean_name = str(name).strip()
            mapping[s] = clean_name
            try: mapping[str(int(s))] = clean_name
            except ValueError: pass
        set_cache(key, mapping)
        print(f"[Store Names] Excel 加载 {len(set(mapping.values()))} 个门店名称")
        return mapping
    except Exception as e:
        print(f"[警告] 加载门店名称失败: {e}")
        return {}


# ============================================================================
# ERPNext 价格加载（带缓存包装）
# ============================================================================

def _load_uploaded_prices(excel_path: str) -> tuple[dict, dict]:
    """从上传的 Excel 价格清单读取物料单价和换算系数。
    支持 '干冻货' 和 '设备材料' 两个 sheet，以及 '盘点单位匹配分析' sheet。
    返回: (prices: {material_code: unit_price}, conversions: {material_code: conv_rate})
    其中 unit_price = 清单单价 ÷ 销售换算系数（得到最小单位单价）
    """
    if not excel_path or not os.path.exists(excel_path):
        return {}, {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # 1) 读取销售换算系数（从"盘点单位匹配分析"）
        conversions = {}
        if "盘点单位匹配分析" in wb.sheetnames:
            ws_conv = wb["盘点单位匹配分析"]
            for row in ws_conv.iter_rows(min_row=2, values_only=True):
                code = row[0]
                conv = row[8]  # 销售换算系数 (I列)
                if code and conv is not None:
                    try:
                        conv_val = float(conv)
                        if conv_val > 0:
                            conversions[str(code).strip()] = conv_val
                    except (ValueError, TypeError):
                        pass
            print(f"[Uploaded Prices] 加载 {len(conversions)} 条换算系数")
        
        # 2) 读取单价并换算
        prices = {}
        for sheet_name in ("干冻货", "设备材料"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = row[0]
                price = row[7] if len(row) > 7 else None
                if code and price is not None and price != '#N/A':
                    try:
                        price_val = float(price)
                        code_str = str(code).strip()
                        # 用最小单位单价 = 盘点单价 ÷ 销售换算系数
                        conv = conversions.get(code_str, 1)
                        prices[code_str] = price_val / conv
                    except (ValueError, TypeError):
                        pass
        print(f"[Uploaded Prices] 从 {excel_path} 加载 {len(prices)} 条最小单位单价")
        return prices, conversions
    except Exception as e:
        print(f"[警告] 加载上传价格清单失败: {e}")
        return {}, {}


def _try_load_erp_prices(price_list: str = None, cache_ttl: int = 3600):
    from bq_reports.utils.erpnext_api import load_erpnext_prices
    key = cache_key("erpnext_prices", {"price_list": price_list or "Standard Buying"})
    
    # 1) 先尝试缓存（正常 TTL）
    cached = get_cache(key, ttl_seconds=cache_ttl)
    if cached is not None:
        print(f"[ERPNext API] 缓存命中: {len(cached)} 条价格")
        return cached
    
    # 2) 缓存未命中，尝试 API
    try:
        prices = load_erpnext_prices(price_list=price_list)
        set_cache(key, prices)
        return prices
    except Exception as e:
        print(f"[警告] ERPNext API 失败: {e}")
    
    # 3) API 也失败，强制读缓存（忽略 TTL，永不过期 fallback）
    cached = get_cache(key, ttl_seconds=99999999)
    if cached is not None:
        print(f"[ERPNext API] API 不可用，使用过期缓存: {len(cached)} 条价格")
        return cached
    
    print("[警告] 未加载到 ERPNext 价格，成本将显示为 0")
    return {}


# ============================================================================
# 外挂 BOM 加载（支持 priority 栈 + 缓存）
#
# 配置形态:
#   bom_sources:                      # 新格式: 多源 + priority
#     - name: colleague_20260513
#       priority: 100                 # 数字大 = 更权威, 优先命中
#       adapter: excel
#       path: ...
#       sheet: ...
#       mapping: {...}
#     - name: customer_20260506
#       priority: 50
#       ...
#   fallback_bom:                     # 老格式 (兼容): 单源, 隐式 priority=50
#     adapter: excel
#     ...
#
# BQ 原生 BOM 是隐式最低层 (priority=0); 客户给的"调整版"应放高 priority,
# 它的 BOM 会覆盖 BQ 原生 (而不是只在 BQ 没 BOM 时兜底).
# ============================================================================

def _load_one_bom_source(source_cfg: dict, cache_ttl: int, label: str = "BOM Source"):
    """加载单个 BOM 源 → {product_name: [(code, mat_name, num, uom), ...]}。"""
    # v2: 应用市场 BOM 替换/删除规则
    key = cache_key("fallback_boms_v2", {"path": source_cfg.get("path", "")})
    cached = get_cache(key, ttl_seconds=cache_ttl)
    if cached is not None:
        print(f"[{label}] 缓存命中: {len(cached)} 个商品")
        return cached

    try:
        adapter = get_adapter(source_cfg["adapter"])
        records = adapter.load(source_cfg)

        boms = {}
        current_product = None

        for r in records:
            product_name = r.get("product_name")
            material_code = r.get("material_code")
            material_name = r.get("material_name")
            bom_num = r.get("qty")
            uom = r.get("unit")

            if product_name:
                current_product = str(product_name).strip()
                if current_product not in boms:
                    boms[current_product] = []

            if current_product and material_code and bom_num is not None:
                code_str = str(material_code).strip()
                if code_str in ("—", "-", "", "None", "null", "(无编号)"):
                    continue
                try:
                    num_val = float(bom_num)
                except (ValueError, TypeError):
                    continue
                existing_codes = {item[0] for item in boms[current_product]}
                if code_str in existing_codes:
                    continue
                uom_mapping = {"克": "g", "g": "g", "个": "pc", "pc": "pc", "份": "pc"}
                std_uom = uom_mapping.get(str(uom).strip(), str(uom).strip() if uom else "")
                boms[current_product].append((
                    code_str,
                    str(material_name or "").strip(),
                    num_val,
                    std_uom,
                ))

        # 对外挂 BOM 应用相同的替换/删除规则（统一口径）
        fb_drop = 0
        fb_replace = 0
        for product_name, recs in list(boms.items()):
            wrapped = [(c, n, num, u, 1.0, 0.0) for c, n, num, u in recs]
            new_wrapped = _apply_bom_overrides(wrapped)
            old_codes = {r[0] for r in recs}
            fb_drop += len(old_codes & BOM_DROP_CODES)
            fb_replace += len(old_codes & set(BOM_REPLACEMENTS.keys()))
            boms[product_name] = [(c, n, num, u) for c, n, num, u, _, _ in new_wrapped]
        if fb_drop or fb_replace:
            print(f"[{label}] 市场规则: 替换 {fb_replace} 处, 删除 {fb_drop} 处")

        set_cache(key, boms)
        print(f"[{label}] 加载 {len(boms)} 个商品的外挂 BOM")
        return boms

    except Exception as e:
        print(f"[{label}] 加载失败: {e}")
        return {}


def _load_bom_layers(config: dict = None):
    """加载所有外挂 BOM 源, 按 priority 降序返回。

    Returns:
        List[(name, priority, boms_dict, match_mode)], priority 大的在前。
        match_mode: "fuzzy" (默认, 5 层模糊匹配) | "exact" (只整 key 精确).
        客户/市场精确列出商品名的层 (e.g. 补充 BOM) 应在 config 里标
        match_mode: exact, 避免短名误命中长 key.
    """
    cfg = config or {}
    cache_ttl = cfg.get("cache", {}).get("fallback_bom_ttl", 86400)

    # 新格式: bom_sources 列表
    sources = cfg.get("bom_sources")
    if not sources:
        # 老格式兼容: 单个 fallback_bom (隐式 priority=50)
        legacy = cfg.get("fallback_bom")
        if not legacy:
            return []
        sources = [{
            "name": legacy.get("name", "fallback_bom"),
            "priority": legacy.get("priority", 50),
            **{k: v for k, v in legacy.items() if k not in ("name", "priority")},
        }]

    layers = []
    for src in sources:
        # 来源名 = 显式 name > path 的 basename > "?"
        # 强制用文件名而不是自己编的标签, 客户看 Excel 能直接定位到源文件
        path = src.get("path", "")
        name = src.get("name") or (os.path.basename(path) if path else "?")
        priority = int(src.get("priority", 0))
        match_mode = src.get("match_mode", "fuzzy")
        boms = _load_one_bom_source(src, cache_ttl, label=f"BOM[{name}]")
        if boms:
            layers.append((name, priority, boms, match_mode))
    layers.sort(key=lambda x: -x[1])
    if layers:
        order = " > ".join(
            f"{n}(p={p}, {len(b)}, {m})" for n, p, b, m in layers)
        print(f"[BOM Layers] 优先级顺序: {order}")
    return layers


# _build_bom_resolver / _match_bom_layered / _find_matched_bom_key / _match_fallback_bom
# 现在统一由 semantic.cogs.bom_match 提供 (import 见文件顶部); 这里保留私有别名仅为
# 兼容老 tests 的 `from bq_reports.profit_margin_report import _xxx`.


def _load_fallback_boms(config: dict = None):
    """向后兼容: 把所有 BOM 层合并成扁平 dict (高优先级覆盖低优先级)。

    新代码请用 _load_bom_layers + _match_bom_layered。
    保留此函数仅为兼容旧 tests / 旧调用点。
    """
    layers = _load_bom_layers(config)
    merged = {}
    # 从低 priority 到高 priority 依次 update, 让高的覆盖低的
    for _name, _p, boms, _mode in reversed(layers):
        merged.update(boms)
    return merged


# ============================================================================
# 商家列表加载（适配器 + 缓存）
# ============================================================================

def _fetch_store_names_from_bq(uuids, project_id):
    """并发查每个 dataset 的 ttpos_setting 拿 store_code/store_name。"""
    from concurrent.futures import ThreadPoolExecutor
    from bq_reports.utils.bq_client import get_bq_client

    def _query(uuid_str):
        try:
            client = get_bq_client(project_id)
            sql = f"""
            SELECT
              JSON_EXTRACT_SCALAR(`values`, '$.store_code') AS code,
              JSON_EXTRACT_SCALAR(`values`, '$.name') AS name
            FROM `{project_id}`.`shop{uuid_str}`.`ttpos_setting`
            WHERE `key` = 'store' AND delete_time = 0
            LIMIT 1
            """
            rows = list(client.query(sql).result())
            if not rows:
                return uuid_str, None, None
            r = rows[0]
            return uuid_str, (r.code or None), (r.name or None)
        except Exception as e:
            print(f"[警告] 查询 shop{uuid_str} 的 store_name 失败: {e}")
            return uuid_str, None, None

    result = {}
    with ThreadPoolExecutor(max_workers=10) as ex:
        for uuid_str, code, name in ex.map(_query, uuids):
            result[uuid_str] = (code, name)
    return result


def _load_merchants(config: dict, store_names: dict, override_path: str = None,
                    project_id: str = None):
    """加载商家列表。门店名优先从 BQ ttpos_setting 实时取，
    fallback 到 store_names Excel 映射，再 fallback 到 "-"。
    门店编号同样优先 BQ store_code，否则用 admin-XXX 解析出的数字。"""
    merchant_cfg = config.get("merchant_list")
    if merchant_cfg:
        cache_ttl = config.get("cache", {}).get("merchant_list_ttl", 86400)
        # v3: 接入 BQ 实时门店名
        key = cache_key("merchants_v3", {"path": merchant_cfg.get("path", ""),
                                          "project": project_id or ""})
        cached = get_cache(key, ttl_seconds=cache_ttl)
        if cached is not None:
            print(f"[Merchants] 缓存命中: {len(cached)} 个")
            return cached

        adapter = get_adapter(merchant_cfg["adapter"])
        records = adapter.load(merchant_cfg)
        raw = []
        for r in records:
            account = r.get("account")
            uuid_str = r.get("uuid")
            if not account or not uuid_str:
                continue
            account = str(account).strip()
            uuid_str = str(uuid_str).strip()
            m = re.search(r'admin-(\d+)@', account)
            store_num_excel = m.group(1) if m else account
            raw.append((account, uuid_str, store_num_excel))

        # 从 BQ 拉真实门店名（覆盖 Excel 映射）
        bq_names = {}
        if project_id and raw:
            print(f"[Merchants] 从 BQ 查询 {len(raw)} 个门店的 store_code/store_name...")
            bq_names = _fetch_store_names_from_bq([r[1] for r in raw], project_id)

        merchants = []
        for account, uuid_str, store_num_excel in raw:
            bq_code, bq_name = bq_names.get(uuid_str, (None, None))
            store_num = bq_code or store_num_excel
            store_name = bq_name
            if not store_name:
                store_name = store_names.get(store_num_excel)
                if not store_name and store_num_excel.isdigit():
                    store_name = store_names.get(str(int(store_num_excel)))
            store_name = store_name or "-"
            merchants.append((account, uuid_str, store_num, store_name))
        set_cache(key, merchants)
        return merchants
    else:
        # 回退：直接读 Excel（兼容旧用法）
        from openpyxl import load_workbook
        path = override_path or "resources/merchants.xlsx"
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        merchants = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 3 and row[1] and row[2]:
                account = str(row[1]).strip()
                uuid_str = str(row[2]).strip()
                m = re.search(r'admin-(\d+)@', account)
                store_num = m.group(1) if m else account
                store_name = store_names.get(store_num, "-")
                merchants.append((account, uuid_str, store_num, store_name))
        wb.close()
        return merchants


# ============================================================================
# SQL 模板（聚合优化版）
# ============================================================================

# 套餐 / 单品订单聚合 — 按 ttpos `CountProductSale` 算法（statistics_product + takeout_order_item）
#
# ttpos 源码: ttpos-server-go/main/app/repository/statistics.go:1933-2143 (CountProductSale)
#
# Shop (statistics_product):
#   sale_num   = SUM(product_num)
#   sale_amount(actual) = SUM(IF(free_num>0 OR give_num>0, 0,
#                  product_final_price * (product_num - refund_num)))
# Takeout (takeout_order_item + takeout_order):
#   sale_num   = SUM(quantity)
#   sale_amount(actual) = SUM(IF(order_state=60, 0, price * quantity))
#   order_state IN (10,20,30,40,60), accepted_time > 0
# 合并：按 product_package_uuid FULL OUTER JOIN
#
# 注意:
#   1. ttpos 不展开 套餐子商品（无 copy_num × unit_num 乘子）— 跟我们之前实现差异大
#   2. cancelled (state=60) 订单 sale_num 计入但 amount 计 0（ttpos 设计，争议）
#   3. 赠品 (free_num/give_num) amount 直接归零
#   4. shop 端退款通过 (product_num - refund_num) 反映
#   5. std_unit_price 保留我们的算法（按销量加权 product_bom.price），ttpos 没有这个概念

# _PROFIT_SALES_TPL is assembled from the semantic layer's CTE factories. The
# entity strings still carry literal `{project}` / `{dataset}` / `{start_ts}` /
# `{end_ts}` placeholders (they pass through f-string interpolation unchanged
# because the outer f-string only resolves its own `{…}` expressions). The final
# `{{product_type}}` is escaped to a literal `{product_type}` so engine.query()
# can still substitute it per-call via `.replace()` below.
def _build_profit_sales_tpl(exclude_test_business: bool = False) -> str:
    """生成 profit margin 主 SQL 模板。exclude_test_business=True 对齐 ttpos 后台口径。"""
    return f"""
WITH
-- 价格拆分：取前3个主要价格档（按销量降序），其余归到"其他"
-- 必须同时覆盖堂食 + 外卖，否则销量/营业额拆分对不上下游 shop_sales+takeout_sales 合并值
{price_breakdown.price_top3_ctes(exclude_test_business=exclude_test_business)},
{sale_line.shop_sales_cte(exclude_test_business=exclude_test_business)},
{takeout_line.takeout_sales_cte(exclude_test_business=exclude_test_business)},
{total_line.merged_cte()}{_PROFIT_SALES_BODY}"""


_PROFIT_SALES_BODY = """
SELECT
  m.item_uuid,
  -- 部分商品名末尾带回车/换行/制表符等不可见字符（前端 trim 显示无异，但 BQ 取出来会带）
  -- 用 REGEXP_REPLACE 去掉首尾不可见字符，避免渲染成 _x000D_ 看似两个不同商品
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pp.name, '$.en'),
    '未知'
  ), r'^\\s+|\\s+$', '') AS item_name,
  m.qty AS qty,
  m.revenue AS revenue,
  m.sales_price AS sales_price,
  -- 毛额 (守恒闭环锚): 来自 merged CTE, 堂食=sales_price, 外卖 state-UNCONDITIONED
  m.gross_amount AS gross_amount,
  m.original_amount AS original_amount,
  m.avg_member_discount AS avg_member_discount,
  m.free_qty AS free_qty,
  m.give_qty AS give_qty,
  m.refund_qty AS refund_qty,
  m.refund_amount AS refund_amount,
  m.cancelled_qty AS cancelled_qty,
  m.cancelled_amount AS cancelled_amount,
  -- 金额恒等式三分项：sales_price = actual + refund + free + give + cancelled + discount
  m.free_amount AS free_amount,
  m.give_amount AS give_amount,
  m.discount_amount AS discount_amount,
  -- 价格拆分：前3个主要价格档 + 其他
  p3.price_1 AS price_1,
  p3.qty_1 AS qty_1,
  p3.price_2 AS price_2,
  p3.qty_2 AS qty_2,
  p3.price_3 AS price_3,
  p3.qty_3 AS qty_3,
  p3.other_qty AS other_price_qty,
  -- 单价取 Shop 商品管理标价 ttpos_product_package.price
  IFNULL(pp.price, 0) AS list_price
FROM merged m
-- 价格拆分 JOIN
LEFT JOIN price_top3 p3 ON p3.item_uuid = m.item_uuid
-- ttpos 导出用 LEFT JOIN，不过滤 pp.delete_time（已删除的商品也算销售）
-- 但本报表必须按 product_type 区分套餐/单品 sheet，所以仍 INNER JOIN，去掉 delete_time 过滤
JOIN `{project}`.`{dataset}`.`ttpos_product_package` pp
  ON pp.uuid = m.item_uuid
WHERE pp.product_type = {product_type}
  AND m.qty > 0
"""

_PROFIT_SALES_TPL = _build_profit_sales_tpl(exclude_test_business=False)
_PROFIT_SALES_TPL_TB = _build_profit_sales_tpl(exclude_test_business=True)

COMBO_ORDERS_SQL = _PROFIT_SALES_TPL.replace("{product_type}", "1")
SINGLE_ORDERS_SQL = _PROFIT_SALES_TPL.replace("{product_type}", "0")
COMBO_ORDERS_SQL_TB = _PROFIT_SALES_TPL_TB.replace("{product_type}", "1")
SINGLE_ORDERS_SQL_TB = _PROFIT_SALES_TPL_TB.replace("{product_type}", "0")

# 产品 BOM + 套餐结构 SQL 真源在 semantic.entities.{bom,combo}
BOM_SQL = bom.bom_sql()
COMBO_STRUCTURE_SQL = combo.combo_structure_sql()


# ============================================================================
# 价格解析 — BOM_UNIT_CORRECTIONS 真源在 semantic.cogs.material_price,
# 这里 re-import (见文件顶部) 仅为兼容老 tests 的 `from bq_reports... import`.
# ============================================================================

# 市场要求的 BOM 物料替换/删除规则（套餐和单品都生效）
# value 为 None 时保留旧 material_name；指定字符串时覆盖名称。
BOM_REPLACEMENTS = {
    "FR01008": ("FR02001", None),
    "VE01001": ("MK01018", None),
}
BOM_DROP_CODES = {"TL99008"}


def _apply_bom_overrides(bom_records):
    """对单个 (store, item) 的 BOM 列表应用替换/删除规则。

    record 元组顺序: (material_code, material_name, bom_num, bom_unit, conv_rate, bq_price)

    两轮合并：先收录所有未被替换的物料（元数据权威），再把被替换的并入 —
    若目标 code 已存在则只累加 bom_num，保留真实物料的 name/unit/conv/price。
    """
    merged = {}
    deferred = []
    for code, name, bom_num, bom_unit, conv_rate, bq_price in bom_records:
        if code in BOM_DROP_CODES:
            continue
        if code in BOM_REPLACEMENTS:
            new_code, new_name = BOM_REPLACEMENTS[code]
            override_name = new_name if new_name is not None else name
            deferred.append((new_code, override_name, bom_num, bom_unit, conv_rate, bq_price))
            continue
        if code in merged:
            prev_name, prev_num, prev_unit, prev_conv, prev_price = merged[code]
            merged[code] = (
                prev_name or name,
                prev_num + bom_num,
                prev_unit or bom_unit,
                prev_conv,
                prev_price,
            )
        else:
            merged[code] = (name, bom_num, bom_unit, conv_rate, bq_price)

    for code, name, bom_num, bom_unit, conv_rate, bq_price in deferred:
        if code in merged:
            prev_name, prev_num, prev_unit, prev_conv, prev_price = merged[code]
            merged[code] = (prev_name, prev_num + bom_num, prev_unit, prev_conv, prev_price)
        else:
            merged[code] = (name, bom_num, bom_unit, conv_rate, bq_price)

    return [(c, n, bn, bu, cr, bp) for c, (n, bn, bu, cr, bp) in merged.items()]


# ============================================================================
# 物料单价 Resolver — 真源在 semantic.cogs.material_price.
# _MaterialPriceLayerProvider / _build_material_price_resolver /
# _resolve_unit_price_with_source 在文件顶部以私有别名 re-import, 仅为
# 兼容老 tests 的 `from bq_reports.profit_margin_report import _xxx`.
# 这里只保留 _resolve_base_unit_price (旧"不带 source"接口) 的薄包装。
# ============================================================================

def _resolve_base_unit_price(material_code, bq_price, uploaded_prices, erp_prices,
                              price_layers=None, strict=False):
    """旧接口: 只返回 price, 不带来源。新代码请用
    semantic.cogs.material_price.resolve_unit_price。"""
    price, _ = _resolve_unit_price_with_source(
        material_code, bq_price, uploaded_prices, erp_prices,
        price_layers=price_layers, strict=strict)
    return price


# ============================================================================
# 物料单价层 (priority 栈, 独立于 BOM 数量栈)
# ============================================================================

def _load_material_price_layers(config: dict = None):
    """从 config.material_price_sources 加载物料价 priority 栈, 支持双索引 (编码 + 名字).

    支持 2 类 adapter:
      - 老 'excel' adapter: 返回 records 列表, mapping 字段 material_code/unit_price
        → 转成 by_code 单 dict 索引
      - 新 'cost_price_taixi' / 'cost_price_import' adapter: 返回 {key: {price, unit, source_tag}}
        → 用 source 配置的 'index_by' 决定是 by_code 还是 by_name
          index_by = 'code'  → 按编码索引
          index_by = 'name'  → 按归一化名字索引 (CostPriceImportAdapter 的输出已经是归一化名)

    Returns: List[Layer], Layer.data 是 dict, 但里面 value 是 {price, unit, source_tag, ...} 结构 (新)
             或 float (老兼容).
    """
    from utils.layered_resource import load_layers

    cfg = config or {}
    cache_ttl = cfg.get("cache", {}).get("material_price_ttl",
                cfg.get("cache", {}).get("fallback_bom_ttl", 86400))
    sources = cfg.get("material_price_sources", [])
    if not sources:
        return []

    def _loader(src):
        key = cache_key("material_prices_v2", {
            "path": src.get("path", ""), "sheet": src.get("sheet", ""),
            "adapter": src.get("adapter", "")})
        cached = get_cache(key, ttl_seconds=cache_ttl)
        if cached is not None:
            return cached

        adapter = get_adapter(src["adapter"])
        records = adapter.load(src)

        if isinstance(records, dict):
            # 新成本价 adapter 直接返回 {key: {price, unit, source_tag, ...}}
            out = records
        else:
            # 老 'excel' adapter 返回 records 列表, 转换成 {code: price} dict
            out = {}
            for r in records:
                code = r.get("material_code")
                price = r.get("unit_price")
                if not code or price is None:
                    continue
                try:
                    out[str(code).strip()] = float(price)
                except (ValueError, TypeError):
                    continue
        set_cache(key, out)
        return out

    layers = load_layers(sources, _loader, label_prefix="MaterialPrice")

    # 给每层标记 index_by (用于 _resolve_unit_price_with_source 选择匹配键)
    for src, layer in zip(sources, layers):
        layer.index_by = src.get("index_by", "code")    # 默认按编码
    return layers


# ============================================================================
# Source 标注 (P2)
#
# 把 BOM/price 来源算好写回 agg_data, 让 validator 能消费 source 信息.
# _build_summary_rows / _build_rows 内部仍会重新算 (报表展示用), 这里独立算
# 一份给 validator, 保证 source 元数据进 check_rows.
# ============================================================================

def _annotate_agg_data_sources(
    agg_data,
    bom_layers,
    uploaded_prices,
    erp_prices,
    price_layers,
    strict_price=False,
):
    """对每个 (store, sku) 算 bom_source / price_source 写回 agg_data['bom_source'] /
    agg_data['price_source']。

    后续 build_rows / build_summary_rows 仍会重新算 (报表展示用，互不污染)，
    但 check_rows 在 check() 之前能拿到 source 信息，让 SOURCE_COVERAGE
    identity 工作。
    """
    bom_layers = bom_layers or []
    # 构造一次共享 resolver, 避免 per-BOM-row 重建 (10 万行 = 重建 10 万次)
    _price_resolver = _build_material_price_resolver(
        uploaded_prices, erp_prices, price_layers, strict=strict_price)
    for key, data in agg_data.items():
        _store_num, _store_name, _item_uuid, item_name = key
        bom_source = "bq_native" if data.get("bom") else None
        price_sources_seen = set()

        matched, layer_name = _match_bom_layered(item_name, bom_layers)
        if matched:
            for code, name, _bom_num, _uom in matched:
                _, p_src = _resolve_unit_price_with_source(
                    code, 0, uploaded_prices, erp_prices,
                    price_layers=price_layers, strict=strict_price,
                    material_name=name, resolver=_price_resolver,
                )
                price_sources_seen.add(p_src)
            bom_source = layer_name
        else:
            for code, _n, _bn, _up, _u in (data.get("bom") or []):
                _, p_src = _resolve_unit_price_with_source(
                    code, 0, uploaded_prices, erp_prices,
                    price_layers=price_layers, strict=strict_price,
                    resolver=_price_resolver,
                )
                price_sources_seen.add(p_src)

        if bom_source is None:
            bom_source = "无"
        price_source_str = (
            " + ".join(sorted(price_sources_seen)) if price_sources_seen else "无"
        )

        data["bom_source"] = bom_source
        data["price_source"] = price_source_str


# ============================================================================
# 套餐结构加载（缓存）
# ============================================================================

def _load_combo_structures(engine, merchants, start_ts, end_ts, config: dict = None):
    """
    查询并缓存每个门店的套餐结构。
    返回: {store_num: {combo_uuid: [(child_uuid, child_num, weight), ...]}}

    v2 (2026-05): 改读 ttpos_product_package_group + _group_item 定义表，
    跨月稳定，不再依赖订单反推；返回 3 元组带份数 + 摊薄权重。
    """
    cfg = config or {}
    cache_ttl = cfg.get("cache", {}).get("combo_structure_ttl", 604800)  # 7天
    # v2: shape 变了 (list[(child, num, weight)] 三元组), bump cache key 避免读旧格式
    key = cache_key("combo_structures_v2", {"count": len(merchants)})
    cached = get_cache(key, ttl_seconds=cache_ttl)
    if cached is not None:
        print(f"[Combo Structure] 缓存命中: {len(cached)} 个门店")
        return cached

    print("[Combo Structure] 从 BQ 查询套餐定义...")
    raw_rows, errors = engine.query(
        sql_template=COMBO_STRUCTURE_SQL,
        merchants=merchants,
        # SQL v2 不读 {start_ts}/{end_ts}（套餐定义跨月稳定），但 engine.query
        # 仍要这俩参数避免 KeyError —— 传死值即可
        start_ts=0,
        end_ts=0,
        workers=10,
        row_proxy_factory=lambda row, acc, num, name: type("RowProxy", (), {
            "__getattr__": lambda self, attr: getattr(row, attr),
            "account": acc, "store_num": num, "store_name": name,
        })(),
        label="套餐结构",
    )

    structures = {}
    for row in raw_rows:
        store_num = row.store_num
        if store_num not in structures:
            structures[store_num] = {}
        combo_uuid = str(row.combo_uuid)
        child_uuid = str(row.child_uuid)
        child_num = float(row.child_num or 0)
        weight = float(row.weight or 1.0)
        if combo_uuid not in structures[store_num]:
            structures[store_num][combo_uuid] = []
        # 同 combo 可能有多个 group，多个 slot，允许同 child 多行（不 dedup）
        structures[store_num][combo_uuid].append((child_uuid, child_num, weight))

    set_cache(key, structures)
    total_combos = sum(len(v) for v in structures.values())
    print(f"[Combo Structure] 加载 {len(structures)} 个门店，共 {total_combos} 个套餐")
    return structures


# ============================================================================
# BOM 加载（缓存）
# ============================================================================

def _load_boms(engine, merchants, config: dict = None):
    """
    查询并缓存每个门店的产品 BOM。
    返回: {store_num: {item_uuid: [(material_code, material_name, bom_num, bom_unit, conv_rate, bq_price), ...]}}
    """
    cfg = config or {}
    cache_ttl = cfg.get("cache", {}).get("bom_ttl", 86400)  # 1天
    # v2: 增加 bom_unit / conversion_rate 字段，旧缓存格式不兼容
    # v3: BOM 加载层去重 (store, item, material)，旧缓存有重复行
    # v4: 应用市场 BOM 替换/删除规则 (FR01008→FR02001, VE01001→MK01018, drop TL99008)
    # v5: 软删商品的 BOM fallback —— pb.delete_time != 0 但全店无 active 时仍纳入
    key = cache_key("boms_v5", {"count": len(merchants)})
    cached = get_cache(key, ttl_seconds=cache_ttl)
    if cached is not None:
        print(f"[BOM] 缓存命中: {len(cached)} 个门店")
        return cached

    print("[BOM] 从 BQ 查询产品 BOM...")
    raw_rows, errors = engine.query(
        sql_template=BOM_SQL,
        merchants=merchants,
        start_ts=0,
        end_ts=2147483647,
        workers=10,
        row_proxy_factory=lambda row, acc, num, name: type("RowProxy", (), {
            "__getattr__": lambda self, attr: getattr(row, attr),
            "account": acc, "store_num": num, "store_name": name,
        })(),
        label="BOM",
    )

    boms = {}
    # ttpos 多规格商品的 product_bom × related_material JOIN 会让同一 material
    # 在同 (store, item) 内重复出现 N 次。在加载层去重，让套餐子商品共用物料的
    # 跨 child 累加逻辑安全（不会被 intra-item 冗余污染）。
    seen_keys = {}
    for row in raw_rows:
        store_num = row.store_num
        item_uuid = str(row.item_uuid)
        material_code = row.material_code
        if not material_code:
            continue
        dedup_key = (store_num, item_uuid, str(material_code))
        if dedup_key in seen_keys:
            continue
        seen_keys[dedup_key] = True
        if store_num not in boms:
            boms[store_num] = {}
        if item_uuid not in boms[store_num]:
            boms[store_num][item_uuid] = []
        boms[store_num][item_uuid].append((
            str(material_code),
            row.material_name or "",
            float(row.bom_num or 0),
            (row.bom_unit or "").strip() or "-",
            float(row.conversion_rate or 1),
            float(row.material_bq_price or 0),
        ))

    # 应用市场 BOM 替换/删除规则
    drop_count = 0
    replace_count = 0
    for store_num, items in boms.items():
        for item_uuid, records in list(items.items()):
            new_records = _apply_bom_overrides(records)
            old_codes = {r[0] for r in records}
            new_codes = {r[0] for r in new_records}
            drop_count += len(old_codes & BOM_DROP_CODES)
            replace_count += len(old_codes & set(BOM_REPLACEMENTS.keys()))
            items[item_uuid] = new_records
    if drop_count or replace_count:
        print(f"[BOM] 市场规则: 替换 {replace_count} 处, 删除 {drop_count} 处")

    set_cache(key, boms)
    total_items = sum(len(v) for v in boms.values())
    print(f"[BOM] 加载 {len(boms)} 个门店，共 {total_items} 个产品的 BOM")
    return boms


# ============================================================================
# 数据聚合（新版：预聚合 orders + 预加载 BOM）
# ============================================================================

def aggregate_with_bom(order_rows, bom_data, combo_structure, uploaded_prices=None, erp_prices=None,
                        mode="combo", price_layers=None, strict_price=False):
    """
    聚合订单和 BOM 数据（中间表版：保留所有原始字段，不预计算）。

    Args:
        order_rows: 引擎返回的订单行（已带 store_num, store_name）
        bom_data: {store_num: {item_uuid: [(material_code, ...), ...]}}
        combo_structure: {store_num: {combo_uuid: [child_uuid, ...]}}
        uploaded_prices: 上传价格清单 {material_code: price}
        erp_prices: ERPNext 价格 {material_code: (price, uom)}
        mode: "combo" 或 "single"
    """
    data = {}
    for row in order_rows:
        store_num = row.store_num
        store_name = row.store_name
        item_uuid = str(row.item_uuid)
        item_name = row.item_name
        qty = float(row.qty or 0)
        revenue = float(row.revenue or 0)
        sales_price = float(getattr(row, "sales_price", None) or 0)
        original_amount = float(getattr(row, "original_amount", None) or 0)
        avg_member_discount = float(getattr(row, "avg_member_discount", None) or 1.0)
        free_qty = float(getattr(row, "free_qty", None) or 0)
        give_qty = float(getattr(row, "give_qty", None) or 0)
        refund_qty = float(getattr(row, "refund_qty", None) or 0)
        refund_amount = float(getattr(row, "refund_amount", None) or 0)
        cancelled_qty = float(getattr(row, "cancelled_qty", None) or 0)
        cancelled_amount = float(getattr(row, "cancelled_amount", None) or 0)
        # 金额恒等式分项（堂食才有非零值）
        free_amount = float(getattr(row, "free_amount", None) or 0)
        give_amount = float(getattr(row, "give_amount", None) or 0)
        discount_amount = float(getattr(row, "discount_amount", None) or 0)
        list_price = float(getattr(row, "list_price", None) or 0)
        price_1 = getattr(row, "price_1", None)
        qty_1 = getattr(row, "qty_1", None)
        price_2 = getattr(row, "price_2", None)
        qty_2 = getattr(row, "qty_2", None)
        price_3 = getattr(row, "price_3", None)
        qty_3 = getattr(row, "qty_3", None)
        other_price_qty = getattr(row, "other_price_qty", None)

        key = (store_num, store_name, item_uuid, item_name)
        if key not in data:
            data[key] = {
                "qty": 0.0,
                "revenue": 0.0,
                "sales_price": 0.0,
                "gross_amount": 0.0,
                "original_amount": 0.0,
                "refund_qty": 0.0,
                "refund_amount": 0.0,
                "cancelled_qty": 0.0,
                "cancelled_amount": 0.0,
                "free_amount": 0.0,
                "give_amount": 0.0,
                "discount_amount": 0.0,
                "avg_member_discount": 0.0,
                "free_qty": 0.0,
                "give_qty": 0.0,
                "list_price": list_price,
                "price_1": price_1,
                "qty_1": qty_1,
                "price_2": price_2,
                "qty_2": qty_2,
                "price_3": price_3,
                "qty_3": qty_3,
                "other_price_qty": other_price_qty,
                "bom": {},
            }
        data[key]["qty"] += qty
        data[key]["revenue"] += revenue
        data[key]["sales_price"] += sales_price
        # 真实列 (sale_line/takeout_line 已投影 gross_amount), 毛额守恒为真校验
        data[key]["gross_amount"] += float(getattr(row, "gross_amount", 0) or 0)
        data[key]["original_amount"] += original_amount
        data[key]["refund_qty"] += refund_qty
        data[key]["refund_amount"] += refund_amount
        data[key]["cancelled_qty"] += cancelled_qty
        data[key]["cancelled_amount"] += cancelled_amount
        data[key]["free_amount"] += free_amount
        data[key]["give_amount"] += give_amount
        data[key]["discount_amount"] += discount_amount
        # 加权平均会员折扣率
        data[key]["avg_member_discount"] += avg_member_discount * qty
        data[key]["free_qty"] += free_qty
        data[key]["give_qty"] += give_qty

    # 归一化加权平均折扣率
    for key, val in data.items():
        if val["qty"] > 0:
            val["avg_member_discount"] = val["avg_member_discount"] / val["qty"]

    # 为每个 item 匹配 BOM — 委托给 semantic.cogs.expand_item_bom.
    # 4 层 priority 栈构造一次, 全店共享 (避免 per-row 重建).
    # bq_price 是 per-BOM-row 的兜底, 不进 Resolver, 由 price_fn 闭包从 row 上读。
    _price_resolver = _build_material_price_resolver(
        uploaded_prices, erp_prices, price_layers, strict=strict_price,
    )

    def _make_price_fn(rows):
        bq_price_by_code = {}
        for r in rows:
            mc = r[0]
            if mc:
                # 同物料多行: 后值覆盖前值, 跟旧 per-row 实现里"最后一次写"对齐.
                bq_price_by_code[mc] = r[5]  # (code, name, num, unit, conv, bq_price)

        def fn(material_code, material_name):
            r = _price_resolver.resolve((material_code, material_name))
            if r is not None:
                return r.value
            if strict_price:
                return 0.0
            return float(bq_price_by_code.get(material_code) or 0)
        return fn

    for key, val in data.items():
        store_num, _store_name, item_uuid, _item_name = key
        store_boms = bom_data.get(store_num, {})
        store_struct = combo_structure.get(store_num, {})

        # 收集"本商品会用到的 BOM 行"以便构造 per-row bq_price fallback
        if mode == "combo":
            relevant_rows = []
            for spec in store_struct.get(item_uuid, []):
                if isinstance(spec, (tuple, list)) and len(spec) == 3:
                    cu = spec[0]
                else:
                    cu = spec
                relevant_rows.extend(store_boms.get(cu, []))
        else:
            relevant_rows = store_boms.get(item_uuid, [])

        val["bom"] = expand_item_bom(
            item_uuid=item_uuid,
            mode=mode,
            store_boms=store_boms,
            store_combo_struct=store_struct,
            price_resolver=_make_price_fn(relevant_rows),
        )

    # 转换为列表格式（保留所有原始字段，利润指标由 Excel 公式计算）
    result = {}
    for key, val in data.items():
        # 衍生指标：净销量（真正卖出去的件数）
        net_qty = (val["qty"] - val["free_qty"] - val["give_qty"]
                   - val["refund_qty"] - val["cancelled_qty"])
        result[key] = {
            "qty": val["qty"],
            "net_qty": net_qty,
            "revenue": val["revenue"],
            "sales_price": val["sales_price"],
            # 真实列透传 — sale_line/takeout_line/total_line 已投影 gross_amount
            "gross_amount": val["gross_amount"],
            "original_amount": val["original_amount"],
            "refund_qty": val["refund_qty"],
            "refund_amount": val["refund_amount"],
            "cancelled_qty": val["cancelled_qty"],
            "cancelled_amount": val["cancelled_amount"],
            "free_amount": val["free_amount"],
            "give_amount": val["give_amount"],
            "discount_amount": val["discount_amount"],
            "avg_member_discount": val["avg_member_discount"],
            "free_qty": val["free_qty"],
            "give_qty": val["give_qty"],
            "list_price": val["list_price"],
            "price_1": val["price_1"],
            "qty_1": val["qty_1"],
            "price_2": val["price_2"],
            "qty_2": val["qty_2"],
            "price_3": val["price_3"],
            "qty_3": val["qty_3"],
            "other_price_qty": val["other_price_qty"],
            "bom": [
                (code, name, bom_num, price, uom)
                for code, (name, bom_num, price, uom) in val["bom"].items()
            ],
        }
    return result


# ============================================================================
# 扁平化行构建
# ============================================================================

def _build_rows(agg_data, mode, bom_layers=None, uploaded_prices=None, erp_prices=None,
                fallback_boms=None, price_layers=None, strict_price=False,
                strict_bom=False):
    """中间表行构建。

    BOM 选取顺序（priority 栈）:
        1. bom_layers 按 priority 从高到低逐层匹配，命中即 override BQ 原生 BOM
        2. 都没命中 → 用 BQ 原生 (data["bom"])
        3. BQ 也无 → 空行 (BOM 列填 "-")

    strict_bom=True 时禁用第 2 步: bom_layers 没命中直接标 BOM来源="无",
    抛弃 BQ 原生 BOM. 跟 strict_price 对称, 用于"BOM 只走客户事实表"场景.

    单价选取 (独立 priority 栈, 跟 BOM 数量解耦):
        price_layers → uploaded → ERPNext → bq_price
        strict_price=True 时只走 price_layers, 缺失 → 0 (审计列标 "无")

    `fallback_boms` 是老接口的兼容参数 (扁平 dict)，仅当未传 bom_layers 时使用。
    """
    # 老接口兼容: 把 fallback_boms 包成单层
    if bom_layers is None and fallback_boms:
        bom_layers = [("fallback_bom", 50, fallback_boms)]
    bom_layers = bom_layers or []
    """
    中间表：只输出原始数据，所有计算交给 Excel。
    列结构（26列）:
      0-2:   门店编号、门店名称、商品名称
      3-12:  当前标价、销量、营业额、标准金额、实收金额、会员折扣率、赠品数量、赠送数量、退款数量、退款金额
      13-19: 价格1、销量1、价格2、销量2、价格3、销量3、其他价格销量
      20-24: BOM物品名称、BOM物品编码、消耗数量、物料单价、单位
      25:    商品UUID(隐藏)
    """
    # 构造一次共享 resolver, 避免 per-BOM-row 重建
    _price_resolver = _build_material_price_resolver(
        uploaded_prices, erp_prices, price_layers, strict=strict_price)
    rows = []
    for (store_num, store_name, item_uuid, item_name), data in sorted(agg_data.items()):
        qty = data["qty"]
        revenue = data["revenue"]
        sales_price = data["sales_price"]
        original_amount = data["original_amount"]
        refund_qty = data["refund_qty"]
        refund_amount = data["refund_amount"]
        cancelled_qty = data.get("cancelled_qty", 0)
        cancelled_amount = data.get("cancelled_amount", 0)
        avg_member_discount = data["avg_member_discount"]
        free_qty = data["free_qty"]
        give_qty = data["give_qty"]
        list_price = data["list_price"]
        price_1 = data.get("price_1")
        qty_1 = data.get("qty_1")
        price_2 = data.get("price_2")
        qty_2 = data.get("qty_2")
        price_3 = data.get("price_3")
        qty_3 = data.get("qty_3")
        other_price_qty = data.get("other_price_qty")
        bom_list = data["bom"]
        if strict_bom:
            # 抛弃 BQ 原生 BOM, 等下要么命中 bom_layers, 要么标"无"
            bom_list = []
            bom_source = None
        else:
            bom_source = "bq_native" if bom_list else None
        price_sources_seen = set()   # 该 SKU 用到的所有单价来源 (审计)

        # 外挂 BOM override (priority 栈, 高 priority 优先, 覆盖 BQ 原生)
        matched, layer_name = _match_bom_layered(item_name, bom_layers)
        if matched:
            bom_list = []
            for code, name, bom_num, uom in matched:
                unit_price, p_src = _resolve_unit_price_with_source(
                    code, 0, uploaded_prices, erp_prices,
                    price_layers=price_layers, strict=strict_price,
                    material_name=name, resolver=_price_resolver)
                price_sources_seen.add(p_src)
                bom_list.append((code, name, bom_num, unit_price, uom or "-"))
            bom_source = layer_name
        elif not strict_bom:
            # BQ 原生 BOM 路径 — 单价已在 aggregate_with_bom 时按 price_layers 解析过
            for code, name, _bn, _up, _u in (data["bom"] or []):
                _, p_src = _resolve_unit_price_with_source(
                    code, 0, uploaded_prices, erp_prices,
                    price_layers=price_layers, strict=strict_price,
                    material_name=name, resolver=_price_resolver)
                price_sources_seen.add(p_src)
        if bom_source is None:
            bom_source = "无"

        # 价来源汇总字符串 (跨 BOM 行去重, 多个用 " + ")
        if price_sources_seen:
            price_source_str = " + ".join(sorted(price_sources_seen))
        else:
            price_source_str = "无"

        # row 末尾扩展槽位（field_index 26-29 = utility 公式列；30-31 = 标价应收/异常损失公式列；
        # 32 = 取消数量、33 = 取消金额、34 = BOM 来源；35 = 物料价来源）。
        tail = [None, None, None, None, None, None,
                round(cancelled_qty, 2), round(cancelled_amount, 2),
                bom_source, price_source_str]

        if not bom_list:
            rows.append([
                store_num, store_name, item_name,
                round(list_price, 2), round(qty, 2),
                round(sales_price, 2), round(original_amount, 2),
                round(revenue, 2), round(avg_member_discount, 4),
                round(free_qty, 2), round(give_qty, 2),
                round(refund_qty, 2), round(refund_amount, 2),
                price_1, qty_1, price_2, qty_2, price_3, qty_3, other_price_qty,
                "-", "-", None, None, "-",
                str(item_uuid),
            ] + tail)
            continue

        for code, name, bom_num, mat_price, uom in bom_list:
            rows.append([
                store_num, store_name, item_name,
                round(list_price, 2), round(qty, 2),
                round(sales_price, 2), round(original_amount, 2),
                round(revenue, 2), round(avg_member_discount, 4),
                round(free_qty, 2), round(give_qty, 2),
                round(refund_qty, 2), round(refund_amount, 2),
                price_1, qty_1, price_2, qty_2, price_3, qty_3, other_price_qty,
                name, code, round(bom_num, 4), round(mat_price, 4), uom or "-",
                str(item_uuid),
            ] + tail)

    return rows


def _build_summary_rows(agg_data, mode, bom_layers=None, uploaded_prices=None, erp_prices=None,
                        fallback_boms=None, price_layers=None, strict_price=False,
                        strict_bom=False):
    """SKU 汇总视角 — BOM 选取顺序跟 _build_rows 一致 (priority 栈 override).

    strict_bom=True 时禁用 BQ 原生 BOM 兜底, 跟 _build_rows 对称.

    `fallback_boms` 仅老接口兼容。
    """
    if bom_layers is None and fallback_boms:
        bom_layers = [("fallback_bom", 50, fallback_boms)]
    bom_layers = bom_layers or []
    """SKU 汇总视角：一个 (店, SKU) 一行，提前算好 总成本/总利润/利润率。

    跟 `_build_rows` 共享 agg_data 输入，但 flatten 方式不同：
      - _build_rows:        N 个 BOM 物料 → N 行（共享 SKU 维度列）
      - _build_summary_rows: BOM 列表压成"单份总成本"标量 → 1 行

    适合客户做透视/汇总。`negative_red` 在 yaml 中触发字体红色。

    Returns: list of [store_num, store_name, item_name, qty, net_qty,
                       sales_price, revenue, per_unit_cost, total_cost,
                       total_profit, margin, item_uuid]
    """
    # 构造一次共享 resolver, 避免 per-BOM-row 重建
    _price_resolver = _build_material_price_resolver(
        uploaded_prices, erp_prices, price_layers, strict=strict_price)
    rows = []
    for (store_num, store_name, item_uuid, item_name), data in sorted(agg_data.items()):
        qty = data["qty"]
        net_qty = data.get("net_qty", qty)
        sales_price = data["sales_price"]
        revenue = data["revenue"]
        bom_list = data["bom"]
        if strict_bom:
            bom_list = []
            bom_source = None
        else:
            bom_source = "bq_native" if bom_list else None
        price_sources_seen = set()

        # 外挂 BOM override (口径跟 _build_rows 一致, 高 priority 覆盖 BQ)
        matched, layer_name = _match_bom_layered(item_name, bom_layers)
        if matched:
            bom_list = []
            for code, name, bom_num, uom in matched:
                unit_price, p_src = _resolve_unit_price_with_source(
                    code, 0, uploaded_prices, erp_prices,
                    price_layers=price_layers, strict=strict_price,
                    material_name=name, resolver=_price_resolver)
                price_sources_seen.add(p_src)
                bom_list.append((code, name, bom_num, unit_price, uom or "-"))
            bom_source = layer_name
        elif not strict_bom:
            for code, name, _bn, _up, _u in (data["bom"] or []):
                _, p_src = _resolve_unit_price_with_source(
                    code, 0, uploaded_prices, erp_prices,
                    price_layers=price_layers, strict=strict_price,
                    material_name=name, resolver=_price_resolver)
                price_sources_seen.add(p_src)
        if bom_source is None:
            bom_source = "无"
        price_source_str = " + ".join(sorted(price_sources_seen)) if price_sources_seen else "无"

        # 单份总成本 = Σ (BOM 消耗 × 物料单价)
        per_unit_cost = sum(bom_num * unit_price
                            for _, _, bom_num, unit_price, _ in bom_list)
        # 总成本 = 单份成本 × 销量（含赠/退/取，因为 BOM 实际消耗了）
        total_cost = per_unit_cost * qty
        # 总利润 = 实收 - 总成本
        total_profit = revenue - total_cost
        margin = total_profit / revenue if revenue > 0 else 0

        rows.append([
            store_num, store_name, item_name,
            round(qty, 2),
            round(net_qty, 2),
            round(sales_price, 2),
            round(revenue, 2),
            round(per_unit_cost, 4),
            round(total_cost, 2),
            round(total_profit, 2),
            round(margin, 4),
            str(item_uuid),
            bom_source,           # field_index 12
            price_source_str,     # field_index 13
        ])
    return rows


# ============================================================================
# 资源配置加载
# ============================================================================

# 唯一活配置. 统一前散落在 resources/wallace.<日期>/config.yaml (导致每次跑
# 报表都可能用错版本); 统一后只有这一份, 新数据文件仍按 resources/wallace.<日期>/
# 归档, 但 config 只改 resources/config.yaml.
_DEFAULT_CONFIG_PATH = os.path.join(
    os.path.dirname(__file__), "..", "resources", "config.yaml")


def _resolve_latest_config() -> str | None:
    """兜底: resources/config.yaml 不存在时, 找 resources/wallace.*/config.yaml
    最新一份 (统一前的旧布局). 正常流程不会走到这里.
    """
    import glob
    pattern = os.path.join(os.path.dirname(__file__), "..",
                            "resources", "wallace.*", "config.yaml")
    matches = sorted(glob.glob(pattern))
    return matches[-1] if matches else None


def load_config(config_path: str = None) -> dict:
    """加载资源配置 YAML.

    --config 没传时, 默认读 resources/config.yaml (唯一活配置).
    该文件不存在时兜底找 resources/wallace.*/config.yaml 最新一份 (旧布局).
    stdout 打印实际用的哪个 — AI / 人审计一目了然.
    """
    import yaml
    if config_path is None:
        if os.path.exists(_DEFAULT_CONFIG_PATH):
            config_path = _DEFAULT_CONFIG_PATH
        else:
            config_path = _resolve_latest_config()
            if config_path:
                print(f"[Config] resources/config.yaml 不存在, 兜底取旧布局: "
                      f"{os.path.relpath(config_path)}")
    if not config_path or not os.path.exists(config_path):
        return {}
    print(f"[Config] 加载: {os.path.relpath(config_path)}")
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


# ============================================================================
# 主流程
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="利润报表导出")
    parser.add_argument("--mode", default="both", choices=["combo", "single", "both"], help="报表模式")
    parser.add_argument("--month", default=None, help="月份，格式 YYYY-MM（与 --start-date/--end-date 互斥）")
    parser.add_argument("--start-date", default=None, help="开始日期，格式 YYYY-MM-DD")
    parser.add_argument("--end-date", default=None, help="结束日期，格式 YYYY-MM-DD")
    parser.add_argument("--merchants", default="resources/merchants.xlsx", help="商家列表 Excel 路径")
    parser.add_argument("--output", default=None, help="输出 Excel 文件路径（默认自动推导）")
    parser.add_argument("--project", default="diyl-407103", help="GCP 项目 ID")
    parser.add_argument("--use-erp-price", action="store_true", default=True, help="启用 ERPNext Item Price 替换 BQ 成本（默认开启）")
    parser.add_argument("--no-erp-price", action="store_true", help="禁用 ERPNext 价格，使用 BQ 内置价格")
    parser.add_argument("--erp-price-list", default=None, help="ERPNext 价格表名称，默认 Standard Buying")
    parser.add_argument("--config", default=None, help="资源配置 YAML 路径")
    parser.add_argument("--column-config", default="resources/reports/profit_margin.yaml", help="列配置 YAML 路径")
    parser.add_argument("--price-list", default=None, help="上传的物料价格清单 Excel 路径（最高优先级）")
    # 单价默认 strict: 只走 material_price_sources (客户成本表),
    # 缺失 → 0 + 来源标 '无 (strict)' + 报表里那行物料标黄.
    # --allow-erp-fallback 反向开关 (应急): 允许 fallback ERPNext.
    parser.add_argument("--allow-erp-fallback", action="store_true",
                        help="允许 fallback ERPNext (默认禁用)。默认: 单价只走 material_price_sources, "
                             "缺失即留空, 报表里那行物料标黄.")
    # BOM 默认 strict: 只走 bom_sources (客户事实表), 缺失 → BOM来源="无".
    # --allow-bq-native-bom 反向开关 (应急): 允许 fallback BQ 内置 product_bom.
    parser.add_argument("--allow-bq-native-bom", action="store_true",
                        help="允许 fallback BQ 内置 BOM (默认禁用)。默认: BOM 只走 bom_sources, "
                             "缺失即标 BOM来源='无', 成本列空.")
    parser.add_argument("--summary", action="store_true",
                        help="汇总视角：每 SKU 一行（不展开 BOM 物料），默认输出 sku_profit_summary.yaml 格式")
    parser.add_argument("--no-cache", action="store_true", help="禁用缓存")
    parser.add_argument("--force", action="store_true",
                        help="强制导出即使校验未通过 (文件将带水印, 不得对外交付)")
    args = parser.parse_args()
    # --summary 自动切换列配置（除非用户显式覆盖）
    if args.summary and args.column_config == "resources/reports/profit_margin.yaml":
        args.column_config = "resources/reports/sku_profit_summary.yaml"

    # 时间范围解析
    if args.month and (args.start_date or args.end_date):
        print("[错误] --month 与 --start-date/--end-date 不能同时使用")
        return 1

    if args.month:
        assert_month_not_frozen(args.month)
        start_ts, end_ts = _month_to_ts_range(args.month)
        range_label = args.month.replace("-", "")
    elif args.start_date and args.end_date:
        start_dt = datetime.strptime(args.start_date, "%Y-%m-%d").replace(tzinfo=BKK_TZ)
        end_dt = datetime.strptime(args.end_date, "%Y-%m-%d").replace(tzinfo=BKK_TZ) + timedelta(days=1)
        range_label = f"{args.start_date.replace('-', '')}_{args.end_date.replace('-', '')}"
        start_ts = int(start_dt.timestamp())
        end_ts = int(end_dt.timestamp())
    else:
        print("[错误] 必须指定 --month 或 --start-date + --end-date")
        return 1

    # 自动推导输出路径; 不传 --output 时每次升版 _v{N}, 跟 profit_by_price 一致
    if args.output:
        output_path = args.output
    else:
        from bq_reports.profit_by_price_report import _next_version_path
        base_dir = Path("exports")
        base_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(_next_version_path(base_dir, f"profit_{range_label}"))

    # 初始化引擎
    engine = ReportEngine(project_id=args.project)

    # 加载资源配置
    config = load_config(args.config)

    # 加载上传价格清单（最高优先级）
    uploaded_prices = {}
    if args.price_list:
        uploaded_prices, _ = _load_uploaded_prices(args.price_list)
        print()

    # 加载 ERPNext 价格（带缓存，fallback）
    erp_prices = {}
    if args.use_erp_price and not args.no_erp_price:
        erp_ttl = 0 if args.no_cache else config.get("cache", {}).get("erp_prices_ttl", 3600)
        erp_prices = _try_load_erp_prices(price_list=args.erp_price_list, cache_ttl=erp_ttl)
        if not erp_prices:
            print("[警告] 未加载到 ERPNext 价格，成本将显示为 0")
        print()

    # 加载门店名称映射
    store_names = _load_store_names(config, client=engine.client)

    # 加载商家列表
    merchants = _load_merchants(config, store_names, override_path=args.merchants,
                                  project_id=args.project)

    range_desc = args.month if args.month else f"{args.start_date} ~ {args.end_date}"
    print(f"模式: {args.mode}, 时间: {range_desc}")
    print(f"时间范围: {start_ts} - {end_ts}")
    print(f"门店数: {len(merchants)}")
    print()

    # 加载外挂 BOM 层 (按 priority 从高到低)
    bom_layers = _load_bom_layers(config)

    # 加载物料单价层 (独立 priority 栈, 跟 BOM 数量解耦)
    price_layers = _load_material_price_layers(config)
    strict_price = not bool(args.allow_erp_fallback)   # 默认 strict
    if strict_price:
        print(f"[Strict Price] 默认严格模式: 物料单价只走外挂成本表, 缺失 → 0 (不 fallback ERP)")
        print(f"               想恢复 ERP fallback 用 --allow-erp-fallback\n")
    else:
        print(f"[ERP Fallback] 允许 fallback ERPNext\n")

    strict_bom = not bool(args.allow_bq_native_bom)   # 默认 strict
    if strict_bom:
        print(f"[Strict BOM] 默认严格模式: BOM 只走 bom_sources, 缺失 → 来源='无' (不 fallback BQ 内置)")
        print(f"             想恢复 BQ 内置 BOM 用 --allow-bq-native-bom\n")
    else:
        print(f"[BQ Native BOM Fallback] 允许 fallback BQ 内置 product_bom\n")

    # 预加载套餐结构（如果 mode 包含 combo）
    combo_structure = {}
    if args.mode in ("combo", "both"):
        combo_structure = _load_combo_structures(engine, merchants, start_ts, end_ts, config)
        print()

    # 预加载 BOM（所有 mode 都需要）
    bom_data = _load_boms(engine, merchants, config)
    print()

    # 确定要处理的 mode
    modes = []
    if args.mode in ("combo", "both"):
        modes.append("combo")
    if args.mode in ("single", "both"):
        modes.append("single")

    # 准备输出
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    import xlsxwriter
    wb = xlsxwriter.Workbook(str(output_path))

    # 测试营业时段过滤 (对齐 ttpos 后台 ExcludeTestBusinessByBillSQL 口径)
    from semantic.dimensions.test_business import get_stores_with_test_business
    tb_stores = get_stores_with_test_business(
        engine.client, [m[1] for m in merchants])
    if tb_stores:
        print(f"\n[Test Business] 排除测试营业时段, 影响 {len(tb_stores)} 店")

    watermarked = False  # 多 mode 循环只打一次水印
    from semantic.validators.gate import (
        add_watermark_sheet_xlsxwriter, validate_and_gate)
    from semantic.validators.identities import FULL_IDENTITIES

    for mode in modes:
        item_label = "套餐" if mode == "combo" else "单品"
        sql_template = COMBO_ORDERS_SQL if mode == "combo" else SINGLE_ORDERS_SQL
        sql_template_tb = COMBO_ORDERS_SQL_TB if mode == "combo" else SINGLE_ORDERS_SQL_TB
        print(f"\n========== 开始处理 {item_label} ==========\n")

        sql_factory = (
            (lambda u: sql_template_tb if u in tb_stores else sql_template)
            if tb_stores else None
        )
        # 并发查询聚合后的订单
        raw_rows, errors = engine.query(
            sql_template=sql_template,
            merchants=merchants,
            start_ts=start_ts,
            end_ts=end_ts,
            workers=10,
            row_proxy_factory=lambda row, acc, num, name: type("RowProxy", (), {
                "__getattr__": lambda self, attr: getattr(row, attr),
                "account": acc, "store_num": num, "store_name": name,
            })(),
            label=item_label,
            sql_template_factory=sql_factory,
        )

        # 聚合（订单 + BOM）
        agg_data = aggregate_with_bom(
            raw_rows, bom_data, combo_structure,
            uploaded_prices=uploaded_prices, erp_prices=erp_prices, mode=mode,
            price_layers=price_layers, strict_price=strict_price,
        )

        # P2: 把 bom_source / price_source 写回 agg_data, 让后续 validator 用
        _annotate_agg_data_sources(
            agg_data, bom_layers, uploaded_prices, erp_prices,
            price_layers, strict_price=strict_price,
        )

        # 扁平化（汇总视角 or BOM 展开视角）
        common = dict(uploaded_prices=uploaded_prices, erp_prices=erp_prices,
                      bom_layers=bom_layers, price_layers=price_layers,
                      strict_price=strict_price, strict_bom=strict_bom)
        if args.summary:
            flat_rows = _build_summary_rows(agg_data, mode, **common)
        else:
            flat_rows = _build_rows(agg_data, mode, **common)

        # 加载列配置并写入 Excel
        sheet_cfg = engine.load_sheet_config(args.column_config, item_label)
        engine.write_sheet(wb, item_label, sheet_cfg, flat_rows)

        print(f"\n[{item_label}] 总明细行数: {len(flat_rows)} 行")
        if not args.summary:
            print(f"[{item_label}] 此为中间表，利润指标请自行在 Excel 中定义")

        # 零容差闸门: 校验恒等式 + 来源完整性 + 业务合理性 (FULL_IDENTITIES)
        # agg_data 已带 bom_source / price_source (P2 annotation)
        # 闸门在 wb.close() 前执行 — 有 MUST_FIX 且无 --force 则 exit(2), 不落盘
        check_rows = [
            {"store_num": store_num, "item_name": item_name, **data}
            for (store_num, _store_name, _uuid, item_name), data
            in agg_data.items()
        ]
        outcome = validate_and_gate(
            check_rows, FULL_IDENTITIES,
            force=args.force, report_name=f"profit_margin/{item_label}",
            row_label=lambda r: f"店 {r['store_num']:>3}  {r['item_name']:<30}")
        if outcome.needs_watermark and not watermarked:
            add_watermark_sheet_xlsxwriter(wb, outcome.watermark_lines())
            watermarked = True

    wb.close()
    print(f"\n输出文件: {output_path}")
    # NOTE: 物料单价为空标黄的后处理已禁用 (openpyxl 重开慢, 见 profit_by_price_report 注释)
    return 0


if __name__ == "__main__":
    sys.exit(main())
