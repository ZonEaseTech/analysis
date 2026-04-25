#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
华莱士门店单品销量按日明细 (BigQuery 版本)

堂食 + 外卖，按日汇总

输出列：门店编号、门店名称、单品名称、单品销量、日期
"""

import os
import sys
import csv
import calendar
import subprocess
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from google.cloud import bigquery
from google.oauth2.credentials import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ID = "diyl-407103"
OUTPUT_DIR = Path(__file__).parent / "output"

STORE_LIST = [
    ("1", "1958987436032000"), ("2", "2269470793728000"), ("3", "2598648160256000"),
    ("4", "2876210421760000"), ("5", "3446618988544000"), ("6", "3870122057728000"),
    ("7", "4149605310464000"), ("8", "4358842359808000"), ("9", "4912616316928000"),
    ("10", "5567347171328000"), ("11", "5999171739648000"), ("12", "6542950670336000"),
    ("13", "6789240201216000"), ("14", "6977459593216000"), ("15", "7191251656704000"),
    ("16", "7400123801600000"), ("17", "7648065888256000"), ("18", "7863653113856000"),
    ("19", "8100551598080000"), ("20", "8501761941504000"), ("21", "8722592047104000"),
    ("22", "2947521978368000"), ("23", "3448951017472000"), ("24", "3782477877248000"),
    ("25", "4024875094016000"), ("26", "4229506797568000"), ("27", "4418766376960000"),
    ("28", "4613872816128000"), ("29", "4805464432640000"), ("30", "5001267122176000"),
    ("31", "5250979205120000"), ("32", "5444022046720000"), ("33", "7600687026176000"),
    ("34", "7813128523776000"), ("35", "8051063001088000"), ("36", "8535580610560000"),
    ("37", "8723170856960000"), ("38", "1515821506560000"), ("39", "3631470387200000"),
    ("40", "5498438983680000"), ("42", "9231705128960000"), ("44", "1379607252992000"),
    ("46", "1745354756096000"), ("47", "1919875551232000"), ("49", "2277263806464000"),
    ("51", "2618629820416000"), ("52", "2788834676736000"), ("53", "2992442970112000"),
    ("54", "3169446793216000"), ("55", "3367514411008000"), ("56", "3662462062592000"),
    ("59", "4197894328320000"), ("61", "3087884357632000"),
]

UUID_SET = {uuid for _, uuid in STORE_LIST}


def log(msg: str):
    print(msg, flush=True)


def setup_proxy():
    proxy_url = os.environ.get("BQ_PROXY")
    if not proxy_url:
        return
    for k in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"]:
        os.environ[k] = proxy_url


def get_creds():
    r = subprocess.run(["gcloud", "auth", "print-access-token"], capture_output=True, text=True, check=True)
    return Credentials(token=r.stdout.strip(), scopes=["https://www.googleapis.com/auth/cloud-platform"])


def get_time_range():
    if len(sys.argv) > 1:
        ym = sys.argv[1]
    else:
        d = date.today().replace(day=1) - timedelta(days=1)
        ym = d.strftime("%Y-%m")
    y, m = map(int, ym.split('-'))
    ld = calendar.monthrange(y, m)[1]
    st = int(datetime(y, m, 1, 0, 0, 0).timestamp())
    et = int(datetime(y, m, ld, 23, 59, 59).timestamp())
    return ym, st, et


def query_daily_sales(client, dataset_id, start_ts, end_ts):
    """查询单品按日销量（堂食+外卖）"""
    # 多语言名称表达式 (BigQuery 语法)
    pp_name_expr = """
    (SELECT STRING_AGG(val, ' / ')
     FROM UNNEST([
        NULLIF(JSON_EXTRACT_SCALAR(pp.name, '$.zh'), 'null'),
        NULLIF(JSON_EXTRACT_SCALAR(pp.name, '$.en'), 'null'),
        NULLIF(JSON_EXTRACT_SCALAR(pp.name, '$.th'), 'null'),
        NULLIF(JSON_EXTRACT_SCALAR(pp.name, '$.ja'), 'null'),
        NULLIF(JSON_EXTRACT_SCALAR(pp.name, '$.ko'), 'null')
     ]) AS val
     WHERE val IS NOT NULL)
    """
    
    # 外卖 UNION（如果存在外卖表）
    to_union = f"""
    UNION ALL
    SELECT
        DATE(TIMESTAMP_SECONDS(
            CASE WHEN tko.order_state = 40 THEN tko.completed_time ELSE tko.accepted_time END
        )) AS sale_date,
        COALESCE({pp_name_expr}, toi.item_name) AS product_name,
        CAST(toi.quantity AS FLOAT64) AS qty
    FROM `{PROJECT_ID}.{dataset_id}.ttpos_takeout_order_item` toi
    INNER JOIN `{PROJECT_ID}.{dataset_id}.ttpos_takeout_order` tko 
        ON tko.uuid = toi.takeout_order_uuid AND tko.delete_time = 0
    LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_product_package` pp 
        ON pp.uuid = toi.ttpos_product_package_uuid AND pp.delete_time = 0
    WHERE toi.delete_time = 0
      AND tko.order_state IN (10, 20, 30, 40, 60)
      AND tko.accepted_time > 0
      AND (
          (tko.order_state = 40 AND tko.completed_time > 0 
           AND tko.completed_time >= {start_ts} AND tko.completed_time <= {end_ts})
          OR
          (tko.order_state != 40 AND tko.accepted_time >= {start_ts} AND tko.accepted_time <= {end_ts})
      )
    """
    
    query = f"""
    WITH dine_sales AS (
        -- 堂食：statistics_product 预聚合表
        SELECT
            DATE(TIMESTAMP_SECONDS(sp.complete_time)) AS sale_date,
            {pp_name_expr} AS product_name,
            sp.product_num AS qty
        FROM `{PROJECT_ID}.{dataset_id}.ttpos_statistics_product` sp
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_product_package` pp 
            ON pp.uuid = sp.product_package_uuid AND pp.delete_time = 0
        WHERE sp.delete_time = 0
          AND sp.complete_time > 0
          AND sp.complete_time >= {start_ts}
          AND sp.complete_time <= {end_ts}
        
        UNION ALL
        
        -- 堂食：套餐子品 product_type=2
        SELECT
            DATE(TIMESTAMP_SECONDS(
                COALESCE(NULLIF(sb.finish_time, 0), so.finish_time)
            )) AS sale_date,
            {pp_name_expr} AS product_name,
            sop.num * sop.copy_num * IF(sop.unit_num = 0, 1, sop.unit_num) AS qty
        FROM `{PROJECT_ID}.{dataset_id}.ttpos_sale_order_product` sop
        INNER JOIN `{PROJECT_ID}.{dataset_id}.ttpos_sale_bill` sb 
            ON sb.uuid = sop.sale_bill_uuid AND sb.delete_time = 0
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_sale_order` so 
            ON so.uuid = sop.sale_order_uuid AND so.delete_time = 0
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_product_package` pp 
            ON pp.uuid = sop.product_package_uuid AND pp.delete_time = 0
        WHERE sop.delete_time = 0
          AND sop.cancel_time = 0
          AND sb.status = 1
          AND sop.product_type = 2
          AND COALESCE(NULLIF(sb.finish_time, 0), so.finish_time) > 0
          AND COALESCE(NULLIF(sb.finish_time, 0), so.finish_time) >= {start_ts}
          AND COALESCE(NULLIF(sb.finish_time, 0), so.finish_time) <= {end_ts}
        
        {to_union}
    )
    SELECT 
        sale_date,
        product_name,
        ROUND(SUM(qty), 4) AS total_qty
    FROM dine_sales
    WHERE qty > 0 AND product_name IS NOT NULL
    GROUP BY sale_date, product_name
    ORDER BY sale_date, product_name
    """
    
    try:
        result = client.query(query).result()
        return [(r.sale_date, r.product_name, r.total_qty) for r in result]
    except Exception as e:
        log(f"  {dataset_id} 查询失败: {e}")
        return []


def get_store_name(client, dataset_id):
    try:
        r = list(client.query(f"""
            SELECT JSON_EXTRACT_SCALAR(values, '$.store_code') as no,
                   JSON_EXTRACT_SCALAR(values, '$.store_name') as name
            FROM `{PROJECT_ID}.{dataset_id}.ttpos_setting`
            WHERE `key` = 'store' AND delete_time = 0 LIMIT 1
        """).result())[0]
        return r.no or "", r.name or ""
    except:
        return "", ""


def process_one(args):
    idx, total, uuid, start_ts, end_ts = args
    dataset_id = f"shop{uuid}"
    
    try:
        client = bigquery.Client(project=PROJECT_ID, credentials=get_creds())
    except Exception as e:
        return {"error": str(e), "idx": idx}
    
    store_no, store_name = get_store_name(client, dataset_id)
    if not store_name:
        store_name = f"门店_{uuid[:12]}"
    
    t0 = time.time()
    rows = query_daily_sales(client, dataset_id, start_ts, end_ts)
    elapsed = time.time() - t0
    
    return {
        "idx": idx,
        "store_no": store_no,
        "store_name": store_name,
        "rows": rows,
        "count": len(rows),
        "elapsed": elapsed
    }


def main():
    t0 = time.time()
    log("=" * 60)
    log("华莱士门店单品销量按日明细 (BigQuery 版本)")
    log("=" * 60)
    
    setup_proxy()
    
    ym, start_ts, end_ts = get_time_range()
    log(f"统计月份: {ym}")
    log(f"时间戳: {start_ts} ~ {end_ts}\n")
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # 获取所有门店数据集
    client = bigquery.Client(project=PROJECT_ID, credentials=get_creds())
    all_datasets = [d.dataset_id for d in client.list_datasets() if d.dataset_id.startswith("shop")]
    shop_datasets = [(i+1, uuid) for i, (_, uuid) in enumerate(STORE_LIST) 
                     if f"shop{uuid}" in all_datasets]
    
    log(f"找到 {len(shop_datasets)} 个门店数据集")
    log(f"开始并发查询 (5并发)...\n")
    
    all_data = []
    total_rows = 0
    query_time = 0
    
    query_start = time.time()
    with ThreadPoolExecutor(max_workers=5) as ex:
        futures = {ex.submit(process_one, (idx, len(shop_datasets), uuid, start_ts, end_ts)): idx 
                   for idx, uuid in shop_datasets}
        
        for fut in as_completed(futures):
            r = fut.result()
            if "error" in r:
                log(f"[{r['idx']}] 错误: {r['error']}")
                continue
            
            log(f"[{r['idx']}/{len(shop_datasets)}] {r['store_name']}: {r['count']} 行 ({r['elapsed']:.2f}s)")
            query_time += r['elapsed']
            
            for row in r["rows"]:
                all_data.append((r["store_no"], r["store_name"], row[0], row[1], row[2]))
            total_rows += r["count"]
    
    query_elapsed = time.time() - query_start
    log(f"\n查询完成: {total_rows:,} 行数据")
    log(f"纯查询耗时: {query_elapsed:.2f} 秒")
    
    # 生成 CSV
    log("\n正在生成 CSV...")
    csv_start = time.time()
    csv_path = OUTPUT_DIR / f"华莱士门店单品销量按日-{ym}.csv"
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(['门店编号', '门店名称', '日期', '单品名称', '单品销量'])
        for row in all_data:
            w.writerow([row[0], row[1], str(row[2]), row[3], f"{row[4]:.4f}"])
    csv_elapsed = time.time() - csv_start
    log(f"CSV 生成耗时: {csv_elapsed:.2f} 秒")
    
    # 生成 Excel
    log("正在生成 Excel...")
    xlsx_start = time.time()
    xlsx_path = OUTPUT_DIR / f"华莱士门店单品销量按日-{ym}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = f"单品销量-{ym}"
    
    headers = ['门店编号', '门店名称', '日期', '单品名称', '单品销量']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin = Border(left=Side(style='thin', color='D9D9D9'),
                  right=Side(style='thin', color='D9D9D9'),
                  top=Side(style='thin', color='D9D9D9'),
                  bottom=Side(style='thin', color='D9D9D9'))
    
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin
        c.alignment = Alignment(horizontal='center')
    
    for ri, row in enumerate(all_data, 2):
        ws.cell(row=ri, column=1, value=row[0]).border = thin
        ws.cell(row=ri, column=2, value=row[1]).border = thin
        ws.cell(row=ri, column=3, value=str(row[2])).border = thin
        ws.cell(row=ri, column=4, value=row[3]).border = thin
        c = ws.cell(row=ri, column=5, value=float(row[4]))
        c.border = thin
        c.number_format = "#,##0.####"
        
        if ri % 10000 == 0:
            log(f"  已写入 {ri-1} 行...")
    
    for i, w in enumerate([12, 28, 12, 40, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    xlsx_elapsed = time.time() - xlsx_start
    log(f"Excel 生成耗时: {xlsx_elapsed:.2f} 秒")
    
    total_elapsed = time.time() - t0
    
    log(f"\n{'='*60}")
    log(f"完成!")
    log(f"总数据行数: {total_rows:,} 行")
    log(f"纯查询耗时: {query_elapsed:.2f} 秒")
    log(f"CSV 生成: {csv_elapsed:.2f} 秒")
    log(f"Excel 生成: {xlsx_elapsed:.2f} 秒")
    log(f"总耗时: {total_elapsed:.2f} 秒 ({total_elapsed/60:.2f} 分钟)")
    log(f"CSV: {csv_path}")
    log(f"Excel: {xlsx_path}")
    log(f"{'='*60}")


if __name__ == "__main__":
    main()
