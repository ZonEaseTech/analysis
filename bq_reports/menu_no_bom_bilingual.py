#!/usr/bin/env python3
"""
Menu 未设置 BOM 菜品导出 —— 中英泰三语版

输出：单 sheet Excel
  列: 商品类型 | 商品名称(中) | 商品名称(英) | 商品名称(泰) | 分类(中) | 分类(英) | 分类(泰)

范围：单品 + 加料（product_type=0 的单品 + product_sauce 表加料）
      未配置 BOM 的商品

Usage:
    venv/bin/python -m bq_reports.menu_no_bom_bilingual --output exports/menu_no_bom.xlsx
"""

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
DEFAULT_DATASET = "shop1958987436032000"

# 单品 —— 未配置 BOM 的（LEFT JOIN 后 material_name IS NULL）
SINGLE_NO_BOM_SQL = """
WITH bom_dedup AS (
  SELECT
    product_package_uuid,
    product_bom_card_uuid,
    MIN(product_flavor_uuid) AS flavor_uuid
  FROM `{project}`.`{dataset}`.`ttpos_product_bom`
  WHERE delete_time = 0 AND product_bom_card_uuid > 0
  GROUP BY product_package_uuid, product_bom_card_uuid
),
pkg_card_count AS (
  SELECT product_package_uuid, COUNT(*) AS card_cnt
  FROM bom_dedup
  GROUP BY product_package_uuid
),
pkg_meta AS (
  SELECT
    pp.uuid AS pkg_uuid,
    REGEXP_REPLACE(COALESCE(
      JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
      JSON_EXTRACT_SCALAR(pp.name, '$.en'),
      ''
    ), r'^\\s+|\\s+$', '') AS base_name,
    COALESCE(
      JSON_EXTRACT_SCALAR(c.name, '$.zh'),
      JSON_EXTRACT_SCALAR(c.name, '$.en'),
      ''
    ) AS category_name
  FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
  LEFT JOIN `{project}`.`{dataset}`.`ttpos_product_category` c
    ON c.uuid = pp.category_uuid AND c.delete_time = 0
  WHERE pp.delete_time = 0 AND pp.product_type = 0
),
duplicate_names AS (
  SELECT base_name
  FROM pkg_meta
  WHERE base_name != ''
  GROUP BY base_name
  HAVING COUNT(*) > 1
)
SELECT
  0 AS sort_type,
  pp.uuid AS product_uuid,
  CASE
    WHEN IFNULL(pcc.card_cnt, 0) > 1 AND pf.uuid IS NOT NULL THEN CONCAT(
      pm.base_name,
      ' (',
      COALESCE(JSON_EXTRACT_SCALAR(pf.name, '$.zh'),
               JSON_EXTRACT_SCALAR(pf.name, '$.en'),
               ''),
      ')'
    )
    WHEN dn.base_name IS NOT NULL AND pm.category_name != '' THEN CONCAT(
      pm.base_name,
      ' (',
      pm.category_name,
      ')'
    )
    ELSE pm.base_name
  END AS product_name,
  JSON_EXTRACT_SCALAR(pp.name, '$.en') AS product_name_en,
  JSON_EXTRACT_SCALAR(pp.name, '$.th') AS product_name_th,
  COALESCE(JSON_EXTRACT_SCALAR(c.name, '$.zh'), '') AS category_name,
  COALESCE(JSON_EXTRACT_SCALAR(c.name, '$.en'), '') AS category_name_en,
  COALESCE(JSON_EXTRACT_SCALAR(c.name, '$.th'), '') AS category_name_th,
  rm.material_uuid AS has_bom
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
LEFT JOIN bom_dedup pb ON pb.product_package_uuid = pp.uuid
LEFT JOIN pkg_card_count pcc ON pcc.product_package_uuid = pp.uuid
LEFT JOIN pkg_meta pm ON pm.pkg_uuid = pp.uuid
LEFT JOIN duplicate_names dn ON dn.base_name = pm.base_name
LEFT JOIN `{project}`.`{dataset}`.`ttpos_product_flavor` pf
  ON pf.uuid = pb.flavor_uuid AND pf.delete_time = 0
LEFT JOIN `{project}`.`{dataset}`.`ttpos_related_material` rm
  ON rm.related_uuid = pb.product_bom_card_uuid AND rm.delete_time = 0
LEFT JOIN `{project}`.`{dataset}`.`ttpos_product_category` c
  ON c.uuid = pp.category_uuid AND c.delete_time = 0
WHERE pp.delete_time = 0 AND pp.product_type = 0
"""

# 加料 —— 未配置 BOM 的
SAUCE_NO_BOM_SQL = """
SELECT
  1 AS sort_type,
  ps.uuid AS product_uuid,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(ps.name, '$.zh'),
    JSON_EXTRACT_SCALAR(ps.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS product_name,
  JSON_EXTRACT_SCALAR(ps.name, '$.en') AS product_name_en,
  JSON_EXTRACT_SCALAR(ps.name, '$.th') AS product_name_th,
  '' AS category_name,
  '' AS category_name_en,
  '' AS category_name_th,
  rm.material_uuid AS has_bom
FROM `{project}`.`{dataset}`.`ttpos_product_sauce` ps
LEFT JOIN `{project}`.`{dataset}`.`ttpos_related_material` rm
  ON rm.related_uuid = ps.product_bom_card_uuid
  AND ps.product_bom_card_uuid > 0
  AND rm.delete_time = 0
WHERE ps.delete_time = 0
"""

TYPE_LABEL = {0: "单品", 1: "加料"}


def query_no_bom_rows(client, dataset: str):
    sql = f"""
    {SINGLE_NO_BOM_SQL.format(project=client.project, dataset=dataset)}
    UNION ALL
    {SAUCE_NO_BOM_SQL.format(project=client.project, dataset=dataset)}
    ORDER BY sort_type, product_name
    """
    rows = list(client.query(sql).result())
    # 过滤：只保留 has_bom IS NULL（即没有配置 BOM）的
    no_bom = []
    seen = set()
    for r in rows:
        if r.has_bom is not None:
            continue
        key = (r.sort_type, r.product_uuid)
        if key in seen:
            continue
        seen.add(key)
        no_bom.append(r)
    return no_bom


def _style_header(ws, headers):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, size=11, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align


def _merge_runs(ws, col_letter: str, start_row: int, end_row: int, value_of):
    if end_row <= start_row:
        return
    run_start = start_row
    run_value = value_of(start_row)
    for r in range(start_row + 1, end_row + 1):
        v = value_of(r)
        if v != run_value:
            if r - 1 > run_start:
                ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{r - 1}")
            run_start = r
            run_value = v
    if end_row > run_start:
        ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{end_row}")


def write_excel(no_bom_rows, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "未设置BOM菜品"

    headers = [
        "商品类型", "商品名称(中)", "商品名称(英)", "商品名称(泰)",
        "分类(中)", "分类(英)", "分类(泰)",
    ]
    _style_header(ws, headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(no_bom_rows, start=2):
        ws.cell(row=r_idx, column=1, value=TYPE_LABEL.get(row.sort_type, "")).alignment = center
        ws.cell(row=r_idx, column=2, value=row.product_name or "").alignment = left
        ws.cell(row=r_idx, column=3, value=row.product_name_en or "").alignment = left
        ws.cell(row=r_idx, column=4, value=row.product_name_th or "").alignment = left
        ws.cell(row=r_idx, column=5, value=row.category_name or "").alignment = center
        ws.cell(row=r_idx, column=6, value=row.category_name_en or "").alignment = center
        ws.cell(row=r_idx, column=7, value=row.category_name_th or "").alignment = center

    if no_bom_rows:
        last_row = 1 + len(no_bom_rows)
        if last_row > 2:
            _merge_runs(ws, "A", 2, last_row,
                        lambda r: TYPE_LABEL.get(no_bom_rows[r - 2].sort_type, ""))

    col_widths = [12, 28, 28, 35, 14, 14, 16]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args():
    p = argparse.ArgumentParser(description="Menu 未设置 BOM 菜品导出（中英泰三语）")
    p.add_argument("--project", default=PROJECT_ID, help="GCP 项目 ID")
    p.add_argument("--dataset", default=DEFAULT_DATASET,
                   help=f"BQ dataset(任意一家分店即可,默认 {DEFAULT_DATASET})")
    p.add_argument("--output", default="exports/menu_no_bom_bilingual.xlsx",
                   help="输出 Excel 文件路径")
    return p.parse_args()


def main():
    args = parse_args()
    setup_proxy()

    print(f"[BQ] project={args.project} dataset={args.dataset}")
    client = get_bq_client(project_id=args.project)

    print("[BQ] 查询未设置 BOM 的单品 + 加料...")
    no_bom = query_no_bom_rows(client, args.dataset)
    print(f"[BQ] 未设置 BOM 共 {len(no_bom)} 个")

    singles = sum(1 for r in no_bom if r.sort_type == 0)
    sauces = sum(1 for r in no_bom if r.sort_type == 1)

    write_excel(no_bom, args.output)
    print(f"[完成] 输出 {args.output}")
    print(f"  单品: {singles} 个")
    print(f"  加料: {sauces} 个")
    return 0


if __name__ == "__main__":
    sys.exit(main())
