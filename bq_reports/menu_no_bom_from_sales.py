#!/usr/bin/env python3
"""
基于销售报表「未设置BOM商品销量」导出干净的 menu 中英泰对照表

逻辑：
  1. 读取上传的「销售业绩+物品消耗表」中的「未设置BOM商品销量」sheet
  2. 用商品名去 BQ 匹配（只保留匹配到的）
  3. 输出中英泰三语 + 商品类型 + 销量汇总

Usage:
    venv/bin/python -m bq_reports.menu_no_bom_from_sales \
        --input /path/to/销售业绩+物品消耗表.xlsx \
        --output exports/menu_no_bom_clean.xlsx
"""

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
DEFAULT_DATASET = "shop1958987436032000"

SQL = """
SELECT
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pp.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS name_zh,
  JSON_EXTRACT_SCALAR(pp.name, '$.en') AS name_en,
  JSON_EXTRACT_SCALAR(pp.name, '$.th') AS name_th,
  CASE WHEN pp.product_type = 0 THEN '单品' WHEN pp.product_type = 1 THEN '套餐' ELSE '其他' END AS ptype
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
WHERE pp.delete_time = 0

UNION ALL

SELECT
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(ps.name, '$.zh'),
    JSON_EXTRACT_SCALAR(ps.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS name_zh,
  JSON_EXTRACT_SCALAR(ps.name, '$.en') AS name_en,
  JSON_EXTRACT_SCALAR(ps.name, '$.th') AS name_th,
  '加料' AS ptype
FROM `{project}`.`{dataset}`.`ttpos_product_sauce` ps
WHERE ps.delete_time = 0
"""


def load_sales_data(input_path: str):
    """读取销售报表中的「未设置BOM商品销量」"""
    wb = load_workbook(input_path)
    ws = wb["未设置BOM商品销量"]

    items = {}  # product_name -> {stores: set, total_qty: int}
    for row in ws.iter_rows(min_row=2, values_only=True):
        store_id, store_name, product_name, qty = row[0], row[1], row[2], row[3]
        if not product_name:
            continue
        pn = product_name.strip()
        if pn not in items:
            items[pn] = {"stores": set(), "total_qty": 0}
        items[pn]["stores"].add(store_id)
        items[pn]["total_qty"] += qty or 0

    return items


def query_bq(client, dataset: str):
    sql = SQL.format(project=client.project, dataset=dataset)
    rows = list(client.query(sql).result())
    return {r.name_zh: r for r in rows if r.name_zh}


def write_excel(matched_items, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "未设置BOM商品(已匹配)"

    headers = ["商品类型", "商品名称(中)", "商品名称(英)", "商品名称(泰)", "门店数", "总销量"]
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, size=11, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r_idx, (pn, data) in enumerate(matched_items, start=2):
        bq_row, sales = data
        ws.cell(row=r_idx, column=1, value=bq_row.ptype).alignment = center
        ws.cell(row=r_idx, column=2, value=bq_row.name_zh or "").alignment = left
        ws.cell(row=r_idx, column=3, value=bq_row.name_en or "").alignment = left
        ws.cell(row=r_idx, column=4, value=bq_row.name_th or "").alignment = left
        ws.cell(row=r_idx, column=5, value=len(sales["stores"])).alignment = center
        ws.cell(row=r_idx, column=6, value=sales["total_qty"]).alignment = center

    col_widths = [12, 30, 35, 40, 10, 10]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args():
    p = argparse.ArgumentParser(description="基于销售报表导出已匹配的未设置BOM商品中英泰对照")
    p.add_argument("--project", default=PROJECT_ID, help="GCP 项目 ID")
    p.add_argument("--dataset", default=DEFAULT_DATASET, help="BQ dataset")
    p.add_argument("--input", required=True, help="销售报表 Excel 路径")
    p.add_argument("--output", default="exports/menu_no_bom_clean.xlsx", help="输出路径")
    return p.parse_args()


def main():
    args = parse_args()
    setup_proxy()

    print(f"[读取] 销售报表: {args.input}")
    sales_items = load_sales_data(args.input)
    print(f"[销售] 共 {len(sales_items)} 种商品")

    client = get_bq_client(project_id=args.project)
    print("[BQ] 查询 menu 数据...")
    bq_map = query_bq(client, args.dataset)
    print(f"[BQ] 共 {len(bq_map)} 条商品")

    matched = {}
    unmatched = []
    for pn, sales in sales_items.items():
        if pn in bq_map:
            matched[pn] = (bq_map[pn], sales)
        else:
            unmatched.append(pn)

    # 按商品类型排序
    sorted_items = sorted(matched.items(), key=lambda x: (x[1][0].ptype, x[0]))

    write_excel(sorted_items, args.output)

    print(f"[完成] 输出 {args.output}")
    print(f"  已匹配: {len(matched)} 种")
    print(f"  未匹配(已剔除): {len(unmatched)} 种")

    singles = sum(1 for _, (r, _) in matched.items() if r.ptype == "单品")
    combos = sum(1 for _, (r, _) in matched.items() if r.ptype == "套餐")
    sauces = sum(1 for _, (r, _) in matched.items() if r.ptype == "加料")
    print(f"  其中: 单品 {singles} 种, 套餐 {combos} 种, 加料 {sauces} 种")

    if unmatched:
        print(f"\n[未匹配商品]:")
        for pn in sorted(unmatched):
            print(f"  - {pn}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
