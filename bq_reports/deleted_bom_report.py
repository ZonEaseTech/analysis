#!/usr/bin/env python3
"""
已删除商品 BOM 导出 —— 套餐 + 单品 合并版

输出：双 sheet Excel
  Sheet 1「已删除套餐组成」：门店名称 | 套餐名称 | 删除日期 | 分组 | 类型 | 单品名称 | 数量 | 是否必选 | 加价
  Sheet 2「已删除单品BOM」：门店名称 | 商品名称 | 删除日期 | 分类 | BOM物品名称 | BOM消耗 | BOM单位

范围：
  - delete_time 在 2025-09-01 至 2026-04-01 之间的已删除商品
  - 关联表不过滤 delete_time（软删记录仍保留在BQ中）

Usage:
    venv/bin/python -m bq_reports.deleted_bom_report --output exports/deleted_bom.xlsx
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
DEFAULT_DATASET = "shop1958987436032000"

DEFAULT_START_TS = 1756684800   # 2025-09-01
DEFAULT_END_TS = 1775001600     # 2026-04-01

REPORT_VERSION = "v2"

STORE_NAME_SQL = """
SELECT
  COALESCE(
    JSON_EXTRACT_SCALAR(name, '$.zh'),
    JSON_EXTRACT_SCALAR(name, '$.en'),
    name
  ) AS company_name
FROM `{project}`.`{dataset}`.`ttpos_company`
WHERE delete_time = 0
LIMIT 1
"""

# ===== 已删除套餐组成 =====
COMBO_SQL = """
SELECT
  pp.uuid AS combo_uuid,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pp.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS combo_name,
  pp.delete_time AS combo_delete_time,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pgrp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pgrp.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS group_name,
  pgrp.group_type AS group_type,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(child.name, '$.zh'),
    JSON_EXTRACT_SCALAR(child.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS child_name,
  gpi.num AS child_num,
  gpi.is_required,
  gpi.add_price
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
JOIN `{project}`.`{dataset}`.`ttpos_product_package_group` pgrp
  ON pgrp.product_package_uuid = pp.uuid
JOIN `{project}`.`{dataset}`.`ttpos_product_package_group_item` gpi
  ON gpi.product_package_group_uuid = pgrp.uuid
JOIN `{project}`.`{dataset}`.`ttpos_product_package` child
  ON child.uuid = gpi.related_uuid
WHERE pp.product_type = 1
  AND pp.delete_time >= {start_ts}
  AND pp.delete_time < {end_ts}
ORDER BY pp.delete_time DESC, pp.uuid, pgrp.uuid, gpi.sort
"""

# ===== 已删除单品 BOM =====
SINGLE_BOM_SQL = """
WITH bom_dedup AS (
  SELECT
    product_package_uuid,
    product_bom_card_uuid,
    MIN(product_flavor_uuid) AS flavor_uuid
  FROM `{project}`.`{dataset}`.`ttpos_product_bom`
  WHERE product_bom_card_uuid > 0
  GROUP BY product_package_uuid, product_bom_card_uuid
),
pkg_card_count AS (
  SELECT product_package_uuid, COUNT(*) AS card_cnt
  FROM bom_dedup
  GROUP BY product_package_uuid
),
pkg_meta AS (
  SELECT
    pp.uuid AS pkg_uuid,
    pp.delete_time,
    REGEXP_REPLACE(COALESCE(
      JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
      JSON_EXTRACT_SCALAR(pp.name, '$.en'),
      ''
    ), r'^\\s+|\\s+$', '') AS base_name,
    COALESCE(
      JSON_EXTRACT_SCALAR(c.name, '$.zh'),
      JSON_EXTRACT_SCALAR(c.name, '$.en'),
      ''
    ) AS category_name
  FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
  LEFT JOIN `{project}`.`{dataset}`.`ttpos_product_category` c
    ON c.uuid = pp.category_uuid
  WHERE pp.product_type = 0
    AND pp.delete_time >= {start_ts}
    AND pp.delete_time < {end_ts}
),
duplicate_names AS (
  SELECT base_name
  FROM pkg_meta
  WHERE base_name != ''
  GROUP BY base_name
  HAVING COUNT(*) > 1
)
SELECT
  pp.uuid AS product_uuid,
  pp.delete_time,
  CASE
    WHEN IFNULL(pcc.card_cnt, 0) > 1 AND pf.uuid IS NOT NULL THEN CONCAT(
      pm.base_name,
      ' (',
      COALESCE(JSON_EXTRACT_SCALAR(pf.name, '$.zh'),
               JSON_EXTRACT_SCALAR(pf.name, '$.en'),
               ''),
      ')'
    )
    WHEN dn.base_name IS NOT NULL AND pm.category_name != '' THEN CONCAT(
      pm.base_name,
      ' (',
      pm.category_name,
      ')'
    )
    ELSE pm.base_name
  END AS product_name,
  pm.category_name,
  COALESCE(JSON_EXTRACT_SCALAR(m.name, '$.zh'),
           JSON_EXTRACT_SCALAR(m.name, '$.en')) AS material_name,
  rm.num AS bom_num,
  COALESCE(JSON_EXTRACT_SCALAR(rm.unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.unit_name, '$.en'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.en')) AS bom_unit
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
LEFT JOIN bom_dedup pb ON pb.product_package_uuid = pp.uuid
LEFT JOIN pkg_card_count pcc ON pcc.product_package_uuid = pp.uuid
LEFT JOIN pkg_meta pm ON pm.pkg_uuid = pp.uuid
LEFT JOIN duplicate_names dn ON dn.base_name = pm.base_name
LEFT JOIN `{project}`.`{dataset}`.`ttpos_product_flavor` pf
  ON pf.uuid = pb.flavor_uuid
LEFT JOIN `{project}`.`{dataset}`.`ttpos_related_material` rm
  ON rm.related_uuid = pb.product_bom_card_uuid
LEFT JOIN `{project}`.`{dataset}`.`ttpos_material` m
  ON m.uuid = rm.material_uuid
WHERE pp.product_type = 0
  AND pp.delete_time >= {start_ts}
  AND pp.delete_time < {end_ts}
ORDER BY pp.delete_time DESC, product_name, material_name
"""


def get_store_name(client, dataset: str) -> str:
    sql = STORE_NAME_SQL.format(project=client.project, dataset=dataset)
    for row in client.query(sql).result():
        name = row.company_name or ""
        return name.strip() if name else ""
    return ""


def query_combo_rows(client, dataset: str, start_ts: int, end_ts: int):
    sql = COMBO_SQL.format(project=client.project, dataset=dataset, start_ts=start_ts, end_ts=end_ts)
    return list(client.query(sql).result())


def query_single_rows(client, dataset: str, start_ts: int, end_ts: int):
    sql = SINGLE_BOM_SQL.format(project=client.project, dataset=dataset, start_ts=start_ts, end_ts=end_ts)
    return list(client.query(sql).result())


def _style_header(ws, headers):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, size=11, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align


def _merge_runs(ws, col_letter: str, start_row: int, end_row: int, value_of):
    if end_row <= start_row:
        return
    run_start = start_row
    run_value = value_of(start_row)
    for r in range(start_row + 1, end_row + 1):
        v = value_of(r)
        if v != run_value:
            if r - 1 > run_start:
                ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{r - 1}")
            run_start = r
            run_value = v
    if end_row > run_start:
        ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{end_row}")


def _write_combo_sheet(ws, rows, store_name: str):
    headers = ["门店名称", "套餐名称", "删除日期", "分组", "类型", "单品名称", "数量", "是否必选", "加价"]
    _style_header(ws, headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=store_name).alignment = center
        ws.cell(row=r_idx, column=2, value=row.combo_name or "").alignment = center

        delete_dt = ""
        if row.combo_delete_time:
            delete_dt = datetime.fromtimestamp(row.combo_delete_time, tz=timezone.utc).strftime("%Y-%m-%d")
        ws.cell(row=r_idx, column=3, value=delete_dt).alignment = center

        ws.cell(row=r_idx, column=4, value=row.group_name or "").alignment = center
        ws.cell(row=r_idx, column=5, value="固定" if row.group_type == 0 else "可选").alignment = center
        ws.cell(row=r_idx, column=6, value=row.child_name or "").alignment = center
        ws.cell(row=r_idx, column=7, value=float(row.child_num) if row.child_num is not None else 0).alignment = center
        ws.cell(row=r_idx, column=8, value="是" if row.is_required else "否").alignment = center
        ws.cell(row=r_idx, column=9, value=float(row.add_price) if row.add_price else 0).alignment = center

    if rows:
        last_row = 1 + len(rows)
        if last_row > 2:
            ws.merge_cells(f"A2:A{last_row}")
            _merge_runs(ws, "B", 2, last_row, lambda r: rows[r - 2].combo_name)
            _merge_runs(ws, "C", 2, last_row,
                        lambda r: f"{rows[r - 2].combo_name}|{rows[r - 2].combo_delete_time}")
            _merge_runs(ws, "D", 2, last_row,
                        lambda r: f"{rows[r - 2].combo_name}|{rows[r - 2].group_name}")
            _merge_runs(ws, "E", 2, last_row,
                        lambda r: f"{rows[r - 2].combo_name}|{rows[r - 2].group_name}|{'固定' if rows[r - 2].group_type == 0 else '可选'}")

    for col_idx, w in enumerate([14, 25, 12, 12, 10, 25, 8, 10, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


def _write_single_sheet(ws, rows, store_name: str):
    headers = ["门店名称", "商品名称", "删除日期", "分类", "BOM物品名称", "BOM消耗", "BOM单位"]
    _style_header(ws, headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=store_name).alignment = center
        ws.cell(row=r_idx, column=2, value=row.product_name or "").alignment = center

        delete_dt = ""
        if row.delete_time:
            delete_dt = datetime.fromtimestamp(row.delete_time, tz=timezone.utc).strftime("%Y-%m-%d")
        ws.cell(row=r_idx, column=3, value=delete_dt).alignment = center

        ws.cell(row=r_idx, column=4, value=row.category_name or "").alignment = center
        ws.cell(row=r_idx, column=5, value=row.material_name or "")
        if row.bom_num is not None:
            ws.cell(row=r_idx, column=6, value=float(row.bom_num))
        ws.cell(row=r_idx, column=7, value=row.bom_unit or "")

    if rows:
        last_row = 1 + len(rows)
        if last_row > 2:
            ws.merge_cells(f"A2:A{last_row}")
            _merge_runs(ws, "B", 2, last_row, lambda r: rows[r - 2].product_name)
            _merge_runs(ws, "C", 2, last_row,
                        lambda r: f"{rows[r - 2].product_name}|{rows[r - 2].delete_time}")
            _merge_runs(ws, "D", 2, last_row,
                        lambda r: f"{rows[r - 2].product_name}|{rows[r - 2].category_name or ''}")

    for col_idx, w in enumerate([14, 30, 12, 18, 28, 10, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


def write_excel(combo_rows, single_rows, store_name: str, output_path: str,
                force: bool = False):
    from semantic.validators.gate import (
        add_watermark_sheet_openpyxl, validate_and_gate)
    from semantic.validators.identities import (
        make_required_fields_identity, make_unique_key_identity)

    outcomes = []

    # Sheet 1「已删除套餐组成」: 已删除套餐的组成明细 — 空=可能合法(当月无已删除套餐), min_rows=0
    combo_check_rows = [
        {"combo_name": r.combo_name or "", "child_name": r.child_name or ""}
        for r in combo_rows
    ]
    outcomes.append(validate_and_gate(
        combo_check_rows,
        [make_required_fields_identity(("combo_name", "child_name"), name="已删除套餐组成必填")],
        force=force, report_name="deleted_bom/combo",
        row_label=lambda r: f"{r.get('combo_name', '')} / {r.get('child_name', '')}",
        min_rows=0,  # 空=可能合法(当月无已删除套餐)
    ))

    # Sheet 2「已删除单品BOM」: 已删除单品及其 BOM 明细 — 空=可能合法(当月无已删除单品), min_rows=0
    uniq_ident, prepare = make_unique_key_identity(
        ("product_name", "material_name"), name="商品+物料主键唯一")
    check_rows = prepare([
        {"product_name": r.product_name or "", "material_name": r.material_name or ""}
        for r in single_rows
    ])
    outcomes.append(validate_and_gate(
        check_rows,
        [make_required_fields_identity(
            ("product_name",), name="已删除BOM必填字段"),
         uniq_ident],
        force=force, report_name="deleted_bom/single",
        row_label=lambda r: f"{r.get('product_name', '')} / {r.get('material_name', '')}",
        min_rows=0,  # 空=可能合法(当月无已删除单品)
    ))

    wb = Workbook()

    ws_combo = wb.active
    ws_combo.title = "已删除套餐组成"
    _write_combo_sheet(ws_combo, combo_rows, store_name)

    ws_single = wb.create_sheet("已删除单品BOM")
    _write_single_sheet(ws_single, single_rows, store_name)

    if any(o.needs_watermark for o in outcomes):
        first_wm = next(o for o in outcomes if o.needs_watermark)
        add_watermark_sheet_openpyxl(wb, first_wm.watermark_lines())

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _inject_version(path: str) -> str:
    p = Path(path)
    suffix = f"_{REPORT_VERSION}"
    if p.stem.endswith(suffix):
        return str(p)
    return str(p.with_name(f"{p.stem}{suffix}{p.suffix}"))


def parse_args():
    p = argparse.ArgumentParser(description="已删除商品 BOM 导出（套餐+单品）")
    p.add_argument("--project", default=PROJECT_ID, help="GCP 项目 ID")
    p.add_argument("--dataset", default=DEFAULT_DATASET,
                   help=f"BQ dataset(任意一家分店即可,默认 {DEFAULT_DATASET})")
    p.add_argument("--start-ts", type=int, default=DEFAULT_START_TS,
                   help=f"删除时间起始 Unix 时间戳(默认 {DEFAULT_START_TS}=2025-09-01)")
    p.add_argument("--end-ts", type=int, default=DEFAULT_END_TS,
                   help=f"删除时间截止 Unix 时间戳(默认 {DEFAULT_END_TS}=2026-04-01)")
    p.add_argument("--output", required=True,
                   help=f"输出 Excel 文件路径(自动追加版本号 _{REPORT_VERSION})")
    p.add_argument("--force", action="store_true",
                   help="强制导出 (有 🔴 违反时打水印而非阻断)")
    return p.parse_args()


def main():
    args = parse_args()
    setup_proxy()

    output_path = _inject_version(args.output)

    start_dt = datetime.fromtimestamp(args.start_ts, tz=timezone.utc).strftime("%Y-%m-%d")
    end_dt = datetime.fromtimestamp(args.end_ts, tz=timezone.utc).strftime("%Y-%m-%d")

    print(f"[BQ] project={args.project} dataset={args.dataset}")
    print(f"[范围] 删除时间: {start_dt} 至 {end_dt}")
    print(f"[版本] {REPORT_VERSION} → 输出路径: {output_path}")
    client = get_bq_client(project_id=args.project)

    store_name = get_store_name(client, args.dataset)
    if not store_name:
        store_name = ""
    print(f"[门店] {store_name}")

    print("[BQ] 查询已删除套餐组成...")
    combo_rows = query_combo_rows(client, args.dataset, args.start_ts, args.end_ts)
    unique_combos = set(r.combo_uuid for r in combo_rows)
    print(f"[BQ] 套餐: {len(combo_rows)} 行, {len(unique_combos)} 个")

    print("[BQ] 查询已删除单品 BOM...")
    single_rows = query_single_rows(client, args.dataset, args.start_ts, args.end_ts)
    # 统计有BOM vs 无BOM
    with_bom = [r for r in single_rows if r.material_name]
    no_bom = []
    seen = set()
    for r in single_rows:
        if not r.material_name and r.product_uuid not in seen:
            no_bom.append(r)
            seen.add(r.product_uuid)
    print(f"[BQ] 单品: {len(single_rows)} 行 (有BOM {len(with_bom)} 行, 无BOM {len(no_bom)} 个)")

    write_excel(combo_rows, single_rows, store_name, output_path, force=args.force)

    print(f"[完成] 输出 {output_path}")
    print(f"  Sheet 1「已删除套餐组成」: {len(combo_rows)} 行, {len(unique_combos)} 个套餐")
    print(f"  Sheet 2「已删除单品BOM」: {len(single_rows)} 行 (含 {len(no_bom)} 个无BOM单品)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
