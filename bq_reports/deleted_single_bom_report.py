#!/usr/bin/env python3
"""
已删除单品/加料 BOM 导出

输出：双 sheet Excel
  Sheet 1「已删除BOM清单」：6 列(门店名称 | 商品名称 | 删除日期 | BOM物品名称 | BOM消耗 | BOM单位)
  Sheet 2「已删除未配置BOM」：3 列(门店名称 | 商品类型 | 商品名称 | 删除日期)

范围：
  - delete_time 在指定范围内的已删除单品 + 加料
  - 关联表不过滤 delete_time（软删记录仍保留在BQ中）

Usage:
    venv/bin/python -m bq_reports.deleted_single_bom_report --output exports/deleted_single_bom.xlsx
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
DEFAULT_DATASET = "shop1958987436032000"

DEFAULT_START_TS = 1756684800   # 2025-09-01
DEFAULT_END_TS = 1775001600     # 2026-04-01

REPORT_VERSION = "v1"

STORE_NAME_SQL = """
SELECT
  COALESCE(
    JSON_EXTRACT_SCALAR(name, '$.zh'),
    JSON_EXTRACT_SCALAR(name, '$.en'),
    name
  ) AS company_name
FROM `{project}`.`{dataset}`.`ttpos_company`
WHERE delete_time = 0
LIMIT 1
"""

# 已删除单品 BOM
# 不过滤 delete_time（软删记录仍保留在BQ中）
SINGLE_BOM_SQL = """
WITH bom_dedup AS (
  SELECT
    product_package_uuid,
    product_bom_card_uuid,
    MIN(product_flavor_uuid) AS flavor_uuid
  FROM `{project}`.`{dataset}`.`ttpos_product_bom`
  WHERE product_bom_card_uuid > 0
  GROUP BY product_package_uuid, product_bom_card_uuid
),
pkg_card_count AS (
  SELECT product_package_uuid, COUNT(*) AS card_cnt
  FROM bom_dedup
  GROUP BY product_package_uuid
),
pkg_meta AS (
  SELECT
    pp.uuid AS pkg_uuid,
    pp.delete_time,
    REGEXP_REPLACE(COALESCE(
      JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
      JSON_EXTRACT_SCALAR(pp.name, '$.en'),
      ''
    ), r'^\\s+|\\s+$', '') AS base_name,
    COALESCE(
      JSON_EXTRACT_SCALAR(c.name, '$.zh'),
      JSON_EXTRACT_SCALAR(c.name, '$.en'),
      ''
    ) AS category_name
  FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
  LEFT JOIN `{project}`.`{dataset}`.`ttpos_product_category` c
    ON c.uuid = pp.category_uuid
  WHERE pp.product_type = 0
    AND pp.delete_time >= {start_ts}
    AND pp.delete_time < {end_ts}
),
duplicate_names AS (
  SELECT base_name
  FROM pkg_meta
  WHERE base_name != ''
  GROUP BY base_name
  HAVING COUNT(*) > 1
)
SELECT
  0 AS sort_type,
  pp.uuid AS product_uuid,
  pp.delete_time,
  CASE
    WHEN IFNULL(pcc.card_cnt, 0) > 1 AND pf.uuid IS NOT NULL THEN CONCAT(
      pm.base_name,
      ' (',
      COALESCE(JSON_EXTRACT_SCALAR(pf.name, '$.zh'),
               JSON_EXTRACT_SCALAR(pf.name, '$.en'),
               ''),
      ')'
    )
    WHEN dn.base_name IS NOT NULL AND pm.category_name != '' THEN CONCAT(
      pm.base_name,
      ' (',
      pm.category_name,
      ')'
    )
    ELSE pm.base_name
  END AS product_name,
  pm.category_name,
  COALESCE(JSON_EXTRACT_SCALAR(m.name, '$.zh'),
           JSON_EXTRACT_SCALAR(m.name, '$.en')) AS material_name,
  rm.num AS bom_num,
  COALESCE(JSON_EXTRACT_SCALAR(rm.unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.unit_name, '$.en'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.en')) AS bom_unit
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
LEFT JOIN bom_dedup pb ON pb.product_package_uuid = pp.uuid
LEFT JOIN pkg_card_count pcc ON pcc.product_package_uuid = pp.uuid
LEFT JOIN pkg_meta pm ON pm.pkg_uuid = pp.uuid
LEFT JOIN duplicate_names dn ON dn.base_name = pm.base_name
LEFT JOIN `{project}`.`{dataset}`.`ttpos_product_flavor` pf
  ON pf.uuid = pb.flavor_uuid
LEFT JOIN `{project}`.`{dataset}`.`ttpos_related_material` rm
  ON rm.related_uuid = pb.product_bom_card_uuid
LEFT JOIN `{project}`.`{dataset}`.`ttpos_material` m
  ON m.uuid = rm.material_uuid
WHERE pp.product_type = 0
  AND pp.delete_time >= {start_ts}
  AND pp.delete_time < {end_ts}
"""

# 已删除加料 BOM
SAUCE_BOM_SQL = """
SELECT
  1 AS sort_type,
  ps.uuid AS product_uuid,
  ps.delete_time,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(ps.name, '$.zh'),
    JSON_EXTRACT_SCALAR(ps.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS product_name,
  '' AS category_name,
  COALESCE(JSON_EXTRACT_SCALAR(m.name, '$.zh'),
           JSON_EXTRACT_SCALAR(m.name, '$.en')) AS material_name,
  rm.num AS bom_num,
  COALESCE(JSON_EXTRACT_SCALAR(rm.unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.unit_name, '$.en'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.zh'),
           JSON_EXTRACT_SCALAR(rm.base_unit_name, '$.en')) AS bom_unit
FROM `{project}`.`{dataset}`.`ttpos_product_sauce` ps
LEFT JOIN `{project}`.`{dataset}`.`ttpos_related_material` rm
  ON rm.related_uuid = ps.product_bom_card_uuid
  AND ps.product_bom_card_uuid > 0
LEFT JOIN `{project}`.`{dataset}`.`ttpos_material` m
  ON m.uuid = rm.material_uuid
WHERE ps.delete_time >= {start_ts}
  AND ps.delete_time < {end_ts}
"""

TYPE_LABEL = {0: "单品", 1: "加料"}


def get_store_name(client, dataset: str) -> str:
    sql = STORE_NAME_SQL.format(project=client.project, dataset=dataset)
    for row in client.query(sql).result():
        name = row.company_name or ""
        return name.strip() if name else ""
    return ""


def query_rows(client, dataset: str, start_ts: int, end_ts: int):
    sql = f"""
    {SINGLE_BOM_SQL.format(project=client.project, dataset=dataset, start_ts=start_ts, end_ts=end_ts)}
    UNION ALL
    {SAUCE_BOM_SQL.format(project=client.project, dataset=dataset, start_ts=start_ts, end_ts=end_ts)}
    ORDER BY sort_type, product_name, material_name
    """
    return list(client.query(sql).result())


def split_rows(rows):
    """拆分为「有 BOM」明细行 vs「无 BOM」商品列表(每商品/加料 1 条)"""
    with_bom = []
    no_bom = []
    seen_no_bom = set()
    for r in rows:
        if r.material_name:
            with_bom.append(r)
        else:
            key = (r.sort_type, r.product_uuid)
            if key in seen_no_bom:
                continue
            seen_no_bom.add(key)
            no_bom.append(r)
    return with_bom, no_bom


def _style_header(ws, headers):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, size=11, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align


def _merge_runs(ws, col_letter: str, start_row: int, end_row: int, value_of):
    if end_row <= start_row:
        return
    run_start = start_row
    run_value = value_of(start_row)
    for r in range(start_row + 1, end_row + 1):
        v = value_of(r)
        if v != run_value:
            if r - 1 > run_start:
                ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{r - 1}")
            run_start = r
            run_value = v
    if end_row > run_start:
        ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{end_row}")


def _write_bom_sheet(ws, with_bom, store_name: str):
    headers = ["门店名称", "商品名称", "删除日期", "BOM物品名称", "BOM消耗", "BOM单位"]
    _style_header(ws, headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(with_bom, start=2):
        ws.cell(row=r_idx, column=1, value=store_name).alignment = center
        ws.cell(row=r_idx, column=2, value=row.product_name or "").alignment = center

        delete_dt = ""
        if row.delete_time:
            delete_dt = datetime.fromtimestamp(row.delete_time, tz=timezone.utc).strftime("%Y-%m-%d")
        ws.cell(row=r_idx, column=3, value=delete_dt).alignment = center

        ws.cell(row=r_idx, column=4, value=row.material_name or "")
        if row.bom_num is not None:
            ws.cell(row=r_idx, column=5, value=float(row.bom_num))
        ws.cell(row=r_idx, column=6, value=row.bom_unit or "")

    if with_bom:
        last_row = 1 + len(with_bom)
        if last_row > 2:
            ws.merge_cells(f"A2:A{last_row}")
            _merge_runs(ws, "B", 2, last_row, lambda r: with_bom[r - 2].product_name)
            _merge_runs(ws, "C", 2, last_row,
                        lambda r: f"{with_bom[r - 2].product_name}|{with_bom[r - 2].delete_time}")

    for col_idx, w in enumerate([14, 30, 12, 28, 10, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


def _write_no_bom_sheet(ws, no_bom, store_name: str):
    headers = ["门店名称", "商品类型", "商品名称", "删除日期"]
    _style_header(ws, headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(no_bom, start=2):
        ws.cell(row=r_idx, column=1, value=store_name).alignment = center
        ws.cell(row=r_idx, column=2, value=TYPE_LABEL.get(row.sort_type, "")).alignment = center
        ws.cell(row=r_idx, column=3, value=row.product_name or "")

        delete_dt = ""
        if row.delete_time:
            delete_dt = datetime.fromtimestamp(row.delete_time, tz=timezone.utc).strftime("%Y-%m-%d")
        ws.cell(row=r_idx, column=4, value=delete_dt).alignment = center

    if no_bom:
        last_row = 1 + len(no_bom)
        if last_row > 2:
            ws.merge_cells(f"A2:A{last_row}")
            _merge_runs(ws, "B", 2, last_row,
                        lambda r: TYPE_LABEL.get(no_bom[r - 2].sort_type, ""))

    for col_idx, w in enumerate([14, 12, 30, 12], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


def write_excel(with_bom, no_bom, store_name: str, output_path: str):
    wb = Workbook()
    ws_bom = wb.active
    ws_bom.title = "已删除BOM清单"
    _write_bom_sheet(ws_bom, with_bom, store_name)

    ws_no = wb.create_sheet("已删除未配置BOM")
    _write_no_bom_sheet(ws_no, no_bom, store_name)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _inject_version(path: str) -> str:
    p = Path(path)
    suffix = f"_{REPORT_VERSION}"
    if p.stem.endswith(suffix):
        return str(p)
    return str(p.with_name(f"{p.stem}{suffix}{p.suffix}"))


def parse_args():
    p = argparse.ArgumentParser(description="已删除单品/加料 BOM 导出")
    p.add_argument("--project", default=PROJECT_ID, help="GCP 项目 ID")
    p.add_argument("--dataset", default=DEFAULT_DATASET,
                   help=f"BQ dataset(任意一家分店即可,默认 {DEFAULT_DATASET})")
    p.add_argument("--start-ts", type=int, default=DEFAULT_START_TS,
                   help=f"删除时间起始 Unix 时间戳(默认 {DEFAULT_START_TS}=2025-09-01)")
    p.add_argument("--end-ts", type=int, default=DEFAULT_END_TS,
                   help=f"删除时间截止 Unix 时间戳(默认 {DEFAULT_END_TS}=2026-04-01)")
    p.add_argument("--output", required=True,
                   help=f"输出 Excel 文件路径(自动追加版本号 _{REPORT_VERSION})")
    return p.parse_args()


def main():
    args = parse_args()
    setup_proxy()

    output_path = _inject_version(args.output)

    start_dt = datetime.fromtimestamp(args.start_ts, tz=timezone.utc).strftime("%Y-%m-%d")
    end_dt = datetime.fromtimestamp(args.end_ts, tz=timezone.utc).strftime("%Y-%m-%d")

    print(f"[BQ] project={args.project} dataset={args.dataset}")
    print(f"[范围] 删除时间: {start_dt} 至 {end_dt}")
    print(f"[版本] {REPORT_VERSION} → 输出路径: {output_path}")
    client = get_bq_client(project_id=args.project)

    store_name = get_store_name(client, args.dataset)
    if not store_name:
        store_name = ""
    print(f"[门店] {store_name}")

    print("[BQ] 查询已删除单品 + 加料 BOM...")
    rows = query_rows(client, args.dataset, args.start_ts, args.end_ts)
    print(f"[BQ] 共 {len(rows)} 行")

    with_bom, no_bom = split_rows(rows)
    write_excel(with_bom, no_bom, store_name, output_path)

    no_bom_singles = sum(1 for r in no_bom if r.sort_type == 0)
    no_bom_sauces = sum(1 for r in no_bom if r.sort_type == 1)
    print(f"[完成] 输出 {output_path}")
    print(f"  Sheet 1「已删除BOM清单」: {len(with_bom)} 行")
    print(f"  Sheet 2「已删除未配置BOM」: {len(no_bom)} 个 (单品 {no_bom_singles} 个, 加料 {no_bom_sauces} 个)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
