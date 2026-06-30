"""新事实表接入模板 — load → resolve → diff → write.

从 scripts/adhoc/onboard_final_bom_202605.py 提取可复用骨架。
适用场景: 客户给了一份新的 BOM/物料价 Excel/CSV, 需要跟现有管线对比后接入。

Usage:
  from bq_reports.shared.onboard_helper import build_onboard_diff
  build_onboard_diff(bom_dict, price_dict, out_path)
"""

from __future__ import annotations
import openpyxl


def build_onboard_diff(bom: dict, price: dict, out_path: str,
                       *, strict: bool = True) -> None:
    """对比新 BOM/价格 与现有管线 layers, 写三 sheet Excel diff。

    Args:
        bom: {商品名: [(物料编码, 物料名, 消耗量, 单位), ...]}
        price: {物料编码: (物料名, 单价, 单位)}
        out_path: DIFF_XLSX 输出路径
        strict: True 时只走客户外挂层, 未命中=(0, '无(strict)')
    """
    from bq_reports.profit_margin_report import (
        load_config, _load_bom_layers, _match_bom_layered,
        _load_material_price_layers, _build_material_price_resolver)

    cfg = load_config()
    layers = _load_bom_layers(cfg)
    price_layers = _load_material_price_layers(cfg)
    resolver = _build_material_price_resolver({}, {}, price_layers, strict=strict)

    existing_products = set()
    for _n, _p, boms, _m in layers:
        existing_products |= set(boms.keys())

    wb = openpyxl.Workbook()

    # Sheet 1: 商品覆盖
    ws1 = wb.active
    ws1.title = "商品覆盖"
    ws1.append(["商品名称", "状态"])
    new_set = set(bom)
    for p in sorted(new_set - existing_products):
        ws1.append([p, "新增 (现有没有)"])
    for p in sorted(existing_products - new_set):
        ws1.append([p, "丢失 (替换后无 BOM)"])
    for p in sorted(new_set & existing_products):
        ws1.append([p, "重叠"])

    # Sheet 2: BOM 内容变动
    ws2 = wb.create_sheet("BOM内容变动")
    ws2.append(["商品名称", "现有来源", "新增物料", "删除物料", "消耗量变化"])
    for prod, recs in bom.items():
        matched, src = _match_bom_layered(prod, layers)
        if not matched:
            continue
        old = {c: q for c, _n, q, _u in matched}
        new = {c: q for c, _n, q, _u in recs}
        added = set(new) - set(old)
        removed = set(old) - set(new)
        qty_chg = {c: (old[c], new[c]) for c in set(old) & set(new)
                   if abs(float(old[c] or 0) - float(new[c] or 0)) > 1e-6}
        if added or removed or qty_chg:
            ws2.append([
                prod, src,
                ", ".join(sorted(added)) or "-",
                ", ".join(sorted(removed)) or "-",
                "; ".join(f"{c}: {o}→{n}" for c, (o, n) in qty_chg.items()) or "-",
            ])

    # Sheet 3: 价格变动
    ws3 = wb.create_sheet("价格变动")
    ws3.append(["物品编号", "物品名称", "现有价", "新文件价", "倍数", "类型"])
    for code, (name, np, _unit) in price.items():
        r = resolver.resolve((code, name))
        if r is None:
            ws3.append([code, name, "-", np, "-", "新增 (现有没有)"])
        elif abs(float(r.value or 0) - float(np or 0)) > 1e-6:
            old_v = float(r.value or 0)
            mult = f"{float(np or 0) / old_v:.1f}×" if old_v else "-"
            ws3.append([code, name, r.value, np, mult, "变动"])
    wb.save(out_path)
    print(f"差异报告: {out_path}")
    print(f"  商品覆盖   {ws1.max_row - 1} 行")
    print(f"  BOM内容变动 {ws2.max_row - 1} 行")
    print(f"  价格变动   {ws3.max_row - 1} 行")
