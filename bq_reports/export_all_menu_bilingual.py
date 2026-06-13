#!/usr/bin/env python3
"""
导出所有菜单商品（单品+套餐+加料）的中英泰三语名称

Usage:
    venv/bin/python -m bq_reports.export_all_menu_bilingual --output exports/all_menu_bilingual.xlsx
"""

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
DEFAULT_DATASET = "shop1958987436032000"

SQL = """
SELECT
  '单品/套餐' AS item_type,
  pp.uuid,
  CASE WHEN pp.product_type = 0 THEN '单品' WHEN pp.product_type = 1 THEN '套餐' ELSE '其他' END AS product_type,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pp.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS name_zh,
  JSON_EXTRACT_SCALAR(pp.name, '$.en') AS name_en,
  JSON_EXTRACT_SCALAR(pp.name, '$.th') AS name_th
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
WHERE pp.delete_time = 0

UNION ALL

SELECT
  '加料' AS item_type,
  ps.uuid,
  '加料' AS product_type,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(ps.name, '$.zh'),
    JSON_EXTRACT_SCALAR(ps.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS name_zh,
  JSON_EXTRACT_SCALAR(ps.name, '$.en') AS name_en,
  JSON_EXTRACT_SCALAR(ps.name, '$.th') AS name_th
FROM `{project}`.`{dataset}`.`ttpos_product_sauce` ps
WHERE ps.delete_time = 0

ORDER BY item_type, name_zh
"""


def query_rows(client, dataset: str):
    sql = SQL.format(project=client.project, dataset=dataset)
    return list(client.query(sql).result())


def write_excel(rows, output_path: str, force: bool = False):
    from semantic.validators.gate import (
        add_watermark_sheet_openpyxl, validate_and_gate)
    from semantic.validators.identities import (
        make_required_fields_identity, make_unique_key_identity)

    uniq_ident, prepare = make_unique_key_identity(("uuid",), name="UUID唯一")
    check_rows = prepare([
        {"name_zh": r.name_zh or "", "uuid": str(r.uuid) if r.uuid else ""}
        for r in rows
    ])
    outcome = validate_and_gate(
        check_rows,
        [make_required_fields_identity(("name_zh",), name="菜单中文名必填"),
         uniq_ident],
        force=force, report_name="export_all_menu_bilingual",
        row_label=lambda r: r.get("name_zh", "") or r.get("uuid", ""),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "菜单中英泰对照"

    headers = ["商品类型", "单品/套餐/加料", "商品名称(中)", "商品名称(英)", "商品名称(泰)", "UUID"]
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, size=11, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=row.item_type).alignment = center
        ws.cell(row=r_idx, column=2, value=row.product_type).alignment = center
        ws.cell(row=r_idx, column=3, value=row.name_zh or "").alignment = left
        ws.cell(row=r_idx, column=4, value=row.name_en or "").alignment = left
        ws.cell(row=r_idx, column=5, value=row.name_th or "").alignment = left
        ws.cell(row=r_idx, column=6, value=row.uuid).alignment = center

    col_widths = [12, 14, 30, 35, 40, 22]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"

    if outcome.needs_watermark:
        add_watermark_sheet_openpyxl(wb, outcome.watermark_lines())

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args():
    p = argparse.ArgumentParser(description="导出所有菜单商品中英泰三语名称")
    p.add_argument("--project", default=PROJECT_ID, help="GCP 项目 ID")
    p.add_argument("--dataset", default=DEFAULT_DATASET, help="BQ dataset")
    p.add_argument("--output", default="exports/all_menu_bilingual.xlsx", help="输出路径")
    p.add_argument("--force", action="store_true",
                   help="强制导出 (有 🔴 违反时打水印而非阻断)")
    return p.parse_args()


def main():
    args = parse_args()
    setup_proxy()

    client = get_bq_client(project_id=args.project)
    rows = query_rows(client, args.dataset)
    write_excel(rows, args.output, force=args.force)
    print(f"[完成] 共 {len(rows)} 条商品，输出 {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
