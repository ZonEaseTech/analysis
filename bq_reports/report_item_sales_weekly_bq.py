#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
单品+套餐明细 销售统计 (BigQuery 版本)

跟 TTPOS UI 商品销售统计 saleMode=4 一致：单品(含套餐拆分) + 套餐整体

行类型：
- 套餐：product_package.product_type=1 (堂食) / takeout_order_item.ttpos_product_type=1 (外卖)
- 单品：product_package.product_type=0 + 套餐子品(product_type=2 拆回到 product_package_uuid) + 外卖单品

输出：单文件双 Sheet（中文 / English），English 版增加 Unit Price / Total Price。

用法:
    venv/bin/python -m bq_reports.report_item_sales_weekly_bq \
        --start 2026-04-20 --end 2026-04-26
    # 不传则默认最近 7 天 (今天-7 ~ 昨天)
"""

import argparse
import csv
import os
import re
import subprocess
import sys
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from google.cloud import bigquery
from google.oauth2.credentials import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ID = "diyl-407103"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "exports"

STORE_LIST = [
    ("1", "1958987436032000"), ("2", "2269470793728000"), ("3", "2598648160256000"),
    ("4", "2876210421760000"), ("5", "3446618988544000"), ("6", "3870122057728000"),
    ("7", "4149605310464000"), ("8", "4358842359808000"), ("9", "4912616316928000"),
    ("10", "5567347171328000"), ("11", "5999171739648000"), ("12", "6542950670336000"),
    ("13", "6789240201216000"), ("14", "6977459593216000"), ("15", "7191251656704000"),
    ("16", "7400123801600000"), ("17", "7648065888256000"), ("18", "7863653113856000"),
    ("19", "8100551598080000"), ("20", "8501761941504000"), ("21", "8722592047104000"),
    ("22", "2947521978368000"), ("23", "3448951017472000"), ("24", "3782477877248000"),
    ("25", "4024875094016000"), ("26", "4229506797568000"), ("27", "4418766376960000"),
    ("28", "4613872816128000"), ("29", "4805464432640000"), ("30", "5001267122176000"),
    ("31", "5250979205120000"), ("32", "5444022046720000"), ("33", "7600687026176000"),
    ("34", "7813128523776000"), ("35", "8051063001088000"), ("36", "8535580610560000"),
    ("37", "8723170856960000"), ("38", "1515821506560000"), ("39", "3631470387200000"),
    ("40", "5498438983680000"), ("42", "9231705128960000"), ("44", "1379607252992000"),
    ("45", "1559157018624000"), ("46", "1745354756096000"), ("47", "1919875551232000"),
    ("48", "2101535051776000"), ("49", "2277263806464000"),
    ("51", "2618629820416000"), ("52", "2788834676736000"), ("53", "2992442970112000"),
    ("54", "3169446793216000"), ("55", "3367514411008000"), ("56", "3662462062592000"),
    ("58", "4053593493504000"), ("59", "4197894328320000"), ("61", "3087884357632000"),
]


def log(msg: str):
    print(msg, flush=True)


def setup_proxy():
    proxy_url = os.environ.get("BQ_PROXY")
    if not proxy_url:
        return
    for k in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"]:
        os.environ[k] = proxy_url


def get_creds():
    r = subprocess.run(
        ["gcloud", "auth", "print-access-token"],
        capture_output=True, text=True, check=True,
    )
    return Credentials(
        token=r.stdout.strip(),
        scopes=["https://www.googleapis.com/auth/cloud-platform"],
    )


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--start", help="起始日期 YYYY-MM-DD (含)")
    p.add_argument("--end", help="结束日期 YYYY-MM-DD (含)")
    p.add_argument("--workers", type=int, default=5, help="并发查询门店数")
    p.add_argument("--version", type=int, help="指定版本号；不传则自动递增")
    a = p.parse_args()
    if a.start and a.end:
        sd = datetime.strptime(a.start, "%Y-%m-%d").date()
        ed = datetime.strptime(a.end, "%Y-%m-%d").date()
    else:
        ed = date.today() - timedelta(days=1)
        sd = ed - timedelta(days=6)
    if sd > ed:
        sys.exit("start 不能晚于 end")
    st = int(datetime(sd.year, sd.month, sd.day, 0, 0, 0).timestamp())
    et = int(datetime(ed.year, ed.month, ed.day, 23, 59, 59).timestamp())
    return sd, ed, st, et, a.workers, a.version


def next_version(out_dir: Path, prefix: str) -> int:
    """扫描 out_dir 下 `{prefix}_v{N}.xlsx`，返回下一个版本号（最小 1）。"""
    pat = re.compile(rf"^{re.escape(prefix)}_v(\d+)\.xlsx$")
    used = []
    if out_dir.exists():
        for p in out_dir.iterdir():
            m = pat.match(p.name)
            if m:
                used.append(int(m.group(1)))
    return (max(used) + 1) if used else 1


def query_shop(client, dataset_id, start_ts, end_ts):
    """
    返回行：(sale_date, row_type, product_uuid, name_zh, name_en, name_th,
            cat_zh, cat_en, cat_th, total_qty, total_amount)
      row_type: '套餐' / '单品'
    """
    # 商品名多语言
    pp_zh = "JSON_EXTRACT_SCALAR(pp.name, '$.zh')"
    pp_en = "JSON_EXTRACT_SCALAR(pp.name, '$.en')"
    pp_th = "JSON_EXTRACT_SCALAR(pp.name, '$.th')"

    # 分类名多语言
    pc_zh = "JSON_EXTRACT_SCALAR(pc.name, '$.zh')"
    pc_en = "JSON_EXTRACT_SCALAR(pc.name, '$.en')"
    pc_th = "JSON_EXTRACT_SCALAR(pc.name, '$.th')"

    # sop.name 也是 JSON 快照
    sop_zh = "JSON_EXTRACT_SCALAR(sop.name, '$.zh')"
    sop_en = "JSON_EXTRACT_SCALAR(sop.name, '$.en')"
    sop_th = "JSON_EXTRACT_SCALAR(sop.name, '$.th')"

    # toi.item_name 通常是已渲染的字符串（非 JSON），优先 join pp，回退 toi.item_name
    toi_zh = (
        "COALESCE(NULLIF(JSON_EXTRACT_SCALAR(pp.name, '$.zh'), 'null'), toi.item_name)"
    )
    toi_en = (
        "COALESCE(NULLIF(JSON_EXTRACT_SCALAR(pp.name, '$.en'), 'null'), "
        "NULLIF(JSON_EXTRACT_SCALAR(pp.name, '$.zh'), 'null'), toi.item_name)"
    )
    toi_th = (
        "COALESCE(NULLIF(JSON_EXTRACT_SCALAR(pp.name, '$.th'), 'null'), toi.item_name)"
    )

    query = f"""
    WITH dine_main AS (
        -- A. 堂食 单品+套餐：statistics_product 已排除 product_type=2
        SELECT
            DATE(TIMESTAMP_SECONDS(sp.complete_time)) AS sale_date,
            IF(pp.product_type = 1, '套餐', '单品') AS row_type,
            sp.product_package_uuid AS product_uuid,
            {pp_zh} AS name_zh,
            {pp_en} AS name_en,
            {pp_th} AS name_th,
            {pc_zh} AS cat_zh,
            {pc_en} AS cat_en,
            {pc_th} AS cat_th,
            CAST(sp.product_num AS FLOAT64) AS qty,
            CAST(COALESCE(sp.product_final_price, 0) * sp.product_num AS FLOAT64) AS amount
        FROM `{PROJECT_ID}.{dataset_id}.ttpos_statistics_product` sp
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_product_package` pp
            ON pp.uuid = sp.product_package_uuid AND pp.delete_time = 0
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_product_category` pc
            ON pc.uuid = pp.category_uuid AND pc.delete_time = 0
        WHERE sp.delete_time = 0
          AND sp.complete_time BETWEEN {start_ts} AND {end_ts}
    ),
    dine_subitem AS (
        -- B. 堂食 套餐子品 → 算到「单品」
        SELECT
            DATE(TIMESTAMP_SECONDS(
                COALESCE(NULLIF(sb.finish_time, 0), so.finish_time)
            )) AS sale_date,
            '单品' AS row_type,
            sop.product_package_uuid AS product_uuid,
            COALESCE({pp_zh}, {sop_zh}) AS name_zh,
            COALESCE({pp_en}, {sop_en}) AS name_en,
            COALESCE({pp_th}, {sop_th}) AS name_th,
            {pc_zh} AS cat_zh,
            {pc_en} AS cat_en,
            {pc_th} AS cat_th,
            CAST(sop.num * sop.copy_num * IF(sop.unit_num = 0, 1, sop.unit_num) AS FLOAT64) AS qty,
            CAST(COALESCE(sop.price, 0) * sop.num * sop.copy_num * IF(sop.unit_num = 0, 1, sop.unit_num) AS FLOAT64) AS amount
        FROM `{PROJECT_ID}.{dataset_id}.ttpos_sale_order_product` sop
        INNER JOIN `{PROJECT_ID}.{dataset_id}.ttpos_sale_bill` sb
            ON sb.uuid = sop.sale_bill_uuid AND sb.delete_time = 0
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_sale_order` so
            ON so.uuid = sop.sale_order_uuid AND so.delete_time = 0
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_product_package` pp
            ON pp.uuid = sop.product_package_uuid AND pp.delete_time = 0
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_product_category` pc
            ON pc.uuid = pp.category_uuid AND pc.delete_time = 0
        WHERE sop.delete_time = 0
          AND sop.cancel_time = 0
          AND sb.status = 1
          AND sop.product_type = 2
          AND COALESCE(NULLIF(sb.finish_time, 0), so.finish_time) BETWEEN {start_ts} AND {end_ts}
    ),
    takeout AS (
        -- C. 外卖
        SELECT
            DATE(TIMESTAMP_SECONDS(
                CASE WHEN tko.order_state = 40 THEN tko.completed_time ELSE tko.accepted_time END
            )) AS sale_date,
            IF(toi.ttpos_product_type = 1, '套餐', '单品') AS row_type,
            toi.ttpos_product_package_uuid AS product_uuid,
            {toi_zh} AS name_zh,
            {toi_en} AS name_en,
            {toi_th} AS name_th,
            {pc_zh} AS cat_zh,
            {pc_en} AS cat_en,
            {pc_th} AS cat_th,
            CAST(toi.quantity AS FLOAT64) AS qty,
            CAST(COALESCE(toi.price, 0) * toi.quantity AS FLOAT64) AS amount
        FROM `{PROJECT_ID}.{dataset_id}.ttpos_takeout_order_item` toi
        INNER JOIN `{PROJECT_ID}.{dataset_id}.ttpos_takeout_order` tko
            ON tko.uuid = toi.takeout_order_uuid AND tko.delete_time = 0
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_product_package` pp
            ON pp.uuid = toi.ttpos_product_package_uuid AND pp.delete_time = 0
        LEFT JOIN `{PROJECT_ID}.{dataset_id}.ttpos_product_category` pc
            ON pc.uuid = pp.category_uuid AND pc.delete_time = 0
        WHERE toi.delete_time = 0
          AND tko.order_state IN (10, 20, 30, 40, 60)
          AND tko.accepted_time > 0
          AND (
              (tko.order_state = 40 AND tko.completed_time BETWEEN {start_ts} AND {end_ts})
              OR
              (tko.order_state != 40 AND tko.accepted_time BETWEEN {start_ts} AND {end_ts})
          )
    ),
    unioned AS (
        SELECT * FROM dine_main
        UNION ALL SELECT * FROM dine_subitem
        UNION ALL SELECT * FROM takeout
    )
    SELECT
        sale_date,
        row_type,
        COALESCE(CAST(product_uuid AS STRING), '') AS product_uuid,
        MAX(name_zh) AS name_zh,
        MAX(name_en) AS name_en,
        MAX(name_th) AS name_th,
        MAX(cat_zh) AS cat_zh,
        MAX(cat_en) AS cat_en,
        MAX(cat_th) AS cat_th,
        ROUND(SUM(qty), 4) AS total_qty,
        ROUND(SUM(amount), 2) AS total_amount
    FROM unioned
    WHERE qty > 0
      AND (name_zh IS NOT NULL OR name_en IS NOT NULL OR name_th IS NOT NULL)
    GROUP BY sale_date, row_type, product_uuid
    ORDER BY sale_date, row_type, product_uuid
    """
    try:
        result = client.query(query).result()
        return [
            (r.sale_date, r.row_type, r.product_uuid,
             r.name_zh, r.name_en, r.name_th,
             r.cat_zh, r.cat_en, r.cat_th, r.total_qty, r.total_amount)
            for r in result
        ]
    except Exception as e:
        log(f"  {dataset_id} 查询失败: {e}")
        return []


def get_store_name(client, dataset_id):
    """门店名在 ttpos_setting key='store' 的 values JSON 里，字段是 name（不是 store_name）。"""
    try:
        r = list(client.query(f"""
            SELECT JSON_EXTRACT_SCALAR(values, '$.store_code') AS code,
                   JSON_EXTRACT_SCALAR(values, '$.name') AS name
            FROM `{PROJECT_ID}.{dataset_id}.ttpos_setting`
            WHERE `key` = 'store' AND delete_time = 0
            LIMIT 1
        """).result())[0]
        return (r.code or "").strip(), (r.name or "").strip()
    except Exception:
        return "", ""


def process_one(args):
    idx, total, store_no_cfg, uuid, start_ts, end_ts = args
    dataset_id = f"shop{uuid}"
    try:
        client = bigquery.Client(project=PROJECT_ID, credentials=get_creds())
    except Exception as e:
        return {"error": str(e), "idx": idx, "store_no": store_no_cfg}

    # 门店编号沿用 STORE_LIST 配置（1~61），门店名称取自 BQ
    store_code, store_name = get_store_name(client, dataset_id)
    store_no = store_no_cfg
    if not store_name:
        store_name = store_code or f"门店_{uuid[:12]}"

    t0 = time.time()
    rows = query_shop(client, dataset_id, start_ts, end_ts)
    elapsed = time.time() - t0
    return {
        "idx": idx, "store_no": store_no, "store_name": store_name,
        "rows": rows, "count": len(rows), "elapsed": elapsed,
    }


def pick_name(name_zh, name_en, name_th, prefer):
    """prefer='zh' or 'en'，按优先级回退。"""
    if prefer == "zh":
        chain = (name_zh, name_en, name_th)
    else:
        chain = (name_en, name_zh, name_th)
    for n in chain:
        if n and n != "null":
            return n
    return ""


def write_sheet(ws, lang, rows):
    """
    rows: (store_no, store_name, sale_date, row_type, category, product_name, qty, amount)
    lang: 'zh' or 'en'
    """
    if lang == "zh":
        headers = ["门店编号", "门店名称", "日期", "类型", "分类", "商品名称", "销量"]
        type_map = {"单品": "单品", "套餐": "套餐"}
        has_amount = False
    else:
        headers = ["Store Code", "Store Name", "Date", "Type", "Category", "Product Name", "Qty", "Unit Price", "Total Price"]
        type_map = {"单品": "Single Item", "套餐": "Combo"}
        has_amount = True

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin
        c.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        if has_amount:
            store_no, store_name, sale_date, row_type, category, product_name, qty, amount = row
        else:
            store_no, store_name, sale_date, row_type, category, product_name, qty = row
            amount = None

        ws.cell(row=ri, column=1, value=store_no).border = thin
        ws.cell(row=ri, column=2, value=store_name).border = thin
        ws.cell(row=ri, column=3, value=str(sale_date)).border = thin
        ws.cell(row=ri, column=4, value=type_map.get(row_type, row_type)).border = thin
        ws.cell(row=ri, column=5, value=category).border = thin
        ws.cell(row=ri, column=6, value=product_name).border = thin
        c = ws.cell(row=ri, column=7, value=float(qty))
        c.border = thin
        c.number_format = "#,##0.####"

        if has_amount and amount is not None:
            unit_price = round(float(amount) / float(qty), 2) if float(qty) != 0 else 0
            c = ws.cell(row=ri, column=8, value=float(unit_price))
            c.border = thin
            c.number_format = "#,##0.00"
            c = ws.cell(row=ri, column=9, value=float(amount))
            c.border = thin
            c.number_format = "#,##0.00"

    if has_amount:
        widths = [12, 28, 12, 10, 18, 45, 12, 12, 12]
    else:
        widths = [12, 28, 12, 10, 18, 45, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def aggregate_per_lang(all_data, prefer):
    """
    all_data 行：(store_no, store_name, sale_date, row_type, product_uuid,
                  name_zh, name_en, name_th, cat_zh, cat_en, cat_th, qty, amount)
    按 (store_no, store_name, sale_date, row_type, group_key) 聚合。
    输出：
      zh -> (store_no, store_name, sale_date, row_type, category, chosen_name, qty)
      en -> (store_no, store_name, sale_date, row_type, category, chosen_name, qty, amount)
    """
    bucket = {}
    amount_bucket = {}
    name_winner = {}
    cat_winner = {}
    for (store_no, store_name, sale_date, row_type,
         product_uuid, name_zh, name_en, name_th,
         cat_zh, cat_en, cat_th, qty, amount) in all_data:
        chosen = pick_name(name_zh, name_en, name_th, prefer)
        cat = pick_name(cat_zh, cat_en, cat_th, prefer)
        gkey = product_uuid or f"name::{chosen}"
        key = (store_no, store_name, sale_date, row_type, gkey)
        bucket[key] = bucket.get(key, 0.0) + float(qty or 0)
        amount_bucket[key] = amount_bucket.get(key, 0.0) + float(amount or 0)
        if key not in name_winner or not name_winner[key]:
            name_winner[key] = chosen
        if key not in cat_winner or not cat_winner[key]:
            cat_winner[key] = cat

    rows = []
    for key, qty in bucket.items():
        if qty <= 0:
            continue
        store_no, store_name, sale_date, row_type, gkey = key
        if prefer == "en":
            rows.append((
                store_no, store_name, sale_date, row_type,
                cat_winner.get(key, ""),
                name_winner.get(key, ""),
                round(qty, 4),
                round(amount_bucket.get(key, 0.0), 2),
            ))
        else:
            rows.append((
                store_no, store_name, sale_date, row_type,
                cat_winner.get(key, ""),
                name_winner.get(key, ""),
                round(qty, 4),
            ))

    def sort_key(r):
        sn = r[0]
        try:
            sn_n = (0, int(sn))
        except Exception:
            sn_n = (1, sn)
        type_order = 0 if r[3] == "套餐" else 1
        return (sn_n, str(r[2]), type_order, r[4], r[5])
    rows.sort(key=sort_key)
    return rows


def main():
    t0 = time.time()
    log("=" * 60)
    log("单品+套餐明细 销售统计 (BigQuery 版本)")
    log("=" * 60)

    setup_proxy()
    sd, ed, start_ts, end_ts, workers, version = parse_args()
    log(f"统计区间: {sd} ~ {ed}")
    log(f"时间戳:   {start_ts} ~ {end_ts}")
    log(f"门店数:   {len(STORE_LIST)}, 并发: {workers}\n")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    suffix_dates = f"{sd.strftime('%Y%m%d')}-{ed.strftime('%Y%m%d')}"
    file_prefix = f"单品套餐销售统计_{suffix_dates}"
    if version is None:
        version = next_version(OUTPUT_DIR, file_prefix)
    log(f"输出版本: v{version}\n")

    client = bigquery.Client(project=PROJECT_ID, credentials=get_creds())
    all_datasets = {
        d.dataset_id for d in client.list_datasets()
        if d.dataset_id.startswith("shop")
    }
    targets = [
        (i + 1, no, uuid)
        for i, (no, uuid) in enumerate(STORE_LIST)
        if f"shop{uuid}" in all_datasets
    ]
    missing = [no for (no, uuid) in STORE_LIST if f"shop{uuid}" not in all_datasets]
    if missing:
        log(f"⚠ 跳过未发现 dataset 的门店: {missing}")
    log(f"待查询门店: {len(targets)}\n")

    all_data = []
    total_rows = 0
    query_time = 0.0
    query_start = time.time()
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {
            ex.submit(process_one, (idx, len(targets), no, uuid, start_ts, end_ts)): idx
            for idx, no, uuid in targets
        }
        for fut in as_completed(futs):
            r = fut.result()
            if "error" in r:
                log(f"[{r['idx']}] 错误: {r['error']}")
                continue
            log(f"[{r['idx']}/{len(targets)}] #{r['store_no']} {r['store_name']}: "
                f"{r['count']} 行 ({r['elapsed']:.2f}s)")
            query_time += r["elapsed"]
            for row in r["rows"]:
                all_data.append((r["store_no"], r["store_name"]) + row)
            total_rows += r["count"]

    query_elapsed = time.time() - query_start
    log(f"\n查询完成: 原始行 {total_rows:,} (累计单店耗时 {query_time:.1f}s, 墙钟 {query_elapsed:.1f}s)")

    rows_zh = aggregate_per_lang(all_data, "zh")
    rows_en = aggregate_per_lang(all_data, "en")
    log(f"中文聚合后: {len(rows_zh):,} 行  /  English: {len(rows_en):,} 行")

    # CSV (中文版)
    csv_path = OUTPUT_DIR / f"{file_prefix}_v{version}.csv"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["门店编号", "门店名称", "日期", "类型", "分类", "商品名称", "销量"])
        for row in rows_zh:
            w.writerow([row[0], row[1], str(row[2]), row[3], row[4], row[5], f"{row[6]:.4f}"])
    log(f"CSV: {csv_path}")

    # Excel：双 Sheet
    xlsx_path = OUTPUT_DIR / f"{file_prefix}_v{version}.xlsx"
    wb = Workbook()
    ws_zh = wb.active
    ws_zh.title = "中文版"
    write_sheet(ws_zh, "zh", rows_zh)
    ws_en = wb.create_sheet("English")
    write_sheet(ws_en, "en", rows_en)
    wb.save(xlsx_path)
    log(f"Excel: {xlsx_path}")

    log(f"\n{'='*60}")
    log(f"完成! 总耗时 {time.time() - t0:.1f} 秒")
    log(f"{'='*60}")


if __name__ == "__main__":
    main()
