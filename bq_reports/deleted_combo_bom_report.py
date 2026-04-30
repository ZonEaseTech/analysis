#!/usr/bin/env python3
"""
已删除套餐 BOM 导出 —— 2025-09 至 2026-03 期间删除的套餐组成明细

输出：单 sheet Excel（格式同 bom_export_report.py 的「套餐组成」）
  列：门店名称 | 套餐名称 | 分组 | 类型 | 单品名称 | 数量 | 是否必选 | 加价 | 删除日期

范围：
  - 商户真实名称取自 ttpos_company.name
  - 只包含 product_type = 1 且 delete_time 在指定范围内的套餐
  - 关联表(package_group, group_item) 不过滤 delete_time（已删除套餐的分组记录也是软删状态）

Usage:
    venv/bin/python -m bq_reports.deleted_combo_bom_report --output exports/deleted_combo_bom.xlsx
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bq_reports.utils.bq_client import get_bq_client, setup_proxy

PROJECT_ID = "diyl-407103"
DEFAULT_DATASET = "shop1958987436032000"

# 默认范围：2025-09-01 至 2026-04-01（不含4月）
DEFAULT_START_TS = 1756684800   # 2025-09-01 00:00:00 UTC
DEFAULT_END_TS = 1775001600     # 2026-04-01 00:00:00 UTC

REPORT_VERSION = "v1"

STORE_NAME_SQL = """
SELECT
  COALESCE(
    JSON_EXTRACT_SCALAR(name, '$.zh'),
    JSON_EXTRACT_SCALAR(name, '$.en'),
    name
  ) AS company_name
FROM `{project}`.`{dataset}`.`ttpos_company`
WHERE delete_time = 0
LIMIT 1
"""

# 已删除套餐的分组明细
# 不过滤 delete_time（软删记录仍保留在BQ中）
DELETED_COMBO_SQL = """
SELECT
  pp.uuid AS combo_uuid,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pp.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS combo_name,
  pp.delete_time AS combo_delete_time,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(pgrp.name, '$.zh'),
    JSON_EXTRACT_SCALAR(pgrp.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS group_name,
  pgrp.group_type AS group_type,
  REGEXP_REPLACE(COALESCE(
    JSON_EXTRACT_SCALAR(child.name, '$.zh'),
    JSON_EXTRACT_SCALAR(child.name, '$.en'),
    ''
  ), r'^\\s+|\\s+$', '') AS child_name,
  gpi.num AS child_num,
  gpi.is_required,
  gpi.add_price
FROM `{project}`.`{dataset}`.`ttpos_product_package` pp
JOIN `{project}`.`{dataset}`.`ttpos_product_package_group` pgrp
  ON pgrp.product_package_uuid = pp.uuid
JOIN `{project}`.`{dataset}`.`ttpos_product_package_group_item` gpi
  ON gpi.product_package_group_uuid = pgrp.uuid
JOIN `{project}`.`{dataset}`.`ttpos_product_package` child
  ON child.uuid = gpi.related_uuid
WHERE pp.product_type = 1
  AND pp.delete_time >= {start_ts}
  AND pp.delete_time < {end_ts}
ORDER BY pp.delete_time DESC, pp.uuid, pgrp.uuid, gpi.sort
"""


def get_store_name(client, dataset: str) -> str:
    sql = STORE_NAME_SQL.format(project=client.project, dataset=dataset)
    for row in client.query(sql).result():
        name = row.company_name or ""
        return name.strip() if name else ""
    return ""


def query_deleted_combo_rows(client, dataset: str, start_ts: int, end_ts: int):
    sql = DELETED_COMBO_SQL.format(
        project=client.project,
        dataset=dataset,
        start_ts=start_ts,
        end_ts=end_ts
    )
    return list(client.query(sql).result())


def _style_header(ws, headers):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, size=11, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align


def _merge_runs(ws, col_letter: str, start_row: int, end_row: int, value_of):
    if end_row <= start_row:
        return
    run_start = start_row
    run_value = value_of(start_row)
    for r in range(start_row + 1, end_row + 1):
        v = value_of(r)
        if v != run_value:
            if r - 1 > run_start:
                ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{r - 1}")
            run_start = r
            run_value = v
    if end_row > run_start:
        ws.merge_cells(f"{col_letter}{run_start}:{col_letter}{end_row}")


def _write_sheet(ws, rows, store_name: str):
    headers = ["门店名称", "套餐名称", "删除日期", "分组", "类型", "单品名称", "数量", "是否必选", "加价"]
    _style_header(ws, headers)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=store_name).alignment = center
        ws.cell(row=r_idx, column=2, value=row.combo_name or "").alignment = center

        delete_dt = ""
        if row.combo_delete_time:
            delete_dt = datetime.fromtimestamp(
                row.combo_delete_time, tz=timezone.utc
            ).strftime("%Y-%m-%d")
        ws.cell(row=r_idx, column=3, value=delete_dt).alignment = center

        ws.cell(row=r_idx, column=4, value=row.group_name or "").alignment = center
        ws.cell(row=r_idx, column=5, value="固定" if row.group_type == 0 else "可选").alignment = center
        ws.cell(row=r_idx, column=6, value=row.child_name or "").alignment = center
        ws.cell(row=r_idx, column=7, value=float(row.child_num) if row.child_num is not None else 0).alignment = center
        ws.cell(row=r_idx, column=8, value="是" if row.is_required else "否").alignment = center
        ws.cell(row=r_idx, column=9, value=float(row.add_price) if row.add_price else 0).alignment = center

    if rows:
        last_row = 1 + len(rows)
        if last_row > 2:
            # A 列(门店名称)整列合并
            ws.merge_cells(f"A2:A{last_row}")
            # B 列(套餐名称+删除日期)相邻同名合并
            _merge_runs(ws, "B", 2, last_row, lambda r: rows[r - 2].combo_name)
            _merge_runs(ws, "C", 2, last_row,
                        lambda r: f"{rows[r - 2].combo_name}|{rows[r - 2].combo_delete_time}")
            # D 列(分组)相邻同组合并
            _merge_runs(ws, "D", 2, last_row,
                        lambda r: f"{rows[r - 2].combo_name}|{rows[r - 2].group_name}")
            # E 列(类型)相邻同类型合并
            _merge_runs(ws, "E", 2, last_row,
                        lambda r: f"{rows[r - 2].combo_name}|{rows[r - 2].group_name}|{'固定' if rows[r - 2].group_type == 0 else '可选'}")

    for col_idx, w in enumerate([14, 25, 12, 12, 10, 25, 8, 10, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


def write_excel(rows, store_name: str, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "已删除套餐组成"
    _write_sheet(ws, rows, store_name)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _inject_version(path: str) -> str:
    p = Path(path)
    suffix = f"_{REPORT_VERSION}"
    if p.stem.endswith(suffix):
        return str(p)
    return str(p.with_name(f"{p.stem}{suffix}{p.suffix}"))


def parse_args():
    p = argparse.ArgumentParser(description="已删除套餐 BOM 导出")
    p.add_argument("--project", default=PROJECT_ID, help="GCP 项目 ID")
    p.add_argument("--dataset", default=DEFAULT_DATASET,
                   help=f"BQ dataset(任意一家分店即可,默认 {DEFAULT_DATASET})")
    p.add_argument("--start-ts", type=int, default=DEFAULT_START_TS,
                   help=f"删除时间起始 Unix 时间戳(默认 {DEFAULT_START_TS}=2025-09-01)")
    p.add_argument("--end-ts", type=int, default=DEFAULT_END_TS,
                   help=f"删除时间截止 Unix 时间戳(默认 {DEFAULT_END_TS}=2026-04-01)")
    p.add_argument("--output", required=True,
                   help=f"输出 Excel 文件路径(自动追加版本号 _{REPORT_VERSION})")
    return p.parse_args()


def main():
    args = parse_args()
    setup_proxy()

    output_path = _inject_version(args.output)

    start_dt = datetime.fromtimestamp(args.start_ts, tz=timezone.utc).strftime("%Y-%m-%d")
    end_dt = datetime.fromtimestamp(args.end_ts, tz=timezone.utc).strftime("%Y-%m-%d")

    print(f"[BQ] project={args.project} dataset={args.dataset}")
    print(f"[范围] 删除时间: {start_dt} 至 {end_dt}")
    print(f"[版本] {REPORT_VERSION} → 输出路径: {output_path}")
    client = get_bq_client(project_id=args.project)

    store_name = get_store_name(client, args.dataset)
    if not store_name:
        store_name = ""
    print(f"[门店] {store_name}")

    print("[BQ] 查询已删除套餐组成...")
    rows = query_deleted_combo_rows(client, args.dataset, args.start_ts, args.end_ts)
    print(f"[BQ] 共 {len(rows)} 行")

    # 统计不同套餐数
    unique_combos = set(r.combo_uuid for r in rows)
    print(f"[BQ] 涉及 {len(unique_combos)} 个不同的已删除套餐")

    write_excel(rows, store_name, output_path)

    print(f"[完成] 输出 {output_path}")
    print(f"  Sheet「已删除套餐组成」: {len(rows)} 行, {len(unique_combos)} 个套餐")
    return 0


if __name__ == "__main__":
    sys.exit(main())
